from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi import BackgroundTasks
from contextlib import asynccontextmanager
import asyncio
import os
import pg_adapter
import numpy as np
import cv2
import ipaddress
try:
    import db
except ModuleNotFoundError as _db_import_error:
    if _db_import_error.name == "asyncpg":
        class _DbFallback:
            """Fallback when optional asyncpg layer is unavailable."""

            async def init_pool(self, min_size=5, max_size=20):
                print("asyncpg not installed; skipping async DB pool init and using pg_adapter.")

            async def close_pool(self):
                print("asyncpg not installed; skipping async DB pool shutdown.")

        db = _DbFallback()
    else:
        raise

try:
    import torch

    torch_available = True
    print("PyTorch is available")
    # Try to add torch/lib DLLs to path for ONNX Runtime GPU support on Windows
    if hasattr(os, "add_dll_directory"):
        torch_lib = os.path.join(os.path.dirname(torch.__file__), "lib")
        if os.path.exists(torch_lib):
            try:
                os.add_dll_directory(torch_lib)
                print(f"Added Torch DLL directory to path: {torch_lib}")
            except Exception as e:
                print(f"Note: Could not add Torch DLL directory to path: {e}")
except ImportError:
    torch_available = False
    print("PyTorch is not available - running in CPU-only mode")
from insightface.app import FaceAnalysis  # type: ignore
from datetime import datetime, timedelta
import hashlib
import secrets
from collections import deque
import threading
import urllib.parse
from concurrent.futures import ThreadPoolExecutor
import functools

# Async concurrency and performance modules
import db_async
from concurrency import acquire_user_lock
from rate_limiter import attendance_rate_limiter, face_register_rate_limiter

# Thread pool for offloading CPU-bound operations (face recognition, etc.)
_cpu_executor = ThreadPoolExecutor(max_workers=20, thread_name_prefix="cpu_worker")

# -------------------------------------------------
# CONFIGURATION - Confidence Thresholds & Security
# -------------------------------------------------
# STRICT thresholds for accurate face recognition
# These settings prioritize accurate identity verification
CONFIDENCE_THRESHOLD = 0.60  # Higher threshold for better accuracy
# Minimum cosine similarity (primary metric)
MIN_COSINE_SIMILARITY = 0.50  # Stricter - face must be more similar
# Maximum acceptable euclidean distance
MAX_EUCLIDEAN_DISTANCE = 1.2  # Lower - face must  be closer match
# Maximum consecutive failed attempts before lockout
MAX_FAILED_ATTEMPTS = 20
# Lockout duration in minutes
LOCKOUT_DURATION_MINUTES = 5
# Liveness detection parameters
LIVENESS_FRAMES = 3
# Minimum number of frames needed
MIN_FRAMES_FOR_LIVENESS = 3
# Photo detection thresholds - MORE LENIENT for real faces
PHOTO_DETECTION_THRESHOLD = 0.05  # Lower = more strict
# Texture analysis threshold for photo detection
TEXTURE_THRESHOLD = 0.03
# Required movement between frames for liveness
MIN_MOVEMENT = 0.002
MAX_MOVEMENT = 0.8
# Anti-spoofing enabled by default
ANTISPOOFING_ENABLED = False
# If True, blocks suspected photos. If False, only logs warnings
ANTISPOOF_STRICT_MODE = False
# Face profile adaptation/training configuration
INSIGHTFACE_BASE_THRESHOLD = 0.68
# Accept near-boundary live matches to reduce false rejects from minor
# capture variance (motion blur/lighting) while keeping a strict main threshold.
INSIGHTFACE_NEAR_MATCH_MARGIN = 0.060
FALLBACK_BASE_THRESHOLD = 0.95
MAX_PROFILE_SAMPLES = 24
PROFILE_SAMPLE_MIN_CONFIDENCE = 0.78
DAILY_TRAIN_HOUR = 2
DAILY_TRAIN_MINUTE = 30

# Lockout storage (in-memory for demo, use Redis in production)
_lockout_storage = {}
_failed_attempts = {}
_lockout_timers = {}
_lockout_timers_lock = threading.Lock()

# Common VPN/Hosting ASNs and IP ranges
KNOWN_VPN_ASNS = {
    "AS14061",  # DigitalOcean
    "AS16509",  # Amazon AWS
    "AS396982",  # Google Cloud
    "AS8075",  # Microsoft Azure
    "AS20473",  # Vultr
    "AS63949",  # Linode
    "AS203020",  # NordVPN
    "AS212238",  # ExpressVPN
    "AS13335",  # Cloudflare
}


# Enhanced VPN/Threat Detection System
class ThreatIntelligenceManager:
    """Manages dynamic IP range updates and threat intelligence."""

    def __init__(self):
        self.ip_ranges = self._load_static_ranges()
        self.threat_cache = LRUCache(max_size=10000, ttl_seconds=3600)  # 1 hour cache
        self.last_update = 0
        self.update_interval = 24 * 3600  # 24 hours

    def _load_static_ranges(self):
        """Load static IP ranges as fallback."""
        return [
            # Cloud providers
            ipaddress.ip_network("3.0.0.0/8"),      # AWS
            ipaddress.ip_network("13.0.0.0/8"),     # AWS
            ipaddress.ip_network("34.0.0.0/8"),     # GCP
            ipaddress.ip_network("35.0.0.0/8"),     # GCP
            ipaddress.ip_network("52.0.0.0/8"),     # AWS
            ipaddress.ip_network("104.16.0.0/12"),  # Cloudflare
            ipaddress.ip_network("172.64.0.0/13"),  # Cloudflare

            # Known VPN providers (static ranges)
            ipaddress.ip_network("185.199.108.0/22"),  # GitHub Pages/VPNs
            ipaddress.ip_network("103.192.152.0/22"),  # NordVPN
            ipaddress.ip_network("185.220.101.0/24"),  # Tor exit nodes
            ipaddress.ip_network("209.222.18.0/24"),   # PIA VPN
            ipaddress.ip_network("198.143.200.0/23"),  # ExpressVPN
            ipaddress.ip_network("192.71.192.0/22"),   # Mullvad
            ipaddress.ip_network("193.32.248.0/22"),   # Mullvad
            ipaddress.ip_network("149.102.0.0/16"),    # ProtonVPN
        ]

    async def update_cloud_ranges(self):
        """Update cloud provider IP ranges from official sources."""
        import aiohttp
        import asyncio

        try:
            async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=30)) as session:
                # AWS IP ranges
                aws_url = "https://ip-ranges.amazonaws.com/ip-ranges.json"
                async with session.get(aws_url) as response:
                    if response.status == 200:
                        data = await response.json()
                        aws_ranges = []
                        for prefix in data.get("prefixes", []):
                            try:
                                aws_ranges.append(ipaddress.ip_network(prefix["ip_prefix"]))
                            except:
                                pass
                        # Update AWS ranges (replace old ones)
                        self.ip_ranges = [r for r in self.ip_ranges if not str(r).startswith(('3.', '13.', '52.'))]
                        self.ip_ranges.extend(aws_ranges[:50])  # Limit to prevent bloat
                        print(f"Updated AWS IP ranges: {len(aws_ranges)} ranges")

                # Cloudflare IP ranges
                cf_url = "https://www.cloudflare.com/ips-v4/"
                async with session.get(cf_url) as response:
                    if response.status == 200:
                        text = await response.text()
                        cf_ranges = []
                        for line in text.strip().split('\n'):
                            try:
                                cf_ranges.append(ipaddress.ip_network(line.strip()))
                            except:
                                pass
                        # Update Cloudflare ranges
                        self.ip_ranges = [r for r in self.ip_ranges if not str(r).startswith(('104.16.', '172.64.'))]
                        self.ip_ranges.extend(cf_ranges)
                        print(f"Updated Cloudflare IP ranges: {len(cf_ranges)} ranges")

            self.last_update = time.time()
            print(f"Threat intelligence updated at {time.ctime(self.last_update)}")

        except Exception as e:
            print(f"Failed to update threat intelligence: {e}")

    async def check_external_threat_api(self, ip_str: str) -> dict:
        """Check IP against external threat intelligence APIs."""
        # This is a placeholder for actual API integration
        # In production, integrate with services like:
        # - AbuseIPDB (https://www.abuseipdb.com/)
        # - IPQualityScore (https://www.ipqualityscore.com/)
        # - IPInfo (https://ipinfo.io/)

        # For now, return a basic assessment
        try:
            ip = ipaddress.ip_address(ip_str)

            # Check for known problematic ranges
            suspicious_ranges = [
                ipaddress.ip_network("185.220.101.0/24"),  # Tor
                ipaddress.ip_network("192.168.0.0/16"),     # Private networks (shouldn't be seen)
                ipaddress.ip_network("10.0.0.0/8"),         # Private networks
                ipaddress.ip_network("172.16.0.0/12"),      # Private networks
            ]

            for net in suspicious_ranges:
                if ip in net:
                    return {
                        "is_threat": True,
                        "threat_type": "suspicious_range",
                        "confidence": 0.8
                    }

            return {
                "is_threat": False,
                "threat_type": None,
                "confidence": 0.1
            }

        except:
            return {
                "is_threat": False,
                "threat_type": "invalid_ip",
                "confidence": 0.0
            }

    def is_vpn_or_hosting_ip(self, ip_str: str) -> tuple[bool, str]:
        """Enhanced VPN/hosting IP detection with threat intelligence."""
        try:
            ip = ipaddress.ip_address(ip_str)

            # Check static ranges first
            for net in self.ip_ranges:
                if ip in net:
                    return True, f"IP in known cloud/VPN range: {net}"

            # Check cached threat assessment
            cache_key = f"threat_{ip_str}"
            cached_result = self.threat_cache.get(cache_key)
            if cached_result:
                return cached_result["is_threat"], cached_result["reason"]

            return False, "IP not in known threat ranges"

        except Exception as e:
            return False, f"IP validation error: {e}"

# Global threat intelligence manager (initialized after LRUCache class)
_threat_manager = None

def is_vpn_ip(ip_str: str) -> bool:
    """Enhanced VPN/hosting IP detection with threat intelligence."""
    try:
        ip = ipaddress.ip_address(ip_str)
        if (
            ip.is_private
            or ip.is_loopback
            or ip.is_link_local
            or ip.is_reserved
            or ip.is_multicast
        ):
            return False
    except Exception:
        return False

    is_vpn, reason = _threat_manager.is_vpn_or_hosting_ip(ip_str)
    if is_vpn:
        print(f"🚫 VPN/Hosting detected: {ip_str} - {reason}")
    return is_vpn


def get_client_ip(request: Request) -> str:
    """Resolve real client IP from common reverse-proxy/CDN headers."""
    candidates = []

    cf_ip = request.headers.get("CF-Connecting-IP")
    if cf_ip:
        candidates.append(cf_ip.strip())

    real_ip = request.headers.get("X-Real-IP")
    if real_ip:
        candidates.append(real_ip.strip())

    xff = request.headers.get("X-Forwarded-For")
    if xff:
        candidates.extend([p.strip() for p in xff.split(",") if p.strip()])

    if request.client and request.client.host:
        candidates.append(request.client.host.strip())

    for ip_str in candidates:
        try:
            ipaddress.ip_address(ip_str)
            return ip_str
        except Exception:
            continue

    return "0.0.0.0"


# -------------------------------------------------
# ENHANCED CACHE MANAGEMENT WITH MEMORY LIMITS
# -------------------------------------------------
import json
import time
from collections import OrderedDict
from functools import lru_cache

class LRUCache:
    """Thread-safe LRU cache with memory limits and optional persistence."""

    def __init__(self, max_size: int = 1000, ttl_seconds: int = None, persist_file: str = None):
        self.max_size = max_size
        self.ttl_seconds = ttl_seconds
        self.persist_file = persist_file
        self.cache = OrderedDict()
        self.access_times = {}
        self.lock = threading.Lock()
        self.hits = 0
        self.misses = 0
        self.evictions = 0

        # Load persisted data if file exists
        if persist_file and os.path.exists(persist_file):
            try:
                with open(persist_file, 'r') as f:
                    data = json.load(f)
                    for key, value in data.items():
                        if not self._is_expired(key):
                            self.cache[key] = value
                            self.access_times[key] = time.time()
                print(f"Loaded {len(self.cache)} items from {persist_file}")
            except Exception as e:
                print(f"Failed to load cache from {persist_file}: {e}")
                # Rename the corrupted file so we start fresh and don't fail next time
                try:
                    corrupted_file = persist_file + ".corrupted"
                    if os.path.exists(corrupted_file):
                        try:
                            os.remove(corrupted_file)
                        except:
                            pass
                    os.rename(persist_file, corrupted_file)
                    print(f"Renamed corrupted cache file to {corrupted_file} and starting fresh.")
                except Exception as backup_err:
                    print(f"Failed to handle corrupted cache file: {backup_err}")

    def _is_expired(self, key: str) -> bool:
        """Check if cache entry is expired."""
        if self.ttl_seconds is None:
            return False
        access_time = self.access_times.get(key, 0)
        return (time.time() - access_time) > self.ttl_seconds

    def get(self, key: str, default=None):
        """Get item from cache with optional default."""
        with self.lock:
            if key in self.cache:
                if self._is_expired(key):
                    del self.cache[key]
                    del self.access_times[key]
                    self.misses += 1
                    return default
                # Move to end (most recently used)
                self.cache.move_to_end(key)
                self.access_times[key] = time.time()
                self.hits += 1
                return self.cache[key]
            self.misses += 1
            return default

    def put(self, key: str, value):
        """Put item in cache."""
        with self.lock:
            if key in self.cache:
                self.cache.move_to_end(key)
            else:
                if len(self.cache) >= self.max_size:
                    # Remove least recently used
                    oldest_key, _ = self.cache.popitem(last=False)
                    if oldest_key in self.access_times:
                        del self.access_times[oldest_key]
                    self.evictions += 1

            self.cache[key] = value
            self.access_times[key] = time.time()

            # Persist if configured
            if self.persist_file:
                self._persist()

    def remove(self, key: str):
        """Remove item from cache."""
        with self.lock:
            if key in self.cache:
                del self.cache[key]
                if key in self.access_times:
                    del self.access_times[key]

    def clear(self):
        """Clear all cache entries."""
        with self.lock:
            self.cache.clear()
            self.access_times.clear()

    def size(self) -> int:
        """Get current cache size."""
        with self.lock:
            return len(self.cache)

    def stats(self) -> dict:
        """Get cache statistics."""
        with self.lock:
            total_requests = self.hits + self.misses
            hit_rate = (self.hits / total_requests * 100) if total_requests > 0 else 0
            return {
                "size": len(self.cache),
                "max_size": self.max_size,
                "hits": self.hits,
                "misses": self.misses,
                "evictions": self.evictions,
                "hit_rate_percent": round(hit_rate, 2)
            }

    def _persist(self):
        """Persist cache to disk."""
        try:
            # Only persist non-expired items
            data = {}
            for key, value in self.cache.items():
                if not self._is_expired(key):
                    # Convert numpy arrays to lists for JSON serialization
                    if hasattr(value, 'tolist'):
                        data[key] = value.tolist()
                    else:
                        data[key] = value

            with open(self.persist_file, 'w') as f:
                json.dump(data, f, default=str)
        except Exception as e:
            print(f"Failed to persist cache to {self.persist_file}: {e}")

# Initialize enhanced caches
_face_profile_cache = LRUCache(
    max_size=500,  # Limit to 500 users to prevent memory bloat
    ttl_seconds=3600 * 24,  # 24 hours TTL
    persist_file=os.path.join(os.path.dirname(__file__), "face_cache.json")
)

_audit_log = deque(maxlen=5000)  # Reduced from 10000 to save memory
_audit_lock = threading.Lock()

# Lockout and failed attempts with size limits
_lockout_storage = LRUCache(max_size=1000, ttl_seconds=3600)  # Auto-expire after 1 hour
_failed_attempts = LRUCache(max_size=1000, ttl_seconds=3600 * 24)  # 24 hours

# Legacy compatibility - provide dict-like access
class CacheDictAdapter:
    """Adapter to provide dict-like access to LRUCache for backward compatibility."""
    def __init__(self, cache: LRUCache):
        self.cache = cache

    def __getitem__(self, key):
        value = self.cache.get(key)
        if value is None:
            raise KeyError(key)
        return value

    def __setitem__(self, key, value):
        self.cache.put(key, value)

    def __delitem__(self, key):
        self.cache.remove(key)

    def __contains__(self, key):
        return self.cache.get(key) is not None

    def get(self, key, default=None):
        value = self.cache.get(key, None)
        return value if value is not None else default

    def pop(self, key, default=None):
        value = self.cache.get(key)
        if value is not None:
            self.cache.remove(key)
            return value
        return default

# Provide backward compatibility
_face_profile_dict = CacheDictAdapter(_face_profile_cache)

# Cache lock for thread safety
_cache_lock = threading.Lock()

# Initialize threat intelligence manager after all classes are defined
_threat_manager = ThreatIntelligenceManager()

# -------------------------------------------------
# STANDARDIZED API RESPONSE HELPERS
# -------------------------------------------------


def success_response(message: str = "Success", data: any = None):
    """Create a standardized success response"""
    return {
        "success": True,
        "message": message,
        "data": data,
        "timestamp": datetime.now().isoformat(),
    }


def error_response(
    message: str = "Error", error_code: str = "ERROR", status_code: int = 400
):
    """Create a standardized error response"""
    return {
        "success": False,
        "error": message,
        "error_code": error_code,
        "timestamp": datetime.now().isoformat(),
    }


def log_audit_event(
    event_type: str = None,
    reg_no: str = None,
    success: bool = False,
    details: str = None,
):
    """Log verification events for audit purposes"""
    with _audit_lock:
        _audit_log.append(
            {
                "timestamp": datetime.now().isoformat(),
                "event_type": event_type,
                "reg_no": reg_no,
                "success": success,
                "details": details,
                "ip": "N/A",  # Can be enhanced to capture actual IP
            }
        )


def check_lockout(reg_no: str = None) -> tuple[bool, int]:
    """Check if user is locked out. Returns (is_locked, remaining_seconds)"""
    key = reg_no or "global"

    # Check global lockout
    lockout_until = _lockout_storage.get(key)
    if lockout_until is not None:
        remaining = (lockout_until - datetime.now()).total_seconds()
        if remaining > 0:
            return True, int(remaining)
        else:
            # Lockout expired
            _lockout_storage.remove(key)

    return False, 0


def record_failed_attempt(reg_no: str):
    """Record a failed verification attempt and trigger lockout if needed"""
    current_count = _failed_attempts.get(reg_no, 0) + 1
    _failed_attempts.put(reg_no, current_count)

    log_audit_event(
        "FAILED_ATTEMPT",
        reg_no,
        False,
        f"Attempt {current_count}/{MAX_FAILED_ATTEMPTS}",
    )

    if current_count >= MAX_FAILED_ATTEMPTS:
        trigger_lockout(reg_no)


def trigger_lockout(reg_no: str):
    """Trigger lockout for a user"""
    lockout_until = datetime.now() + timedelta(minutes=LOCKOUT_DURATION_MINUTES)
    _lockout_storage.put(reg_no, lockout_until)

    log_audit_event(
        "LOCKOUT", reg_no, False, f"Locked out until {lockout_until.isoformat()}"
    )

    # Schedule automatic unlock
    def remove_lockout():
        _lockout_storage.remove(reg_no)
        _failed_attempts.remove(reg_no)

    with _lockout_timers_lock:
        if reg_no in _lockout_timers:
            _lockout_timers[reg_no].cancel()

        _lockout_timers[reg_no] = threading.Timer(
            LOCKOUT_DURATION_MINUTES * 60, remove_lockout
        )
        _lockout_timers[reg_no].start()


def clear_failed_attempts(reg_no: str):
    """Clear failed attempts after successful verification"""
    _failed_attempts.remove(reg_no)


def get_audit_logs(reg_no: str = None, limit: int = 100):
    """Get audit logs, optionally filtered by reg_no"""
    with _audit_lock:
        logs = list(_audit_log)

    if reg_no:
        logs = [log for log in logs if log["reg_no"] == reg_no]

    return logs[-limit:]


# -------------------------------------------------
# APP LIFESPAN - Startup and shutdown tasks
# -------------------------------------------------
async def _init_database_tables():
    """Startup compatibility hook.
    Tables/indexes are initialized during module load in this codebase.
    """
    return


async def _warm_face_embedding_cache_async():
    """Async compatibility wrapper for synchronous cache warmup."""
    await asyncio.to_thread(_warm_face_embedding_cache)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan context manager for startup/shutdown events."""
    print("🚀 Initializing database connection pool...")
    await db.init_pool(min_size=5, max_size=20)

    # Initialize database tables and data
    print("📊 Setting up database tables...")
    await _init_database_tables()

    # Warm up face embedding cache
    print("🤖 Warming up face recognition cache...")
    await _warm_face_embedding_cache_async()

    # Update threat intelligence
    print("🛡️  Updating threat intelligence...")
    try:
        await _threat_manager.update_cloud_ranges()
    except Exception as e:
        print(f"⚠️  Threat intelligence update failed: {e}")

    # Start background threat intelligence updates
    async def periodic_threat_updates():
        """Periodically update threat intelligence in background."""
        while True:
            try:
                await asyncio.sleep(_threat_manager.update_interval)
                print("🔄 Updating threat intelligence...")
                await _threat_manager.update_cloud_ranges()
            except Exception as e:
                print(f"⚠️  Background threat update failed: {e}")
                await asyncio.sleep(3600)  # Retry in 1 hour

    # Start background materialized view refresh
    async def periodic_mv_refresh():
        """Periodically refresh materialized views for reporting."""
        while True:
            try:
                await asyncio.sleep(3600 * 6)  # Every 6 hours
                print("🔄 Refreshing materialized views...")
                # Note: In production, use a proper async database call
                # For now, this is a placeholder
                print("Materialized view refresh completed")
            except Exception as e:
                print(f"⚠️  Materialized view refresh failed: {e}")
                await asyncio.sleep(3600)  # Retry in 1 hour

    # Start the background tasks
    threat_update_task = asyncio.create_task(periodic_threat_updates())
    mv_refresh_task = asyncio.create_task(periodic_mv_refresh())

    # Check and process leave expiry on startup
    print("⏳ Checking CL/EL expiry dates...")
    try:
        _check_and_expire_leaves(triggered_by="startup")
    except Exception as e:
        print(f"⚠️ Startup leave expiry check failed: {e}")

    # Start periodic leave expiry check background task (runs daily)
    async def periodic_leave_expiry():
        while True:
            try:
                await asyncio.sleep(86400)  # Check every 24 hours
                _check_and_expire_leaves(triggered_by="periodic_job")
            except Exception as e:
                print(f"⚠️ Periodic leave expiry check failed: {e}")
                await asyncio.sleep(3600)

    leave_expiry_task = asyncio.create_task(periodic_leave_expiry())

    print("✅ Application startup complete")

    yield

    print("🛑 Shutting down application...")
    threat_update_task.cancel()
    mv_refresh_task.cancel()
    leave_expiry_task.cancel()
    try:
        await threat_update_task
        await mv_refresh_task
        await leave_expiry_task
    except asyncio.CancelledError:
        pass

    await db.close_pool()
    print("✅ Application shutdown complete")

# -------------------------------------------------
# APP INIT
# -------------------------------------------------
app = FastAPI(title="Automated Face Attendance System", lifespan=lifespan)

# Add CORS middleware for web access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins (you can restrict this in production)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# VPN Detection Middleware
@app.middleware("http")
async def vpn_detection_middleware(request: Request, call_next):
    # Skip VPN check for settings, auth and public endpoints
    skip_paths = [
        "/settings/allow_any_network",
        "/admin/settings",
        "/api/login",
        "/api/auth",
        "/docs",
        "/openapi.json",
        # Allow face registration flows even when VPN blocking is enabled.
        # These endpoints already have strong token/identity checks.
        "/staff/face/register",
        "/other_staff/face/register",
        "/hod/face/register",
        "/admin/face/register",
    ]

    if any(request.url.path.startswith(path) for path in skip_paths):
        return await call_next(request)

    # Only check if VPN blocking is enabled
    if _app_settings.get("enforce_vpn_blocking", True):
        client_ip = get_client_ip(request)

        # Debug: Log client IP
        print(f"[VPN] Client IP: {client_ip}")

        if is_vpn_ip(client_ip):
            print(f"[VPN] Blocked VPN connection from {client_ip}")
            return JSONResponse(
                status_code=403,
                content={
                    "success": False,
                    "error": "VPN detected. Please turn off your VPN connection to continue.",
                    "error_code": "VPN_DETECTED",
                    "timestamp": datetime.now().isoformat(),
                },
            )

    return await call_next(request)


# -------------------------------------------------
# GLOBAL EXCEPTION HANDLERS - Return JSON errors instead of HTML
# -------------------------------------------------
from fastapi.responses import JSONResponse
from fastapi import Request
import traceback


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """Catch all unhandled exceptions and return JSON error response"""
    # Log the full error
    print(f"🚨 UNHANDLED EXCEPTION: {str(exc)}")
    print(traceback.format_exc())

    return JSONResponse(
        status_code=500,
        content={
            "success": False,
            "error": str(exc),
            "error_code": "INTERNAL_ERROR",
            "message": "An internal error occurred. Please try again or contact support.",
            "timestamp": datetime.now().isoformat(),
        },
    )


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Handle HTTP exceptions and return JSON error response"""
    return JSONResponse(
        status_code=exc.status_code,
        content={
            "success": False,
            "error": exc.detail,
            "error_code": f"HTTP_{exc.status_code}",
            "timestamp": datetime.now().isoformat(),
        },
    )


# -------------------------------------------------
# APP SETTINGS STORAGE
# -------------------------------------------------
# In-memory settings storage (use database or Redis in production)
_app_settings = {
    "allow_any_network": True,
    "college_ssid": "",
    "enforce_geo_fence": True,
    "enforce_app_geo_fence": True,
    "enforce_vpn_blocking": True,
    "profile_password": "admin123",
}

_academic_settings = {}
_academic_settings_last_refresh = 0.0
_ACADEMIC_SETTINGS_REFRESH_SECONDS = 5  # Max age before reloading from DB

# Web geofence polygons (lat, lng) - supports multiple outer boundaries
_geo_fence_outer_polygons = [
    [
        (11.040730, 77.073717),
        (11.040865, 77.075121),
        (11.039733, 77.075201),
        (11.039529, 77.075786),
        (11.038500, 77.075892),
        (11.038551, 77.073616),
    ]
]
# Backward compatibility alias (first outer polygon)
_geo_fence_polygon = _geo_fence_outer_polygons[0]

# Movement limit range polygons (lat, lng)
_geo_fence_limit_range_polygons = []


def _point_in_polygon(
    lat: float, lng: float, polygon: list[tuple[float, float]]
) -> bool:
    """Ray casting point-in-polygon. lng=x, lat=y."""
    inside = False
    n = len(polygon)

    for i in range(n):
        j = (i + 1) % n
        yi, xi = polygon[i]
        yj, xj = polygon[j]

        # Check if point's y (lat) is within the edge's y range
        if (yi > lat) != (yj > lat):
            # Calculate x intersection at this y
            dy = yj - yi
            if abs(dy) < 0.0000001:
                continue  # Skip horizontal edges

            t = (lat - yi) / dy
            x_intersect = xi + t * (xj - xi)

            # Check if point's x (lng) is to the left of intersection
            if lng < x_intersect:
                inside = not inside

    return inside


def _point_in_any_polygon(
    lat: float, lng: float, polygons: list[list[tuple[float, float]]]
) -> bool:
    for polygon in polygons:
        if len(polygon) >= 3 and _point_in_polygon(lat, lng, polygon):
            return True
    return False


def get_ip_location(ip: str) -> tuple[float | None, float | None]:
    """Fetch latitude and longitude for a public IP with a 1.5s timeout."""
    import urllib.request
    import json
    try:
        ip_obj = ipaddress.ip_address(ip)
        if ip_obj.is_private or ip_obj.is_loopback:
            return None, None
        
        url = f"http://ip-api.com/json/{ip}?fields=status,lat,lon"
        with urllib.request.urlopen(url, timeout=1.5) as response:
            data = json.loads(response.read().decode('utf-8'))
            if data.get("status") == "success":
                return data.get("lat"), data.get("lon")
    except Exception as e:
        print(f"[IP Geolocation] Failed to fetch location for {ip}: {e}")
    return None, None


_cached_server_public_ip = None
_last_ip_fetch_time = 0

def get_server_public_ip() -> str:
    global _cached_server_public_ip, _last_ip_fetch_time
    import time
    import urllib.request
    now = time.time()
    if _cached_server_public_ip is None or (now - _last_ip_fetch_time > 300):
        try:
            req = urllib.request.Request(
                "https://api.ipify.org",
                headers={'User-Agent': 'Mozilla/5.0'}
            )
            with urllib.request.urlopen(req, timeout=3.0) as response:
                _cached_server_public_ip = response.read().decode('utf-8').strip()
                _last_ip_fetch_time = now
        except Exception as e:
            print(f"[IP Verify] Failed to fetch server public IP: {e}")
    return _cached_server_public_ip

def _enforce_web_geofence(form: dict, client_ip: str) -> bool:
    """Enforce geo-fence rules for web and app clients independently based on settings.

    Geofence enforcement is the ONLY concern of this function.  WiFi / network
    enforcement is handled separately by check_wifi(), which is called earlier
    in each attendance endpoint.

    Platform rules:
      web  -> reads enforce_geo_fence setting
      app  -> reads enforce_app_geo_fence setting

    If the relevant setting is OFF this function is a no-op for that platform.
    When enabled, the client must supply client_lat / client_lng in the form
    data and must be inside the configured outer geofence polygon.
    """
    platform = form.get("client_platform", "app")  # Default to app if not specified

    # Select the correct setting key for the requesting platform
    is_web = (platform == "web")
    enforce_key = "enforce_geo_fence" if is_web else "enforce_app_geo_fence"

    # Skip geofence check if it is disabled for this platform
    if not _app_settings.get(enforce_key, True):
        return False  # geofence not active for this platform

    lat_raw = form.get("client_lat")
    lng_raw = form.get("client_lng")

    if lat_raw is None or lng_raw is None:
        raise HTTPException(
            status_code=400,
            detail="Location is required to mark attendance.",
        )
    try:
        lat = float(lat_raw)
        lng = float(lng_raw)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid location format.")

    if not _point_in_any_polygon(lat, lng, _geo_fence_outer_polygons):
        raise HTTPException(
            status_code=403,
            detail="You are outside the allowed location.",
        )

    return True  # geofence was active and passed


@app.get("/admin/cache/stats")
async def get_cache_stats(request: Request):
    """Get cache statistics - Admin only"""
    verify_admin_token(request)

    return success_response("Cache statistics retrieved", {
        "face_profiles": _face_profile_cache.stats(),
        "lockouts": _lockout_storage.stats(),
        "failed_attempts": _failed_attempts.stats(),
        "audit_logs": {
            "size": len(_audit_log),
            "max_size": 5000
        }
    })


@app.get("/admin/cache/clear")
async def clear_cache(request: Request):
    """Clear all caches - Admin only (emergency use)"""
    verify_admin_token(request)

    _face_profile_cache.clear()
    _lockout_storage.clear()
    _failed_attempts.clear()
    _audit_log.clear()

    log_audit_event(
        "CACHE_CLEARED",
        "admin",
        True,
        "All caches cleared manually"
    )

    return success_response("All caches cleared successfully")


@app.post("/admin/threat-intelligence/update")
async def update_threat_intelligence(request: Request):
    """Manually update threat intelligence from external sources - Admin only"""
    verify_admin_token(request)

    try:
        await _threat_manager.update_cloud_ranges()
        log_audit_event(
            "THREAT_UPDATE",
            "admin",
            True,
            "Threat intelligence updated manually"
        )
        return success_response("Threat intelligence updated successfully")
    except Exception as e:
        log_audit_event(
            "THREAT_UPDATE_FAILED",
            "admin",
            False,
            str(e)
        )
        return error_response(f"Failed to update threat intelligence: {e}")


@app.get("/admin/threat-intelligence/check-ip/{ip}")
async def check_ip_threat(ip: str, request: Request):
    """Check if an IP is considered a threat - Admin only"""
    verify_admin_token(request)

    try:
        is_vpn, reason = _threat_manager.is_vpn_or_hosting_ip(ip)
        threat_info = await _threat_manager.check_external_threat_api(ip)

        return success_response("IP threat check completed", {
            "ip": ip,
            "is_vpn_hosting": is_vpn,
            "vpn_reason": reason,
            "threat_assessment": threat_info
        })
    except Exception as e:
        return error_response(f"IP check failed: {e}")


@app.get("/admin/threat-intelligence/stats")
async def get_threat_stats(request: Request):
    """Get threat intelligence statistics - Admin only"""
    verify_admin_token(request)

    return success_response("Threat intelligence statistics", {
        "ip_ranges_count": len(_threat_manager.ip_ranges),
        "threat_cache_size": _threat_manager.threat_cache.size(),
        "threat_cache_stats": _threat_manager.threat_cache.stats(),
        "last_update": time.ctime(_threat_manager.last_update) if _threat_manager.last_update else "Never",
        "update_interval_hours": _threat_manager.update_interval / 3600
    })


@app.get("/admin/database/stats")
async def get_database_stats(request: Request):
    """Get database performance statistics - Admin only"""
    verify_admin_token(request)

    try:
        # Get basic database statistics
        user_count = cursor.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        attendance_count = cursor.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
        cache_stats = _face_profile_cache.stats()

        # Get query performance if available
        query_stats = {"message": "Query optimizer not available"}

        return success_response("Database statistics retrieved", {
            "user_count": user_count,
            "attendance_records": attendance_count,
            "face_profiles_cached": cache_stats["size"],
            "cache_hit_rate_percent": cache_stats["hit_rate_percent"],
            "query_performance": query_stats,
            "database_indexes": "Comprehensive indexing implemented"
        })
    except Exception as e:
        return error_response(f"Failed to get database stats: {e}")


@app.post("/admin/database/analyze-slow-queries")
async def analyze_slow_queries(request: Request):
    """Analyze potentially slow queries - Admin only"""
    verify_admin_token(request)

    try:
        analysis = {
            "recommendations": [
                "Consider adding more composite indexes for complex queries",
                "Monitor query execution times in production",
                "Use EXPLAIN ANALYZE for query optimization",
                "Consider partitioning large tables by date"
            ],
            "current_indexes": [
                "idx_users_reg_no, idx_users_username, idx_users_role, idx_users_dept",
                "idx_attendance_reg_no_timestamp, idx_attendance_dept_timestamp",
                "idx_daily_attendance_status_reg_no_date",
                "idx_face_embedding_samples_reg_no_created"
            ]
        }

        return success_response("Query analysis completed", analysis)
    except Exception as e:
        return error_response(f"Query analysis failed: {e}")


@app.get("/admin/performance/metrics")
async def get_performance_metrics(request: Request):
    """Get comprehensive performance metrics - Admin only"""
    verify_admin_token(request)

    try:
        import psutil
        import os

        # System metrics
        memory = psutil.virtual_memory()
        cpu_percent = psutil.cpu_percent(interval=1)

        # Application metrics
        cache_stats = _face_profile_cache.stats()

        # Database connection info (simplified)
        db_connections = "Thread-local connections via pg_adapter"

        return success_response("Performance metrics retrieved", {
            "system": {
                "memory_usage_percent": memory.percent,
                "memory_used_mb": memory.used / 1024 / 1024,
                "cpu_usage_percent": cpu_percent
            },
            "application": {
                "cache_hit_rate_percent": cache_stats["hit_rate_percent"],
                "cache_size": cache_stats["size"],
                "cache_max_size": cache_stats["max_size"],
                "cache_evictions": cache_stats["evictions"]
            },
            "database": {
                "connection_pooling": db_connections,
                "indexes_created": "25+ performance indexes",
                "query_caching": "LRU cache with 30min TTL"
            }
        })
    except Exception as e:
        return error_response(f"Failed to get performance metrics: {e}")


@app.get("/admin/settings")
async def get_settings(request: Request):
    """Get all application settings - requires admin authentication"""
    verify_admin_token(request)
    return {"success": True, "settings": _app_settings}


@app.post("/admin/settings")
async def update_settings(request: Request):
    """Update application settings - requires admin authentication"""
    try:
        body = await request.json()

        # Verify admin authentication using existing function
        try:
            admin_user = verify_admin_token(request)
        except HTTPException as e:
            raise e
        except Exception:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Verify profile password
        profile_password = body.get("profile_password")
        stored_password = _app_settings.get("profile_password", "admin123")
        if not profile_password or profile_password != stored_password:
            raise HTTPException(status_code=403, detail="Invalid profile password")

        # Update settings and track changes
        changes = []
        if "allow_any_network" in body:
            new_value = bool(body["allow_any_network"])
            if _app_settings.get("allow_any_network", False) != new_value:
                _app_settings["allow_any_network"] = new_value
                save_system_config("allow_any_network", str(new_value))
                status = "disabled" if new_value else "enabled"
                changes.append(f"Network access restrictions {status}")
        if "college_ssid" in body:
            new_value = str(body["college_ssid"]).strip()
            if _app_settings.get("college_ssid", "") != new_value:
                _app_settings["college_ssid"] = new_value
                save_system_config("college_ssid", new_value)
                changes.append(f"College SSID updated to '{new_value}'")
        if "enforce_geo_fence" in body:
            new_value = bool(body["enforce_geo_fence"])
            if _app_settings.get("enforce_geo_fence", True) != new_value:
                _app_settings["enforce_geo_fence"] = new_value
                save_system_config("enforce_geo_fence", str(new_value))
                status = "enabled" if new_value else "disabled"
                changes.append(f"Geo-fencing {status}")
        if "enforce_app_geo_fence" in body:
            new_value = bool(body["enforce_app_geo_fence"])
            if _app_settings.get("enforce_app_geo_fence", True) != new_value:
                _app_settings["enforce_app_geo_fence"] = new_value
                save_system_config("enforce_app_geo_fence", str(new_value))
                status = "enabled" if new_value else "disabled"
                changes.append(f"App geo-fencing {status}")
        if "enforce_vpn_blocking" in body:
            new_value = bool(body["enforce_vpn_blocking"])
            if _app_settings.get("enforce_vpn_blocking", True) != new_value:
                _app_settings["enforce_vpn_blocking"] = new_value
                save_system_config("enforce_vpn_blocking", str(new_value))
                status = "enabled" if new_value else "disabled"
                changes.append(f"VPN blocking {status}")

        message = ", ".join(changes) if changes else "No settings were changed"

        return {
            "success": True,
            "message": message,
            "settings": _app_settings,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating settings: {e}")
        raise HTTPException(status_code=500, detail="Failed to update settings")


@app.post("/admin/change_profile_password")
async def change_profile_password(request: Request):
    """Change settings profile password - requires admin authentication"""
    try:
        admin_user = verify_admin_token(request)
    except HTTPException as e:
        raise e
    except Exception:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        body = await request.json()
        current_password = body.get("current_password")
        new_password = body.get("new_password")

        if not current_password or not new_password:
            raise HTTPException(status_code=400, detail="Current and new passwords are required")
        if len(new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters long")

        stored_password = _app_settings.get("profile_password", "admin123")
        if current_password != stored_password:
            raise HTTPException(status_code=403, detail="Incorrect current profile password")

        _app_settings["profile_password"] = new_password
        save_system_config("profile_password", new_password)

        return {"success": True, "message": "Profile password updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error changing profile password: {e}")
        raise HTTPException(status_code=500, detail="Failed to change profile password")


@app.get("/api/health")
@app.head("/api/health")
async def health_check():
    """Health check endpoint for VPN / Load balancer / client"""
    return {"status": "healthy"}


@app.get("/api/test-vpn")
async def test_vpn(request: Request):
    """Test endpoint to check if current IP is detected as VPN"""
    client_ip = get_client_ip(request)

    is_vpn = is_vpn_ip(client_ip)
    return {
        "client_ip": client_ip,
        "is_vpn_detected": is_vpn,
        "vpn_blocking_enabled": _app_settings.get("enforce_vpn_blocking", True),
    }


@app.get("/settings/allow_any_network")
async def get_network_setting():
    """Public endpoint to check if any network is allowed (for client-side validation)"""
    return {
        "allow_any_network": _app_settings.get("allow_any_network", False),
        "college_ssid": _app_settings.get("college_ssid", ""),
        "enforce_geo_fence": _app_settings.get("enforce_geo_fence", True),
        "enforce_app_geo_fence": _app_settings.get("enforce_app_geo_fence", True),
        "enforce_vpn_blocking": _app_settings.get("enforce_vpn_blocking", True),
    }


@app.get("/check_vpn")
async def check_vpn(request: Request):
    """Check if client IP is from a known VPN or hosting provider"""
    client_ip = get_client_ip(request)

    # If VPN blocking is disabled, always return false
    if not _app_settings.get("enforce_vpn_blocking", True):
        return {"vpn_detected": False, "client_ip": client_ip}

    # Check if IP is in known VPN ranges
    vpn_detected = is_vpn_ip(client_ip)

    return {
        "vpn_detected": vpn_detected,
        "client_ip": client_ip,
        "message": "VPN detected" if vpn_detected else "No VPN detected",
    }




# -------------------------------------------------
# CASUAL LEAVE (CL) MANAGEMENT ENDPOINTS
# -------------------------------------------------
from datetime import datetime
import calendar


def get_current_month():
    """Get current month in YYYY-MM format"""
    return datetime.now().strftime("%Y-%m")


def get_previous_month():
    """Get previous month in YYYY-MM format"""
    today = datetime.now()
    first_day_of_month = today.replace(day=1)
    last_day_of_prev_month = first_day_of_month - timedelta(days=1)
    return last_day_of_prev_month.strftime("%Y-%m")


def initialize_cl_for_user(reg_no, user_name, dept, role):
    """Initialize CL for a user if not exists for current month"""
    current_month = get_current_month()

    # Check if CL record exists for current month
    cursor.execute(
        """
        SELECT id, current_month_cl_available, accumulated_cl, cl_used_current_month
        FROM casual_leave
        WHERE reg_no = ? AND current_month = ?
    """,
        (reg_no, current_month),
    )

    existing_record = cursor.fetchone()

    if not existing_record:
        # Check previous month for accumulated CL
        previous_month = get_previous_month()
        cursor.execute(
            """
            SELECT accumulated_cl, cl_used_current_month
            FROM casual_leave
            WHERE reg_no = ? AND current_month = ?
        """,
            (reg_no, previous_month),
        )

        prev_record = cursor.fetchone()
        accumulated_cl = 0

        if prev_record:
            prev_accumulated, prev_used = prev_record
            # If previous month CL was not used, carry it over
            if prev_used == 0:
                accumulated_cl = min(
                    prev_accumulated + 1, 2
                )  # Max 2 CL can be accumulated
            else:
                accumulated_cl = 0

        # Insert new CL record for current month
        cursor.execute(
            """
            INSERT INTO casual_leave (reg_no, user_name, dept, role, current_month, 
                                       current_month_cl_available, accumulated_cl, cl_used_current_month)
            VALUES (?, ?, ?, ?, ?, 1, ?, 0)
        """,
            (reg_no, user_name, dept, role, current_month, accumulated_cl),
        )
        conn.commit()

    return current_month


@app.get("/admin/cl/all")
async def get_all_cl(request: Request):
    """Get CL for all staff members - Admin only"""
    verify_admin_token(request)

    current_month = get_current_month()

    # Get CL from users table
    cursor.execute("""
        SELECT reg_no, name, dept, role, 'user' as user_type
        FROM users
        WHERE role IN ('hod', 'staff')
        ORDER BY dept, name
    """)
    users = cursor.fetchall()

    # Get CL from other_staff table
    cursor.execute("""
        SELECT reg_no, name, dept, role, 'other_staff' as user_type
        FROM other_staff
        ORDER BY dept, name
    """)
    other_staff = cursor.fetchall()

    # Combine all staff
    all_staff = users + other_staff

    cl_data = []
    for staff in all_staff:
        reg_no, name, dept, role, user_type = staff

        # Ensure CL is initialized
        initialize_cl_for_user(reg_no, name, dept, role)

        # Get current CL data
        cursor.execute(
            """
            SELECT current_month_cl_available, accumulated_cl, cl_used_current_month, last_updated
            FROM casual_leave
            WHERE reg_no = ? AND current_month = ?
        """,
            (reg_no, current_month),
        )

        cl_record = cursor.fetchone()
        if cl_record:
            cl_available, accumulated, used, last_updated = cl_record
            total_available = cl_available + accumulated
        else:
            cl_available, accumulated, used, total_available = 1, 0, 0, 1

        cl_data.append(
            {
                "reg_no": reg_no,
                "name": name,
                "dept": dept,
                "role": role,
                "user_type": user_type,
                "current_month_cl_available": cl_available,
                "accumulated_cl": accumulated,
                "cl_used_current_month": used,
                "total_cl_available": total_available,
                "last_updated": last_updated,
            }
        )

    return {"success": True, "data": cl_data, "current_month": current_month}


@app.get("/cl/status/{reg_no}")
async def get_cl_status(reg_no: str):
    """Get CL status for a specific user"""
    current_month = get_current_month()

    # Get user info
    cursor.execute("SELECT name, dept, role FROM users WHERE reg_no = ?", (reg_no,))
    user = cursor.fetchone()

    if not user:
        cursor.execute(
            "SELECT name, dept, role FROM other_staff WHERE reg_no = ?", (reg_no,)
        )
        user = cursor.fetchone()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    name, dept, role = user

    # Ensure CL is initialized
    initialize_cl_for_user(reg_no, name, dept, role)

    # Get CL data
    cursor.execute(
        """
        SELECT current_month_cl_available, accumulated_cl, cl_used_current_month
        FROM casual_leave
        WHERE reg_no = ? AND current_month = ?
    """,
        (reg_no, current_month),
    )

    cl_record = cursor.fetchone()
    if cl_record:
        cl_available, accumulated, used = cl_record
        total_available = cl_available + accumulated
    else:
        cl_available, accumulated, used, total_available = 1, 0, 0, 1

    return {
        "success": True,
        "data": {
            "reg_no": reg_no,
            "name": name,
            "dept": dept,
            "role": role,
            "current_month_cl_available": cl_available,
            "accumulated_cl": accumulated,
            "cl_used_current_month": used,
            "total_cl_available": total_available,
            "current_month": current_month,
        },
    }


@app.post("/admin/cl/use")
async def use_cl(request: Request):
    """Use a CL for a user - typically called when leave is approved"""
    try:
        body = await request.json()
        reg_no = body.get("reg_no")
    except:
        raise HTTPException(status_code=400, detail="Invalid request body")

    if not reg_no:
        raise HTTPException(status_code=400, detail="reg_no is required")

    current_month = get_current_month()

    # Get user info
    cursor.execute("SELECT name, dept, role FROM users WHERE reg_no = ?", (reg_no,))
    user = cursor.fetchone()

    if not user:
        cursor.execute(
            "SELECT name, dept, role FROM other_staff WHERE reg_no = ?", (reg_no,)
        )
        user = cursor.fetchone()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    name, dept, role = user

    # Ensure CL is initialized
    initialize_cl_for_user(reg_no, name, dept, role)

    # Get current CL data
    cursor.execute(
        """
        SELECT current_month_cl_available, accumulated_cl, cl_used_current_month
        FROM casual_leave
        WHERE reg_no = ? AND current_month = ?
    """,
        (reg_no, current_month),
    )

    cl_record = cursor.fetchone()
    if not cl_record:
        raise HTTPException(status_code=400, detail="CL not initialized for user")

    cl_available, accumulated, used = cl_record
    total_available = cl_available + accumulated

    # Check if CL is available
    if total_available <= 0:
        raise HTTPException(status_code=400, detail="No CL available")

    # Use CL - first use current month CL, then accumulated
    if cl_available > 0:
        new_cl_available = cl_available - 1
        new_accumulated = accumulated
    else:
        new_cl_available = 0
        new_accumulated = accumulated - 1

    new_used = used + 1

    # Update CL record
    cursor.execute(
        """
        UPDATE casual_leave
        SET current_month_cl_available = ?, accumulated_cl = ?, cl_used_current_month = ?, 
            last_updated = CURRENT_TIMESTAMP
        WHERE reg_no = ? AND current_month = ?
    """,
        (new_cl_available, new_accumulated, new_used, reg_no, current_month),
    )
    conn.commit()

    return {
        "success": True,
        "message": "CL used successfully",
        "data": {
            "reg_no": reg_no,
            "current_month_cl_available": new_cl_available,
            "accumulated_cl": new_accumulated,
            "cl_used_current_month": new_used,
            "total_cl_available": new_cl_available + new_accumulated,
        },
    }


@app.post("/admin/cl/reset")
async def reset_monthly_cl(request: Request):
    """Reset CL for all staff - Admin only"""
    verify_admin_token(request)

    try:
        body = await request.json()
        target_month = body.get("month", get_current_month())
    except:
        target_month = get_current_month()

    # Get all users
    cursor.execute(
        "SELECT reg_no, name, dept, role FROM users WHERE role IN ('hod', 'staff')"
    )
    users = cursor.fetchall()

    # Get all other_staff
    cursor.execute("SELECT reg_no, name, dept, role FROM other_staff")
    other_staff = cursor.fetchall()

    all_staff = users + other_staff

    reset_count = 0
    for staff in all_staff:
        reg_no, name, dept, role = staff

        # Calculate accumulated CL from previous month
        previous_month = get_previous_month()
        cursor.execute(
            """
            SELECT accumulated_cl, cl_used_current_month
            FROM casual_leave
            WHERE reg_no = ? AND current_month = ?
        """,
            (reg_no, previous_month),
        )

        prev_record = cursor.fetchone()
        accumulated_cl = 0

        if prev_record:
            prev_accumulated, prev_used = prev_record
            # If previous month CL was not used, carry it over (max 2)
            if prev_used == 0:
                accumulated_cl = min(prev_accumulated + 1, 2)

        # Update or insert CL record for target month
        cursor.execute(
            """
            INSERT INTO casual_leave (reg_no, user_name, dept, role, current_month, 
                                       current_month_cl_available, accumulated_cl, cl_used_current_month)
            VALUES (?, ?, ?, ?, ?, 1, ?, 0)
            ON CONFLICT(reg_no, current_month) DO UPDATE SET
                current_month_cl_available = 1,
                accumulated_cl = excluded.accumulated_cl,
                cl_used_current_month = 0,
                last_updated = CURRENT_TIMESTAMP
        """,
            (reg_no, name, dept, role, target_month, accumulated_cl),
        )

        reset_count += 1

    conn.commit()

    return {
        "success": True,
        "message": f"CL reset successfully for {reset_count} staff members",
        "target_month": target_month,
    }


@app.post("/admin/cl/adjust")
async def adjust_cl(request: Request):
    """Manually adjust CL balance for a user - Admin only"""
    verify_admin_token(request)

    try:
        body = await request.json()
        reg_no = body.get("reg_no")
        current_month_cl = body.get("current_month_cl_available")
        accumulated_cl = body.get("accumulated_cl")
        used_cl = body.get("used_cl")
    except:
        raise HTTPException(status_code=400, detail="Invalid request body")

    if not reg_no:
        raise HTTPException(status_code=400, detail="reg_no is required")

    current_month = get_current_month()

    # Validate inputs
    if current_month_cl is not None and (current_month_cl < 0 or current_month_cl > 2):
        raise HTTPException(
            status_code=400, detail="current_month_cl_available must be 0-2"
        )
    if accumulated_cl is not None and (accumulated_cl < 0 or accumulated_cl > 2):
        raise HTTPException(status_code=400, detail="accumulated_cl must be 0-2")
    if used_cl is not None and (used_cl < 0 or used_cl > 20):
        raise HTTPException(status_code=400, detail="used_cl must be 0-20")

    # Get user info
    cursor.execute("SELECT name, dept, role FROM users WHERE reg_no = ?", (reg_no,))
    user = cursor.fetchone()

    if not user:
        cursor.execute(
            "SELECT name, dept, role FROM other_staff WHERE reg_no = ?", (reg_no,)
        )
        user = cursor.fetchone()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    name, dept, role = user

    # Ensure CL record exists
    initialize_cl_for_user(reg_no, name, dept, role)

    # Build update query based on provided fields
    update_fields = []
    update_values = []

    if current_month_cl is not None:
        update_fields.append("current_month_cl_available = ?")
        update_values.append(current_month_cl)

    if accumulated_cl is not None:
        update_fields.append("accumulated_cl = ?")
        update_values.append(accumulated_cl)

    if used_cl is not None:
        update_fields.append("cl_used_current_month = ?")
        update_values.append(used_cl)

    if update_fields:
        update_fields.append("last_updated = CURRENT_TIMESTAMP")
        update_values.extend([reg_no, current_month])

        query = f"""
            UPDATE casual_leave
            SET {", ".join(update_fields)}
            WHERE reg_no = ? AND current_month = ?
        """
        cursor.execute(query, update_values)

    conn.commit()

    # Get updated values
    cursor.execute(
        """
        SELECT current_month_cl_available, accumulated_cl, cl_used_current_month
        FROM casual_leave
        WHERE reg_no = ? AND current_month = ?
    """,
        (reg_no, current_month),
    )

    cl_record = cursor.fetchone()
    cl_available, accumulated, used = cl_record

    return {
        "success": True,
        "message": "CL adjusted successfully",
        "data": {
            "reg_no": reg_no,
            "current_month_cl_available": cl_available,
            "accumulated_cl": accumulated,
            "cl_used_current_month": used,
            "total_cl_available": cl_available + accumulated,
        },
    }


# -------------------------------------------------
# ATTENDANCE DURATION SETTINGS ENDPOINTS
# -------------------------------------------------


@app.get("/admin/attendance/duration")
async def get_attendance_duration_settings(request: Request):
    """Get all attendance duration settings - Admin only"""
    verify_admin_token(request)

    cursor.execute("""
        SELECT id, slot_number, start_time, duration_minutes, is_enabled, created_at, updated_at, slot_type, slot_half
        FROM attendance_duration_settings
        ORDER BY slot_number ASC
    """)
    rows = cursor.fetchall()

    settings = []
    for row in rows:
        settings.append(
            {
                "id": row[0],
                "slot_number": row[1],
                "start_time": row[2],
                "duration_minutes": row[3],
                "is_enabled": bool(row[4]),
                "created_at": row[5],
                "updated_at": row[6],
                "slot_type": row[7] if len(row) > 7 and row[7] else "check_in",
                "slot_half": row[8] if len(row) > 8 and row[8] else "full_day",
            }
        )

    hd_settings = _get_half_day_settings()
    session_boundaries = {
        "first_half_start": hd_settings["first_half_start"],
        "first_half_end": hd_settings["first_half_end"],
        "second_half_start": hd_settings["second_half_start"],
        "second_half_end": hd_settings["second_half_end"],
    }
    auto_expansion = {
        "auto_expand_checkin_enabled": hd_settings["auto_expand_checkin_enabled"],
        "auto_expand_checkout_enabled": hd_settings["auto_expand_checkout_enabled"],
        "auto_expand_fn_minutes": hd_settings["auto_expand_fn_minutes"],
        "auto_expand_an_minutes": hd_settings["auto_expand_an_minutes"],
        "auto_expand_minutes": hd_settings["auto_expand_minutes"],
        "require_fn_check_out": hd_settings["require_fn_check_out"],
    }

    return {
        "success": True,
        "data": settings,
        "session_boundaries": session_boundaries,
        "auto_expansion": auto_expansion,
    }



@app.post("/admin/attendance/duration")
async def save_attendance_duration_settings(request: Request):
    """Save attendance duration settings - Admin only"""
    admin_user = verify_admin_token(request)

    try:
        data = await request.json()
        settings = data.get("settings", [])
        session_boundaries = data.get("session_boundaries")
    except:
        raise HTTPException(status_code=400, detail="Invalid request body")

    if not settings:
        raise HTTPException(status_code=400, detail="No settings provided")

    # Helper to convert HH:MM to minutes since midnight
    def to_minutes(time_str):
        try:
            parts = time_str.split(":")
            return int(parts[0]) * 60 + int(parts[1])
        except Exception:
            raise HTTPException(status_code=400, detail=f"Invalid time format: {time_str}")

    # Process and save master session boundaries if provided
    current_boundaries = _get_half_day_settings()
    fh_start_str = current_boundaries["first_half_start"]
    fh_end_str = current_boundaries["first_half_end"]
    sh_start_str = current_boundaries["second_half_start"]
    sh_end_str = current_boundaries["second_half_end"]

    if session_boundaries and isinstance(session_boundaries, dict):
        fh_start_str = session_boundaries.get("first_half_start", fh_start_str)
        fh_end_str = session_boundaries.get("first_half_end", fh_end_str)
        sh_start_str = session_boundaries.get("second_half_start", sh_start_str)
        sh_end_str = session_boundaries.get("second_half_end", sh_end_str)

        fh_start_m = to_minutes(fh_start_str)
        fh_end_m = to_minutes(fh_end_str)
        sh_start_m = to_minutes(sh_start_str)
        sh_end_m = to_minutes(sh_end_str)

        if fh_start_m >= fh_end_m:
            raise HTTPException(status_code=400, detail=f"First Half start time ({fh_start_str}) must be before end time ({fh_end_str}).")
        if sh_start_m >= sh_end_m:
            raise HTTPException(status_code=400, detail=f"Second Half start time ({sh_start_str}) must be before end time ({sh_end_str}).")
        if max(fh_start_m, sh_start_m) < min(fh_end_m, sh_end_m):
            raise HTTPException(status_code=400, detail=f"First Half boundary ({fh_start_str} - {fh_end_str}) overlaps with Second Half boundary ({sh_start_str} - {sh_end_str}).")

        _save_leave_setting("first_half_start", fh_start_str, admin_user["name"])
        _save_leave_setting("first_half_end", fh_end_str, admin_user["name"])
        _save_leave_setting("second_half_start", sh_start_str, admin_user["name"])
        _save_leave_setting("second_half_end", sh_end_str, admin_user["name"])

    fh_boundary_start = to_minutes(fh_start_str)
    fh_boundary_end = to_minutes(fh_end_str)
    sh_boundary_start = to_minutes(sh_start_str)
    sh_boundary_end = to_minutes(sh_end_str)


    # Validate against custom CCL settings
    cursor.execute("""
        SELECT ccl_date, early_enabled, early_start, early_end, late_enabled, late_start, late_end 
        FROM ccl_custom_dates
    """)
    custom_dates = cursor.fetchall()

    for setting in settings:
        slot_number = setting.get("slot_number")
        start_time = setting.get("start_time")
        duration_minutes = int(setting.get("duration_minutes", 30))
        slot_type = setting.get("slot_type", "check_in")
        is_enabled_val = setting.get("is_enabled", True)
        is_enabled = (
            is_enabled_val is True
            or is_enabled_val == 1
            or is_enabled_val == "true"
        )

        if not slot_number or not start_time or not is_enabled:
            continue

        slot_start_min = to_minutes(start_time)
        slot_end_min = slot_start_min + duration_minutes
        slot_half = setting.get("slot_half", "full_day")

        # Validate slot boundary against master session boundaries
        if slot_half == "first_half":
            if slot_start_min < fh_boundary_start or slot_end_min > fh_boundary_end:
                err_msg = (
                    f"Slot {slot_number} ({start_time} - {duration_minutes}m) assigned to First Half "
                    f"exceeds the First Half session boundary ({fh_start_str} - {fh_end_str}). "
                    f"Please adjust the start time or duration."
                )
                raise HTTPException(status_code=400, detail=err_msg)
        elif slot_half == "second_half":
            if slot_start_min < sh_boundary_start or slot_end_min > sh_boundary_end:
                err_msg = (
                    f"Slot {slot_number} ({start_time} - {duration_minutes}m) assigned to Second Half "
                    f"exceeds the Second Half session boundary ({sh_start_str} - {sh_end_str}). "
                    f"Please adjust the start time or duration."
                )
                raise HTTPException(status_code=400, detail=err_msg)

        for c_date in custom_dates:
            c_date_str = c_date[0]
            if not isinstance(c_date_str, str):
                c_date_str = c_date_str.strftime("%Y-%m-%d")
            c_early_enabled = bool(c_date[1])
            c_early_start = c_date[2] or "07:00"
            c_early_end = c_date[3] or "08:00"
            c_late_enabled = bool(c_date[4])
            c_late_start = c_date[5] or "17:00"
            c_late_end = c_date[6] or "18:00"

            ccl_early_start = to_minutes(c_early_start)
            ccl_early_end = to_minutes(c_early_end)
            ccl_late_start = to_minutes(c_late_start)
            ccl_late_end = to_minutes(c_late_end)

            if slot_type == "check_in" and c_early_enabled:
                if max(ccl_early_start, slot_start_min) < min(ccl_early_end, slot_end_min):
                    err_msg = (f"The Check-In Slot {slot_number} ({start_time} - {start_time} + {duration_minutes}m) overlaps with "
                               f"the Early Check-In CCL window ({c_early_start} - {c_early_end}) configured for {c_date_str}. "
                               f"Please set the slot timing outside the CCL window.")
                    raise HTTPException(status_code=400, detail=err_msg)
            elif slot_type == "check_out" and c_late_enabled:
                if max(ccl_late_start, slot_start_min) < min(ccl_late_end, slot_end_min):
                    err_msg = (f"The Check-Out Slot {slot_number} ({start_time} - {start_time} + {duration_minutes}m) overlaps with "
                               f"the Late Check-Out CCL window ({c_late_start} - {c_late_end}) configured for {c_date_str}. "
                               f"Please set the slot timing outside the CCL window.")
                    raise HTTPException(status_code=400, detail=err_msg)

    # Validate that First Half and Second Half slots do not overlap each other
    fh_intervals = []
    sh_intervals = []
    for setting in settings:
        s_num = setting.get("slot_number")
        s_start = setting.get("start_time")
        s_dur = int(setting.get("duration_minutes", 30))
        s_half = setting.get("slot_half", "full_day")
        s_en = setting.get("is_enabled", True)
        is_en = (s_en is True or s_en == 1 or s_en == "true")
        if not s_start or not is_en:
            continue
        start_m = to_minutes(s_start)
        end_m = start_m + s_dur
        if s_half == "first_half":
            fh_intervals.append((start_m, end_m, s_num, s_start))
        elif s_half == "second_half":
            sh_intervals.append((start_m, end_m, s_num, s_start))

    for fh_start, fh_end, fh_num, fh_t in fh_intervals:
        for sh_start, sh_end, sh_num, sh_t in sh_intervals:
            if max(fh_start, sh_start) < min(fh_end, sh_end):
                err_msg = (
                    f"First Half Slot {fh_num} ({fh_t}) overlaps with Second Half Slot {sh_num} ({sh_t}). "
                    f"First Half and Second Half slots cannot cross over each other's duration."
                )
                raise HTTPException(status_code=400, detail=err_msg)


    # Validate slot durations (Max 360 mins / 6 hours)
    for setting in settings:
        dur = int(setting.get("duration_minutes", 30))
        if dur > 360:
            raise HTTPException(
                status_code=400,
                detail=f"Slot {setting.get('slot_number')} duration ({dur} mins) exceeds maximum limit of 360 minutes (6 hours)."
            )
        if dur <= 0:
            raise HTTPException(
                status_code=400,
                detail=f"Slot {setting.get('slot_number')} duration must be greater than 0 minutes."
            )

    # Validate auto extension durations (Max 60 mins) if provided
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    if "auto_expansion" in body:
        auto_exp = body["auto_expansion"]
        fn_ext = int(auto_exp.get("auto_expand_fn_minutes", 5))
        an_ext = int(auto_exp.get("auto_expand_an_minutes", 5))
        def_ext = int(auto_exp.get("auto_expand_minutes", 5))
        if fn_ext > 60:
            raise HTTPException(
                status_code=400,
                detail=f"FN Extension duration ({fn_ext} mins) exceeds maximum limit of 60 minutes."
            )
        if an_ext > 60:
            raise HTTPException(
                status_code=400,
                detail=f"AN Extension duration ({an_ext} mins) exceeds maximum limit of 60 minutes."
            )
        if def_ext > 60:
            raise HTTPException(
                status_code=400,
                detail=f"Extension duration ({def_ext} mins) exceeds maximum limit of 60 minutes."
            )

    try:
        # Delete existing settings and insert new ones
        cursor.execute("DELETE FROM attendance_duration_settings")

        for setting in settings:
            slot_number = setting.get("slot_number")
            start_time = setting.get("start_time")
            duration_minutes = int(setting.get("duration_minutes", 30))
            slot_type = setting.get("slot_type", "check_in")
            is_enabled_val = setting.get("is_enabled", True)
            is_enabled = (
                1
                if (
                    is_enabled_val is True
                    or is_enabled_val == 1
                    or is_enabled_val == "true"
                )
                else 0
            )

            if not slot_number or not start_time:
                continue

            slot_half = setting.get("slot_half", "full_day")

            cursor.execute(
                """
                INSERT INTO attendance_duration_settings 
                (slot_number, start_time, duration_minutes, is_enabled, created_by, slot_type, slot_half)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
                (
                    slot_number,
                    start_time,
                    duration_minutes,
                    is_enabled,
                    admin_user["name"],
                    slot_type,
                    slot_half,
                ),
            )

        conn.commit()

        # Save auto expansion settings if provided
        body = await request.json() if request.headers.get("content-type") == "application/json" else {}
        if "auto_expansion" in body:
            auto_exp = body["auto_expansion"]
            if "auto_expand_checkin_enabled" in auto_exp:
                val = "true" if auto_exp["auto_expand_checkin_enabled"] else "false"
                _save_leave_setting("auto_expand_checkin_enabled", val, admin_user["name"])
            if "auto_expand_checkout_enabled" in auto_exp:
                val_out = "true" if auto_exp["auto_expand_checkout_enabled"] else "false"
                _save_leave_setting("auto_expand_checkout_enabled", val_out, admin_user["name"])
            if "auto_expand_fn_minutes" in auto_exp:
                mins_fn = str(int(auto_exp["auto_expand_fn_minutes"]))
                _save_leave_setting("auto_expand_fn_minutes", mins_fn, admin_user["name"])
            if "auto_expand_an_minutes" in auto_exp:
                mins_an = str(int(auto_exp["auto_expand_an_minutes"]))
                _save_leave_setting("auto_expand_an_minutes", mins_an, admin_user["name"])
            if "auto_expand_minutes" in auto_exp:
                mins = str(int(auto_exp["auto_expand_minutes"]))
                _save_leave_setting("auto_expand_minutes", mins, admin_user["name"])
            if "require_fn_check_out" in auto_exp:
                req_fn = "true" if auto_exp["require_fn_check_out"] else "false"
                _save_leave_setting("require_fn_check_out", req_fn, admin_user["name"])

        # Auto-enable half_day_enabled when any slot is configured for FN or AN
        has_half_day_slots = any(
            s.get("slot_half") in ("first_half", "second_half") and
            (s.get("is_enabled") is True or s.get("is_enabled") == 1 or s.get("is_enabled") == "true")
            for s in settings
        )
        _save_leave_setting(
            "half_day_enabled",
            "true" if has_half_day_slots else "false",
            admin_user["name"],
        )
        conn.commit()

        log_audit_event(
            "ATTENDANCE_DURATION_UPDATED",
            admin_user["reg_no"],
            True,
            f"Attendance duration settings updated: {len(settings)} slots",
        )

        return {
            "success": True,
            "message": f"Duration settings saved successfully for {len(settings)} slots",
        }
    except Exception as e:
        print(f"Error saving duration settings: {e}")
        raise HTTPException(status_code=500, detail="Failed to save duration settings")


def _calculate_effective_slot_duration(slot_type: str, slot_half: str, base_duration: int) -> tuple[int, bool, int]:
    """
    Calculate effective slot duration applying auto-extension for FN and AN check-in & check-out slots.
    Returns: (effective_duration, is_auto_extended, applied_ext_mins)
    """
    hd_settings = _get_half_day_settings()
    auto_expand_checkin = hd_settings.get("auto_expand_checkin_enabled", True)
    auto_expand_checkout = hd_settings.get("auto_expand_checkout_enabled", True)
    auto_expand_fn_mins = hd_settings.get("auto_expand_fn_minutes", 5)
    auto_expand_an_mins = hd_settings.get("auto_expand_an_minutes", 5)
    auto_expand_default_mins = hd_settings.get("auto_expand_minutes", 5)

    if (slot_type == "check_in" and auto_expand_checkin) or (slot_type == "check_out" and auto_expand_checkout):
        if slot_half == "first_half":
            applied_ext_mins = auto_expand_fn_mins
        elif slot_half == "second_half":
            applied_ext_mins = auto_expand_an_mins
        else:
            applied_ext_mins = auto_expand_default_mins
        return base_duration + applied_ext_mins, True, applied_ext_mins

    return base_duration, False, 0


@app.get("/admin/attendance/duration/check")
async def check_attendance_window():
    """Check if current time is within any attendance window - Public endpoint"""

    cursor.execute("""
        SELECT slot_number, start_time, duration_minutes, is_enabled, slot_type, slot_half
        FROM attendance_duration_settings
        WHERE is_enabled = 1
        ORDER BY slot_number ASC
    """)
    rows = cursor.fetchall()

    if not rows:
        ccl_slot_type = get_active_ccl_slot_type()
        if ccl_slot_type:
            settings = get_ccl_settings_for_date()
            if ccl_slot_type == "check_in":
                start_time = settings.get("early_check_in_start", "07:00")
                end_time = settings.get("early_check_in_end", "08:00")
            else:
                start_time = settings.get("late_check_out_start", "17:00")
                end_time = settings.get("late_check_out_end", "18:00")
                
            def time_to_min(t_str):
                p = t_str.split(":")
                return int(p[0]) * 60 + int(p[1])
                
            now_time = datetime.now()
            curr_min = now_time.hour * 60 + now_time.minute
            remaining = max(0, time_to_min(end_time) - curr_min)
            
            return {
                "allowed": True,
                "message": f"Within CCL window ({ccl_slot_type})",
                "current_slot": -1,
                "slot_type": ccl_slot_type,
                "start_time": start_time,
                "end_time": end_time,
                "remaining_minutes": remaining,
            }
            
        settings = get_ccl_settings_for_date()
        early_enabled = settings["early_check_in_ccl_enabled"]
        late_enabled = settings["late_check_out_ccl_enabled"]
        
        if early_enabled or late_enabled:
            return {
                "allowed": False,
                "message": "Outside attendance window",
                "current_slot": None,
                "slot_type": "check_in",
                "available_slots": [],
            }
            
        return {
            "allowed": True,
            "message": "No duration restrictions",
            "current_slot": None,
            "slot_type": "check_in",
        }

    current_time = datetime.now()
    current_time_str = current_time.strftime("%H:%M")
    hd_settings = _get_half_day_settings()
    auto_expand_checkin = hd_settings.get("auto_expand_checkin_enabled", True)
    auto_expand_checkout = hd_settings.get("auto_expand_checkout_enabled", True)
    auto_expand_fn_mins = hd_settings.get("auto_expand_fn_minutes", 5)
    auto_expand_an_mins = hd_settings.get("auto_expand_an_minutes", 5)
    auto_expand_default_mins = hd_settings.get("auto_expand_minutes", 5)

    for row in rows:
        slot_number = row[0]
        start_time = row[1]  # Format: HH:MM
        duration_minutes = row[2]
        slot_type = row[4] if len(row) > 4 and row[4] else "check_in"
        slot_half = row[5] if len(row) > 5 and row[5] else "full_day"

        effective_duration, is_auto_extended, applied_ext_mins = _calculate_effective_slot_duration(slot_type, slot_half, duration_minutes)

        # Parse start time
        start_hour, start_minute = map(int, start_time.split(":"))
        start_datetime = current_time.replace(
            hour=start_hour, minute=start_minute, second=0, microsecond=0
        )
        end_datetime = start_datetime + timedelta(minutes=effective_duration)

        # Check if current time is within the window
        if start_datetime <= current_time < end_datetime:
            return {
                "allowed": True,
                "message": f"Within attendance window {slot_number}" + (" (Auto Extended)" if is_auto_extended else ""),
                "current_slot": slot_number,
                "slot_type": slot_type,
                "slot_half": slot_half,
                "start_time": start_time,
                "end_time": end_datetime.strftime("%H:%M"),
                "remaining_minutes": int(
                    (end_datetime - current_time).total_seconds() / 60
                ),
                "is_auto_extended": is_auto_extended,
                "extended_minutes": applied_ext_mins if is_auto_extended else 0,
            }

    # If not in normal slot, check active CCL slot
    ccl_slot_type = get_active_ccl_slot_type()
    if ccl_slot_type:
        settings = get_ccl_settings_for_date()
        if ccl_slot_type == "check_in":
            start_time = settings.get("early_check_in_start", "07:00")
            end_time = settings.get("early_check_in_end", "08:00")
        else:
            start_time = settings.get("late_check_out_start", "17:00")
            end_time = settings.get("late_check_out_end", "18:00")
            
        def time_to_min(t_str):
            p = t_str.split(":")
            return int(p[0]) * 60 + int(p[1])
            
        now_time = datetime.now()
        curr_min = now_time.hour * 60 + now_time.minute
        remaining = max(0, time_to_min(end_time) - curr_min)
        
        return {
            "allowed": True,
            "message": f"Within CCL window ({ccl_slot_type})",
            "current_slot": -1,
            "slot_type": ccl_slot_type,
            "start_time": start_time,
            "end_time": end_time,
            "remaining_minutes": remaining,
        }

    # Check if before first slot or after last slot
    return {
        "allowed": False,
        "message": "Outside attendance window",
        "current_slot": None,
        "slot_type": "check_in",
        "available_slots": [
            {
                "slot_number": row[0],
                "start_time": row[1],
                "duration_minutes": row[2],
                "slot_type": row[4] if len(row) > 4 and row[4] else "check_in",
            }
            for row in rows
        ],
    }


# -------------------------------------------------
# CCL (COMPENSATORY CASUAL LEAVE) SETTINGS & BALANCES
# -------------------------------------------------

@app.get("/admin/ccl/settings")
async def get_ccl_settings(request: Request):
    """Get CCL settings - Admin only"""
    verify_admin_token(request)
    
    cursor.execute("SELECT key, value FROM ccl_settings")
    rows = cursor.fetchall()
    
    settings = {}
    for row in rows:
        settings[row[0]] = row[1]
        
    return {
        "success": True,
        "data": {
            "early_check_in_ccl_enabled": settings.get("early_check_in_ccl_enabled", "false") == "true",
            "late_check_out_ccl_enabled": settings.get("late_check_out_ccl_enabled", "false") == "true",
            "early_check_in_start": settings.get("early_check_in_start", "07:00"),
            "early_check_in_end": settings.get("early_check_in_end", "08:00"),
            "late_check_out_start": settings.get("late_check_out_start", "17:00"),
            "late_check_out_end": settings.get("late_check_out_end", "18:00")
        }
    }


@app.get("/admin/ccl/custom-dates")
async def get_ccl_custom_dates(request: Request):
    """Get all custom CCL dates - Admin only"""
    verify_admin_token(request)
    try:
        cursor.execute("""
            SELECT id, ccl_date, early_enabled, early_start, early_end, early_duration, late_enabled, late_start, late_end, late_duration 
            FROM ccl_custom_dates 
            ORDER BY ccl_date ASC
        """)
        rows = cursor.fetchall()
        
        custom_dates = []
        for r in rows:
            date_val = r[1]
            if not isinstance(date_val, str):
                date_val = date_val.strftime("%Y-%m-%d")
            custom_dates.append({
                "id": r[0],
                "date": date_val,
                "early_enabled": bool(r[2]),
                "early_start": r[3] or "07:00",
                "early_end": r[4] or "08:00",
                "early_duration": r[5] or 60,
                "late_enabled": bool(r[6]),
                "late_start": r[7] or "17:00",
                "late_end": r[8] or "18:00",
                "late_duration": r[9] or 60
            })
        return {"success": True, "data": custom_dates}
    except Exception as e:
        print(f"[CCL CUSTOM DATES GET ERROR] {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch custom dates: {e}")


@app.post("/admin/ccl/custom-dates")
async def save_ccl_custom_dates(request: Request):
    """Save or update custom CCL dates (handles batch) - Admin only"""
    verify_admin_token(request)
    try:
        body = await request.json()
        print(f"[CCL CUSTOM DATES POST] Received body: {body}")
        
        dates = body.get("dates", [])
        early_enabled = body.get("early_enabled", False)
        early_start = body.get("early_start", "07:00")
        early_end = body.get("early_end", "08:00")
        early_duration = body.get("early_duration", 60)
        late_enabled = body.get("late_enabled", False)
        late_start = body.get("late_start", "17:00")
        late_end = body.get("late_end", "18:00")
        late_duration = body.get("late_duration", 60)
        
        def to_minutes(time_str):
            try:
                parts = time_str.split(":")
                return int(parts[0]) * 60 + int(parts[1])
            except Exception:
                raise HTTPException(status_code=400, detail=f"Invalid time format: {time_str}")

        if to_minutes(early_start) >= to_minutes(early_end):
            raise HTTPException(status_code=400, detail="Early check-in start must be before end time")
        if to_minutes(late_start) >= to_minutes(late_end):
            raise HTTPException(status_code=400, detail="Late check-out start must be before end time")

        # Validate against duration settings (normal timing)
        if early_enabled or late_enabled:
            cursor.execute("""
                SELECT slot_number, start_time, duration_minutes, slot_type 
                FROM attendance_duration_settings 
                WHERE is_enabled = 1
            """)
            duration_slots = cursor.fetchall()
            
            ccl_early_start = to_minutes(early_start)
            ccl_early_end = to_minutes(early_end)
            ccl_late_start = to_minutes(late_start)
            ccl_late_end = to_minutes(late_end)
            
            for slot in duration_slots:
                slot_num = slot[0]
                slot_start_time = slot[1]
                slot_dur = slot[2]
                slot_type = slot[3] or "check_in"
                
                slot_start_min = to_minutes(slot_start_time)
                slot_end_min = slot_start_min + slot_dur
                
                # Check overlap logic
                if slot_type == "check_in" and early_enabled:
                    if max(ccl_early_start, slot_start_min) < min(ccl_early_end, slot_end_min):
                        err_msg = (f"The proposed Early Check-In window ({early_start} - {early_end}) overlaps with "
                                   f"normal Check-In Slot {slot_num} ({slot_start_time} - {slot_start_time} + {slot_dur}m). "
                                   f"Please set the CCL early timing outside normal check-in timing.")
                        raise HTTPException(status_code=400, detail=err_msg)
                elif slot_type == "check_out" and late_enabled:
                    if max(ccl_late_start, slot_start_min) < min(ccl_late_end, slot_end_min):
                        err_msg = (f"The proposed Late Check-Out window ({late_start} - {late_end}) overlaps with "
                                   f"normal Check-Out Slot {slot_num} ({slot_start_time} - {slot_start_time} + {slot_dur}m). "
                                   f"Please set the CCL late timing outside normal check-out timing.")
                        raise HTTPException(status_code=400, detail=err_msg)

        for d_str in dates:
            cursor.execute("""
                INSERT INTO ccl_custom_dates 
                (ccl_date, early_enabled, early_start, early_end, early_duration, late_enabled, late_start, late_end, late_duration)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (ccl_date) DO UPDATE SET
                    early_enabled = EXCLUDED.early_enabled,
                    early_start = EXCLUDED.early_start,
                    early_end = EXCLUDED.early_end,
                    early_duration = EXCLUDED.early_duration,
                    late_enabled = EXCLUDED.late_enabled,
                    late_start = EXCLUDED.late_start,
                    late_end = EXCLUDED.late_end,
                    late_duration = EXCLUDED.late_duration
            """, (d_str, early_enabled, early_start, early_end, early_duration, late_enabled, late_start, late_end, late_duration))
            
        return {"success": True, "message": f"Successfully saved settings for {len(dates)} dates."}
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"[CCL CUSTOM DATES POST ERROR] {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save custom dates: {e}")


@app.delete("/admin/ccl/custom-dates/{date_str}")
async def delete_ccl_custom_date(date_str: str, request: Request):
    """Delete custom CCL date settings - Admin only"""
    verify_admin_token(request)
    try:
        cursor.execute("DELETE FROM ccl_custom_dates WHERE ccl_date = ?", (date_str,))
        return {"success": True, "message": f"Successfully deleted custom settings for {date_str}."}
    except Exception as e:
        print(f"[CCL CUSTOM DATES DELETE ERROR] {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete custom date: {e}")


@app.post("/admin/ccl/settings")
async def save_ccl_settings(request: Request):
    """Save CCL settings with validation - Admin only"""
    verify_admin_token(request)
    
    try:
        body = await request.json()
        print(f"[CCL SETTINGS] Received body: {body}")
        early_check_in_ccl_enabled = body.get("early_check_in_ccl_enabled", False)
        late_check_out_ccl_enabled = body.get("late_check_out_ccl_enabled", False)
        early_check_in_start = body.get("early_check_in_start", "07:00")
        early_check_in_end = body.get("early_check_in_end", "08:00")
        late_check_out_start = body.get("late_check_out_start", "17:00")
        late_check_out_end = body.get("late_check_out_end", "18:00")
    except Exception as e:
        print(f"[CCL SETTINGS ERROR] Failed to parse JSON body: {e}")
        raise HTTPException(status_code=400, detail="Invalid request body")
        
    # Helper to convert HH:MM to minutes since midnight
    def to_minutes(time_str):
        try:
            parts = time_str.split(":")
            return int(parts[0]) * 60 + int(parts[1])
        except Exception:
            raise HTTPException(status_code=400, detail=f"Invalid time format: {time_str}")
            
    ccl_early_start = to_minutes(early_check_in_start)
    ccl_early_end = to_minutes(early_check_in_end)
    ccl_late_start = to_minutes(late_check_out_start)
    ccl_late_end = to_minutes(late_check_out_end)
    
    if ccl_early_start >= ccl_early_end:
        print("[CCL SETTINGS ERROR] Early check-in start must be before end time")
        raise HTTPException(status_code=400, detail="Early check-in start must be before end time")
        
    if ccl_late_start >= ccl_late_end:
        print("[CCL SETTINGS ERROR] Late check-out start must be before end time")
        raise HTTPException(status_code=400, detail="Late check-out start must be before end time")
        
    # Validation against duration settings (normal timing)
    if early_check_in_ccl_enabled or late_check_out_ccl_enabled:
        cursor.execute("""
            SELECT slot_number, start_time, duration_minutes, slot_type 
            FROM attendance_duration_settings 
            WHERE is_enabled = 1
        """)
        duration_slots = cursor.fetchall()
        
        for slot in duration_slots:
            slot_num = slot[0]
            slot_start_time = slot[1]
            slot_dur = slot[2]
            slot_type = slot[3] or "check_in"
            
            slot_start_min = to_minutes(slot_start_time)
            slot_end_min = slot_start_min + slot_dur
            
            # Check overlap logic
            if slot_type == "check_in" and early_check_in_ccl_enabled:
                # Check early check-in window overlap with check-in slots
                if max(ccl_early_start, slot_start_min) < min(ccl_early_end, slot_end_min):
                    err_msg = (f"The proposed Early Check-In window ({early_check_in_start} - {early_check_in_end}) overlaps with "
                               f"normal Check-In Slot {slot_num} ({slot_start_time} - {slot_start_time} + {slot_dur}m). "
                               f"Please set the CCL early timing outside normal check-in timing.")
                    print(f"[CCL SETTINGS ERROR] {err_msg}")
                    raise HTTPException(
                        status_code=400,
                        detail=err_msg
                    )
            elif slot_type == "check_out" and late_check_out_ccl_enabled:
                # Check late check-out window overlap with check-out slots
                if max(ccl_late_start, slot_start_min) < min(ccl_late_end, slot_end_min):
                    err_msg = (f"The proposed Late Check-Out window ({late_check_out_start} - {late_check_out_end}) overlaps with "
                               f"normal Check-Out Slot {slot_num} ({slot_start_time} - {slot_start_time} + {slot_dur}m). "
                               f"Please set the CCL late timing outside normal check-out timing.")
                    print(f"[CCL SETTINGS ERROR] {err_msg}")
                    raise HTTPException(
                        status_code=400,
                        detail=err_msg
                    )

    # Save to database
    early_enabled_str = "true" if early_check_in_ccl_enabled else "false"
    late_enabled_str = "true" if late_check_out_ccl_enabled else "false"
    cursor.execute("UPDATE ccl_settings SET value = ? WHERE key = 'early_check_in_ccl_enabled'", (early_enabled_str,))
    cursor.execute("UPDATE ccl_settings SET value = ? WHERE key = 'late_check_out_ccl_enabled'", (late_enabled_str,))
    cursor.execute("UPDATE ccl_settings SET value = ? WHERE key = 'early_check_in_start'", (early_check_in_start,))
    cursor.execute("UPDATE ccl_settings SET value = ? WHERE key = 'early_check_in_end'", (early_check_in_end,))
    cursor.execute("UPDATE ccl_settings SET value = ? WHERE key = 'late_check_out_start'", (late_check_out_start,))
    cursor.execute("UPDATE ccl_settings SET value = ? WHERE key = 'late_check_out_end'", (late_check_out_end,))
    conn.commit()
    
    return {"success": True, "message": "CCL settings updated successfully"}


# ─────────────────────────────────────────────────────────────────────────────
# LEAVE SETTINGS — Half-day, CL Expiry, EL Expiry (Admin)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/admin/leave-settings")
async def get_leave_settings_api(request: Request):
    """Get all leave settings including half-day config, CL/EL expiry — Admin only."""
    verify_admin_token(request)
    _get_leave_settings(force=True)
    hd_settings = _get_half_day_settings()
    return {
        "success": True,
        "settings": {
            "half_day_enabled": hd_settings["half_day_enabled"],
            "first_half_label": hd_settings["first_half_label"],
            "second_half_label": hd_settings["second_half_label"],
            "cl_monthly_allocation": hd_settings["cl_monthly_allocation"],
            "cl_expiry_date": hd_settings["cl_expiry_date"],
            "el_expiry_date": hd_settings["el_expiry_date"],
        },
    }



@app.post("/admin/leave-settings")
async def save_leave_settings_api(request: Request):
    """Save leave settings (half-day config, CL/EL expiry dates) — Admin only."""
    admin = verify_admin_token(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request body")

    admin_name = admin.get("name", "admin") if admin else "admin"
    saved = []

    # Boolean/string settings
    if "half_day_enabled" in body:
        val = "true" if body["half_day_enabled"] else "false"
        _save_leave_setting("half_day_enabled", val, admin_name)
        saved.append("half_day_enabled")

    if "first_half_label" in body:
        _save_leave_setting("first_half_label", str(body["first_half_label"]), admin_name)
        saved.append("first_half_label")

    if "second_half_label" in body:
        _save_leave_setting("second_half_label", str(body["second_half_label"]), admin_name)
        saved.append("second_half_label")

    if "cl_monthly_allocation" in body:
        try:
            alloc = int(body["cl_monthly_allocation"])
            if alloc < 0:
                raise ValueError
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="cl_monthly_allocation must be a non-negative integer")
        _save_leave_setting("cl_monthly_allocation", str(alloc), admin_name)
        saved.append("cl_monthly_allocation")

    if "cl_expiry_date" in body:
        val = (body["cl_expiry_date"] or "").strip()
        # Validate date format if provided
        if val:
            try:
                datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="cl_expiry_date must be YYYY-MM-DD or empty")
        _save_leave_setting("cl_expiry_date", val, admin_name)
        saved.append("cl_expiry_date")

    if "el_expiry_date" in body:
        val = (body["el_expiry_date"] or "").strip()
        if val:
            try:
                datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="el_expiry_date must be YYYY-MM-DD or empty")
        _save_leave_setting("el_expiry_date", val, admin_name)
        saved.append("el_expiry_date")

    conn.commit()
    _get_leave_settings(force=True)  # Refresh cache
    log_audit_event("LEAVE_SETTINGS_UPDATED", admin_name, True, f"Keys updated: {saved}")

    return {"success": True, "message": "Leave settings saved successfully", "updated_keys": saved}


@app.post("/admin/cl/expire")
async def expire_cl_now(request: Request):
    """Immediately zero out all CL balances (admin-triggered expiry) — Admin only."""
    admin = verify_admin_token(request)
    admin_name = admin.get("name", "admin") if admin else "admin"
    today_str = datetime.now().strftime("%Y-%m-%d")
    try:
        # Fetch non-zero CL records before zeroing out to record in expired_leaves
        cursor.execute("""
            SELECT reg_no, user_name, dept, role, (current_month_cl_available + accumulated_cl) as total_cl
            FROM casual_leave
            WHERE current_month_cl_available > 0 OR accumulated_cl > 0
        """)
        cl_rows = cursor.fetchall()
        for row in cl_rows:
            cursor.execute("""
                INSERT INTO expired_leaves (reg_no, user_name, dept, role, leave_type, expired_amount, expiry_date, reason)
                VALUES (?, ?, ?, ?, 'Casual Leave (CL)', ?, ?, 'Manually expired by Admin')
            """, (row[0], row[1], row[2], row[3], float(row[4]), today_str))

        cursor.execute(
            """
            UPDATE casual_leave
            SET current_month_cl_available = 0,
                accumulated_cl = 0,
                last_updated = CURRENT_TIMESTAMP
            WHERE current_month_cl_available > 0 OR accumulated_cl > 0
        """
        )
        count = cursor.rowcount or 0
        conn.commit()
        log_audit_event("CL_MANUALLY_EXPIRED", admin_name, True, f"CL zeroed for {count} records")
        return {
            "success": True,
            "message": f"CL balances expired successfully for {count} staff record(s).",
            "records_affected": count,
        }
    except Exception as e:
        print(f"[CL Expiry] Error: {e}")
        raise HTTPException(status_code=500, detail=f"CL expiry failed: {e}")


@app.post("/admin/el/expire")
async def expire_el_now(request: Request):
    """Immediately zero out all EL balances (admin-triggered expiry) — Admin only."""
    admin = verify_admin_token(request)
    admin_name = admin.get("name", "admin") if admin else "admin"
    today_str = datetime.now().strftime("%Y-%m-%d")
    try:
        # Fetch non-zero EL records before zeroing out to record in expired_leaves
        cursor.execute("""
            SELECT reg_no, user_name, dept, role, balance
            FROM earned_leave
            WHERE balance > 0
        """)
        el_rows = cursor.fetchall()
        for row in el_rows:
            cursor.execute("""
                INSERT INTO expired_leaves (reg_no, user_name, dept, role, leave_type, expired_amount, expiry_date, reason)
                VALUES (?, ?, ?, ?, 'Earned Leave (EL)', ?, ?, 'Manually expired by Admin')
            """, (row[0], row[1], row[2], row[3], float(row[4]), today_str))

        cursor.execute(
            """
            UPDATE earned_leave
            SET balance = 0.0,
                updated_at = CURRENT_TIMESTAMP
            WHERE balance > 0
        """
        )
        count = cursor.rowcount or 0
        conn.commit()
        log_audit_event("EL_MANUALLY_EXPIRED", admin_name, True, f"EL zeroed for {count} records")
        return {
            "success": True,
            "message": f"Earned Leave balances expired successfully for {count} staff record(s).",
            "records_affected": count,
        }
    except Exception as e:
        print(f"[EL Expiry] Error: {e}")
        raise HTTPException(status_code=500, detail=f"EL expiry failed: {e}")


@app.get("/leave/expired")
async def get_expired_leaves(request: Request, reg_no: str = None, dept: str = None):
    """Get historical list of expired leaves.
    Admin gets all/filtered by reg_no/dept; regular users get their own expired leaves.
    """
    user = verify_any_user_token(request)
    is_admin = user.get("role") == "admin"

    try:
        # Check and process automated expiry against set cl_expiry_date and el_expiry_date
        _process_auto_leave_expiry()

        query = """
            SELECT id, reg_no, user_name, dept, role, leave_type, expired_amount, expiry_date, reason, created_at
            FROM expired_leaves
        """
        params = []
        conditions = []

        if not is_admin:
            conditions.append("reg_no = ?")
            params.append(user["reg_no"])
        else:
            if reg_no:
                conditions.append("reg_no = ?")
                params.append(reg_no)
            if dept and dept != "All":
                conditions.append("dept = ?")
                params.append(dept)

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        query += " ORDER BY expiry_date DESC, created_at DESC"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        expired_list = [
            {
                "id": r[0],
                "reg_no": r[1],
                "user_name": r[2],
                "dept": r[3],
                "role": r[4],
                "leave_type": r[5],
                "expired_amount": float(r[6]),
                "expiry_date": str(r[7]),
                "reason": r[8],
                "created_at": _ts(r[9]),
            }
            for r in rows
        ]

        return {
            "success": True,
            "expired_leaves": expired_list,
            "count": len(expired_list),
        }
    except Exception as e:
        print(f"[get_expired_leaves] Error: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch expired leaves")


def _process_auto_leave_expiry():
    """Automated check if today reached or passed cl_expiry_date or el_expiry_date."""
    try:
        s = _get_half_day_settings()
        cl_exp = s.get("cl_expiry_date")
        el_exp = s.get("el_expiry_date")
        today_str = datetime.now().strftime("%Y-%m-%d")

        # Process CL Expiry
        if cl_exp and today_str >= cl_exp:
            cursor.execute("""
                SELECT reg_no, user_name, dept, role, (current_month_cl_available + accumulated_cl) as total_cl
                FROM casual_leave
                WHERE current_month_cl_available > 0 OR accumulated_cl > 0
            """)
            cl_rows = cursor.fetchall()
            for row in cl_rows:
                cursor.execute("""
                    INSERT INTO expired_leaves (reg_no, user_name, dept, role, leave_type, expired_amount, expiry_date, reason)
                    VALUES (?, ?, ?, ?, 'Casual Leave (CL)', ?, ?, 'Automated expiry date reached')
                """, (row[0], row[1], row[2], row[3], float(row[4]), cl_exp))

            cursor.execute("""
                UPDATE casual_leave
                SET current_month_cl_available = 0, accumulated_cl = 0, last_updated = CURRENT_TIMESTAMP
                WHERE current_month_cl_available > 0 OR accumulated_cl > 0
            """)
            _save_leave_setting("cl_expiry_date", "", "system")

        # Process EL Expiry
        if el_exp and today_str >= el_exp:
            cursor.execute("""
                SELECT reg_no, user_name, dept, role, balance
                FROM earned_leave
                WHERE balance > 0
            """)
            el_rows = cursor.fetchall()
            for row in el_rows:
                cursor.execute("""
                    INSERT INTO expired_leaves (reg_no, user_name, dept, role, leave_type, expired_amount, expiry_date, reason)
                    VALUES (?, ?, ?, ?, 'Earned Leave (EL)', ?, ?, 'Automated expiry date reached')
                """, (row[0], row[1], row[2], row[3], float(row[4]), el_exp))

            cursor.execute("""
                UPDATE earned_leave
                SET balance = 0.0, updated_at = CURRENT_TIMESTAMP
                WHERE balance > 0
            """)
            _save_leave_setting("el_expiry_date", "", "system")

        conn.commit()
    except Exception as e:
        print(f"[_process_auto_leave_expiry] Error: {e}")


@app.post("/admin/half-day/mark-absent")
async def trigger_eod_marking(request: Request):
    """Manually trigger the end-of-day half-day absent marking for a given date — Admin only."""
    verify_admin_token(request)
    try:
        body = await request.json()
        date_str = body.get("date", datetime.now().strftime("%Y-%m-%d"))
        # Validate date
        datetime.strptime(date_str, "%Y-%m-%d")
    except Exception:
        date_str = datetime.now().strftime("%Y-%m-%d")

    _mark_halves_absent_for_day(date_str)
    return {"success": True, "message": f"End-of-day marking triggered for {date_str}"}


@app.get("/admin/ccl/balances")

async def get_ccl_balances(request: Request):
    """Get all staff members' Earned Leave (CCL) balances - Admin only"""
    verify_admin_token(request)
    
    try:
        # Ensure everyone in users and other_staff is in earned_leave
        cursor.execute("""
            SELECT reg_no, name, dept, role FROM users WHERE role IN ('hod', 'staff')
            UNION ALL
            SELECT reg_no, name, dept, role FROM other_staff
        """)
        all_users = cursor.fetchall()
        
        # Initialize missing balances
        for row in all_users:
            reg_no, name, dept, role = row
            cursor.execute("SELECT 1 FROM earned_leave WHERE reg_no = %s", (reg_no,))
            if not cursor.fetchone():
                cursor.execute("""
                    INSERT INTO earned_leave (reg_no, user_name, dept, role, balance)
                    VALUES (%s, %s, %s, %s, 0.0)
                """, (reg_no, name, dept, role))
                
        cursor.execute("SELECT reg_no, user_name, dept, role, balance, updated_at FROM earned_leave ORDER BY dept, user_name")
        rows = cursor.fetchall()
        
        balances = []
        for row in rows:
            balances.append({
                "reg_no": row[0],
                "name": row[1],
                "dept": row[2],
                "role": row[3],
                "balance": float(row[4]) if row[4] is not None else 0.0,
                "updated_at": str(row[5]) if row[5] else None
            })
            
        return {"success": True, "data": balances}
    except Exception as e:
        print(f"ERROR in get_ccl_balances: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load balances: {str(e)}")


@app.post("/admin/ccl/adjust")
async def adjust_ccl_balance(request: Request):
    """Manually adjust CCL/Earned Leave balance for a user - Admin only"""
    verify_admin_token(request)
    
    try:
        body = await request.json()
        reg_no = body.get("reg_no")
        adjustment = float(body.get("adjustment", 0.0))
        balance = body.get("balance") if body.get("balance") is not None else body.get("new_balance")
    except:
        raise HTTPException(status_code=400, detail="Invalid request body")
        
    if not reg_no:
        raise HTTPException(status_code=400, detail="reg_no is required")


    # Get user info
    cursor.execute("SELECT name, dept, role FROM users WHERE reg_no = ?", (reg_no,))
    user = cursor.fetchone()
    if not user:
        cursor.execute("SELECT name, dept, role FROM other_staff WHERE reg_no = ?", (reg_no,))
        user = cursor.fetchone()
        
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    name, dept, role = user
    
    # Ensure entry exists
    cursor.execute("SELECT balance FROM earned_leave WHERE reg_no = ?", (reg_no,))
    record = cursor.fetchone()
    if not record:
        cursor.execute("""
            INSERT INTO earned_leave (reg_no, user_name, dept, role, balance)
            VALUES (?, ?, ?, ?, 0.0)
        """, (reg_no, name, dept, role))
        current_bal = 0.0
    else:
        current_bal = float(record[0])
        
    if balance is not None:
        new_bal = float(balance)
    else:
        new_bal = current_bal + adjustment
        
    if new_bal < 0:
        new_bal = 0.0
        
    eff_adjustment = new_bal - current_bal
    
    cursor.execute("""
        UPDATE earned_leave 
        SET balance = ?, updated_at = CURRENT_TIMESTAMP
        WHERE reg_no = ?
    """, (new_bal, reg_no))
    
    if eff_adjustment != 0.0:
        cursor.execute("""
            INSERT INTO ccl_earned_history (reg_no, name, dept, date, time, slot_type, earned_points)
            VALUES (?, ?, ?, ?, ?, 'adjustment', ?)
        """, (
            reg_no, 
            name, 
            dept, 
            datetime.now().strftime("%Y-%m-%d"), 
            datetime.now().strftime("%H:%M:%S"), 
            eff_adjustment
        ))
        
    conn.commit()
    
    return {
        "success": True,
        "message": f"Earned Leave balance updated successfully to {new_bal}",
        "data": {
            "reg_no": reg_no,
            "balance": new_bal
        }
    }


@app.delete("/admin/ccl/balances/{reg_no}")
async def delete_ccl_balance(request: Request, reg_no: str):
    """Delete (or reset) a user's Earned Leave (CCL) balance - Admin only"""
    verify_admin_token(request)
    
    try:
        cursor.execute("DELETE FROM earned_leave WHERE reg_no = ?", (reg_no,))
        conn.commit()
        return {"success": True, "message": f"CCL Balance record deleted for user {reg_no}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete record: {str(e)}")


@app.post("/admin/ccl/sync")
async def sync_ccl_balances(request: Request):
    """Synchronize CCL balances and history, ensuring no discrepancies - Admin only"""
    verify_admin_token(request)
    
    try:
        # 1. Ensure all users exist in earned_leave
        cursor.execute("""
            SELECT reg_no, name, dept, role FROM users WHERE role IN ('hod', 'staff')
            UNION ALL
            SELECT reg_no, name, dept, role FROM other_staff
        """)
        all_users = cursor.fetchall()
        
        for row in all_users:
            reg_no, name, dept, role = row
            cursor.execute("SELECT 1 FROM earned_leave WHERE reg_no = ?", (reg_no,))
            if not cursor.fetchone():
                cursor.execute("""
                    INSERT INTO earned_leave (reg_no, user_name, dept, role, balance)
                    VALUES (?, ?, ?, ?, 0.0)
                """, (reg_no, name, dept, role))
                
        # 2. Get history sums
        cursor.execute("""
            SELECT reg_no, COALESCE(SUM(earned_points), 0.0) 
            FROM ccl_earned_history 
            GROUP BY reg_no
        """)
        history_sums = {r[0]: float(r[1]) for r in cursor.fetchall()}
        
        # 3. Get current balances
        cursor.execute("SELECT reg_no, user_name, dept, balance FROM earned_leave")
        balances = cursor.fetchall()
        
        sync_count = 0
        now_date = datetime.now().strftime("%Y-%m-%d")
        now_time = datetime.now().strftime("%H:%M:%S")
        
        for row in balances:
            reg_no, name, dept, bal = row
            bal = float(bal)
            hist_sum = history_sums.get(reg_no, 0.0)
            
            if abs(bal - hist_sum) > 0.001:
                # Update earned_leave balance to match history sum
                cursor.execute("""
                    UPDATE earned_leave
                    SET balance = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE reg_no = ?
                """, (hist_sum, reg_no))
                sync_count += 1
                
        conn.commit()
        return {"success": True, "message": f"Successfully synchronized {sync_count} user balances with history."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


@app.get("/admin/ccl/history")
async def get_ccl_history(request: Request):
    """Get all CCL earnings history - Admin only"""
    verify_admin_token(request)
    
    cursor.execute("""
        SELECT id, reg_no, name, dept, date, time, slot_type, earned_points, created_at
        FROM ccl_earned_history
        ORDER BY created_at DESC
        LIMIT 1000
    """)
    rows = cursor.fetchall()
    
    history = []
    for row in rows:
        history.append({
            "id": row[0],
            "reg_no": row[1],
            "name": row[2],
            "dept": row[3],
            "date": str(row[4]),
            "time": str(row[5]),
            "slot_type": row[6],
            "earned_points": float(row[7]),
            "created_at": row[8]
        })
        
    return {"success": True, "data": history}


@app.get("/ccl/status/{reg_no}")
async def get_ccl_status(reg_no: str):
    """Get CCL balance for a specific user"""
    # Get user info
    cursor.execute("SELECT name, dept, role FROM users WHERE reg_no = ?", (reg_no,))
    user = cursor.fetchone()
    if not user:
        cursor.execute("SELECT name, dept, role FROM other_staff WHERE reg_no = ?", (reg_no,))
        user = cursor.fetchone()
        
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    name, dept, role = user
    
    cursor.execute("SELECT balance FROM earned_leave WHERE reg_no = ?", (reg_no,))
    row = cursor.fetchone()
    if not row:
        cursor.execute("""
            INSERT INTO earned_leave (reg_no, user_name, dept, role, balance)
            VALUES (?, ?, ?, ?, 0.0)
        """, (reg_no, name, dept, role))
        conn.commit()
        balance = 0.0
    else:
        balance = float(row[0])
        
    return {
        "success": True,
        "data": {
            "reg_no": reg_no,
            "name": name,
            "dept": dept,
            "role": role,
            "earned_leave_available": balance
        }
    }


# -------------------------------------------------
# ACADEMICS / ACADEMIC YEAR SETTINGS
# -------------------------------------------------


def _build_academics_response():
    """Build the standard academics response dict including ranges."""
    _load_academic_settings_from_storage()
    ranges = _get_academic_ranges()
    overrides = _academic_settings.get("holiday_overrides", {}) or {}
    if ranges:
        # Compute holidays across all ranges
        all_holidays = []
        for r in ranges:
            start = _parse_date_or_none(r["start"])
            end = _parse_date_or_none(r["end"])
            if start and end:
                all_holidays.extend(_build_academic_holidays(start, end))
        return {
            "academic_ranges": ranges,
            "academic_year_start": ranges[0]["start"],
            "academic_year_end": ranges[-1]["end"],
            "holidays": all_holidays,
            "holiday_overrides": overrides,
        }
    # Fallback to single range
    start_date = _parse_date_or_none(_academic_settings.get("academic_year_start"))
    end_date = _parse_date_or_none(_academic_settings.get("academic_year_end"))
    holidays = _build_academic_holidays(start_date, end_date) if start_date and end_date else []
    return {
        "academic_ranges": [],
        "academic_year_start": start_date.isoformat() if start_date else None,
        "academic_year_end": end_date.isoformat() if end_date else None,
        "holidays": holidays,
        "holiday_overrides": overrides,
    }


@app.get("/admin/academics")
async def get_academics_settings(request: Request):
    """Get academic year settings and derived holiday calendar."""
    verify_admin_token(request)
    return {"success": True, "data": _build_academics_response()}


@app.get("/academics/current")
async def get_current_academics():
    """Get current academic year dates (public, no auth required)."""
    ranges = _get_academic_ranges()
    if ranges:
        return {
            "academic_ranges": ranges,
            "academic_year_start": ranges[0]["start"],
            "academic_year_end": ranges[-1]["end"],
        }
    start = _parse_date_or_none(_academic_settings.get("academic_year_start"))
    end = _parse_date_or_none(_academic_settings.get("academic_year_end"))
    return {
        "academic_ranges": [],
        "academic_year_start": start.isoformat() if start else None,
        "academic_year_end": end.isoformat() if end else None,
    }


@app.post("/admin/academics")
async def save_academics_settings(request: Request):
    """Persist academic year settings and holiday overrides.
    Accepts either academic_year_start/end (single range) or academic_ranges (multiple ranges).
    """
    admin_user = verify_admin_token(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request body")

    ranges = body.get("academic_ranges")
    if ranges:
        # Multiple ranges mode
        if not isinstance(ranges, list) or len(ranges) == 0:
            raise HTTPException(status_code=400, detail="academic_ranges must be a non-empty list")
        normalized_ranges = []
        for i, r in enumerate(ranges):
            start_str = r.get("start")
            end_str = r.get("end")
            if not start_str or not end_str:
                raise HTTPException(status_code=400, detail=f"Range {i}: both start and end are required")
            try:
                start_dt = datetime.strptime(start_str, "%Y-%m-%d").date()
                end_dt = datetime.strptime(end_str, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Range {i}: invalid date format")
            if start_dt > end_dt:
                raise HTTPException(status_code=400, detail=f"Range {i}: start must be before end")
            normalized_ranges.append({"start": start_str, "end": end_str})
        _academic_settings["academic_ranges"] = normalized_ranges
        # Clear single-range fields for consistency
        _academic_settings.pop("academic_year_start", None)
        _academic_settings.pop("academic_year_end", None)
    else:
        # Single range mode (backward compatible)
        start_date = body.get("academic_year_start")
        end_date = body.get("academic_year_end")

        if not start_date or not end_date:
            raise HTTPException(
                status_code=400,
                detail="Either academic_ranges or both academic_year_start and academic_year_end are required",
            )

        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")

        if start_dt > end_dt:
            raise HTTPException(
                status_code=400,
                detail="Academic year start date must be before the end date",
            )

        _academic_settings["academic_year_start"] = start_date
        _academic_settings["academic_year_end"] = end_date
        _academic_settings.pop("academic_ranges", None)

    overrides = body.get("holiday_overrides", {})
    normalized_overrides = {}
    if isinstance(overrides, dict):
        for date_key, item in overrides.items():
            if not date_key:
                continue
            status = str((item or {}).get("status", "")).lower()
            if status not in {"holiday", "working_day"}:
                continue
            normalized_overrides[date_key] = {
                "status": status,
                "reason": str((item or {}).get("reason", "")).strip(),
            }

    _academic_settings["holiday_overrides"] = normalized_overrides
    _save_academic_settings_to_storage()

    log_audit_event(
        "ACADEMIC_SETTINGS_UPDATED",
        admin_user["reg_no"],
        True,
        f"Academic settings updated: ranges={len(normalized_ranges) if ranges else 1}, overrides={len(normalized_overrides)}",
    )

    return {
        "success": True,
        "message": "Academic settings saved successfully",
        "data": _build_academics_response(),
    }


# -------------------------------------------------
# AUTHENTICATION UTILITIES
# -------------------------------------------------
import bcrypt


def hash_password(password: str) -> str:
    """Hash a password using bcrypt"""
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(password: str, hashed: str) -> bool:
    """Verify a password against its hash"""
    return bcrypt.checkpw(password.encode("utf-8"), hashed.encode("utf-8"))


def is_user_suspended(username: str, is_other_staff: bool = False) -> bool:
    """Check if a user is suspended by querying their suspended status from the database by username or reg_no"""
    try:
        table = "other_staff" if is_other_staff else "users"
        # Check by username
        cursor.execute(f"SELECT COALESCE(suspended, FALSE) FROM {table} WHERE username = %s", (username,))
        row = cursor.fetchone()
        if not row:
            # Check by reg_no (case-insensitive)
            cursor.execute(f"SELECT COALESCE(suspended, FALSE) FROM {table} WHERE LOWER(reg_no) = LOWER(%s)", (username,))
            row = cursor.fetchone()
        return row[0] if row else False
    except Exception as e:
        print(f"Error checking user suspension status: {e}")
        return False


def get_user_by_username(username: str):
    """Get user by username"""
    cursor.execute(
        "SELECT id, username, password_hash, reg_no, name, dept, role, is_active, face_embedding, current_device_id, out_permission_enabled, out_permission_expiry, created_by, created_at, updated_at, embedding, can_reregister, suspended FROM users WHERE username = ?",
        (username,),
    )
    return cursor.fetchone()


def get_user_by_reg_no(reg_no: str):
    """Get user by registration number (case-insensitive)"""
    cursor.execute(
        "SELECT id, username, password_hash, reg_no, name, dept, role, is_active, face_embedding, current_device_id, out_permission_enabled, out_permission_expiry, created_by, created_at, updated_at, embedding, can_reregister, suspended FROM users WHERE LOWER(reg_no) = LOWER(?)",
        (reg_no,),
    )
    return cursor.fetchone()


def delete_user_data_by_reg_no(reg_no: str):
    """Delete all attendance, leaves, location records, and face samples for a user"""
    # Delete from audit logs & notifications referring to leave requests first
    cursor.execute("""
        DELETE FROM leave_request_audit_log 
        WHERE leave_request_id IN (
            SELECT id FROM leave_requests WHERE user_reg_no = ?
        )
    """, (reg_no,))
    cursor.execute("""
        DELETE FROM admin_notifications 
        WHERE notification_type = 'leave_request' AND related_id IN (
            SELECT CAST(id AS varchar) FROM leave_requests WHERE user_reg_no = ?
        )
    """, (reg_no,))

    # Now delete primary records safely
    tables_to_delete = [
        ("attendance", "reg_no"),
        ("daily_attendance_status", "reg_no"),
        ("face_embedding_samples", "reg_no"),
        ("casual_leave", "reg_no"),
        ("user_location_logs", "reg_no"),
        ("user_latest_locations", "reg_no"),
        ("other_staff_attendance", "reg_no"),
        ("leave_requests", "user_reg_no"),
        ("face_reregister_requests", "staff_reg_no"),
        ("user_locations", "reg_no"),
        ("attendance_locations", "reg_no"),
        ("geofence_breach_monitoring", "reg_no"),
        ("students", "reg_no"),
        ("earned_leave", "reg_no"),
        ("ccl_earned_history", "reg_no"),
        ("users", "reg_no"),
        ("other_staff", "reg_no"),
    ]
    
    for table, col in tables_to_delete:
        try:
            cursor.execute(f"DELETE FROM {table} WHERE {col} = ?", (reg_no,))
        except Exception as e:
            print(f"[SAFE DELETE] Table {table} delete skipped/failed: {e}")


# -------------------------------------------------
# OTHER STAFF HELPER FUNCTIONS
# -------------------------------------------------
OTHER_STAFF_ROLES = (
    "principal",
    "placement_staff",
    "lab_technician",
    "system_admin",
    "office_staff",
)


def get_other_staff_by_username(username: str):
    """Get other_staff by username"""
    cursor.execute("SELECT * FROM other_staff WHERE username = ?", (username,))
    return cursor.fetchone()


def get_other_staff_by_reg_no(reg_no: str):
    """Get other_staff by registration number (case-insensitive)"""
    cursor.execute(
        "SELECT * FROM other_staff WHERE LOWER(reg_no) = LOWER(?)", (reg_no,)
    )
    return cursor.fetchone()


def get_other_staff_by_id(staff_id: int):
    """Get other_staff by ID"""
    cursor.execute(
        "SELECT id, username, password_hash, reg_no, name, dob, role, dept, embedding, can_reregister, created_at, created_by FROM other_staff WHERE id = ?",
        (staff_id,),
    )
    return cursor.fetchone()


def get_default_department_for_role(role: str) -> str | None:
    """Return the default department for roles that should never have an empty department."""
    role_defaults = {
        "principal": "Administration",
        "placement_staff": "Placement Staff",
        "lab_technician": "Lab Technician",
        "system_admin": "System Admin",
        "office_staff": "Office Staff",
    }
    return role_defaults.get((role or "").strip().lower())


# -------------------------------------------------
# AUTHENTICATION ENDPOINTS
# -------------------------------------------------
@app.post("/login")
async def login(request: Request):
    """Login with username and password - checks both users and other_staff tables"""
    try:
        # Check for VPN before processing login
        client_ip = get_client_ip(request)
        if _app_settings.get("enforce_vpn_blocking", True) and is_vpn_ip(client_ip):
            raise HTTPException(
                status_code=403,
                detail="VPN detected. Please turn off your VPN connection to continue.",
            )

        data = await request.json()
        username = data.get("username")
        password = data.get("password")

        if not username or not password:
            raise HTTPException(status_code=400, detail="Missing credentials")

        # First, try to get user from 'users' table (admin, hod, staff)
        user = get_user_by_username(username)

        # If not found by username, try by reg_no
        if not user:
            user = get_user_by_reg_no(username)

        is_other_staff = False

        print(f"[LOGIN DEBUG] Looking for user: {username}")
        print(f"[LOGIN DEBUG] Found in 'users' table: {user is not None}")

        # If not found in users table, try 'other_staff' table by username
        if not user:
            other_staff = get_other_staff_by_username(username)
            print(
                f"[LOGIN DEBUG] Found in 'other_staff' table by username: {other_staff is not None}"
            )
            if other_staff:
                # other_staff table structure: id, username, password_hash, reg_no, name, dob, role, dept, embedding, can_reregister, created_at, created_by
                user = (
                    other_staff[0],  # id
                    other_staff[1],  # username
                    other_staff[2],  # password_hash (index 2)
                    other_staff[3],  # reg_no
                    other_staff[4],  # name
                    other_staff[7],  # dept (index 7)
                    other_staff[6],  # role (index 6)
                )
                is_other_staff = True
                print(
                    f"[LOGIN DEBUG] Mapped other_staff user - role: {other_staff[6]}, dept: {other_staff[7]}"
                )
            else:
                # Try other_staff by reg_no
                other_staff = get_other_staff_by_reg_no(username)
                print(
                    f"[LOGIN DEBUG] Found in 'other_staff' table by reg_no: {other_staff is not None}"
                )
                if other_staff:
                    user = (
                        other_staff[0],  # id
                        other_staff[1],  # username
                        other_staff[2],  # password_hash (index 2)
                        other_staff[3],  # reg_no
                        other_staff[4],  # name
                        other_staff[7],  # dept (index 7)
                        other_staff[6],  # role (index 6)
                    )
                    is_other_staff = True
                    print(
                        f"[LOGIN DEBUG] Mapped other_staff user by reg_no - role: {other_staff[6]}, dept: {other_staff[7]}"
                    )

        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Verify password
        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check if user is suspended
        if is_user_suspended(user[1], is_other_staff):
            raise HTTPException(
                status_code=403,
                detail="Account suspended. Please contact the administrator.",
            )

        device_id = data.get("device_id")

        # Get user role from database
        user_role = user[6]

        # Update current_device_id in appropriate table
        if is_other_staff:
            cursor.execute("UPDATE other_staff SET current_device_id = ? WHERE username = ?", (device_id, user[1]))
        else:
            cursor.execute("UPDATE users SET current_device_id = ? WHERE username = ?", (device_id, user[1]))
        conn.commit()

        # Create token
        import base64

        token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode(
            "utf-8"
        )

        print(
            f"[LOGIN DEBUG] Login SUCCESS for {username}, role: {user_role}, is_other_staff: {is_other_staff}"
        )

        # Build response based on user type
        if is_other_staff:
            # For other_staff users, get additional info
            other_staff = get_other_staff_by_username(username)
            face_registered = (
                other_staff[8] is not None if len(other_staff) > 8 else False
            )

            return {
                "message": "Login successful",
                "token": token,
                "user": {
                    "id": user[0],
                    "username": user[1],
                    "regNo": user[3],
                    "name": user[4],
                    "dept": user[5],
                    "role": user_role,
                    "face_registered": face_registered,
                },
            }
        else:
            # For regular users (admin, hod, staff)
            return {
                "message": "Login successful",
                "token": token,
                "user": {
                    "id": user[0],
                    "username": user[1],
                    "regNo": user[3],
                    "name": user[4],
                    "dept": user[5],
                    "role": user_role,
                },
            }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Login error: {e}")
        raise HTTPException(status_code=500, detail="Login failed")


@app.post("/login_face")
async def login_face(request: Request):
    """Login with face recognition (existing functionality)"""
    # This would integrate with existing face recognition
    # For now, return error to indicate face login not implemented
    raise HTTPException(status_code=501, detail="Face login not implemented")


# -------------------------------------------------
# USER MANAGEMENT ENDPOINTS (Role-based)
# -------------------------------------------------
@app.post("/users/create")
async def create_user(request: Request):
    """Create a new user (Admin and HOD only)"""
    try:
        data = await request.json()
        username = data.get("username")
        password = data.get("password")
        reg_no = data.get("reg_no")
        name = data.get("name")
        dept = data.get("dept")
        role = data.get("role")
        created_by = data.get("created_by")

        # Validate required fields
        if not all([username, password, name, dept, role]):
            print(
                f"[DEBUG] Missing required fields - username: {username}, password: {password}, name: {name}, dept: {dept}, role: {role}"
            )
            raise HTTPException(status_code=400, detail="Missing required fields")

        # Validate role
        allowed_roles = [
            "admin",
            "hod",
            "staff",
            "principal",
            "vice_chancellor",
            "director",
            "dean",
        ]
        print(
            f"[DEBUG] /users/create Role validation - received: '{role}', allowed: {allowed_roles}, is_valid: {role in allowed_roles}"
        )
        if role not in allowed_roles:
            raise HTTPException(status_code=400, detail="Invalid role")

        # Check if user already exists
        if get_user_by_username(username):
            raise HTTPException(status_code=400, detail="Username already exists")

        # Auto-generate reg_no if not provided
        if not reg_no or reg_no.strip() == "":
            if role == "hod":
                prefix = "HOD"
            elif role == "staff":
                prefix = "STAFF"
            elif role == "principal":
                prefix = "PRINCIPAL"
            elif role == "vice_chancellor":
                prefix = "VC"
            elif role == "director":
                prefix = "DIR"
            elif role == "dean":
                prefix = "DEAN"
            else:
                prefix = "USR"

            cursor.execute("SELECT COUNT(*) FROM users WHERE role = ?", (role,))
            count = cursor.fetchone()[0]
            reg_no = f"{prefix}_{str(count + 1).zfill(4)}"

            while get_user_by_reg_no(reg_no):
                count += 1
                reg_no = f"{prefix}_{str(count).zfill(4)}"
        else:
            if get_user_by_reg_no(reg_no):
                raise HTTPException(
                    status_code=400, detail="Registration number already exists"
                )

        # Hash password
        password_hash = hash_password(password)

        # Insert user
        cursor.execute(
            """
            INSERT INTO users (username, password_hash, reg_no, name, dept, role, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
            (username, password_hash, reg_no, name, dept, role, created_by),
        )
        conn.commit()

        return {
            "message": "User created successfully",
            "user": {
                "username": username,
                "reg_no": reg_no,
                "name": name,
                "dept": dept,
                "role": role,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Create user error: {e}")
        raise HTTPException(status_code=500, detail="Failed to create user")


@app.get("/users")
async def get_users(role: str = None):
    """Get all users or filter by role - optimized with indexes"""
    try:
        if role:
            # Uses idx_users_role index
            cursor.execute(
                "SELECT id, username, reg_no, name, dept, role, created_at FROM users WHERE role = ? ORDER BY name",
                (role,),
            )
        else:
            # Uses default ordering by creation time, can be optimized further if needed
            cursor.execute(
                "SELECT id, username, reg_no, name, dept, role, created_at FROM users ORDER BY created_at DESC"
            )

        rows = cursor.fetchall()
        users_list = []
        for row in rows:
            users_list.append(
                {
                    "id": row[0],
                    "username": row[1],
                    "regNo": row[2],
                    "name": row[3],
                    "dept": row[4],
                    "role": row[5],
                    "createdAt": row[6],
                }
            )
        return {"users": users_list}
    except Exception as e:
        print(f"Error fetching users: {e}")
        return {"users": []}


# -------------------------------------------------
# DATABASE - PostgreSQL with connection pooling
# -------------------------------------------------
# pg_adapter.cursor is a thread-local cursor that auto-manages connections
conn = pg_adapter.cursor
cursor = pg_adapter.cursor

# System configuration table and helper functions
cursor.execute("""
    CREATE TABLE IF NOT EXISTS system_config (
        key VARCHAR(255) PRIMARY KEY,
        value TEXT NOT NULL,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
""")

def save_system_config(key: str, value: str):
    try:
        cursor.execute("SELECT 1 FROM system_config WHERE key = ?", (key,))
        exists = cursor.fetchone()
        if exists:
            cursor.execute("UPDATE system_config SET value = ?, updated_at = CURRENT_TIMESTAMP WHERE key = ?", (value, key))
        else:
            cursor.execute("INSERT INTO system_config (key, value) VALUES (?, ?)", (key, value))
    except Exception as e:
        print(f"Error saving config to database: {e}")

def load_system_config():
    try:
        cursor.execute("SELECT key, value FROM system_config")
        rows = cursor.fetchall()
        for key, value in rows:
            if value.lower() == "true":
                _app_settings[key] = True
            elif value.lower() == "false":
                _app_settings[key] = False
            else:
                _app_settings[key] = value
        print("Loaded settings from database:", _app_settings)
    except Exception as e:
        print(f"Error loading config from database: {e}")


def _load_academic_settings_from_storage(force=False):
    """Load academic year configuration from system_config into in-memory dict.
    
    Uses a timestamp-based refresh gate so the DB is not queried on every request.
    Pass force=True to bypass the gate (e.g. after saving new settings).
    """
    global _academic_settings, _academic_settings_last_refresh
    now = time.time()
    if not force and (now - _academic_settings_last_refresh) < _ACADEMIC_SETTINGS_REFRESH_SECONDS:
        return

    try:
        cursor.execute(
            "SELECT value FROM system_config WHERE key = ?",
            ("academic_settings",),
        )
        row = cursor.fetchone()
        if row and row[0]:
            data = json.loads(row[0])
            _academic_settings = dict(data)
        _academic_settings_last_refresh = now
    except Exception as e:
        print(f"Error loading academic settings from database: {e}")


def _save_academic_settings_to_storage():
    """Persist academic year configuration to system_config."""
    save_system_config("academic_settings", json.dumps(_academic_settings))
    _load_academic_settings_from_storage(force=True)


def _parse_date_or_none(value):
    if not value:
        return None
    if isinstance(value, str):
        return datetime.strptime(value, "%Y-%m-%d").date()
    return value


def _build_academic_holidays(start_date, end_date):
    """Return holiday rows between start and end dates with Sunday defaults."""
    if not start_date or not end_date or start_date > end_date:
        return []

    overrides = _academic_settings.get("holiday_overrides", {}) or {}
    rows = []
    current = start_date
    while current <= end_date:
        date_str = current.isoformat()
        is_sunday = current.weekday() == 6
        override = overrides.get(date_str, {})
        override_status = (override.get("status") or "").lower()
        if override_status == "working_day":
            effective_status = "working_day"
        elif override_status == "holiday":
            effective_status = "holiday"
        else:
            effective_status = "holiday" if is_sunday else "working_day"

        rows.append(
            {
                "date": date_str,
                "is_sunday": is_sunday,
                "status": effective_status,
                "effective_status": effective_status,
                "reason": override.get("reason", ""),
                "source": "override" if date_str in overrides else ("sunday_default" if is_sunday else "working_day"),
            }
        )
        current += timedelta(days=1)
    return rows


def _get_academic_ranges():
    """Return list of {start, end} dicts from academic settings.
    Falls back to single academic_year_start/end if set.
    """
    _load_academic_settings_from_storage()
    ranges = _academic_settings.get("academic_ranges", []) or []

    if ranges:
        return ranges
    start = _academic_settings.get("academic_year_start")
    end = _academic_settings.get("academic_year_end")
    if start and end:
        return [{"start": start, "end": end}]
    return []


def _is_in_academic_ranges(date_str: str) -> bool:
    """Check if a date string falls within any academic range."""
    ranges = _get_academic_ranges()
    if not ranges:
        return True  # no ranges = always valid
    for r in ranges:
        if r["start"] <= date_str <= r["end"]:
            return True
    return False


def _get_academic_date_range():
    """Return (earliest_start_str, latest_end_str) from all academic ranges."""
    ranges = _get_academic_ranges()
    if not ranges:
        return None, None
    starts = [r["start"] for r in ranges]
    ends = [r["end"] for r in ranges]
    return min(starts), max(ends)


def _clamp_to_academic_year(start_str: str, end_str: str):
    """Clamp date range to academic year boundaries if set."""
    acad_start, acad_end = _get_academic_date_range()
    if not acad_start or not acad_end:
        return start_str, end_str
    if start_str < acad_start:
        start_str = acad_start
    if end_str > acad_end:
        end_str = acad_end
    return start_str, end_str


def _cap_end_to_today(end_dt: datetime):
    """Cap end datetime to today to avoid counting future days in absent calculation."""
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    return min(end_dt, today)


def _get_holiday_dates_in_range(start_str: str, end_str: str) -> set:
    """Return set of date strings that are holidays within the given range."""
    holidays = set()
    start_dt = datetime.strptime(start_str, "%Y-%m-%d")
    end_dt = datetime.strptime(end_str, "%Y-%m-%d")
    d = start_dt
    while d <= end_dt:
        date_str = d.strftime("%Y-%m-%d")
        _, _, is_holiday = _academic_status_for_date(date_str)
        if is_holiday:
            holidays.add(date_str)
        d += timedelta(days=1)
    return holidays


def _get_full_attendance_dates(
    reg_no: str, start_date: str, end_date: str, is_other_staff: bool = False
) -> set:
    """Return set of date strings where user is counted as present."""
    hd = _get_half_day_settings()
    if hd["half_day_enabled"]:
        # In half-day mode, rely on daily_attendance_status records
        cursor.execute(
            """
            SELECT date FROM daily_attendance_status
            WHERE reg_no = %s AND status IN ('Present', 'Half Day')
              AND date >= %s AND date <= %s
        """,
            (reg_no, start_date, end_date),
        )
        return {str(r[0])[:10] for r in cursor.fetchall()}

    table = "other_staff_attendance" if is_other_staff else "attendance"
    cursor.execute(
        f"""
        SELECT DATE(timestamp) AS d,
               COUNT(CASE WHEN status = 'check_in' THEN 1 END) as ins,
               COUNT(CASE WHEN status = 'check_out' THEN 1 END) as outs
        FROM {table}
        WHERE reg_no = %s
          AND DATE(timestamp) >= %s
          AND DATE(timestamp) <= %s
        GROUP BY DATE(timestamp)
    """,
        (reg_no, start_date, end_date),
    )
    result = set()
    current_date_str = datetime.now().strftime("%Y-%m-%d")
    for row in cursor.fetchall():
        d, ins, outs = row[0], row[1] or 0, row[2] or 0
        if hasattr(d, "strftime"):
            date_str = d.strftime("%Y-%m-%d")
        else:
            date_str = str(d)[:10]

        if date_str == current_date_str:
            if ins > 0 or outs > 0:
                result.add(date_str)
        else:
            if ins > 0 and outs > 0:
                result.add(date_str)

    # Exclude any dates explicitly marked 'Absent' in daily_attendance_status
    cursor.execute(
        """
        SELECT date FROM daily_attendance_status
        WHERE reg_no = %s AND status = 'Absent' AND date >= %s AND date <= %s
    """,
        (reg_no, start_date, end_date),
    )
    absent_dates = {str(r[0])[:10] for r in cursor.fetchall()}
    return result - absent_dates



def _academic_status_for_date(date_str: str) -> tuple[str, str | None, bool]:
    """Return effective academic status for a date.

    Returns:
        status, reason, is_holiday
    """
    _load_academic_settings_from_storage()

    try:
        current_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return "working_day", None, False

    if not _is_in_academic_ranges(date_str):
        return "outside_academic_year", "Outside academic year", True

    overrides = _academic_settings.get("holiday_overrides", {}) or {}
    override = overrides.get(date_str, {})
    override_status = str(override.get("status", "")).lower()
    if override_status == "working_day":
        return "working_day", override.get("reason"), False
    if override_status == "holiday":
        return "holiday", override.get("reason"), True
    if current_date.weekday() == 6:
        return "holiday", None, True
    return "working_day", None, False


def _ensure_daily_holiday_status(reg_no: str, name: str, dept: str, date_str: str, reason: str | None = None):
    """Persist a holiday status for a specific date."""
    cursor.execute(
        """
        INSERT INTO daily_attendance_status
        (reg_no, name, dept, date, status, leave_type, marked_by, marked_at)
        VALUES (?, ?, ?, ?, 'Holiday', NULL, 'Academics Calendar', CURRENT_TIMESTAMP)
        ON CONFLICT (reg_no, date) DO UPDATE SET
            status = 'Holiday',
            leave_type = NULL,
            leave_request_id = NULL,
            marked_by = 'Academics Calendar',
            marked_at = CURRENT_TIMESTAMP
    """,
        (reg_no, name, dept, date_str),
    )
    if reason:
        log_audit_event("ACADEMIC_HOLIDAY_APPLIED", reg_no, True, f"{date_str}: {reason}")


_load_academic_settings_from_storage()

# Load configuration values from database to overwrite/populate default _app_settings
load_system_config()


# ─────────────────────────────────────────────────────────────────────────────
# HALF-DAY ATTENDANCE SPLIT — Helper functions
# ─────────────────────────────────────────────────────────────────────────────

_leave_settings_cache = {}
_leave_settings_last_refresh = 0
_LEAVE_SETTINGS_REFRESH_SECONDS = 60


def _get_leave_settings(force: bool = False) -> dict:
    """Load all leave_settings rows into a dict (cached, refreshed every 60s)."""
    global _leave_settings_cache, _leave_settings_last_refresh
    now = time.time()
    if not force and (now - _leave_settings_last_refresh) < _LEAVE_SETTINGS_REFRESH_SECONDS:
        return _leave_settings_cache
    try:
        cursor.execute("SELECT key, value FROM leave_settings")
        rows = cursor.fetchall()
        _leave_settings_cache = {r[0]: r[1] for r in rows} if rows else {}
        _leave_settings_last_refresh = now
    except Exception as e:
        print(f"[leave_settings] Error loading: {e}")
    return _leave_settings_cache


def _get_half_day_settings() -> dict:
    """Return half-day config dict with typed values.
    
    half_day_enabled is auto-detected: True if any enabled slot in
    attendance_duration_settings has slot_half = 'first_half' or 'second_half'.
    """
    s = _get_leave_settings()
    default_3m = (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d")

    # Auto-detect half_day_enabled from slot configuration
    half_day_from_settings = str(s.get("half_day_enabled", "false")).lower() == "true"
    half_day_from_slots = False
    try:
        cursor.execute(
            """
            SELECT COUNT(*) FROM attendance_duration_settings
            WHERE is_enabled = 1 AND slot_half IN ('first_half', 'second_half')
            """
        )
        row = cursor.fetchone()
        half_day_from_slots = (row[0] > 0) if row else False
    except Exception:
        pass

    return {
        "half_day_enabled": half_day_from_settings or half_day_from_slots,
        "first_half_label": s.get("first_half_label", "First Half"),
        "second_half_label": s.get("second_half_label", "Second Half"),
        "cl_monthly_allocation": int(s.get("cl_monthly_allocation", "1") or "1"),
        "cl_expiry_date": s.get("cl_expiry_date") or default_3m,
        "el_expiry_date": s.get("el_expiry_date") or default_3m,
        "auto_expand_checkin_enabled": str(s.get("auto_expand_checkin_enabled", "true")).lower() == "true",
        "auto_expand_checkout_enabled": str(s.get("auto_expand_checkout_enabled", "true")).lower() == "true",
        "auto_expand_fn_minutes": int(s.get("auto_expand_fn_minutes", s.get("auto_expand_minutes", "5")) or "5"),
        "auto_expand_an_minutes": int(s.get("auto_expand_an_minutes", s.get("auto_expand_minutes", "5")) or "5"),
        "auto_expand_minutes": int(s.get("auto_expand_minutes", "5") or "5"),
        "require_fn_check_out": str(s.get("require_fn_check_out", "false")).lower() == "true",
        "first_half_start": s.get("first_half_start", "08:30"),
        "first_half_end": s.get("first_half_end", "13:00"),
        "second_half_start": s.get("second_half_start", "13:00"),
        "second_half_end": s.get("second_half_end", "17:30"),
    }



def _save_leave_setting(key: str, value: str, updated_by: str = "system"):
    """Upsert a single key in leave_settings and invalidate cache."""
    global _leave_settings_last_refresh
    cursor.execute("SELECT 1 FROM leave_settings WHERE key = %s", (key,))
    if cursor.fetchone():
        cursor.execute(
            "UPDATE leave_settings SET value = %s, updated_at = CURRENT_TIMESTAMP, updated_by = %s WHERE key = %s",
            (value, updated_by, key),
        )
    else:
        cursor.execute(
            "INSERT INTO leave_settings (key, value, updated_by) VALUES (%s, %s, %s)",
            (key, value, updated_by),
        )
    _leave_settings_last_refresh = 0  # invalidate cache


def _compute_daily_status(
    first_half: str | None,
    second_half: str | None,
    target_date: str | None = None,
    current_dt: datetime | None = None,
) -> str:
    """Compute overall daily status from first and second half statuses.

    Rules:
    - Both Present                                → 'Present'
    - One Present + one Pending (window active)   → 'Half Day' (FN/AN Present • pending other)
    - One Present + one Absent (window passed)    → 'Half Day'
    - One Present + one Leave                     → 'Half Day'
    - Both Leave                                  → 'Leave'
    - Both Absent                                 → 'Absent'
    - Both Pending (window active/future date)    → 'Pending'
    """
    if current_dt is None:
        current_dt = datetime.now()

    today_str = current_dt.strftime("%Y-%m-%d")
    hd_settings = _get_half_day_settings()

    # Determine if session windows have passed
    fh_passed = True
    sh_passed = True

    if target_date:
        if target_date > today_str:
            fh_passed = False
            sh_passed = False
        elif target_date == today_str:
            now_time = current_dt.time()
            try:
                fh_end_time = datetime.strptime(
                    hd_settings.get("first_half_end", "13:00"), "%H:%M"
                ).time()
                sh_end_time = datetime.strptime(
                    hd_settings.get("second_half_end", "17:30"), "%H:%M"
                ).time()
                fh_passed = now_time > fh_end_time
                sh_passed = now_time > sh_end_time
            except Exception:
                cutoff = datetime.strptime("13:00", "%H:%M").time()
                fh_passed = now_time > cutoff
                sh_passed = now_time > cutoff

    # Evaluate un-scanned sessions (None or Pending)
    if not first_half or first_half == "Pending":
        fh = "Absent" if fh_passed else "Pending"
    else:
        fh = first_half.strip()

    if not second_half or second_half == "Pending":
        sh = "Absent" if sh_passed else "Pending"
    else:
        sh = second_half.strip()

    pair = (fh, sh)

    if pair == ("Present", "Present"):
        return "Present"
    if "Present" in pair:
        return "Half Day"
    if pair == ("Leave", "Leave"):
        return "Leave"
    if "Leave" in pair:
        return "Half Day"
    if pair == ("Absent", "Absent"):
        return "Absent"
    if "Absent" in pair:
        return "Half Day"

    # Both Pending
    return "Pending"




def _compute_half_day_value(status: str) -> float:
    """Return numeric day-equivalent for a daily status (for stats)."""
    if status == "Present":
        return 1.0
    if status == "Half Day":
        return 0.5
    return 0.0


def _compute_attendance_value_from_halves(first_half: str | None, second_half: str | None) -> float:
    """Calculate attendance value (0.0, 0.5, or 1.0) based on half-day statuses.
    
    Args:
        first_half: Status of first half ('Present', 'Absent', 'Leave', or None)
        second_half: Status of second half ('Present', 'Absent', 'Leave', or None)
    
    Returns:
        float: Attendance value (1.0 for both present, 0.5 for one present, 0.0 otherwise)
    """
    fh = (first_half or "Absent").strip()
    sh = (second_half or "Absent").strip()
    
    value = 0.0
    if fh == "Present":
        value += 0.5
    if sh == "Present":
        value += 0.5
    
    return value


def _format_scan_status(status_val: str | None, timestamp_val) -> str:
    """Format the scan status to show if it is first half or second half check in/out."""
    if not status_val:
        return "Present"
    
    # Extract action based on raw DB status value
    sv = status_val.lower().strip()
    if sv == "check_out":
        action = "Check-out"
    elif sv == "check_in":
        action = "Check-in"
    else:
        # Unknown raw value – just title-case it
        return status_val.replace('_', ' ').title()
    
    # Try to determine half based on time (cutoff at 13:00 = 1 PM)
    try:
        if isinstance(timestamp_val, str):
            # Normalise: take first 19 chars, replace T with space
            ts_norm = timestamp_val[:19].replace("T", " ")
            dt = datetime.strptime(ts_norm, "%Y-%m-%d %H:%M:%S")
        else:
            dt = timestamp_val
        
        if dt.hour < 13:
            return f"First Half {action}"
        else:
            return f"Second Half {action}"
    except Exception:
        # Fallback — still return the correct action label
        return action


def _mark_halves_absent_for_day(date_str: str):
    """End-of-day job: for every user, set NULL halves to 'Absent' and recompute status.

    Should be called after the second-half window closes for that date.
    Only runs when half_day_enabled = True.
    """
    hd = _get_half_day_settings()
    if not hd["half_day_enabled"]:
        return

    try:
        # Find all users who have an attendance record that is still open (half status NULL)
        cursor.execute(
            """
            SELECT reg_no, name, dept, first_half_status, second_half_status
            FROM daily_attendance_status
            WHERE date = %s
              AND (first_half_status IS NULL OR second_half_status IS NULL)
        """,
            (date_str,),
        )
        rows = cursor.fetchall()
        updated = 0
        for row in rows:
            reg_no, name, dept, fh, sh = row
            new_fh = fh if fh else "Absent"
            new_sh = sh if sh else "Absent"
            new_status = _compute_daily_status(new_fh, new_sh)
            # Calculate attendance value
            attendance_value = _compute_attendance_value_from_halves(new_fh, new_sh)
            
            cursor.execute(
                """
                UPDATE daily_attendance_status
                SET first_half_status = %s,
                    second_half_status = %s,
                    status = %s,
                    attendance_value = %s,
                    marked_at = CURRENT_TIMESTAMP
                WHERE reg_no = %s AND date = %s
            """,
                (new_fh, new_sh, new_status, attendance_value, reg_no, date_str),
            )
            updated += 1

        # Also mark users with NO record at all as Absent (insert rows)
        cursor.execute(
            """
            SELECT u.reg_no, u.name, u.dept
            FROM users u
            WHERE NOT EXISTS (
                SELECT 1 FROM daily_attendance_status das
                WHERE das.reg_no = u.reg_no AND das.date = %s
            )
              AND u.role IN ('staff', 'hod')
        """,
            (date_str,),
        )
        missing = cursor.fetchall()
        for reg_no, name, dept in missing:
            cursor.execute(
                """
                INSERT INTO daily_attendance_status
                  (reg_no, name, dept, date, status, first_half_status, second_half_status, attendance_value, marked_by, marked_at)
                VALUES (%s, %s, %s, %s, 'Absent', 'Absent', 'Absent', 0.0, 'System EOD', CURRENT_TIMESTAMP)
                ON CONFLICT (reg_no, date) DO NOTHING
            """,
                (reg_no, name, dept, date_str),
            )
            updated += 1

        conn.commit()
        print(f"[EOD] Half-day absent marking done for {date_str}: {updated} records updated.")
    except Exception as e:
        print(f"[EOD] Error in _mark_halves_absent_for_day({date_str}): {e}")


def _check_and_expire_leaves(triggered_by: str = "system"):
    """Check CL/EL expiry dates; zero out balances if today >= expiry date."""
    hd = _get_leave_settings(force=True)
    today_str = datetime.now().strftime("%Y-%m-%d")
    results = {"cl_expired": 0, "el_expired": 0}

    # CL expiry
    cl_expiry = hd.get("cl_expiry_date") or ""
    if cl_expiry and today_str >= cl_expiry:
        try:
            cursor.execute(
                """
                UPDATE casual_leave
                SET current_month_cl_available = 0,
                    accumulated_cl = 0,
                    last_updated = CURRENT_TIMESTAMP
                WHERE current_month_cl_available > 0 OR accumulated_cl > 0
            """
            )
            results["cl_expired"] = cursor.rowcount or 0
            conn.commit()
            log_audit_event("CL_EXPIRED", triggered_by, True, f"CL expired on {today_str} (expiry: {cl_expiry})")
            print(f"[Leave Expiry] CL expired for {results['cl_expired']} records.")
        except Exception as e:
            print(f"[Leave Expiry] CL expiry error: {e}")

    # EL expiry
    el_expiry = hd.get("el_expiry_date") or ""
    if el_expiry and today_str >= el_expiry:
        try:
            cursor.execute(
                """
                UPDATE earned_leave
                SET balance = 0.0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE balance > 0
            """
            )
            results["el_expired"] = cursor.rowcount or 0
            conn.commit()
            log_audit_event("EL_EXPIRED", triggered_by, True, f"EL expired on {today_str} (expiry: {el_expiry})")
            print(f"[Leave Expiry] EL expired for {results['el_expired']} records.")
        except Exception as e:
            print(f"[Leave Expiry] EL expiry error: {e}")

    return results


def _init_db_schema():
    """Initialize database schema with multi-worker safety.

    Uses PostgreSQL advisory lock so only one worker runs DDL.
    Other workers skip since tables/indexes use IF NOT EXISTS.
    """

    lock_acquired = False
    try:
        cursor.execute("SELECT pg_try_advisory_lock(123456789)")
        lock_acquired = cursor.fetchone()[0]
    except Exception:
        pass

    if not lock_acquired:
        print("Schema init skipped (another worker is handling it)")
        return

    try:
        _run_ddl()
    except Exception as e:
        if "tuple concurrently updated" in str(e):
            print(f"DDL concurrency conflict (non-critical): {e}")
        else:
            raise
    finally:
        try:
            cursor.execute("SELECT pg_advisory_unlock(123456789)")
        except Exception:
            pass


def _run_ddl():
    """Execute all DDL statements (tables, indexes, materialized views)."""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS user_location_logs (
            id SERIAL PRIMARY KEY,
            reg_no VARCHAR(64) NOT NULL,
            username VARCHAR(120),
            name VARCHAR(160),
            dept VARCHAR(160),
            role VARCHAR(80),
            latitude DECIMAL(10, 8) NOT NULL,
            longitude DECIMAL(11, 8) NOT NULL,
            accuracy_meters DECIMAL(10, 2),
            speed_mps DECIMAL(10, 2),
            heading_deg DECIMAL(10, 2),
            altitude_m DECIMAL(10, 2),
            source VARCHAR(50) DEFAULT 'gps',
            app_state VARCHAR(20) DEFAULT 'foreground',
            is_mocked BOOLEAN DEFAULT FALSE,
            device_id VARCHAR(120),
            captured_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            server_received_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_user_location_logs_reg_time ON user_location_logs (reg_no, captured_at DESC)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_user_location_logs_server_time ON user_location_logs (server_received_at DESC)")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS user_latest_locations (
            reg_no VARCHAR(64) PRIMARY KEY,
            username VARCHAR(120),
            name VARCHAR(160),
            dept VARCHAR(160),
            role VARCHAR(80),
            latitude DECIMAL(10, 8) NOT NULL,
            longitude DECIMAL(11, 8) NOT NULL,
            accuracy_meters DECIMAL(10, 2),
            speed_mps DECIMAL(10, 2),
            heading_deg DECIMAL(10, 2),
            altitude_m DECIMAL(10, 2),
            source VARCHAR(50) DEFAULT 'gps',
            app_state VARCHAR(20) DEFAULT 'foreground',
            is_mocked BOOLEAN DEFAULT FALSE,
            device_id VARCHAR(120),
            captured_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_seen_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            force_update_requested BOOLEAN DEFAULT FALSE
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_user_latest_locations_last_seen ON user_latest_locations (last_seen_at DESC)")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS other_staff (
            id SERIAL PRIMARY KEY,
            username VARCHAR(120) UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            reg_no VARCHAR(64) UNIQUE NOT NULL,
            name VARCHAR(160) NOT NULL,
            dob DATE,
            role VARCHAR(80) NOT NULL,
            dept VARCHAR(160) NOT NULL,
            embedding BYTEA,
            can_reregister BOOLEAN DEFAULT FALSE,
            current_device_id VARCHAR(255),
            suspended BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by VARCHAR(120)
        )
    """)
    for idx in [
        "CREATE INDEX IF NOT EXISTS idx_other_staff_reg_no ON other_staff (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_username ON other_staff (username)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_role ON other_staff (role)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_dept ON other_staff (dept)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_reg_no_lower ON other_staff (LOWER(reg_no))",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_dept_role ON other_staff (dept, role)",
    ]:
        cursor.execute(idx)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS other_staff_attendance (
            id SERIAL PRIMARY KEY,
            reg_no VARCHAR(64) NOT NULL,
            name VARCHAR(160) NOT NULL,
            dept VARCHAR(160) NOT NULL,
            role VARCHAR(80) NOT NULL,
            "timestamp" TIMESTAMP NOT NULL,
            status VARCHAR(20) DEFAULT 'check_in',
            device_id VARCHAR(120),
            location VARCHAR(255)
        )
    """)
    for idx in [
        "CREATE INDEX IF NOT EXISTS idx_other_staff_attendance_reg_no ON other_staff_attendance (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_attendance_timestamp ON other_staff_attendance (timestamp DESC)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_attendance_reg_no_timestamp ON other_staff_attendance (reg_no, timestamp DESC)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_attendance_dept_timestamp ON other_staff_attendance (dept, timestamp DESC)",
    ]:
        cursor.execute(idx)

    # ── MORNING AND EVENING ATTENDANCE TABLES ─────────────────────────
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS morning_attendance (
            id SERIAL PRIMARY KEY,
            reg_no VARCHAR(64) NOT NULL,
            name VARCHAR(160) NOT NULL,
            dept VARCHAR(160) NOT NULL,
            date DATE NOT NULL,
            in_time TIME,
            status VARCHAR(20) DEFAULT 'Present',
            attendance_value DECIMAL(3,2) DEFAULT 0.5,
            marked_by VARCHAR(120),
            marked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(reg_no, date)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS evening_attendance (
            id SERIAL PRIMARY KEY,
            reg_no VARCHAR(64) NOT NULL,
            name VARCHAR(160) NOT NULL,
            dept VARCHAR(160) NOT NULL,
            date DATE NOT NULL,
            in_time TIME,
            status VARCHAR(20) DEFAULT 'Present',
            attendance_value DECIMAL(3,2) DEFAULT 0.5,
            marked_by VARCHAR(120),
            marked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(reg_no, date)
        )
    """)
    for idx in [
        "CREATE INDEX IF NOT EXISTS idx_morning_attendance_reg_no ON morning_attendance (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_morning_attendance_date ON morning_attendance (date DESC)",
        "CREATE INDEX IF NOT EXISTS idx_evening_attendance_reg_no ON evening_attendance (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_evening_attendance_date ON evening_attendance (date DESC)",
    ]:
        cursor.execute(idx)

    # Migration columns
    for col_sql in [
        "ALTER TABLE user_latest_locations ADD COLUMN IF NOT EXISTS boundary_warning BOOLEAN DEFAULT FALSE",
        "ALTER TABLE user_latest_locations ADD COLUMN IF NOT EXISTS warning_message VARCHAR(255)",
        "ALTER TABLE user_latest_locations ADD COLUMN IF NOT EXISTS first_left_boundary_at TIMESTAMP",
        "ALTER TABLE user_latest_locations ADD COLUMN IF NOT EXISTS force_update_requested BOOLEAN DEFAULT FALSE",
        "ALTER TABLE user_location_logs ADD COLUMN IF NOT EXISTS boundary_warning BOOLEAN DEFAULT FALSE",
        "ALTER TABLE user_location_logs ADD COLUMN IF NOT EXISTS warning_message VARCHAR(255)",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS current_device_id VARCHAR(255)",
        "ALTER TABLE other_staff ADD COLUMN IF NOT EXISTS current_device_id VARCHAR(255)",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS suspended BOOLEAN DEFAULT FALSE",
        "ALTER TABLE other_staff ADD COLUMN IF NOT EXISTS suspended BOOLEAN DEFAULT FALSE",
        "ALTER TABLE attendance_duration_settings ADD COLUMN IF NOT EXISTS slot_type VARCHAR(20) DEFAULT 'check_in'",
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'check_in'",
        "ALTER TABLE other_staff_attendance ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'check_in'",
        "ALTER TABLE daily_attendance_status ADD COLUMN IF NOT EXISTS absent_reason VARCHAR(255)",
    ]:
        try:
            cursor.execute(col_sql)
        except Exception:
            pass

    print("Creating performance indexes...")
    for idx in [
        "CREATE INDEX IF NOT EXISTS idx_users_reg_no ON users (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_users_username ON users (username)",
        "CREATE INDEX IF NOT EXISTS idx_users_role ON users (role)",
        "CREATE INDEX IF NOT EXISTS idx_users_dept ON users (dept)",
        "CREATE INDEX IF NOT EXISTS idx_users_reg_no_lower ON users (LOWER(reg_no))",
        "CREATE INDEX IF NOT EXISTS idx_users_dept_role ON users (dept, role)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_reg_no ON other_staff (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_username ON other_staff (username)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_role ON other_staff (role)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_dept ON other_staff (dept)",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_reg_no_lower ON other_staff (LOWER(reg_no))",
        "CREATE INDEX IF NOT EXISTS idx_other_staff_dept_role ON other_staff (dept, role)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_reg_no ON attendance (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_timestamp ON attendance (timestamp DESC)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_dept ON attendance (dept)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_reg_no_timestamp ON attendance (reg_no, timestamp DESC)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_dept_timestamp ON attendance (dept, timestamp DESC)",
        "CREATE INDEX IF NOT EXISTS idx_daily_attendance_status_reg_no ON daily_attendance_status (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_daily_attendance_status_date ON daily_attendance_status (date DESC)",
        "CREATE INDEX IF NOT EXISTS idx_daily_attendance_status_reg_no_date ON daily_attendance_status (reg_no, date DESC)",
        "CREATE INDEX IF NOT EXISTS idx_daily_attendance_status_dept_date ON daily_attendance_status (dept, date DESC)",
        "CREATE INDEX IF NOT EXISTS idx_face_embedding_samples_reg_no ON face_embedding_samples (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_face_embedding_samples_source ON face_embedding_samples (source_table)",
        "CREATE INDEX IF NOT EXISTS idx_face_embedding_samples_reg_no_created ON face_embedding_samples (reg_no, created_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_casual_leave_reg_no ON casual_leave (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_casual_leave_current_month ON casual_leave (current_month)",
        "CREATE INDEX IF NOT EXISTS idx_casual_leave_reg_no_month ON casual_leave (reg_no, current_month)",
        "CREATE INDEX IF NOT EXISTS idx_user_location_logs_reg_no ON user_location_logs (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_user_location_logs_captured_at ON user_location_logs (captured_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_user_location_logs_reg_no_captured ON user_location_logs (reg_no, captured_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_user_latest_locations_reg_no ON user_latest_locations (reg_no)",
        "CREATE INDEX IF NOT EXISTS idx_user_latest_locations_dept ON user_latest_locations (dept)",
    ]:
        cursor.execute(idx)

    cursor.execute("""
        CREATE MATERIALIZED VIEW IF NOT EXISTS mv_attendance_summary AS
        SELECT DATE(timestamp) as attendance_date, dept,
               COUNT(*) as total_attendance, COUNT(DISTINCT reg_no) as unique_users
        FROM attendance
        WHERE timestamp >= CURRENT_DATE - INTERVAL '90 days'
        GROUP BY DATE(timestamp), dept
        ORDER BY attendance_date DESC, dept
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_mv_attendance_summary_date_dept ON mv_attendance_summary (attendance_date DESC, dept)")
    cursor.execute("""
        CREATE OR REPLACE FUNCTION refresh_attendance_summary()
        RETURNS void AS $$
        BEGIN REFRESH MATERIALIZED VIEW mv_attendance_summary; END;
        $$ LANGUAGE plpgsql;
    """)

    # New CCL Feature tables
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ccl_settings (
            key VARCHAR(50) PRIMARY KEY,
            value VARCHAR(255)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ccl_custom_dates (
            id SERIAL PRIMARY KEY,
            ccl_date DATE UNIQUE NOT NULL,
            early_enabled BOOLEAN DEFAULT FALSE,
            early_start VARCHAR(5) DEFAULT '07:00',
            early_end VARCHAR(5) DEFAULT '08:00',
            early_duration INT DEFAULT 60,
            late_enabled BOOLEAN DEFAULT FALSE,
            late_start VARCHAR(5) DEFAULT '17:00',
            late_end VARCHAR(5) DEFAULT '18:00',
            late_duration INT DEFAULT 60
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS earned_leave (
            reg_no VARCHAR(64) PRIMARY KEY,
            user_name VARCHAR(160) NOT NULL,
            dept VARCHAR(160) NOT NULL,
            role VARCHAR(80) NOT NULL,
            balance DECIMAL(10, 2) DEFAULT 0.0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ccl_earned_history (
            id SERIAL PRIMARY KEY,
            reg_no VARCHAR(64) NOT NULL,
            name VARCHAR(160) NOT NULL,
            dept VARCHAR(160) NOT NULL,
            date DATE NOT NULL,
            time TIME NOT NULL,
            slot_type VARCHAR(30) NOT NULL,
            earned_points DECIMAL(10, 2) DEFAULT 1.0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ccl_earned_history_reg_no ON ccl_earned_history (reg_no)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ccl_earned_history_date ON ccl_earned_history (date DESC)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_earned_leave_dept ON earned_leave (dept)")

    # Initialize default CCL configurations
    cursor.execute("SELECT 1 FROM ccl_settings WHERE key = 'early_check_in_ccl_enabled'")
    if not cursor.fetchone():
        cursor.execute("INSERT INTO ccl_settings (key, value) VALUES ('early_check_in_ccl_enabled', 'false')")
        cursor.execute("INSERT INTO ccl_settings (key, value) VALUES ('late_check_out_ccl_enabled', 'false')")
        cursor.execute("INSERT INTO ccl_settings (key, value) VALUES ('early_check_in_start', '07:00')")
        cursor.execute("INSERT INTO ccl_settings (key, value) VALUES ('early_check_in_end', '08:00')")
        cursor.execute("INSERT INTO ccl_settings (key, value) VALUES ('late_check_out_start', '17:00')")
        cursor.execute("INSERT INTO ccl_settings (key, value) VALUES ('late_check_out_end', '18:00')")

    # ─────────────────────────────────────────────────────────────
    # HALF-DAY ATTENDANCE SPLIT & LEAVE EXPIRY — Schema additions
    # ─────────────────────────────────────────────────────────────

    # 1. Half-day status columns on daily_attendance_status
    for _col_sql in [
        "ALTER TABLE daily_attendance_status ADD COLUMN IF NOT EXISTS first_half_status VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE daily_attendance_status ADD COLUMN IF NOT EXISTS second_half_status VARCHAR(20) DEFAULT NULL",
        "ALTER TABLE daily_attendance_status ADD COLUMN IF NOT EXISTS first_half_in_time TIME DEFAULT NULL",
        "ALTER TABLE daily_attendance_status ADD COLUMN IF NOT EXISTS first_half_out_time TIME DEFAULT NULL",
        "ALTER TABLE daily_attendance_status ADD COLUMN IF NOT EXISTS second_half_in_time TIME DEFAULT NULL",
        "ALTER TABLE daily_attendance_status ADD COLUMN IF NOT EXISTS second_half_out_time TIME DEFAULT NULL",
        # slot_half on attendance_duration_settings: 'first_half' | 'second_half' | 'full_day'
        "ALTER TABLE attendance_duration_settings ADD COLUMN IF NOT EXISTS slot_half VARCHAR(20) DEFAULT 'full_day'",
        # Half-day fields on leave_requests
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS is_half_day BOOLEAN DEFAULT FALSE",
        "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS which_half VARCHAR(10) DEFAULT NULL",
    ]:
        try:
            cursor.execute(_col_sql)
        except Exception:
            pass

    # 2. leave_settings table — stores CL/EL expiry dates and half-day toggle
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leave_settings (
            key VARCHAR(100) PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_by VARCHAR(120)
        )
    """)

    # 3. expired_leaves table — stores historical logs of expired leaves (CL, EL, etc.)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS expired_leaves (
            id SERIAL PRIMARY KEY,
            reg_no VARCHAR(64) NOT NULL,
            user_name VARCHAR(160) NOT NULL,
            dept VARCHAR(160) NOT NULL,
            role VARCHAR(80) DEFAULT 'staff',
            leave_type VARCHAR(50) NOT NULL,
            expired_amount DECIMAL(10, 2) NOT NULL,
            expiry_date DATE NOT NULL,
            reason VARCHAR(255) DEFAULT 'System automated expiry',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_expired_leaves_reg_no ON expired_leaves (reg_no)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_expired_leaves_dept ON expired_leaves (dept)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_expired_leaves_expiry_date ON expired_leaves (expiry_date DESC)")

    # Initialize default leave settings if not present
    _leave_setting_defaults = {
        "half_day_enabled": "false",
        "first_half_label": "First Half",
        "second_half_label": "Second Half",
        "cl_monthly_allocation": "1",
        "cl_expiry_date": "",   # empty = no expiry
        "el_expiry_date": "",   # empty = no expiry
        "auto_expand_checkin_enabled": "true",  # Default enabled for check-in
        "auto_expand_checkout_enabled": "true", # Default enabled for check-out
        "auto_expand_fn_minutes": "5",          # Separate extension duration for Forenoon (FN)
        "auto_expand_an_minutes": "5",          # Separate extension duration for Afternoon (AN)
        "auto_expand_minutes": "5",             # Default fallback extension
    }
    for _key, _val in _leave_setting_defaults.items():
        cursor.execute("SELECT 1 FROM leave_settings WHERE key = %s", (_key,))
        if not cursor.fetchone():
            cursor.execute(
                "INSERT INTO leave_settings (key, value) VALUES (%s, %s)",
                (_key, _val),
            )

    print("Database indexes and materialized views created successfully")


_init_db_schema()


def _default_outer_coords():
    return [
        (11.040730, 77.073717),
        (11.040865, 77.075121),
        (11.039733, 77.075201),
        (11.039529, 77.075786),
        (11.038500, 77.075892),
        (11.038551, 77.073616),
    ]


def _default_inner_coords():
    return [
        (11.039537, 77.075328),
        (11.039554, 77.075895),
        (11.038858, 77.075912),
        (11.038501, 77.074908),
    ]


def _insert_polygon_group(polygon_type: str, polygon_group: int, coords):
    for i, (lat, lng) in enumerate(coords):
        cursor.execute(
            """
            INSERT INTO geo_fence_coordinates_v2 (polygon_type, polygon_group, latitude, longitude, point_order)
            VALUES (?, ?, ?, ?, ?)
        """,
            (polygon_type, polygon_group, lat, lng, i),
        )


def _load_geo_fence_polygons_from_db():
    """Load grouped polygons from v2 table."""
    grouped = {"outer": {}, "inner": {}, "limit_range": {}}
    cursor.execute("""
        SELECT polygon_type, polygon_group, latitude, longitude, point_order
        FROM geo_fence_coordinates_v2
        ORDER BY polygon_type, polygon_group, point_order
    """)
    for row in cursor.fetchall():
        ptype = row[0]
        pgroup = int(row[1])
        lat = float(row[2])
        lng = float(row[3])
        grouped.setdefault(ptype, {}).setdefault(pgroup, []).append((lat, lng))

    outer = [
        grouped["outer"][k]
        for k in sorted(grouped["outer"].keys())
        if len(grouped["outer"][k]) >= 3
    ]
    inner = [
        grouped["inner"][k]
        for k in sorted(grouped["inner"].keys())
        if len(grouped["inner"][k]) >= 3
    ]
    limit_range = [
        grouped.get("limit_range", {}).get(k, [])
        for k in sorted(grouped.get("limit_range", {}).keys())
        if len(grouped.get("limit_range", {}).get(k, [])) >= 3
    ]
    return outer, inner, limit_range


# Initialize/migrate geofence data into v2 if empty
cursor.execute("SELECT COUNT(*) FROM geo_fence_coordinates_v2")
if cursor.fetchone()[0] == 0:
    migrated = False
    try:
        cursor.execute("""
            SELECT polygon_type, latitude, longitude, point_order
            FROM geo_fence_coordinates
            ORDER BY polygon_type, point_order
        """)
        old_rows = cursor.fetchall()
        if old_rows:
            by_type = {"outer": [], "inner": []}
            for row in old_rows:
                by_type[row[0]].append((float(row[1]), float(row[2])))
            if len(by_type["outer"]) >= 3:
                _insert_polygon_group("outer", 1, by_type["outer"])
            if len(by_type["inner"]) >= 3:
                _insert_polygon_group("inner", 1, by_type["inner"])
            migrated = True
    except Exception:
        migrated = False

    if not migrated:
        _insert_polygon_group("outer", 1, _default_outer_coords())
        _insert_polygon_group("inner", 1, _default_inner_coords())

# Refresh in-memory geofence polygons from database
try:
    _outer, _inner, _limit_range = _load_geo_fence_polygons_from_db()
    if _outer:
        _geo_fence_outer_polygons = _outer
        _geo_fence_polygon = _outer[0]
    if _limit_range:
        _geo_fence_limit_range_polygons = _limit_range
except Exception as e:
    print(f"Failed to load geofence polygons from DB: {e}")

# -------------------------------------------------
# GEO FENCE MANAGEMENT ENDPOINTS
# -------------------------------------------------


@app.get("/admin/geo-fence")
async def get_geo_fence_coordinates(request: Request):
    """Get all geo fence coordinates - Admin only"""
    verify_admin_token(request)
    try:
        outer_polygons, inner_polygons, limit_range_polygons = _load_geo_fence_polygons_from_db()
        outer_coords = (
            [[float(p[0]), float(p[1])] for p in outer_polygons[0]]
            if outer_polygons
            else []
        )
        inner_coords = (
            [[float(p[0]), float(p[1])] for p in inner_polygons[0]]
            if inner_polygons
            else []
        )
        limit_range_coords = (
            [[float(p[0]), float(p[1])] for p in limit_range_polygons[0]]
            if limit_range_polygons
            else []
        )
        outer_polygons_json = [
            [[float(p[0]), float(p[1])] for p in poly] for poly in outer_polygons
        ]
        inner_polygons_json = [
            [[float(p[0]), float(p[1])] for p in poly] for poly in inner_polygons
        ]
        limit_range_polygons_json = [
            [[float(p[0]), float(p[1])] for p in poly] for poly in limit_range_polygons
        ]

        return success_response(
            "Geo fence coordinates retrieved successfully",
            {
                "outer_polygon": outer_coords,
                "inner_polygon": inner_coords,
                "limit_range_polygon": limit_range_coords,
                "outer_polygons": outer_polygons_json,
                "inner_polygons": inner_polygons_json,
                "limit_range_polygons": limit_range_polygons_json,
            },
        )
    except Exception as e:
        print(f"Error getting geo fence coordinates: {e}")
        return error_response("Failed to retrieve geo fence coordinates", "DB_ERROR")


@app.post("/admin/geo-fence")
async def update_geo_fence_coordinates(request: Request):
    """Update geo fence coordinates - Admin only"""
    admin_user = verify_admin_token(request)

    try:
        body = await request.json()
        outer_polygons = body.get("outer_polygons")
        inner_polygons = body.get("inner_polygons")
        limit_range_polygons = body.get("limit_range_polygons")

        if outer_polygons is None:
            outer_polygon = body.get("outer_polygon", [])
            outer_polygons = [outer_polygon] if isinstance(outer_polygon, list) else []
        if inner_polygons is None:
            inner_polygon = body.get("inner_polygon", [])
            if isinstance(inner_polygon, list) and inner_polygon:
                inner_polygons = [inner_polygon]
            else:
                inner_polygons = []
        if limit_range_polygons is None:
            limit_range_polygon = body.get("limit_range_polygon", [])
            if isinstance(limit_range_polygon, list) and limit_range_polygon:
                limit_range_polygons = [limit_range_polygon]
            else:
                limit_range_polygons = []

        if not isinstance(outer_polygons, list) or not outer_polygons:
            return error_response(
                "At least one outer polygon is required", "INVALID_INPUT"
            )
        if not isinstance(inner_polygons, list):
            inner_polygons = []
        if not isinstance(limit_range_polygons, list):
            limit_range_polygons = []

        # Validate coordinate format
        def validate_coordinates(polygon, name):
            for i, point in enumerate(polygon):
                if not isinstance(point, list) or len(point) != 2:
                    return f"{name} point {i} must be [latitude, longitude]"
                try:
                    lat, lng = float(point[0]), float(point[1])
                    if not (-90 <= lat <= 90):
                        return f"{name} point {i} has invalid latitude {lat}"
                    if not (-180 <= lng <= 180):
                        return f"{name} point {i} has invalid longitude {lng}"
                except (ValueError, TypeError):
                    return f"{name} point {i} has invalid coordinate format"
            return None

        for g, polygon in enumerate(outer_polygons):
            if not isinstance(polygon, list) or len(polygon) < 3:
                return error_response(
                    f"Outer polygon {g + 1} must have at least 3 points",
                    "INVALID_INPUT",
                )
            error_msg = validate_coordinates(polygon, f"Outer polygon {g + 1}")
            if error_msg:
                return error_response(error_msg, "INVALID_COORDINATES")

        for g, polygon in enumerate(inner_polygons):
            if not isinstance(polygon, list) or len(polygon) < 3:
                return error_response(
                    f"Inner polygon {g + 1} must have at least 3 points",
                    "INVALID_INPUT",
                )
            error_msg = validate_coordinates(polygon, f"Inner polygon {g + 1}")
            if error_msg:
                return error_response(error_msg, "INVALID_COORDINATES")

        for g, polygon in enumerate(limit_range_polygons):
            if not isinstance(polygon, list) or len(polygon) < 3:
                return error_response(
                    f"Limit range polygon {g + 1} must have at least 3 points",
                    "INVALID_INPUT",
                )
            error_msg = validate_coordinates(polygon, f"Limit range polygon {g + 1}")
            if error_msg:
                return error_response(error_msg, "INVALID_COORDINATES")

        cursor.execute(
            "DELETE FROM geo_fence_coordinates_v2 WHERE polygon_type IN ('outer','inner','limit_range')"
        )
        for group_idx, polygon in enumerate(outer_polygons, start=1):
            for i, point in enumerate(polygon):
                cursor.execute(
                    """
                    INSERT INTO geo_fence_coordinates_v2
                    (polygon_type, polygon_group, latitude, longitude, point_order, updated_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                """,
                    (
                        "outer",
                        group_idx,
                        point[0],
                        point[1],
                        i,
                        admin_user.get("name", "admin"),
                    ),
                )
        for group_idx, polygon in enumerate(inner_polygons, start=1):
            for i, point in enumerate(polygon):
                cursor.execute(
                    """
                    INSERT INTO geo_fence_coordinates_v2
                    (polygon_type, polygon_group, latitude, longitude, point_order, updated_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                """,
                    (
                        "inner",
                        group_idx,
                        point[0],
                        point[1],
                        i,
                        admin_user.get("name", "admin"),
                    ),
                )
        for group_idx, polygon in enumerate(limit_range_polygons, start=1):
            for i, point in enumerate(polygon):
                cursor.execute(
                    """
                    INSERT INTO geo_fence_coordinates_v2
                    (polygon_type, polygon_group, latitude, longitude, point_order, updated_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                """,
                    (
                        "limit_range",
                        group_idx,
                        point[0],
                        point[1],
                        i,
                        admin_user.get("name", "admin"),
                    ),
                )

        global \
            _geo_fence_outer_polygons, \
            _geo_fence_inner_polygons, \
            _geo_fence_limit_range_polygons, \
            _geo_fence_polygon, \
            _geo_fence_inner_polygon
        _geo_fence_outer_polygons = [
            [(float(p[0]), float(p[1])) for p in poly] for poly in outer_polygons
        ]
        _geo_fence_inner_polygons = [
            [(float(p[0]), float(p[1])) for p in poly] for poly in inner_polygons
        ]
        _geo_fence_limit_range_polygons = [
            [(float(p[0]), float(p[1])) for p in poly] for poly in limit_range_polygons
        ]
        _geo_fence_polygon = (
            _geo_fence_outer_polygons[0] if _geo_fence_outer_polygons else []
        )
        _geo_fence_inner_polygon = (
            _geo_fence_inner_polygons[0] if _geo_fence_inner_polygons else []
        )

        log_audit_event(
            "GEO_FENCE_UPDATED",
            admin_user["reg_no"],
            True,
            f"Updated geo fence: {len(outer_polygons)} outer, {len(inner_polygons)} inner, {len(limit_range_polygons)} limit range boundaries",
        )

        return success_response(
            "Geo fence coordinates updated successfully",
            {
                "outer_polygon": outer_polygons[0] if outer_polygons else [],
                "inner_polygon": inner_polygons[0] if inner_polygons else [],
                "limit_range_polygon": limit_range_polygons[0] if limit_range_polygons else [],
                "outer_polygons": outer_polygons,
                "inner_polygons": inner_polygons,
                "limit_range_polygons": limit_range_polygons,
            },
        )

    except Exception as e:
        print(f"Error updating geo fence coordinates: {e}")
        return error_response("Failed to update geo fence coordinates", "UPDATE_ERROR")


@app.get("/geo-fence/public")
async def get_geo_fence_public():
    """Get geo fence coordinates for public access (used by mobile app)"""
    try:
        outer_polygons, inner_polygons, limit_range_polygons = _load_geo_fence_polygons_from_db()
        outer_coords = (
            [[float(p[0]), float(p[1])] for p in outer_polygons[0]]
            if outer_polygons
            else []
        )
        inner_coords = (
            [[float(p[0]), float(p[1])] for p in inner_polygons[0]]
            if inner_polygons
            else []
        )
        limit_range_coords = (
            [[float(p[0]), float(p[1])] for p in limit_range_polygons[0]]
            if limit_range_polygons
            else []
        )
        outer_polygons_json = [
            [[float(p[0]), float(p[1])] for p in poly] for poly in outer_polygons
        ]
        inner_polygons_json = [
            [[float(p[0]), float(p[1])] for p in poly] for poly in inner_polygons
        ]
        limit_range_polygons_json = [
            [[float(p[0]), float(p[1])] for p in poly] for poly in limit_range_polygons
        ]

        return {
            "success": True,
            "outer_polygon": outer_coords,
            "inner_polygon": inner_coords,
            "limit_range_polygon": limit_range_coords,
            "outer_polygons": outer_polygons_json,
            "inner_polygons": inner_polygons_json,
            "limit_range_polygons": limit_range_polygons_json,
        }
    except Exception as e:
        print(f"Error getting public geo fence coordinates: {e}")
        return {"success": False, "error": "Failed to retrieve geo fence coordinates"}


def _parse_client_timestamp(raw_value):
    """Parse client timestamp safely, fallback to server now."""
    if not raw_value:
        return datetime.now()
    try:
        if isinstance(raw_value, datetime):
            return raw_value
        ts = str(raw_value).strip().replace("Z", "+00:00")
        return datetime.fromisoformat(ts)
    except Exception:
        return datetime.now()


@app.post("/location/update")
async def update_user_location(request: Request):
    user = verify_user_token(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    # Verify device session token to ensure only the latest active device is allowed to upload/track
    client_device_id = body.get("device_id")
    reg_no = user.get("reg_no")
    
    # Try looking in users table
    cursor.execute("SELECT current_device_id FROM users WHERE reg_no = ?", (reg_no,))
    db_row = cursor.fetchone()
    db_device_id = db_row[0] if db_row else None
    
    # If not found in users, check other_staff
    if db_device_id is None:
        cursor.execute("SELECT current_device_id FROM other_staff WHERE reg_no = ?", (reg_no,))
        other_row = cursor.fetchone()
        db_device_id = other_row[0] if other_row else None

    # Mismatch of device_id means the user logged in elsewhere
    if db_device_id is not None and client_device_id is not None:
        if client_device_id != db_device_id:
            raise HTTPException(status_code=403, detail="Device session mismatch. Tracking suspended on this device.")

    try:
        latitude = float(body.get("latitude"))
        longitude = float(body.get("longitude"))
    except Exception:
        raise HTTPException(
            status_code=400, detail="Latitude and longitude are required"
        )

    if not (-90 <= latitude <= 90):
        raise HTTPException(
            status_code=400, detail="Latitude must be between -90 and 90"
        )
    if not (-180 <= longitude <= 180):
        raise HTTPException(
            status_code=400, detail="Longitude must be between -180 and 180"
        )

    accuracy = body.get("accuracy_meters")
    speed = body.get("speed_mps")
    heading = body.get("heading_deg")
    altitude = body.get("altitude_m")
    source = (body.get("source") or "gps").strip()[:50]
    app_state = (body.get("app_state") or "foreground").strip()[:20]
    is_mocked = bool(body.get("is_mocked", False))
    device_id = (body.get("device_id") or "").strip()[:120] or None
    captured_at = _parse_client_timestamp(body.get("captured_at"))

    # Check if user has marked attendance today
    reg_no = user.get("reg_no")
    cursor.execute(
        "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE",
        (reg_no,)
    )
    has_attendance = cursor.fetchone()[0] > 0

    # Query out permission settings
    cursor.execute(
        "SELECT out_permission_enabled, out_permission_expiry FROM users WHERE reg_no = ?",
        (reg_no,)
    )
    user_row = cursor.fetchone()
    out_permitted = False
    if user_row:
        enabled = bool(user_row[0])
        expiry_str = user_row[1]
        if enabled:
            if expiry_str:
                try:
                    expiry = datetime.fromisoformat(expiry_str)
                    if datetime.now() < expiry:
                        out_permitted = True
                except Exception:
                    out_permitted = True
            else:
                out_permitted = True

    boundary_warning = False
    warning_message = None
    first_left_boundary_at = None

    if has_attendance and not out_permitted and _geo_fence_limit_range_polygons:
        # Get existing first_left_boundary_at from DB
        cursor.execute("SELECT first_left_boundary_at FROM user_latest_locations WHERE reg_no = ?", (reg_no,))
        row = cursor.fetchone()
        if row and row[0]:
            if isinstance(row[0], datetime):
                first_left_boundary_at = row[0]
            else:
                try:
                    first_left_boundary_at = datetime.fromisoformat(str(row[0]).replace("Z", "+00:00").split(".")[0])
                except Exception:
                    first_left_boundary_at = datetime.now()

        # Check if point is inside any limit range polygon
        inside_limit = _point_in_any_polygon(latitude, longitude, _geo_fence_limit_range_polygons)
        if inside_limit:
            first_left_boundary_at = None
        else:
            boundary_warning = True
            if first_left_boundary_at is None:
                first_left_boundary_at = datetime.now()
                warning_message = "You are outside the permitted movement boundary. Please return within 3 minutes to avoid being marked absent."
            else:
                now_naive = datetime.now()
                if first_left_boundary_at.tzinfo is not None:
                    first_left_boundary_at = first_left_boundary_at.replace(tzinfo=None)
                elapsed = (now_naive - first_left_boundary_at).total_seconds()
                if elapsed >= 180:
                    warning_message = "You have been outside the boundary for more than 3 minutes and have been marked absent."
                    current_date = now_naive.strftime("%Y-%m-%d")
                    cursor.execute(
                        """
                        UPDATE daily_attendance_status 
                        SET status = 'Absent', marked_by = 'Geofence System', marked_at = CURRENT_TIMESTAMP, absent_reason = ?
                        WHERE reg_no = ? AND date = ?
                        """,
                        ("Outside permitted movement boundary for more than 3 minutes.", reg_no, current_date),
                    )
                else:
                    remaining = int(180 - elapsed)
                    warning_message = f"You are outside the permitted movement boundary. Please return within {remaining} seconds to avoid being marked absent."

    try:
        cursor.execute(
            """
            INSERT INTO user_location_logs
            (reg_no, username, name, dept, role, latitude, longitude, accuracy_meters, speed_mps, heading_deg, altitude_m,
             source, app_state, is_mocked, device_id, captured_at, server_received_at, boundary_warning, warning_message)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?, ?)
        """,
            (
                user.get("reg_no"),
                user.get("username"),
                user.get("name"),
                user.get("dept"),
                user.get("role"),
                latitude,
                longitude,
                accuracy,
                speed,
                heading,
                altitude,
                source,
                app_state,
                is_mocked,
                device_id,
                captured_at,
                boundary_warning,
                warning_message,
            ),
        )

        cursor.execute(
            """
            INSERT INTO user_latest_locations
            (reg_no, username, name, dept, role, latitude, longitude, accuracy_meters, speed_mps, heading_deg, altitude_m,
             source, app_state, is_mocked, device_id, captured_at, last_seen_at, boundary_warning, warning_message, first_left_boundary_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?, ?, ?)
            ON CONFLICT (reg_no) DO UPDATE SET
                username = EXCLUDED.username,
                name = EXCLUDED.name,
                dept = EXCLUDED.dept,
                role = EXCLUDED.role,
                latitude = EXCLUDED.latitude,
                longitude = EXCLUDED.longitude,
                accuracy_meters = EXCLUDED.accuracy_meters,
                speed_mps = EXCLUDED.speed_mps,
                heading_deg = EXCLUDED.heading_deg,
                altitude_m = EXCLUDED.altitude_m,
                source = EXCLUDED.source,
                app_state = EXCLUDED.app_state,
                is_mocked = EXCLUDED.is_mocked,
                device_id = EXCLUDED.device_id,
                captured_at = EXCLUDED.captured_at,
                last_seen_at = CURRENT_TIMESTAMP,
                boundary_warning = EXCLUDED.boundary_warning,
                warning_message = EXCLUDED.warning_message,
                first_left_boundary_at = EXCLUDED.first_left_boundary_at
        """,
            (
                user.get("reg_no"),
                user.get("username"),
                user.get("name"),
                user.get("dept"),
                user.get("role"),
                latitude,
                longitude,
                accuracy,
                speed,
                heading,
                altitude,
                source,
                app_state,
                is_mocked,
                device_id,
                captured_at,
                boundary_warning,
                warning_message,
                first_left_boundary_at,
            ),
        )

        # Reset force update request
        cursor.execute("UPDATE user_latest_locations SET force_update_requested = FALSE WHERE reg_no = ?", (reg_no,))

        return {
            "success": True,
            "message": "Location updated",
            "warning": warning_message,
            "boundary_warning": boundary_warning,
            "server_timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        print(f"Location update error for {user.get('reg_no')}: {e}")
        raise HTTPException(status_code=500, detail="Failed to store location")


@app.post("/location/sync_offline")
async def sync_offline_locations(request: Request):
    user = verify_user_token(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    logs = body.get("logs", [])
    if not logs:
        return {"success": True, "message": "No logs to sync", "rule_violation": False}

    reg_no = user.get("reg_no")
    client_device_id = body.get("device_id")

    # Verify device session token
    cursor.execute("SELECT current_device_id FROM users WHERE reg_no = ?", (reg_no,))
    db_row = cursor.fetchone()
    db_device_id = db_row[0] if db_row else None

    if db_device_id is None:
        cursor.execute("SELECT current_device_id FROM other_staff WHERE reg_no = ?", (reg_no,))
        other_row = cursor.fetchone()
        db_device_id = other_row[0] if other_row else None

    if db_device_id is not None and client_device_id is not None:
        if client_device_id != db_device_id:
            raise HTTPException(status_code=403, detail="Device session mismatch. Tracking suspended on this device.")

    # Sort logs by captured_at to process sequentially
    def get_captured_at(log):
        try:
            return _parse_client_timestamp(log.get("captured_at"))
        except Exception:
            return datetime.min

    try:
        logs.sort(key=get_captured_at)
    except Exception as e:
        print(f"Sorting error: {e}")

    # Check if user has marked attendance today
    cursor.execute(
        "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE",
        (reg_no,)
    )
    has_attendance = cursor.fetchone()[0] > 0

    # Query out permission settings
    cursor.execute(
        "SELECT out_permission_enabled, out_permission_expiry FROM users WHERE reg_no = ?",
        (reg_no,)
    )
    user_row = cursor.fetchone()
    out_permitted_base = False
    out_permission_expiry = None
    if user_row:
        out_permitted_base = bool(user_row[0])
        if user_row[1]:
            try:
                out_permission_expiry = datetime.fromisoformat(user_row[1])
            except Exception:
                pass

    rule_violation = False
    violation_reason = None

    outside_start = None
    gps_off_start = None

    # Get existing first_left_boundary_at from DB to persist boundary timer state
    cursor.execute("SELECT first_left_boundary_at FROM user_latest_locations WHERE reg_no = ?", (reg_no,))
    db_row = cursor.fetchone()
    if db_row and db_row[0]:
        if isinstance(db_row[0], datetime):
            outside_start = db_row[0]
        else:
            try:
                outside_start = datetime.fromisoformat(str(db_row[0]).replace("Z", "+00:00").split(".")[0])
            except Exception:
                outside_start = None

    boundary_warning = False
    warning_message = None

    for log in logs:
        try:
            latitude = float(log.get("latitude", 0.0))
            longitude = float(log.get("longitude", 0.0))
        except (ValueError, TypeError):
            continue

        accuracy = log.get("accuracy_meters")
        speed = log.get("speed_mps")
        heading = log.get("heading_deg")
        altitude = log.get("altitude_m")
        source = (log.get("source") or "gps").strip()[:50]
        app_state = (log.get("app_state") or "background").strip()[:20]
        is_mocked = bool(log.get("is_mocked", False))
        device_id = (log.get("device_id") or "").strip()[:120] or None
        captured_at = _parse_client_timestamp(log.get("captured_at"))
        gps_enabled = bool(log.get("gps_enabled", True))

        # Determine out_permitted dynamically for the log timestamp
        out_permitted = False
        if out_permitted_base:
            if out_permission_expiry:
                try:
                    expiry_check = out_permission_expiry
                    captured_at_compare = captured_at
                    if expiry_check.tzinfo is not None and captured_at_compare.tzinfo is None:
                        expiry_check = expiry_check.replace(tzinfo=None)
                    elif expiry_check.tzinfo is None and captured_at_compare.tzinfo is not None:
                        captured_at_compare = captured_at_compare.replace(tzinfo=None)
                    
                    if captured_at_compare < expiry_check:
                        out_permitted = True
                except Exception:
                    out_permitted = True
            else:
                out_permitted = True

        # Check Rules:
        # Rule 1: Mock Location
        if is_mocked:
            rule_violation = True
            violation_reason = "Mock location usage detected."

        # Rule 2: GPS/Location Disabled
        if not gps_enabled:
            if gps_off_start is None:
                gps_off_start = captured_at
            else:
                elapsed_gps = (captured_at - gps_off_start).total_seconds()
                if elapsed_gps >= 180:
                    rule_violation = True
                    violation_reason = "Location/GPS services were disabled for more than 3 minutes."
        else:
            gps_off_start = None

        # Rule 3: Geofence limits
        boundary_warning = False
        warning_message = None
        if has_attendance and not out_permitted and _geo_fence_limit_range_polygons and gps_enabled:
            inside_limit = _point_in_any_polygon(latitude, longitude, _geo_fence_limit_range_polygons)
            if inside_limit:
                outside_start = None
            else:
                boundary_warning = True
                if outside_start is None:
                    outside_start = captured_at
                    warning_message = "You are outside the permitted movement boundary."
                else:
                    # Strip tzinfo for datetime arithmetic if needed
                    t_outside = outside_start.replace(tzinfo=None) if outside_start.tzinfo else outside_start
                    t_captured = captured_at.replace(tzinfo=None) if captured_at.tzinfo else captured_at
                    elapsed_boundary = (t_captured - t_outside).total_seconds()
                    if elapsed_boundary >= 180:
                        rule_violation = True
                        violation_reason = "You were outside the permitted boundary for more than 3 minutes."
                        warning_message = "You have been outside the boundary for more than 3 minutes."
                    else:
                        remaining = int(180 - elapsed_boundary)
                        warning_message = f"You are outside the permitted movement boundary. Return in {remaining}s."

        # Write to log tables
        try:
            cursor.execute(
                """
                INSERT INTO user_location_logs
                (reg_no, username, name, dept, role, latitude, longitude, accuracy_meters, speed_mps, heading_deg, altitude_m,
                 source, app_state, is_mocked, device_id, captured_at, server_received_at, boundary_warning, warning_message)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?, ?)
            """,
                (
                    user.get("reg_no"),
                    user.get("username"),
                    user.get("name"),
                    user.get("dept"),
                    user.get("role"),
                    latitude,
                    longitude,
                    accuracy,
                    speed,
                    heading,
                    altitude,
                    source,
                    app_state,
                    is_mocked,
                    device_id,
                    captured_at.isoformat(),
                    boundary_warning,
                    warning_message,
                ),
            )

            cursor.execute(
                """
                INSERT INTO user_latest_locations
                (reg_no, username, name, dept, role, latitude, longitude, accuracy_meters, speed_mps, heading_deg, altitude_m,
                 source, app_state, is_mocked, device_id, captured_at, last_seen_at, boundary_warning, warning_message, first_left_boundary_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?, ?, ?)
                ON CONFLICT (reg_no) DO UPDATE SET
                    username = EXCLUDED.username,
                    name = EXCLUDED.name,
                    dept = EXCLUDED.dept,
                    role = EXCLUDED.role,
                    latitude = EXCLUDED.latitude,
                    longitude = EXCLUDED.longitude,
                    accuracy_meters = EXCLUDED.accuracy_meters,
                    speed_mps = EXCLUDED.speed_mps,
                    heading_deg = EXCLUDED.heading_deg,
                    altitude_m = EXCLUDED.altitude_m,
                    source = EXCLUDED.source,
                    app_state = EXCLUDED.app_state,
                    is_mocked = EXCLUDED.is_mocked,
                    device_id = EXCLUDED.device_id,
                    captured_at = EXCLUDED.captured_at,
                    last_seen_at = CURRENT_TIMESTAMP,
                    boundary_warning = EXCLUDED.boundary_warning,
                    warning_message = EXCLUDED.warning_message,
                    first_left_boundary_at = EXCLUDED.first_left_boundary_at
            """,
                (
                    user.get("reg_no"),
                    user.get("username"),
                    user.get("name"),
                    user.get("dept"),
                    user.get("role"),
                    latitude,
                    longitude,
                    accuracy,
                    speed,
                    heading,
                    altitude,
                    source,
                    app_state,
                    is_mocked,
                    device_id,
                    captured_at.isoformat(),
                    boundary_warning,
                    warning_message,
                    outside_start.isoformat() if outside_start else None,
                ),
            )
        except Exception as log_err:
            print(f"Failed to log offline point: {log_err}")

    # Mark absent if user violated rules
    if rule_violation:
        current_date = datetime.now().strftime("%Y-%m-%d")
        cursor.execute(
            """
            UPDATE daily_attendance_status 
            SET status = 'Absent', marked_by = 'Geofence System (Offline)', marked_at = CURRENT_TIMESTAMP, absent_reason = ?
            WHERE reg_no = ? AND date = ?
            """,
            (violation_reason or "Offline tracking rule violation.", reg_no, current_date),
        )

    # Reset force update request
    cursor.execute("UPDATE user_latest_locations SET force_update_requested = FALSE WHERE reg_no = ?", (reg_no,))

    return {
        "success": True,
        "rule_violation": rule_violation,
        "message": violation_reason or "Offline logs synchronized successfully.",
        "boundary_warning": boundary_warning,
        "warning": warning_message,
    }





@app.get("/staff/tracking-status")
async def get_staff_tracking_status(request: Request):
    """Check if tracking should be active for the current user today.
    
    Tracking is active ONLY when the user has marked check-in AND has NOT yet marked check-out today.
    This ensures location sharing runs from check-in to check-out only, and is OFF at all other times.
    """
    user = verify_user_token(request)
    if user.get("role") == "admin":
        return {"success": True, "tracking_active": False, "reason": "admin_exemption"}
    reg_no = user.get("reg_no")
    client_device_id = request.query_params.get("device_id")

    # If client passed a device_id, verify that it matches what's stored in db
    if client_device_id:
        cursor.execute("SELECT current_device_id FROM users WHERE reg_no = ?", (reg_no,))
        db_row = cursor.fetchone()
        db_device_id = db_row[0] if db_row else None
        
        if db_device_id is None:
            cursor.execute("SELECT current_device_id FROM other_staff WHERE reg_no = ?", (reg_no,))
            other_row = cursor.fetchone()
            db_device_id = other_row[0] if other_row else None
            
        if db_device_id is not None and client_device_id != db_device_id:
            return {"success": True, "tracking_active": False, "reason": "device_mismatch"}

    try:
        # Check if user has checked in today (attendance table)
        cursor.execute(
            "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
            (reg_no,)
        )
        has_check_in = cursor.fetchone()[0] > 0

        # Check if user has checked out today (attendance table)
        cursor.execute(
            "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_out'",
            (reg_no,)
        )
        has_check_out = cursor.fetchone()[0] > 0

        # Also check other_staff_attendance table if not found in main table
        if not has_check_in:
            cursor.execute(
                "SELECT COUNT(*) FROM other_staff_attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
                (reg_no,)
            )
            has_check_in = cursor.fetchone()[0] > 0

        if not has_check_out:
            cursor.execute(
                "SELECT COUNT(*) FROM other_staff_attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_out'",
                (reg_no,)
            )
            has_check_out = cursor.fetchone()[0] > 0

        # Tracking is active ONLY if checked in AND NOT checked out
        tracking_active = has_check_in and not has_check_out

        # Check if force update is requested
        cursor.execute("SELECT force_update_requested FROM user_latest_locations WHERE reg_no = ?", (reg_no,))
        row = cursor.fetchone()
        force_update = bool(row[0]) if row else False

        return {
            "success": True,
            "tracking_active": tracking_active,
            "has_check_in": has_check_in,
            "has_check_out": has_check_out,
            "force_update": force_update,
        }

    except Exception as e:
        print(f"Error checking tracking status: {e}")
        return {"success": False, "tracking_active": False}


@app.post("/admin/user/out-permission")
async def update_user_out_permission(request: Request):
    """Grant/revoke out permission for a user (Admin/HOD only)."""
    admin = verify_admin_token(request)
    try:
        body = await request.json()
        reg_no = body.get("reg_no")
        enabled = bool(body.get("enabled", False))
        expires_at = body.get("expires_at")

        if not reg_no:
            return error_response("reg_no is required", "INVALID_INPUT")

        cursor.execute(
            """
            UPDATE users 
            SET out_permission_enabled = ?, out_permission_expiry = ? 
            WHERE reg_no = ?
            """,
            (enabled, expires_at, reg_no)
        )

        log_audit_event(
            "OUT_PERMISSION_UPDATED",
            admin["reg_no"],
            True,
            f"Updated out permission for {reg_no}: enabled={enabled}",
        )
        return success_response("Out permission updated successfully")
    except Exception as e:
        print(f"Error updating out permission: {e}")
        return error_response("Failed to update out permission", "UPDATE_ERROR")


@app.post("/admin/locations/force-update/{reg_no}")
async def request_force_location_update(request: Request, reg_no: str):
    """Request an instant location update for a specific user (Admin/HOD only)."""
    verify_admin_token(request)
    
    # Verify the user exists in latest locations
    cursor.execute("SELECT COUNT(*) FROM user_latest_locations WHERE reg_no = ?", (reg_no,))
    if cursor.fetchone()[0] == 0:
        raise HTTPException(status_code=404, detail="User latest location record not found. User may not be checked in today.")
        
    try:
        cursor.execute(
            "UPDATE user_latest_locations SET force_update_requested = TRUE WHERE reg_no = ?",
            (reg_no,)
        )
        return {"success": True, "message": f"Force location update requested for user {reg_no}."}
    except Exception as e:
        print(f"Error requesting force location update: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/locations/live")
async def admin_get_live_locations(
    request: Request,
    minutes: int = 180,
    include_stale: bool = False,
    online_only: bool = False,
    online_window_minutes: int = 60,
    inside_outer_only: bool = False,
):
    """Get latest known location for all users (Admin/HOD only)."""
    verify_admin_token(request)

    # Keep live locations visible for the whole day window:
    # from today's 12:00 AM up to next day's 12:00 AM.
    now = datetime.now()
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    next_day_start = day_start + timedelta(days=1)

    minutes = max(1, min(minutes, 60 * 24 * 30))
    online_window_minutes = max(1, min(online_window_minutes, 120))

    if include_stale and not online_only:
        cursor.execute(
            """
            SELECT ul.reg_no, ul.username, ul.name, ul.dept, ul.role, ul.latitude, ul.longitude, ul.accuracy_meters,
                   ul.speed_mps, ul.heading_deg, ul.altitude_m, ul.source, ul.app_state, ul.is_mocked,
                   ul.device_id, ul.captured_at, ul.last_seen_at, ul.boundary_warning, ul.warning_message,
                   u.out_permission_enabled, u.out_permission_expiry
            FROM user_latest_locations ul
            LEFT JOIN users u ON ul.reg_no = u.reg_no
            WHERE ul.last_seen_at >= ? AND ul.last_seen_at < ?
            ORDER BY ul.last_seen_at DESC
        """,
            (day_start, next_day_start),
        )
    elif online_only:
        cursor.execute(
            """
            SELECT ul.reg_no, ul.username, ul.name, ul.dept, ul.role, ul.latitude, ul.longitude, ul.accuracy_meters,
                   ul.speed_mps, ul.heading_deg, ul.altitude_m, ul.source, ul.app_state, ul.is_mocked,
                   ul.device_id, ul.captured_at, ul.last_seen_at, ul.boundary_warning, ul.warning_message,
                   u.out_permission_enabled, u.out_permission_expiry
            FROM user_latest_locations ul
            LEFT JOIN users u ON ul.reg_no = u.reg_no
            WHERE ul.last_seen_at >= ? AND ul.last_seen_at < ?
            ORDER BY ul.last_seen_at DESC
        """,
            (day_start, next_day_start),
        )
    else:
        cursor.execute(
            """
            SELECT ul.reg_no, ul.username, ul.name, ul.dept, ul.role, ul.latitude, ul.longitude, ul.accuracy_meters,
                   ul.speed_mps, ul.heading_deg, ul.altitude_m, ul.source, ul.app_state, ul.is_mocked,
                   ul.device_id, ul.captured_at, ul.last_seen_at, ul.boundary_warning, ul.warning_message,
                   u.out_permission_enabled, u.out_permission_expiry
            FROM user_latest_locations ul
            LEFT JOIN users u ON ul.reg_no = u.reg_no
            WHERE ul.last_seen_at >= ? AND ul.last_seen_at < ?
            ORDER BY ul.last_seen_at DESC
        """,
            (day_start, next_day_start),
        )

    rows = cursor.fetchall()
    locations = []
    for row in rows:
        lat = float(row[5])
        lng = float(row[6])
        if inside_outer_only and not _point_in_any_polygon(
            lat, lng, _geo_fence_outer_polygons
        ):
            continue
        locations.append(
            {
                "reg_no": row[0],
                "username": row[1],
                "name": row[2],
                "dept": row[3],
                "role": row[4],
                "latitude": lat,
                "longitude": lng,
                "accuracy_meters": float(row[7]) if row[7] is not None else None,
                "speed_mps": float(row[8]) if row[8] is not None else None,
                "heading_deg": float(row[9]) if row[9] is not None else None,
                "altitude_m": float(row[10]) if row[10] is not None else None,
                "source": row[11],
                "app_state": row[12],
                "is_mocked": bool(row[13]),
                "device_id": row[14],
                "captured_at": _ts(row[15]),
                "last_seen_at": _ts(row[16]),
                "boundary_warning": bool(row[17]) if row[17] is not None else False,
                "warning_message": row[18],
                "out_permission_enabled": bool(row[19]) if row[19] is not None else False,
                "out_permission_expiry": row[20],
            }
        )

    return {
        "success": True,
        "count": len(locations),
        "minutes": minutes,
        "include_stale": include_stale,
        "online_only": online_only,
        "online_window_minutes": online_window_minutes,
        "inside_outer_only": inside_outer_only,
        "locations": locations,
    }


@app.get("/admin/locations/history")
async def admin_get_location_history(request: Request, reg_no: str, date: str = None, limit: int = 500):
    """Get recent location history for a user (Admin/HOD only)."""
    verify_admin_token(request)
    limit = max(1, min(limit, 5000))

    if date:
        try:
            parsed_date = datetime.strptime(date, "%Y-%m-%d")
            day_start = parsed_date.replace(hour=0, minute=0, second=0, microsecond=0)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format, use YYYY-MM-DD")
    else:
        # Filter to only show the path for that day alone (current local day)
        now = datetime.now()
        day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
    next_day_start = day_start + timedelta(days=1)

    cursor.execute(
        """
        SELECT reg_no, username, name, dept, role, latitude, longitude, accuracy_meters,
               speed_mps, heading_deg, altitude_m, source, app_state, is_mocked,
               device_id, captured_at, server_received_at, boundary_warning, warning_message
        FROM user_location_logs
        WHERE reg_no = ? AND captured_at >= ? AND captured_at < ?
        ORDER BY captured_at DESC
        LIMIT ?
    """,
        (reg_no, day_start, next_day_start, limit),
    )
    rows = cursor.fetchall()

    history = [
        {
            "reg_no": row[0],
            "username": row[1],
            "name": row[2],
            "dept": row[3],
            "role": row[4],
            "latitude": float(row[5]),
            "longitude": float(row[6]),
            "accuracy_meters": float(row[7]) if row[7] is not None else None,
            "speed_mps": float(row[8]) if row[8] is not None else None,
            "heading_deg": float(row[9]) if row[9] is not None else None,
            "altitude_m": float(row[10]) if row[10] is not None else None,
            "source": row[11],
            "app_state": row[12],
            "is_mocked": bool(row[13]),
            "device_id": row[14],
            "captured_at": _ts(row[15]),
            "server_received_at": _ts(row[16]),
            "boundary_warning": bool(row[17]) if row[17] is not None else False,
            "warning_message": row[18],
        }
        for row in rows
    ]

    return {
        "success": True,
        "reg_no": reg_no,
        "count": len(history),
        "history": history,
    }


# -------------------------------------------------
# FACE RECOGNITION SYSTEM - GUARANTEED TO WORK
# -------------------------------------------------
print("Initializing face recognition system...")

# Global variables
face_app = None
use_fallback = False
face_cascade = None
face_cascade_alt = None
face_cascade_profile = None
# Retry counter to prevent infinite recursion in extract_face
extract_face_recursion_depth = 0

# Re-export the check_image_quality from later in the file for backwards compatibility
# The actual implementation is defined later in this file


def create_fallback_embedding(face_img):
    """Create a simple but reliable fallback embedding"""
    try:
        # Resize to standard size
        face_resized = cv2.resize(face_img, (128, 128))

        # Convert to grayscale
        gray = cv2.cvtColor(face_resized, cv2.COLOR_BGR2GRAY)

        # Create a simple but unique embedding based on image content
        features = []

        # Divide image into regions and extract features
        for i in range(0, 128, 32):
            for j in range(0, 128, 32):
                region = gray[i : i + 32, j : j + 32]
                features.append(np.mean(region))
                features.append(np.std(region))
                features.append(np.max(region))
                features.append(np.min(region))

        # Add overall image features
        features.append(np.mean(gray))
        features.append(np.std(gray))
        features.append(np.max(gray))
        features.append(np.min(gray))

        # Create hash-based unique signature
        face_hash = hashlib.sha256(gray.tobytes()).hexdigest()
        # Convert hash to numeric features
        for i in range(0, len(face_hash), 4):
            hex_val = face_hash[i : i + 4]
            if len(hex_val) == 4:
                features.append(int(hex_val, 16) / 65535.0)

        # Convert to numpy array
        embedding = np.array(features, dtype=np.float32)

        # Ensure consistent size (512)
        target_size = 512
        if len(embedding) < target_size:
            embedding = np.pad(embedding, (0, target_size - len(embedding)), "constant")
        elif len(embedding) > target_size:
            embedding = embedding[:target_size]

        # Normalize to unit vector
        norm = np.linalg.norm(embedding)
        if norm > 0:
            embedding = embedding / norm

        print(f"Created fallback embedding with {len(features)} features")
        return embedding

    except Exception as e:
        print(f"Fallback embedding error: {e}")
        # Create unique random embedding based on time
        np.random.seed(int(datetime.now().timestamp() * 1000))
        return np.random.rand(512).astype(np.float32)


def calculate_similarity(embedding1, embedding2):
    """Calculate robust similarity between two embeddings"""
    try:
        # Multiple similarity measures for better accuracy
        # 1. Cosine similarity
        cosine_sim = np.dot(embedding1, embedding2) / (
            np.linalg.norm(embedding1) * np.linalg.norm(embedding2)
        )

        # 2. Euclidean distance (inverted)
        euclidean_dist = np.linalg.norm(embedding1 - embedding2)
        euclidean_sim = 1 / (1 + euclidean_dist)

        # 3. Manhattan distance (inverted)
        manhattan_dist = np.sum(np.abs(embedding1 - embedding2))
        manhattan_sim = 1 / (1 + manhattan_dist)

        # 4. Correlation coefficient
        corr = (
            np.corrcoef(embedding1, embedding2)[0, 1]
            if np.std(embedding1) > 0 and np.std(embedding2) > 0
            else 0
        )

        # Weighted combination (cosine is most important)
        combined_sim = (
            0.5 * cosine_sim
            + 0.2 * euclidean_sim
            + 0.15 * manhattan_sim
            + 0.15 * max(0, corr)
        )

        return combined_sim, cosine_sim, euclidean_dist, manhattan_dist, corr

    except Exception as e:
        print(f"Similarity calculation error: {e}")
        return 0.0, 0.0, float("inf"), 0.0, 0.0


# -------------------------------------------------
# LIVENESS DETECTION - Anti-spoofing for photo detection
# -------------------------------------------------
def detect_liveness(frames: list) -> tuple[bool, str]:
    """
    Detect if the face is real or a spoof attempt using multiple methods.
    Returns (is_live, reason)
    """
    global ANTISPOOFING_ENABLED

    # If antispoofing is disabled, allow all
    if not ANTISPOOFING_ENABLED:
        return True, "Anti-spoofing disabled"

    if len(frames) < MIN_FRAMES_FOR_LIVENESS:
        # With fewer frames, do single-image analysis
        if len(frames) == 1:
            return detect_single_image_liveness(frames[0])
        return True, f"Insufficient frames ({len(frames)}), using single-image analysis"

    try:
        # Method 1: Movement analysis between frames
        movement_scores = []
        for i in range(1, len(frames)):
            diff = cv2.absdiff(frames[i - 1], frames[i])
            movement = np.mean(diff) / 255.0
            movement_scores.append(movement)

        avg_movement = np.mean(movement_scores)

        # Check if movement is too little (photo) or too much (video replay attack)
        if avg_movement < MIN_MOVEMENT:
            return (
                False,
                f"No movement detected - possible photo (movement: {avg_movement:.4f})",
            )

        if avg_movement > MAX_MOVEMENT:
            return (
                False,
                f"Unusual movement pattern - possible video replay (movement: {avg_movement:.4f})",
            )

        # Method 2: Analyze brightness/color consistency (photos have uniform lighting)
        brightness_scores = []
        for frame in frames:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            brightness = np.std(gray) / (np.mean(gray) + 1e-6)
            brightness_scores.append(brightness)

        brightness_variation = np.std(brightness_scores)

        # Real faces have natural brightness variation
        if brightness_variation < 0.01:
            return False, f"Unnatural brightness consistency - possible photo"

        # Method 3: Texture analysis - photos have different texture patterns
        for frame in frames:
            is_live, reason = detect_single_image_liveness(frame)
            if not is_live:
                return False, f"Photo detected in frame: {reason}"

        # All checks passed
        return (
            True,
            f"Liveness verified (movement: {avg_movement:.4f}, brightness variation: {brightness_variation:.4f})",
        )

    except Exception as e:
        print(f"Liveness detection error: {e}")
        # On error, be strict - deny for security
        return False, f"Liveness check failed: {str(e)}"


def detect_single_image_liveness(img) -> tuple[bool, str]:
    """
    Legacy function - now delegates to comprehensive analysis.
    Maintained for backward compatibility.
    """
    global ANTISPOOFING_ENABLED

    if not ANTISPOOFING_ENABLED:
        return True, "Anti-spoofing disabled"

    result = analyze_image_comprehensive(img)
    return result.get("is_live", False), result.get("liveness_reason", "Analysis failed")


def extract_face_features(img):
    """Extract face with enhanced features for verification"""
    face = extract_face(img)

    if face is None:
        return None

    # Add additional quality metrics
    if face.embedding is not None:
        # Calculate embedding statistics
        emb_std = np.std(face.embedding)
        emb_mean = np.mean(face.embedding)

        # Set detection score based on embedding quality
        face.det_score = min(0.9, emb_std * 2 + 0.5)

    return face


def _warm_face_embedding_cache():
    """Load all user embeddings and samples into memory at startup."""
    global _face_profile_cache
    with _cache_lock:
        _face_profile_cache.clear()
        loaded_count = 0
        for source_table in ("users", "other_staff"):
            try:
                cursor.execute(
                    f"SELECT reg_no, name, dept, role, embedding FROM {source_table}"
                )
                for reg_no, name, dept, role, emb_blob in cursor.fetchall():
                    if emb_blob is not None:
                        emb = np.frombuffer(emb_blob, dtype=np.float32).copy()
                        norm = np.linalg.norm(emb)
                        if norm > 0:
                            emb = emb / norm
                            _face_profile_cache.put(reg_no, {
                                "primary": emb,
                                "samples": [],
                                "samples_loaded": False,
                                "source_table": source_table,
                                "name": name,
                                "dept": dept,
                                "role": role,
                            })
                            loaded_count += 1
                print(f"  Loaded {loaded_count} profiles from {source_table}")
            except Exception as e:
                print(f"  Warning: Could not load cache from {source_table}: {e}")

        # Pre-load samples for all users (limit to prevent memory bloat)
        try:
            cursor.execute(f"""
                SELECT reg_no, source_table, embedding
                FROM face_embedding_samples
                ORDER BY created_at DESC, id DESC
            """)
            sample_map = {}
            for reg_no, src_table, emb_blob in cursor.fetchall():
                if reg_no not in sample_map:
                    sample_map[reg_no] = []
                if len(sample_map[reg_no]) < MAX_PROFILE_SAMPLES:
                    sample = np.frombuffer(emb_blob, dtype=np.float32).copy()
                    norm = np.linalg.norm(sample)
                    if norm > 0:
                        sample = sample / norm
                        sample_map[reg_no].append(sample)

            sample_count = 0
            for reg_no, samples in sample_map.items():
                profile = _face_profile_cache.get(reg_no)
                if profile:
                    profile["samples"] = samples
                    profile["samples_loaded"] = True
                    _face_profile_cache.put(reg_no, profile)  # Update cache
                    sample_count += 1

            print(f"  Pre-loaded samples for {sample_count} users")
        except Exception as e:
            print(f"  Warning: Could not pre-load samples: {e}")

        print(f"Face embedding cache warmed: {_face_profile_cache.size()} users")


def _cache_get_profile(reg_no: str):
    """Get user profile from cache, falling back to DB if not cached."""
    profile = _face_profile_cache.get(reg_no)
    if profile is not None:
        return profile

    # Cache miss — try to load from DB and add to cache dynamically
    try:
        cursor.execute(
            "SELECT reg_no, name, dept, role, embedding FROM users WHERE reg_no = %s",
            (reg_no,),
        )
        row = cursor.fetchone()
        if row and row[4] is not None:
            emb = np.frombuffer(row[4], dtype=np.float32).copy()
            norm = np.linalg.norm(emb)
            if norm > 0:
                emb = emb / norm
            profile = {
                "primary": emb,
                "samples": [],
                "samples_loaded": False,
                "source_table": "users",
                "name": row[1],
                "dept": row[2],
                "role": row[3],
            }
            _face_profile_cache.put(reg_no, profile)
            print(f"Cache miss resolved: loaded {reg_no} ({row[1]}) from DB")
            return profile

        cursor.execute(
            "SELECT reg_no, name, dept, role, embedding FROM other_staff WHERE reg_no = %s",
            (reg_no,),
        )
        row = cursor.fetchone()
        if row and row[4] is not None:
            emb = np.frombuffer(row[4], dtype=np.float32).copy()
            norm = np.linalg.norm(emb)
            if norm > 0:
                emb = emb / norm
            profile = {
                "primary": emb,
                "samples": [],
                "samples_loaded": False,
                "source_table": "other_staff",
                "name": row[1],
                "dept": row[2],
                "role": row[3],
            }
            _face_profile_cache.put(reg_no, profile)
            print(f"Cache miss resolved: loaded {reg_no} ({row[1]}) from DB")
            return profile
    except Exception as e:
        print(f"Cache fallback error for {reg_no}: {e}")

    return None


def _cache_get_candidates(reg_no: str):
    """Get all candidate embeddings for a user from cache, loading samples from DB if needed."""
    profile = _cache_get_profile(reg_no)
    if profile is None:
        return []

    candidates = []
    if profile["primary"] is not None:
        candidates.append(profile["primary"])

    # Load samples from DB if not already loaded
    if not profile.get("samples_loaded", False):
        try:
            source_table = profile.get("source_table", "users")
            cursor.execute(
                """
                SELECT embedding FROM face_embedding_samples
                WHERE reg_no = %s AND source_table = %s
                ORDER BY created_at DESC, id DESC
                LIMIT %s
                """,
                (reg_no, source_table, MAX_PROFILE_SAMPLES),
            )
            rows = cursor.fetchall()
            for (emb_blob,) in rows:
                sample = np.frombuffer(emb_blob, dtype=np.float32).copy()
                norm = np.linalg.norm(sample)
                if norm > 0:
                    sample = sample / norm
                    candidates.append(sample)
                    profile["samples"].append(sample)
            profile["samples_loaded"] = True
        except Exception as e:
            print(f"Error loading samples for {reg_no}: {e}")
    else:
        for sample in profile.get("samples", []):
            candidates.append(sample)

    # Deduplicate
    unique = []
    seen = set()
    for emb in candidates:
        key = emb.tobytes()
        if key not in seen:
            seen.add(key)
            unique.append(emb)
    return unique


def _cache_update_primary(reg_no: str, embedding: np.ndarray, user_info: dict = None):
    """Update the primary embedding for a user in cache.
    If user not in cache, adds them dynamically.
    """
    norm = np.linalg.norm(embedding)
    if norm > 0:
        embedding = embedding / norm
    profile = _face_profile_cache.get(reg_no)
    if profile:
        profile["primary"] = embedding.copy()
    else:
        # User not in cache — add them dynamically
        # This happens when a user registers their face for the first time
        # after server startup, or a new user is created
        if user_info:
            _face_profile_cache.put(reg_no, {
                "primary": embedding.copy(),
                "samples": [],
                "samples_loaded": False,
                "source_table": user_info.get("source_table", "users"),
                "name": user_info.get("name", ""),
                "dept": user_info.get("dept", ""),
                "role": user_info.get("role", "staff"),
            })
            print(
                f"Cache updated: added {reg_no} ({user_info.get('name')}) dynamically"
            )
        else:
            # Try to fetch user info from DB
            try:
                cursor.execute(
                    "SELECT reg_no, name, dept, role FROM users WHERE reg_no = %s",
                    (reg_no,),
                )
                row = cursor.fetchone()
                if row:
                    _face_profile_cache.put(reg_no, {
                        "primary": embedding.copy(),
                        "samples": [],
                        "samples_loaded": False,
                        "source_table": "users",
                        "name": row[1],
                        "dept": row[2],
                        "role": row[3],
                    })
                    print(f"Cache updated: added {reg_no} ({row[1]}) dynamically")
                    return
                cursor.execute(
                    "SELECT reg_no, name, dept, role FROM other_staff WHERE reg_no = %s",
                    (reg_no,),
                )
                row = cursor.fetchone()
                if row:
                    _face_profile_cache.put(reg_no, {
                        "primary": embedding.copy(),
                        "samples": [],
                        "samples_loaded": False,
                        "source_table": "other_staff",
                        "name": row[1],
                        "dept": row[2],
                        "role": row[3],
                    })
                    print(f"Cache updated: added {reg_no} ({row[1]}) dynamically")
                    return
            except Exception as e:
                print(f"Warning: Could not add {reg_no} to cache: {e}")
            print(
                f"Warning: Cannot update cache for {reg_no} - not in cache and not found in DB"
            )


def _cache_add_sample(reg_no: str, embedding: np.ndarray):
    """Add a new sample to the user's cached samples list."""
    norm = np.linalg.norm(embedding)
    if norm > 0:
        embedding = embedding / norm
    profile = _face_profile_cache.get(reg_no)
    if profile:
        profile["samples"].insert(0, embedding.copy())
        profile["samples"] = profile["samples"][:MAX_PROFILE_SAMPLES]
        profile["samples_loaded"] = True


def _cache_invalidate_user(reg_no: str):
    """Remove a user's profile from cache (forces reload on next access)."""
    _face_profile_cache.remove(reg_no)


def _normalize_embedding(embedding: np.ndarray):
    """Return unit-normalized embedding or None if invalid."""
    if embedding is None:
        return None
    emb = embedding.astype(np.float32)
    if emb.size == 0:
        return None
    norm = np.linalg.norm(emb)
    if norm == 0:
        return None
    return emb / norm


def _save_face_embedding_sample(
    reg_no: str,
    source_table: str,
    embedding: np.ndarray,
    sample_type: str = "registration",
    confidence: float = None,
):
    """Persist a normalized embedding sample and keep only recent samples."""
    try:
        norm_embedding = _normalize_embedding(embedding)
        if norm_embedding is None:
            return
        cursor.execute(
            """
            INSERT INTO face_embedding_samples (reg_no, source_table, embedding, sample_type, confidence)
            VALUES (?, ?, ?, ?, ?)
        """,
            (reg_no, source_table, norm_embedding.tobytes(), sample_type, confidence),
        )
        cursor.execute(
            """
            DELETE FROM face_embedding_samples
            WHERE id IN (
                SELECT id
                FROM face_embedding_samples
                WHERE reg_no = ? AND source_table = ?
                ORDER BY created_at DESC, id DESC
                LIMIT ALL OFFSET %s
            )
        """,
            (reg_no, source_table, MAX_PROFILE_SAMPLES),
        )
        conn.commit()
        _cache_add_sample(reg_no, norm_embedding)
    except Exception as e:
        print(f"Sample save warning for {reg_no}: {e}")


def _get_candidate_embeddings(
    reg_no: str, source_table: str, primary_embedding: np.ndarray
):
    """Return list of candidate embeddings (primary + recent samples)."""
    candidates = []
    primary = _normalize_embedding(primary_embedding)
    if primary is not None:
        candidates.append(primary)

    try:
        cursor.execute(
            """
            SELECT embedding
            FROM face_embedding_samples
            WHERE reg_no = ? AND source_table = ?
            ORDER BY created_at DESC, id DESC
            LIMIT ?
        """,
            (reg_no, source_table, MAX_PROFILE_SAMPLES),
        )
        rows = cursor.fetchall()
        for (emb_blob,) in rows:
            sample = np.frombuffer(emb_blob, dtype=np.float32)
            if primary is not None and len(sample) != len(primary):
                continue
            sample = _normalize_embedding(sample)
            if sample is not None:
                candidates.append(sample)
    except Exception as e:
        print(f"Sample load warning for {reg_no}: {e}")

    # Remove duplicates while preserving order
    unique_candidates = []
    seen = set()
    for emb in candidates:
        key = emb.tobytes()
        if key in seen:
            continue
        seen.add(key)
        unique_candidates.append(emb)
    return unique_candidates


def _retrain_face_profiles():
    """
    Build updated profile embeddings from stored samples.
    Runs in background on startup and daily schedule.
    """
    run_conn = None
    run_cursor = None
    run_id = None
    updated_profiles = 0
    try:
        run_cursor = pg_adapter.cursor
        run_cursor.execute(
            "INSERT INTO face_training_runs (status, notes) VALUES ('running', %s)",
            ("Daily face profile training started",),
        )
        run_id = run_cursor.lastrowid

        for source_table in ("users", "other_staff"):
            run_cursor.execute(f"SELECT reg_no, embedding FROM {source_table}")
            for reg_no, primary_blob in run_cursor.fetchall():
                vectors = []
                primary_vector = None

                if primary_blob is not None:
                    primary_vector = np.frombuffer(primary_blob, dtype=np.float32)
                    primary_vector = _normalize_embedding(primary_vector)
                    if primary_vector is not None:
                        vectors.append(primary_vector)

                run_cursor.execute(
                    """
                    SELECT embedding
                    FROM face_embedding_samples
                    WHERE reg_no = ? AND source_table = ?
                    ORDER BY created_at DESC, id DESC
                    LIMIT ?
                """,
                    (reg_no, source_table, MAX_PROFILE_SAMPLES),
                )
                sample_rows = run_cursor.fetchall()

                for (emb_blob,) in sample_rows:
                    sample = np.frombuffer(emb_blob, dtype=np.float32)
                    if primary_vector is not None and len(sample) != len(
                        primary_vector
                    ):
                        continue
                    sample = _normalize_embedding(sample)
                    if sample is not None:
                        vectors.append(sample)

                if not vectors:
                    continue

                # Mean profile then normalize
                profile = np.mean(np.vstack(vectors), axis=0).astype(np.float32)
                profile = _normalize_embedding(profile)
                if profile is None:
                    continue

                run_cursor.execute(
                    f"UPDATE {source_table} SET embedding = ? WHERE reg_no = ?",
                    (profile.tobytes(), reg_no),
                )
                updated_profiles += 1

        # Cleanup very old samples
        run_cursor.execute("""
            DELETE FROM face_embedding_samples
            WHERE created_at < NOW() - INTERVAL '120 days'
        """)

        run_cursor.execute(
            """
            UPDATE face_training_runs
            SET status = 'completed',
                completed_at = CURRENT_TIMESTAMP,
                updated_profiles = %s,
                notes = %s
            WHERE id = %s
        """,
            (updated_profiles, f"Updated {updated_profiles} profiles", run_id),
        )
        print(f"Daily face training complete: updated {updated_profiles} profiles")
        _warm_face_embedding_cache()
    except Exception as e:
        print(f"Daily face training failed: {e}")
        if run_id is not None:
            try:
                run_cursor.execute(
                    """
                    UPDATE face_training_runs
                    SET status = 'failed',
                        completed_at = CURRENT_TIMESTAMP,
                        notes = %s
                    WHERE id = %s
                """,
                    (str(e), run_id),
                )
            except Exception:
                pass
    finally:
        pass  # pg_adapter manages connections automatically


_face_training_timer = None


def _schedule_next_daily_training():
    global _face_training_timer
    now = datetime.now()
    next_run = now.replace(
        hour=DAILY_TRAIN_HOUR, minute=DAILY_TRAIN_MINUTE, second=0, microsecond=0
    )
    if next_run <= now:
        next_run = next_run + timedelta(days=1)
    delay_seconds = max(30, int((next_run - now).total_seconds()))

    def _runner():
        try:
            _retrain_face_profiles()
        finally:
            _schedule_next_daily_training()

    _face_training_timer = threading.Timer(delay_seconds, _runner)
    _face_training_timer.daemon = True
    _face_training_timer.start()
    print(f"Next daily face training scheduled at {next_run.isoformat()}")


def start_daily_face_training_scheduler():
    # Warm-up run in background to initialize profile means from existing samples.
    threading.Thread(target=_retrain_face_profiles, daemon=True).start()
    _schedule_next_daily_training()


# -------------------------------------------------
# SECURE FACE VERIFICATION - Using InsightFace with simple cosine similarity
# -------------------------------------------------
def verify_face_identity(
    reg_no: str, query_embedding: np.ndarray
) -> tuple[bool, float, str]:
    """
    Verify if the query face matches the enrolled identity for the given reg_no.
    Uses primary profile + recent successful samples to improve robustness.
    Returns (verified, confidence_score, reason)
    """
    profile = _cache_get_profile(reg_no)
    if profile is None:
        return False, 0.0, f"User {reg_no} not found"

    if profile["primary"] is None:
        return False, 0.0, f"No face data enrolled for {reg_no}"

    query_embedding = _normalize_embedding(query_embedding)
    if query_embedding is None:
        return False, 0.0, "Invalid query embedding"

    candidates = _cache_get_candidates(reg_no)
    if not candidates:
        return False, 0.0, "No valid profile embeddings available"

    best_similarity = -1.0
    best_index = -1
    for idx, candidate in enumerate(candidates):
        if len(candidate) != len(query_embedding):
            continue
        sim = float(np.dot(query_embedding, candidate))
        if sim > best_similarity:
            best_similarity = sim
            best_index = idx

    if best_similarity < 0:
        return False, 0.0, "Embedding dimension mismatch"

    if use_fallback:
        threshold = FALLBACK_BASE_THRESHOLD
    else:
        threshold = (
            INSIGHTFACE_BASE_THRESHOLD
            if len(candidates) > 1
            else max(INSIGHTFACE_BASE_THRESHOLD, 0.70)
        )

    print(f"DEBUG: Verification for {reg_no}")
    print(f"  Candidates: {len(candidates)}")
    print(
        f"  Best cosine similarity: {best_similarity:.4f} (candidate #{best_index + 1})"
    )
    print(f"  Threshold: {threshold:.4f}, fallback={use_fallback}")

    if best_similarity >= threshold:
        return True, best_similarity, "Face verified successfully"

    # Near-match band: helps avoid frequent false negatives for genuine users
    # when similarity is only slightly below threshold.
    if (
        not use_fallback
        and best_similarity >= (threshold - INSIGHTFACE_NEAR_MATCH_MARGIN)
    ):
        print(
            f"  Near-match accepted: {best_similarity:.4f} within margin "
            f"{INSIGHTFACE_NEAR_MATCH_MARGIN:.4f}"
        )
        return True, best_similarity, "Face verified successfully (near-match)"

    reason = "Face does not match - Please try again"
    return False, 0.0, reason


# Thread lock for thread-safe InsightFace inference
_face_app_lock = threading.Lock()

try:
    print("Trying InsightFace buffalo_s...")
    if torch_available:
        gpu_available = torch.cuda.is_available()
        if gpu_available:
            try:
                import onnxruntime as ort
                available_providers = ort.get_available_providers()
                if "CUDAExecutionProvider" in available_providers:
                    providers = ["CUDAExecutionProvider", "CPUExecutionProvider"]
                    ctx_id = 0
                    print("GPU available, using GPU for face analysis")
                else:
                    gpu_available = False
                    providers = ["CPUExecutionProvider"]
                    ctx_id = -1
                    print("ONNX Runtime: CUDAExecutionProvider not available in package. Using CPU.")
            except Exception as e:
                gpu_available = False
                providers = ["CPUExecutionProvider"]
                ctx_id = -1
                print(f"ONNX Runtime check failed ({e}). Using CPU.")
        else:
            providers = ["CPUExecutionProvider"]
            ctx_id = -1
            print("GPU not available, using CPU for face analysis")
    else:
        gpu_available = False
        providers = ["CPUExecutionProvider"]
        ctx_id = -1
        print("PyTorch not available, using CPU for face analysis")
    face_app = FaceAnalysis(name="buffalo_s", providers=providers)
    face_app.prepare(ctx_id=ctx_id, det_size=(640, 640))

    # Warm up the model with a dummy inference to reduce first-request latency
    try:
        dummy_img = np.zeros((640, 640, 3), dtype=np.uint8)
        face_app.get(dummy_img)
        print("Face recognition model warmed up successfully")
    except Exception as e:
        print(f"Model warm-up failed (non-critical): {e}")

    # InsightFace is initialized - the models were loaded successfully
    # The fallback check below was wrong - it tested with a synthetic image that has no face
    # Since the models loaded, InsightFace should work with real face images
    print("InsightFace initialized successfully!")
    print("Note: Face embeddings will be generated from actual camera images")
    use_fallback = False  # InsightFace is initialized correctly

except Exception as e:
    print(f"InsightFace failed: {e}")
    print("🔄 Using fallback face recognition system")
    use_fallback = True

# Log which face recognition mode is active
if use_fallback:
    print("=" * 60)
    print("⚠️  WARNING: Using FALLBACK face recognition!")
    print("⚠️  Fallback uses simple image features, NOT real face embeddings!")
    print("⚠️  This is UNSECURE - ANY face may match ANY other face!")
    print("⚠️  Install onnxruntime and retry for proper face recognition!")
    print("=" * 60)
else:
    print("=" * 60)
    print("Using INSIGHTFACE for face recognition!")
    print("This provides proper face embeddings for secure verification")
    print("=" * 60)

if use_fallback:
    # Use OpenCV for face detection as fallback
    try:
        import cv2

        # Try to load Haar cascade with better parameters
        face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
        )

        # Also load alternative cascades for better detection
        face_cascade_alt = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_alt.xml"
        )
        face_cascade_profile = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_profileface.xml"
        )

        print("Multiple OpenCV face detectors loaded")
    except Exception as e:
        print(f"OpenCV face detection failed: {e}")
        face_cascade = None
        face_cascade_alt = None
        face_cascade_profile = None

print("Face recognition system ready!")
print("Warming up face embedding cache...")
_warm_face_embedding_cache()
start_daily_face_training_scheduler()

# -------------------------------------------------
# WIFI / IP RESTRICTION
# -------------------------------------------------
# No IP restriction - server accessible from any network (public/ngrok)
ALLOWED_IP_PREFIX = None


def check_wifi(request: Request):
    """Enforce college-network-only access when WiFi restriction is enabled.

    Applies to app clients only — web clients rely on geofence for location
    enforcement instead.  The Flutter app sends the X-Client-Platform header
    so this function can distinguish web vs app requests.

    When WiFi is required (allow_any_network is False), app clients must
    originate from a private/RFC-1918 IP address (college intranet).  If the
    server is reached via a Cloudflare tunnel the real client IP is taken from
    the CF-Connecting-IP / X-Forwarded-For headers by get_client_ip().
    """
    # Skip if WiFi requirement is disabled (admin toggled off)
    if _app_settings.get("allow_any_network", True):
        return

    # Identify platform — Flutter app sends X-Client-Platform: app
    # Web browser clients send X-Client-Platform: web (or no header)
    platform = request.headers.get("X-Client-Platform", "app").lower()

    # Web clients are exempt from WiFi check — they use geofence instead
    if platform == "web":
        return

    # App clients must be on the college local network (private IP)
    client_ip = get_client_ip(request)
    try:
        ip_obj = ipaddress.ip_address(client_ip)
        if ip_obj.is_private or ip_obj.is_loopback:
            return  # On local / college network — allow
    except Exception:
        pass

    # Client has a public IP — not on the college WiFi
    raise HTTPException(
        status_code=403,
        detail="Access denied. You must be connected to the college WiFi network to mark attendance.",
    )


# -------------------------------------------------
# IMAGE PREPROCESSING
# -------------------------------------------------
def analyze_image_comprehensive(img):
    """
    Comprehensive image analysis combining quality checks and liveness detection.
    Performs all analyses in a single pass to avoid redundant operations.
    Returns a dict with all metrics, quality assessment, and liveness detection.
    """
    try:
        # Convert to grayscale once for all analyses
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        h, w = gray.shape

        # Quality metrics
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
        mean_brightness = np.mean(gray)
        contrast = gray.std()

        # Liveness detection metrics
        sobelx = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
        sobely = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
        gradient_magnitude = np.sqrt(sobelx**2 + sobely**2)
        gradient_mean = np.mean(gradient_magnitude)

        edges = cv2.Canny(gray, 50, 150)
        edge_density = np.sum(edges > 0) / (h * w)

        # Color channel analysis for liveness
        b, g, r = cv2.split(img)
        rg_diff = np.mean(np.abs(r.astype(float) - g.astype(float)))
        gb_diff = np.mean(np.abs(g.astype(float) - b.astype(float)))
        color_std = np.std([np.mean(r), np.mean(g), np.mean(b)])

        # Quality assessment
        warnings = []
        quality_score = 100

        # Blurriness check
        if laplacian_var < 100:
            warnings.append("⚠️ Image is too blurry - please hold camera steady")
            quality_score -= 30
        elif laplacian_var < 200:
            warnings.append("⚠️ Image is slightly blurry")
            quality_score -= 10

        # Brightness check
        if mean_brightness < 50:
            warnings.append("⚠️ Image is too dark - please increase lighting")
            quality_score -= 25
        elif mean_brightness > 200:
            warnings.append("⚠️ Image is too bright - please reduce lighting")
            quality_score -= 20
        elif mean_brightness < 80:
            warnings.append("⚠️ Image is dim - consider better lighting")
            quality_score -= 10

        # Contrast check
        if contrast < 30:
            warnings.append("⚠️ Low contrast - image may be too flat")
            quality_score -= 15

        # Liveness assessment
        liveness_warnings = []

        if ANTISPOOFING_ENABLED:
            if gradient_mean < 4.0:
                liveness_warnings.append(f"Low gradient - possible flat photo/screen (gradient: {gradient_mean:.2f})")

            if edge_density < 0.005:
                liveness_warnings.append(f"Very low edge density - possible smooth photo/screen (edges: {edge_density:.4f})")
            elif edge_density > 0.85:
                liveness_warnings.append(f"Very high edge density - possible digital screen noise (edges: {edge_density:.4f})")

            if rg_diff < 1.5 and gb_diff < 1.5 and color_std < 2.0:
                liveness_warnings.append("Unusual color distribution - possible digital screen/photo replay")

            # FFT Moire, glare detection, and flat color gamut check are bypassed to permanently prevent close-range (<3 meter) false positives.

        is_live = len(liveness_warnings) == 0
        is_poor_quality = quality_score < 60

        return {
            # Quality metrics
            "blur_score": laplacian_var,
            "brightness": mean_brightness,
            "contrast": contrast,
            "quality_score": max(0, quality_score),
            "quality_warnings": warnings,
            "is_poor_quality": is_poor_quality,

            # Liveness metrics
            "gradient_mean": gradient_mean,
            "edge_density": edge_density,
            "color_consistency": color_std,
            "liveness_warnings": liveness_warnings,
            "is_live": is_live,
            "liveness_reason": "Real face detected" if is_live else "; ".join(liveness_warnings),

            # Combined assessment
            "overall_pass": not is_poor_quality and is_live,
            "all_warnings": warnings + liveness_warnings
        }

    except Exception as e:
        print(f"Comprehensive image analysis error: {e}")
        return {
            "error": str(e),
            "overall_pass": False,
            "quality_warnings": ["Analysis failed"],
            "liveness_warnings": ["Analysis failed"],
            "all_warnings": ["Image analysis failed - please try again"]
        }


def check_image_quality(img):
    """
    Legacy function - now delegates to comprehensive analysis.
    Maintained for backward compatibility.
    """
    result = analyze_image_comprehensive(img)
    return {
        "blur_score": result.get("blur_score", 0),
        "brightness": result.get("brightness", 0),
        "contrast": result.get("contrast", 0),
        "quality_score": result.get("quality_score", 0),
        "warnings": result.get("quality_warnings", []),
        "is_poor_quality": result.get("is_poor_quality", True),
    }


def preprocess_image_data(img_np_or_bytes):
    if isinstance(img_np_or_bytes, bytes):
        nparr = np.frombuffer(img_np_or_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    else:
        img = img_np_or_bytes

    if img is None:
        raise HTTPException(status_code=400, detail="Invalid image data")

    # Fix front-camera mirroring
    img = cv2.flip(img, 1)

    h, w, _ = img.shape
    max_size = 1024
    if max(h, w) > max_size:
        scale = max_size / max(h, w)
        img = cv2.resize(img, (int(w * scale), int(h * scale)))

    # Gamma Correction - more subtle for better quality
    gamma = 1.2
    invGamma = 1.0 / gamma
    table = np.array(
        [((i / 255.0) ** invGamma) * 255 for i in np.arange(0, 256)]
    ).astype("uint8")
    img = cv2.LUT(img, table)

    return img


# -------------------------------------------------
# FACE EXTRACTION - GUARANTEED TO WORK
# -------------------------------------------------
def extract_face(img, _recursion_depth=0):
    """
    Extract face from image with quality checking and recursion protection.

    Args:
        img: Input image (BGR format)
        _recursion_depth: Internal counter to prevent infinite loops (don't set manually)

    Returns:
        Face object with embedding, or None if extraction fails
    """
    print("--- Starting Face Extraction ---")
    global use_fallback, extract_face_recursion_depth

    # Check recursion depth to prevent infinite loop
    if _recursion_depth > 2:
        print("Maximum recursion depth reached. Face detection failing repeatedly.")
        print(
            "⚠️  Possible causes: poor camera quality, no face in image, or system misconfiguration"
        )
        return None

    # Note: Comprehensive image analysis is now done before extract_face is called

    if use_fallback:
        # Use OpenCV face detection with multiple methods
        if face_cascade is None:
            print("Face detection not available - please contact administrator")
            return None

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # Try different detection parameters
        all_faces = []

        # Method 1: Default parameters
        faces1 = face_cascade.detectMultiScale(gray, 1.1, 3, minSize=(30, 30))
        print(f"Method 1 detected {len(faces1)} faces")
        all_faces.extend(faces1)

        # Method 2: More sensitive
        faces2 = face_cascade.detectMultiScale(gray, 1.05, 5, minSize=(20, 20))
        print(f"Method 2 detected {len(faces2)} faces")
        all_faces.extend(faces2)

        # Method 3: Alternative cascade if available
        if face_cascade_alt is not None:
            faces3 = face_cascade_alt.detectMultiScale(gray, 1.1, 4, minSize=(30, 30))
            print(f"Method 3 (alt) detected {len(faces3)} faces")
            all_faces.extend(faces3)

        # Method 4: Profile faces if available
        if face_cascade_profile is not None:
            faces4 = face_cascade_profile.detectMultiScale(
                gray, 1.1, 4, minSize=(30, 30)
            )
            print(f"Method 4 (profile) detected {len(faces4)} faces")
            all_faces.extend(faces4)

        if len(all_faces) == 0:
            print("No faces detected with any OpenCV method")
            # Last resort: create a fake face region from center of image
            h, w = gray.shape
            center_x, center_y = w // 2, h // 2
            fake_face = (max(0, center_x - 100), max(0, center_y - 100), 200, 200)
            all_faces.append(fake_face)
            print("🔄 Using center region as fallback face")

        # Get the largest face
        largest_face = max(all_faces, key=lambda f: f[2] * f[3])
        x, y, w, h = largest_face

        print(f"Best face detected at: {x}, {y}, {w}, {h}")

        # Ensure face region is valid
        x = max(0, x)
        y = max(0, y)
        x2 = min(gray.shape[1], x + w)
        y2 = min(gray.shape[0], y + h)

        # Extract face region
        face_region = img[y:y2, x:x2]

        # Create fallback embedding
        embedding = create_fallback_embedding(face_region)

        # Create a face-like object
        class FaceResult:
            def __init__(self, bbox, embedding):
                self.bbox = np.array([x, y, x2, y2], dtype=float)
                self.det_score = 0.8  # Fixed score
                self.embedding = embedding

        face_result = FaceResult([x, y, x2, y2], embedding)
        print(f"Fallback embedding created, shape: {embedding.shape}")
        print(
            "⚠️  WARNING: Using fallback embedding - NOT secure for face verification!"
        )
        print(
            "⚠️  Fallback uses image statistics which cannot distinguish between faces!"
        )
        return face_result

    else:
        # Use InsightFace
        # Only check Original (upright) rotation to ensure high performance and prevent long loading times
        rotations = [None]
        rot_names = ["Original"]

        for i, rot in enumerate(rotations):
            temp = img if rot is None else cv2.rotate(img, rot)

            try:
                with _face_app_lock:
                    faces = face_app.get(temp)
                if len(faces) == 0:
                    continue

                # Pick largest face
                face = max(
                    faces,
                    key=lambda f: (f.bbox[2] - f.bbox[0]) * (f.bbox[3] - f.bbox[1]),
                )

                if face.embedding is not None and len(face.embedding) > 0:
                    print(f"InsightFace embedding found in rotation {rot_names[i]}")
                    return face
                else:
                    print(f"No embedding in rotation {rot_names[i]}")

            except Exception as e:
                print(f"Error in rotation {rot_names[i]}: {e}")
                continue

        print("InsightFace failed to detect any face.")
        return None


# -------------------------------------------------
# DEBUG UTILS
# -------------------------------------------------
DEBUG_DIR = "debug_images"
if not os.path.exists(DEBUG_DIR):
    os.makedirs(DEBUG_DIR)


def save_debug_image(img, prefix="debug"):
    timestamp = datetime.now().strftime("%H%M%S")
    filename = f"{DEBUG_DIR}/{prefix}_{timestamp}.jpg"
    cv2.imwrite(filename, img)
    print(f"Saved debug image: {filename}")


# -------------------------------------------------
# CONFIG ENDPOINT
# -------------------------------------------------
@app.get("/config")
async def get_config():
    """Get face recognition configuration and status"""
    return {
        "face_recognition_mode": "fallback" if use_fallback else "insightface",
        "is_secure": not use_fallback,
        "confidence_threshold": 0.95 if use_fallback else 0.70,
        "warning": "Fallback mode is not secure for face recognition"
        if use_fallback
        else None,
        "antispoofing_enabled": ANTISPOOFING_ENABLED,
    }


# -------------------------------------------------
# ANTISPOOFING CONFIGURATION
# -------------------------------------------------
@app.post("/admin/config/antispoofing")
async def toggle_antispoofing(request: Request, enabled: bool = True):
    """
    Enable or disable antispoofing (liveness detection).
    Admin only endpoint.
    """
    verify_admin_token(request)

    global ANTISPOOFING_ENABLED
    ANTISPOOFING_ENABLED = enabled

    status = "enabled" if enabled else "disabled"
    print(f"⚠️  Anti-spoofing {status} by admin")
    log_audit_event("ANTISPOOFING_TOGGLE", "admin", True, f"Anti-spoofing {status}")

    return {
        "message": f"Anti-spoofing {status} successfully",
        "antispoofing_enabled": ANTISPOOFING_ENABLED,
    }


# -------------------------------------------------
# MARK ATTENDANCE - SECURE VERSION
# -------------------------------------------------
def process_attendance_background(
    reg_no: str, img_bytes: bytes, user_role: str, form_data: dict, request_data: dict
):
    """Background task to process attendance verification and marking."""
    try:
        print(f"🔄 Starting background attendance processing for {reg_no} (request: {request_data.get('request_id', 'unknown')})")

        # Process the attendance verification (this is the heavy lifting)
        # Note: This runs in a separate thread, so synchronous DB operations are fine
        try:
            result = _secure_verify_and_mark(reg_no, img_bytes, user_role, form_data)
        except HTTPException as e:
            # Expected verification/business-rule failures should not be treated as crashes.
            status_code = int(getattr(e, "status_code", 500) or 500)
            detail = str(getattr(e, "detail", "Attendance processing failed"))
            result = {"success": False, "error": detail, "status_code": status_code}

            if status_code in (400, 401, 403):
                print(f"⚠️ Background attendance validation failed for {reg_no}: {detail}")
            else:
                # Preserve signal for truly unexpected HTTP exceptions.
                raise

        # Log the result
        if result.get("success"):
            print(f"✅ Background attendance processed successfully for {reg_no}")
            log_audit_event(
                "ATTENDANCE_BACKGROUND_SUCCESS",
                reg_no,
                True,
                f"Background processing completed successfully"
            )
        else:
            print(f"❌ Background attendance failed for {reg_no}: {result.get('error', 'Unknown error')}")
            log_audit_event(
                "ATTENDANCE_BACKGROUND_FAILED",
                reg_no,
                False,
                f"Background processing failed: {result.get('error', 'Unknown error')}"
            )

    except Exception as e:
        print(f"💥 Background attendance processing error for {reg_no}: {e}")
        import traceback
        traceback.print_exc()
        log_audit_event(
            "ATTENDANCE_BACKGROUND_ERROR",
            reg_no,
            False,
            f"Background processing error: {str(e)}"
        )


def _attendance_sync_work(
    reg_no, user_role, form_data, active_slot_type, img_bytes, active_slot=None, active_slot_half=None
):
    """Run the blocking attendance marking work in a thread pool.
    
    This function contains all the synchronous DB and face recognition
    operations extracted from the async endpoint. It runs in a thread pool
    executor to avoid blocking the event loop.
    """
    return _secure_verify_and_mark(reg_no, img_bytes, user_role, form_data, active_slot_type, active_slot, active_slot_half)


def get_ccl_settings_for_date(date_str: str = None) -> dict:
    if date_str is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
    try:
        cursor.execute("""
            SELECT early_enabled, early_start, early_end, late_enabled, late_start, late_end, early_duration, late_duration
            FROM ccl_custom_dates 
            WHERE ccl_date = ?
        """, (date_str,))
        row = cursor.fetchone()
        if row:
            return {
                "early_check_in_ccl_enabled": bool(row[0]) if row[0] is not None else False,
                "early_check_in_start": row[1] or "07:00",
                "early_check_in_end": row[2] or "08:00",
                "early_duration": int(row[6]) if row[6] is not None else 60,
                "late_check_out_ccl_enabled": bool(row[3]) if row[3] is not None else False,
                "late_check_out_start": row[4] or "17:00",
                "late_check_out_end": row[5] or "18:00",
                "late_duration": int(row[7]) if row[7] is not None else 60
            }
    except Exception as e:
        print(f"Error checking custom CCL dates: {e}")

    try:
        cursor.execute("SELECT key, value FROM ccl_settings")
        ccl_settings = {r[0]: r[1] for r in cursor.fetchall()}
    except Exception as e:
        print(f"Error fetching global ccl_settings: {e}")
        ccl_settings = {}
        
    return {
        "early_check_in_ccl_enabled": ccl_settings.get("early_check_in_ccl_enabled", "false") == "true",
        "early_check_in_start": ccl_settings.get("early_check_in_start", "07:00"),
        "early_check_in_end": ccl_settings.get("early_check_in_end", "08:00"),
        "early_duration": 60,
        "late_check_out_ccl_enabled": ccl_settings.get("late_check_out_ccl_enabled", "false") == "true",
        "late_check_out_start": ccl_settings.get("late_check_out_start", "17:00"),
        "late_check_out_end": ccl_settings.get("late_check_out_end", "18:00"),
        "late_duration": 60
    }


def get_active_ccl_slot_type(date_str: str = None) -> str:
    """Returns 'check_in' or 'check_out' if current time falls in enabled CCL windows, else None."""
    try:
        if date_str is None:
            date_str = datetime.now().strftime("%Y-%m-%d")
        settings = get_ccl_settings_for_date(date_str)
        
        early_enabled = settings["early_check_in_ccl_enabled"]
        late_enabled = settings["late_check_out_ccl_enabled"]
        
        if not (early_enabled or late_enabled):
            return None
            
        now_time = datetime.now()
        curr_min = now_time.hour * 60 + now_time.minute
        
        def time_to_min(t_str):
            p = t_str.split(":")
            return int(p[0]) * 60 + int(p[1])
            
        if early_enabled:
            early_start = time_to_min(settings["early_check_in_start"])
            early_end = time_to_min(settings["early_check_in_end"])
            if early_start <= curr_min < early_end:
                return "check_in"
                
        if late_enabled:
            late_start = time_to_min(settings["late_check_out_start"])
            late_end = time_to_min(settings["late_check_out_end"])
            if late_start <= curr_min < late_end:
                return "check_out"
    except Exception as e:
        print(f"Error checking active CCL slot: {e}")
    return None


@app.post("/mark_attendance")
async def mark_attendance_secure(
    background_tasks: BackgroundTasks,
    request: Request,
    reg_no: str = Form(None),
    client_platform: str = Form("app"),
    client_lat: str = Form(None),
    client_lng: str = Form(None),
    image: UploadFile = File(...),
):
    """Secure attendance marking endpoint with concurrency and rate limiting."""
    today_str = datetime.now().strftime("%Y-%m-%d")
    academic_status, academic_reason, is_holiday = _academic_status_for_date(today_str)
    if is_holiday:
        if reg_no is None:
            reg_no = request.query_params.get("reg_no")
        if reg_no is None:
            content_type = request.headers.get("content-type", "").lower()
            if "application/json" in content_type:
                try:
                    body = await request.json()
                    if body:
                        reg_no = body.get("reg_no")
                except:
                    pass
        if reg_no:
            user_row = None
            cursor.execute(
                """
                SELECT reg_no, name, dept FROM users WHERE reg_no = ?
                UNION ALL
                SELECT reg_no, name, dept FROM other_staff WHERE reg_no = ?
            """,
                (reg_no, reg_no),
            )
            user_row = cursor.fetchone()
            if user_row:
                _ensure_daily_holiday_status(
                    user_row[0], user_row[1], user_row[2], today_str, academic_reason
                )
        raise HTTPException(
            status_code=403,
            detail="Attendance is blocked today because the date is marked as a holiday in the academic calendar.",
        )

    cursor.execute("""
        SELECT slot_number, start_time, duration_minutes, is_enabled, slot_type, slot_half
        FROM attendance_duration_settings
        WHERE is_enabled = 1
        ORDER BY slot_number ASC
    """)
    duration_rows = cursor.fetchall()

    active_slot = None
    active_slot_type = "check_in"
    active_slot_half = None  # 'first_half' | 'second_half' | None

    if duration_rows:
        current_time = datetime.now()
        allowed = False

        for row in duration_rows:
            slot_number = row[0]
            start_time = row[1]
            duration_minutes = row[2]
            slot_type = row[4] if len(row) > 4 and row[4] else "check_in"
            slot_half = row[5] if len(row) > 5 and row[5] else "full_day"
            effective_duration, _, _ = _calculate_effective_slot_duration(slot_type, slot_half, duration_minutes)

            start_hour, start_minute = map(int, start_time.split(":"))
            start_datetime = current_time.replace(
                hour=start_hour, minute=start_minute, second=0, microsecond=0
            )
            end_datetime = start_datetime + timedelta(minutes=effective_duration)

            if start_datetime <= current_time < end_datetime:
                allowed = True
                active_slot = slot_number
                active_slot_type = slot_type
                active_slot_half = slot_half if slot_half in ("first_half", "second_half") else None
                break

        if not allowed:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                allowed = True
                active_slot = -1
                active_slot_type = ccl_slot_type
                
        if not allowed:
            slots_info = ", ".join(
                [f"Slot {row[0]} ({row[4] if len(row) > 4 and row[4] else 'check_in'} / {row[5] if len(row) > 5 and row[5] else 'full_day'}): {row[1]} ({row[2]} min)" for row in duration_rows]
            )
            raise HTTPException(
                status_code=403,
                detail=f"Attendance marking is not allowed at this time. Available slots: {slots_info}",
            )
    else:
        settings = get_ccl_settings_for_date()
        early_enabled = settings["early_check_in_ccl_enabled"]
        late_enabled = settings["late_check_out_ccl_enabled"]
        
        if early_enabled or late_enabled:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                active_slot = -1
                active_slot_type = ccl_slot_type
            else:
                raise HTTPException(
                    status_code=403,
                    detail="Attendance marking is not allowed at this time. (Outside CCL window)",
                )

    if reg_no is None:
        reg_no = request.query_params.get("reg_no")

    content_type = request.headers.get("content-type", "").lower()
    if reg_no is None and "application/json" in content_type:
        try:
            body = await request.json()
            if body:
                reg_no = body.get("reg_no")
        except:
            pass

    if reg_no:
        is_locked, remaining = check_lockout(reg_no)
        if is_locked:
            log_audit_event(
                "LOCKOUT_ATTEMPT", reg_no, False, f"Locked out for {remaining}s"
            )
            raise HTTPException(
                status_code=423,
                detail=f"Account locked due to multiple failed attempts. Try again in {remaining} seconds.",
            )

        if active_slot is not None:
            if active_slot_type == "check_out":
                cursor.execute(
                    "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
                    (reg_no,)
                )
                cin_count = cursor.fetchone()[0]
                if cin_count == 0:
                    cursor.execute(
                        "SELECT COUNT(*) FROM other_staff_attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
                        (reg_no,)
                    )
                    cin_count = cursor.fetchone()[0]
                if cin_count == 0:
                    raise HTTPException(
                        status_code=400,
                        detail="You cannot Check-Out without checking in first for today."
                    )

            _hd_cfg = _get_half_day_settings()
            _req_fn_checkout = _hd_cfg.get("require_fn_check_out", False)

            if active_slot_half == "first_half":
                if active_slot_type == "check_out" and not _req_fn_checkout:
                    raise HTTPException(
                        status_code=400,
                        detail="Morning (FN) Check-Out is currently cancelled/bypassed by system policy. No check-out scan is required."
                    )

                cursor.execute(
                    """
                    SELECT COUNT(*) FROM attendance 
                    WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                    """,
                    (reg_no, active_slot_type)
                )
                dup_count = cursor.fetchone()[0]
                if dup_count == 0:
                    cursor.execute(
                        """
                        SELECT COUNT(*) FROM other_staff_attendance 
                        WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                        """,
                        (reg_no, active_slot_type)
                    )
                    dup_count = cursor.fetchone()[0]

                if dup_count > 0:
                    slot_name = "Check-In" if active_slot_type == "check_in" else "Check-Out"
                    raise HTTPException(
                        status_code=403,
                        detail=f"You have already marked Morning (FN) {slot_name} attendance for today."
                    )
            elif active_slot_half == "second_half":
                cursor.execute(
                    """
                    SELECT COUNT(*) FROM attendance 
                    WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                    """,
                    (reg_no, active_slot_type)
                )
                dup_count = cursor.fetchone()[0]
                if dup_count == 0:
                    cursor.execute(
                        """
                        SELECT COUNT(*) FROM other_staff_attendance 
                        WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                        """,
                        (reg_no, active_slot_type)
                    )
                    dup_count = cursor.fetchone()[0]

                if dup_count > 0:
                    slot_name = "Check-In" if active_slot_type == "check_in" else "Check-Out"
                    raise HTTPException(
                        status_code=403,
                        detail=f"You have already marked Afternoon (AN) {slot_name} attendance for today."
                    )
            else:

                # Legacy full-day mode duplicate check
                cursor.execute(
                    """
                    SELECT COUNT(*) FROM daily_attendance_status
                    WHERE reg_no = ? AND date = ? AND status = 'Present'
                    """,
                    (reg_no, today_str)
                )
                is_present_today = cursor.fetchone()[0] > 0

                if is_present_today and active_slot_type == "check_in":
                    raise HTTPException(
                        status_code=403,
                        detail="You are already marked present for today."
                    )

                cursor.execute(
                    """
                    SELECT COUNT(*) FROM attendance 
                    WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                    """,
                    (reg_no, active_slot_type)
                )
                dup_count = cursor.fetchone()[0]

                if dup_count == 0:
                    cursor.execute(
                        """
                        SELECT COUNT(*) FROM other_staff_attendance 
                        WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                        """,
                        (reg_no, active_slot_type)
                    )
                    dup_count = cursor.fetchone()[0]

                if dup_count > 0:
                    raise HTTPException(
                        status_code=403,
                        detail=f"You have already marked {active_slot_type.replace('_', ' ')} attendance for today."
                    )

    form_data = None
    try:
        form_data = await request.form()
    except:
        pass

    platform = client_platform or "app"
    if form_data is not None and form_data.get("client_platform"):
        platform = form_data.get("client_platform")
    else:
        hdr_platform = request.headers.get("X-Client-Platform")
        if hdr_platform:
            platform = hdr_platform.lower()

    lat = client_lat
    if lat is None:
        lat = request.query_params.get("client_lat")
    if lat is None and form_data is not None:
        lat = form_data.get("client_lat")

    lng = client_lng
    if lng is None:
        lng = request.query_params.get("client_lng")
    if lng is None and form_data is not None:
        lng = form_data.get("client_lng")

    validation_form = {
        "client_platform": platform,
        "client_lat": lat,
        "client_lng": lng,
        "reg_no": reg_no
    }

    # Run geofence first — if it's enabled and passes, WiFi check is redundant
    try:
        geofence_active = _enforce_web_geofence(validation_form, get_client_ip(request))
        if not geofence_active:
            check_wifi(request)
    except HTTPException as e:
        if reg_no:
            record_failed_attempt(reg_no)
        raise

    form_data = validation_form

    img_bytes = await image.read()

    if len(img_bytes) == 0:
        raise HTTPException(status_code=400, detail="Empty image")

    if reg_no is None:
        log_audit_event(
            "ATTENDANCE_NO_REG_NO", None, False, "No registration number provided"
        )
        raise HTTPException(
            status_code=400,
            detail="Registration number is required to mark attendance. Please provide your staff ID.",
        )

    logged_in_user = None
    user_role = None

    try:
        logged_in_user = verify_staff_token(request)
        user_role = "staff"
    except HTTPException:
        pass

    if logged_in_user is None:
        try:
            logged_in_user = verify_admin_token(request)
            user_role = "admin"
        except HTTPException:
            pass

    if logged_in_user is None:
        try:
            logged_in_user = verify_hod_token(request)
            user_role = "hod"
        except HTTPException:
            pass

    if logged_in_user is None:
        try:
            logged_in_user = verify_user_token(request)
            user_role = (
                logged_in_user.get("role", "staff") if logged_in_user else "staff"
            )
        except HTTPException:
            pass

    if logged_in_user is None:
        log_audit_event(
            "ATTENDANCE_NO_TOKEN",
            reg_no,
            False,
            "No valid authentication token provided",
        )
        raise HTTPException(
            status_code=401, detail="Valid authentication required to mark attendance."
        )

    if logged_in_user and logged_in_user.get("reg_no") != reg_no:
        log_audit_event(
            "ATTENDANCE_MISMATCH",
            reg_no,
            False,
            f"Token user {logged_in_user.get('reg_no')} tried to mark attendance for {reg_no}",
        )
        raise HTTPException(
            status_code=403,
            detail="You can only mark attendance for yourself. Please use your own credentials.",
        )

    # ---- RATE LIMITING ----
    allowed, retry_after = attendance_rate_limiter.is_allowed(reg_no, "mark_attendance")
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail=f"Too many requests. Please wait {retry_after} seconds before trying again.",
        )

    # ---- PER-USER CONCURRENCY GUARD ----
    # Prevents multiple simultaneous verification requests for the same user
    # (e.g., when a user opens multiple tabs and the camera is accessed twice)
    async with acquire_user_lock(reg_no):
        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(
            _cpu_executor,
            _attendance_sync_work,
            reg_no, user_role, form_data, active_slot_type, img_bytes, active_slot, active_slot_half
        )
        return result


def _ts(val):
    """Convert timestamp value to string for JSON serialization."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def _date_str(val):
    """Convert date value to string for JSON serialization."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)


def _secure_verify_and_mark(
    reg_no: str, img_bytes: bytes, user_role: str = "staff", form_data=None,
    slot_type: str = "check_in", active_slot: int = None, active_slot_half: str = None
):
    """Internal function for secure verification with identity binding."""
    # Check lockout status
    is_locked, remaining = check_lockout(reg_no)
    if is_locked:
        log_audit_event(
            "LOCKOUT_ATTEMPT", reg_no, False, f"Locked out for {remaining}s"
        )
        raise HTTPException(
            status_code=423,
            detail=f"Account locked due to multiple failed attempts. Try again in {remaining} seconds.",
        )

    # SECURE DOUBLE CHECK: Enforce duplicate check for the active slot
    today_date_str = datetime.now().strftime("%Y-%m-%d")

    if active_slot_half == "first_half":
        _hd_cfg = _get_half_day_settings()
        _req_fn_checkout = _hd_cfg.get("require_fn_check_out", False)
        chk_slot_type = slot_type or "check_in"
        if chk_slot_type == "check_out" and not _req_fn_checkout:
            raise HTTPException(
                status_code=400,
                detail="Morning (FN) Check-Out is currently cancelled/bypassed by system policy. No check-out scan is required."
            )

        cursor.execute(
            """
            SELECT COUNT(*) FROM attendance 
            WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
            """,
            (reg_no, chk_slot_type)
        )
        dup_count = cursor.fetchone()[0]
        if dup_count == 0:
            cursor.execute(
                """
                SELECT COUNT(*) FROM other_staff_attendance 
                WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                """,
                (reg_no, chk_slot_type)
            )
            dup_count = cursor.fetchone()[0]

        if dup_count > 0:
            slot_name = "Check-In" if chk_slot_type == "check_in" else "Check-Out"
            raise HTTPException(
                status_code=403,
                detail=f"You have already marked Morning (FN) {slot_name} attendance for today."
            )
    elif active_slot_half == "second_half":
        chk_slot_type = slot_type or "check_in"
        cursor.execute(
            """
            SELECT COUNT(*) FROM attendance 
            WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
            """,
            (reg_no, chk_slot_type)
        )
        dup_count = cursor.fetchone()[0]
        if dup_count == 0:
            cursor.execute(
                """
                SELECT COUNT(*) FROM other_staff_attendance 
                WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                """,
                (reg_no, chk_slot_type)
            )
            dup_count = cursor.fetchone()[0]

        if dup_count > 0:
            slot_name = "Check-In" if chk_slot_type == "check_in" else "Check-Out"
            raise HTTPException(
                status_code=403,
                detail=f"You have already marked Afternoon (AN) {slot_name} attendance for today."
            )
    else:

        # Legacy full-day check_in/check_out mode
        chk_slot_type = slot_type or "check_in"
        cursor.execute(
            """
            SELECT COUNT(*) FROM daily_attendance_status
            WHERE reg_no = ? AND date = ? AND status = 'Present'
            """,
            (reg_no, today_date_str)
        )
        is_present_today = cursor.fetchone()[0] > 0

        if is_present_today and chk_slot_type == "check_in":
            raise HTTPException(
                status_code=403,
                detail="You are already marked present for today."
            )

        cursor.execute(
            """
            SELECT COUNT(*) FROM attendance 
            WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
            """,
            (reg_no, chk_slot_type)
        )
        dup_count = cursor.fetchone()[0]

        if dup_count == 0:
            cursor.execute(
                """
                SELECT COUNT(*) FROM other_staff_attendance 
                WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = ?
                """,
                (reg_no, chk_slot_type)
            )
            dup_count = cursor.fetchone()[0]

        if dup_count > 0:
            raise HTTPException(
                status_code=403,
                detail=f"You have already marked {chk_slot_type.replace('_', ' ')} attendance for today."
            )


    # SECURE DOUBLE CHECK: Enforce geofencing and WiFi checks directly at the validation level
    if form_data:
        try:
            # We don't have request here directly but we can enforce IP/SSID and coordinates from form_data
            # Enforce geofence on the form_data coordinates
            from fastapi import Request
            # Check geofencing on the resolved coordinates in form_data
            _enforce_web_geofence(form_data, "0.0.0.0")  # client_ip is secondary for geofence coordinates check
        except HTTPException as e:
            record_failed_attempt(reg_no)
            raise e

    # Check if user exists in cache (backed by users/other_staff tables)
    profile = _cache_get_profile(reg_no)

    user_type = None
    name = None
    dept = None
    embedding = None
    is_other_staff = False

    if profile is not None:
        user_type = profile["role"]
        name = profile["name"]
        dept = profile["dept"]
        embedding = profile["primary"]
        is_other_staff = profile["source_table"] == "other_staff"
    else:
        # Cache miss - fall back to DB lookup (user may have been added after startup)
        cursor.execute(
            "SELECT reg_no, name, dept, embedding, role FROM users WHERE reg_no = ?",
            (reg_no,),
        )
        user = cursor.fetchone()

        if user is not None:
            user_type = user[4]
            name = user[1]
            dept = user[2]
            embedding = user[3]
            is_other_staff = False
        else:
            cursor.execute(
                "SELECT reg_no, name, dept, embedding, role FROM other_staff WHERE reg_no = ?",
                (reg_no,),
            )
            other_user = cursor.fetchone()

            if other_user is not None:
                user_type = other_user[4]
                name = other_user[1]
                dept = other_user[2]
                embedding = other_user[3]
                is_other_staff = True
            else:
                log_audit_event(
                    "UNKNOWN_USER",
                    reg_no,
                    False,
                    "User not found in users or other_staff table",
                )
                raise HTTPException(
                    status_code=404,
                    detail="User not found. Please check your registration number.",
                )

    # Warning if using fallback mode
    if use_fallback:
        print(
            "⚠️  WARNING: Using FALLBACK mode for attendance - HIGH RISK OF FALSE POSITIVES!"
        )
        log_audit_event(
            "ATTENDANCE_FALLBACK_MODE", reg_no, False, "Using insecure fallback mode"
        )

    # CRITICAL: Check if face is registered before allowing attendance
    if embedding is None:
        log_audit_event(
            "FACE_NOT_REGISTERED",
            reg_no,
            False,
            f"Staff/HOD/Admin attempted attendance without face registration",
        )
        raise HTTPException(
            status_code=403,
            detail="Face not registered. Please register your face first before marking attendance. Contact your HOD or admin for assistance.",
        )

    try:
        img = preprocess_image_data(img_bytes)
    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Preprocessing error: {e}")
        log_audit_event("PREPROCESSING_ERROR", reg_no, False, str(e))
        raise HTTPException(status_code=400, detail="Image processing failed")

    # Comprehensive image analysis (quality + liveness in one pass)
    analysis = analyze_image_comprehensive(img)

    # Early rejection for poor quality images
    if analysis.get("is_poor_quality", False):
        quality_msg = "; ".join(analysis.get("quality_warnings", []))
        print(f"[ATTENDANCE FAILURE] RegNo: {reg_no} | Reason: Image quality too poor ({quality_msg}) | Quality Score: {analysis['quality_score']:.2f}")
        log_audit_event("QUALITY_FAILED", reg_no, False, quality_msg)
        raise HTTPException(
            status_code=400,
            detail="Image quality is too poor. Please improve lighting and camera stability.",
        )

    face = extract_face(img)

    if face is None:
        save_debug_image(img, f"fail_detection_{reg_no}")
        print(f"[ATTENDANCE FAILURE] RegNo: {reg_no} | Reason: Face not detected | Quality Score: {analysis['quality_score']:.2f}")
        log_audit_event("NO_FACE_DETECTED", reg_no, False)
        raise HTTPException(
            status_code=400,
            detail="Face is unable to detect. Please adjust your position and ensure good lighting.",
        )

    # =====================================================
    # LIVENESS CHECK - Use results from comprehensive analysis
    # =====================================================
    is_live = analysis.get("is_live", True)
    liveness_reason = analysis.get("liveness_reason", "Analysis completed")

    if not is_live:
        if ANTISPOOF_STRICT_MODE:
            save_debug_image(img, f"liveness_fail_{reg_no}")
            print(f"[ATTENDANCE FAILURE] RegNo: {reg_no} | Reason: Liveness check failed ({liveness_reason}) | Quality Score: {analysis['quality_score']:.2f}")
            log_audit_event("LIVENESS_FAILED", reg_no, False, liveness_reason)
            raise HTTPException(
                status_code=400,
                detail=f"Liveness check failed: {liveness_reason}. Please ensure you are using a live camera feed, not a photo or screen.",
            )
        else:
            # In non-strict mode, just log warning but allow
            log_audit_event("LIVENESS_WARNING", reg_no, False, liveness_reason)

    query_embedding = face.embedding.astype(np.float32)

    # Verify face identity against enrolled data
    verified, confidence, reason = verify_face_identity(reg_no, query_embedding)

    if not verified:
        print(f"[ATTENDANCE FAILURE] RegNo: {reg_no} | Reason: Face verification failed ({reason}) | Liveness: {'PASS' if is_live else 'WARNING'} ({liveness_reason}) | Quality Score: {analysis['quality_score']:.2f}")
        record_failed_attempt(reg_no)

        # Check if now locked out
        is_locked, remaining = check_lockout(reg_no)
        lockout_msg = f" Account locked for {remaining}s." if is_locked else ""

        log_audit_event("VERIFICATION_FAILED", reg_no, False, reason)
        raise HTTPException(
            status_code=401,
            detail=f"Face does not match - Please try again.{lockout_msg}",
        )

    # Verification successful - clear failed attempts
    clear_failed_attempts(reg_no)

    # Adaptive profile update: learn from high-confidence successful attendance.
    if not use_fallback and confidence >= PROFILE_SAMPLE_MIN_CONFIDENCE:
        _save_face_embedding_sample(
            reg_no=reg_no,
            source_table="other_staff" if is_other_staff else "users",
            embedding=query_embedding,
            sample_type="attendance_success",
            confidence=confidence,
        )

    # Check if this scan falls in an active CCL window (which rewards Earned Leave only)
    is_ccl_scan = (active_slot == -1)
    is_early_check_in = False
    is_late_check_out = False
    
    if is_ccl_scan:
        try:
            settings = get_ccl_settings_for_date()
            early_enabled = settings["early_check_in_ccl_enabled"]
            late_enabled = settings["late_check_out_ccl_enabled"]

            if early_enabled or late_enabled:
                now_time = datetime.now()
                curr_min = now_time.hour * 60 + now_time.minute
                
                def time_to_min(t_str):
                    p = t_str.split(":")
                    return int(p[0]) * 60 + int(p[1])

                ins_slot_type = slot_type or "check_in"
                if ins_slot_type == "check_in" and early_enabled:
                    early_start_str = settings.get("early_check_in_start", "07:00")
                    early_end_str = settings.get("early_check_in_end", "08:00")
                    if time_to_min(early_start_str) <= curr_min < time_to_min(early_end_str):
                        is_early_check_in = True
                elif ins_slot_type == "check_out" and late_enabled:
                    late_start_str = settings.get("late_check_out_start", "17:00")
                    late_end_str = settings.get("late_check_out_end", "18:00")
                    if time_to_min(late_start_str) <= curr_min < time_to_min(late_end_str):
                        is_late_check_out = True
        except Exception as e:
            print(f"Error evaluating is_ccl_scan details: {e}")

    # Use the values we extracted
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ins_slot_type = slot_type or "check_in"
    # Insert attendance into the correct table based on user type
    if is_other_staff:
        # Insert into other_staff_attendance table
        cursor.execute(
            """
            INSERT INTO other_staff_attendance (reg_no, name, dept, role, "timestamp", status)
            VALUES (?, ?, ?, ?, ?, ?)
        """,
            (reg_no, name, dept, user_type, timestamp, ins_slot_type),
        )
    else:
        # Insert into regular attendance table
        cursor.execute(
            """
            INSERT INTO attendance (reg_no, name, dept, "timestamp", status)
            VALUES (?, ?, ?, ?, ?)
        """,
            (reg_no, name, dept, timestamp, ins_slot_type),
        )
    conn.commit()

    # Sync with daily_attendance_status table (used by admin analysis tab)
    current_date = datetime.now().strftime("%Y-%m-%d")
    time_now_str = datetime.now().strftime("%H:%M:%S")
    sync_slot_type = slot_type or "check_in"

    # Determine slot_half — use the value passed directly from mark_attendance
    # (active_slot_half is already resolved from the slot's slot_half column by the caller)
    _hd_settings = _get_half_day_settings()
    _half_day_enabled = _hd_settings["half_day_enabled"]
    # Use the passed active_slot_half (already resolved by mark_attendance endpoint)
    _slot_half = active_slot_half  # 'first_half' | 'second_half' | None


    try:
        cursor.execute(
            """
            SELECT id, status, leave_type, leave_request_id, first_half_status, second_half_status, first_half_in_time, second_half_in_time
            FROM daily_attendance_status 
            WHERE reg_no = ? AND date = ?
        """,
            (reg_no, current_date),
        )
        existing_status = cursor.fetchone()

        if _half_day_enabled:
            # ── HALF-DAY MODE ──────────────────────────────────────────────────
            effective_half = _slot_half
            if not effective_half:
                try:
                    now_time = datetime.now().time()
                    fh_start = datetime.strptime(_hd_settings.get("first_half_start", "08:30"), "%H:%M").time()
                    fh_end = datetime.strptime(_hd_settings.get("first_half_end", "13:00"), "%H:%M").time()
                    sh_start = datetime.strptime(_hd_settings.get("second_half_start", "13:00"), "%H:%M").time()
                    sh_end = datetime.strptime(_hd_settings.get("second_half_end", "17:30"), "%H:%M").time()
                    
                    if fh_start <= now_time <= fh_end:
                        effective_half = "first_half"
                    elif sh_start <= now_time <= sh_end:
                        effective_half = "second_half"
                    else:
                        cutoff = datetime.strptime("13:00", "%H:%M").time()
                        effective_half = "first_half" if now_time < cutoff else "second_half"
                except Exception:
                    cutoff = datetime.strptime("13:00", "%H:%M").time()
                    effective_half = "first_half" if datetime.now().time() < cutoff else "second_half"

            # Insert or update into the correct separate session table
            if effective_half == "first_half":
                cursor.execute(
                    """
                    INSERT INTO morning_attendance (reg_no, name, dept, date, in_time, status, attendance_value, marked_by, marked_at)
                    VALUES (?, ?, ?, ?, ?, 'Present', 0.5, 'Attendance System', CURRENT_TIMESTAMP)
                    ON CONFLICT (reg_no, date) DO UPDATE SET
                        in_time = EXCLUDED.in_time,
                        status = 'Present',
                        attendance_value = 0.5,
                        marked_by = 'Attendance System',
                        marked_at = CURRENT_TIMESTAMP
                    """,
                    (reg_no, name, dept, current_date, time_now_str),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO evening_attendance (reg_no, name, dept, date, in_time, status, attendance_value, marked_by, marked_at)
                    VALUES (?, ?, ?, ?, ?, 'Present', 0.5, 'Attendance System', CURRENT_TIMESTAMP)
                    ON CONFLICT (reg_no, date) DO UPDATE SET
                        in_time = EXCLUDED.in_time,
                        status = 'Present',
                        attendance_value = 0.5,
                        marked_by = 'Attendance System',
                        marked_at = CURRENT_TIMESTAMP
                    """,
                    (reg_no, name, dept, current_date, time_now_str),
                )

            # Query current state from both session tables
            cursor.execute("SELECT status, attendance_value FROM morning_attendance WHERE reg_no = ? AND date = ?", (reg_no, current_date))
            _m_row = cursor.fetchone()
            cursor.execute("SELECT status, attendance_value FROM evening_attendance WHERE reg_no = ? AND date = ?", (reg_no, current_date))
            _e_row = cursor.fetchone()

            new_fh = _m_row[0] if _m_row else None
            new_sh = _e_row[0] if _e_row else None
            m_val = float(_m_row[1]) if _m_row and _m_row[1] is not None else 0.0
            e_val = float(_e_row[1]) if _e_row and _e_row[1] is not None else 0.0
            attendance_value = m_val + e_val

            new_status = _compute_daily_status(new_fh, new_sh, target_date=current_date)
            # Determine in_time or out_time based on sync_slot_type
            fh_in_time = time_now_str if (effective_half == "first_half" and sync_slot_type == "check_in") else (str(existing_status[6]) if existing_status and existing_status[6] else None)
            fh_out_time = time_now_str if (effective_half == "first_half" and sync_slot_type == "check_out") else None
            sh_in_time = time_now_str if (effective_half == "second_half" and sync_slot_type == "check_in") else (str(existing_status[7]) if existing_status and existing_status[7] else None)
            sh_out_time = time_now_str if (effective_half == "second_half" and sync_slot_type == "check_out") else None

            # Keep daily_attendance_status table fully in sync
            cursor.execute(
                """
                INSERT INTO daily_attendance_status
                (reg_no, name, dept, date, status,
                 first_half_status, second_half_status,
                 first_half_in_time, first_half_out_time, second_half_in_time, second_half_out_time,
                 attendance_value, marked_by, marked_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Attendance System', CURRENT_TIMESTAMP)
                ON CONFLICT (reg_no, date) DO UPDATE SET
                    first_half_status = EXCLUDED.first_half_status,
                    second_half_status = EXCLUDED.second_half_status,
                    first_half_in_time = COALESCE(EXCLUDED.first_half_in_time, daily_attendance_status.first_half_in_time),
                    first_half_out_time = COALESCE(EXCLUDED.first_half_out_time, daily_attendance_status.first_half_out_time),
                    second_half_in_time = COALESCE(EXCLUDED.second_half_in_time, daily_attendance_status.second_half_in_time),
                    second_half_out_time = COALESCE(EXCLUDED.second_half_out_time, daily_attendance_status.second_half_out_time),
                    status = EXCLUDED.status,
                    attendance_value = EXCLUDED.attendance_value,
                    marked_by = 'Attendance System',
                    marked_at = CURRENT_TIMESTAMP
                """,
                (reg_no, name, dept, current_date, new_status, new_fh, new_sh, fh_in_time, fh_out_time, sh_in_time, sh_out_time, attendance_value)
            )


            # Log attendance with value
            log_audit_event(
                "ATTENDANCE_MARKED",
                reg_no,
                True,
                f"Attendance marked via {sync_slot_type} in {effective_half}. Value: {attendance_value}"
            )
        else:
            # ── LEGACY CHECK_IN / CHECK_OUT MODE ───────────────────────────────
            if existing_status:
                old_status = existing_status[1]
                # On check-in: set status to Present, in_time, and clear leave tags.
                # On check-out: set status to Present and out_time.
                if sync_slot_type == "check_in":
                    cursor.execute(
                        """
                        UPDATE daily_attendance_status 
                        SET status = 'Present', in_time = ?, attendance_value = 1.0, leave_type = NULL, leave_request_id = NULL,
                            marked_by = 'Attendance System', marked_at = CURRENT_TIMESTAMP
                        WHERE reg_no = ? AND date = ?
                    """,
                        (time_now_str, reg_no, current_date),
                    )
                else:
                    # Check-out — complete the attendance
                    cursor.execute(
                        """
                        UPDATE daily_attendance_status 
                        SET status = 'Present', out_time = ?, attendance_value = 1.0, leave_type = NULL, leave_request_id = NULL,
                            marked_by = 'Attendance System', marked_at = CURRENT_TIMESTAMP
                        WHERE reg_no = ? AND date = ?
                    """,
                        (time_now_str, reg_no, current_date),
                    )
                new_status = "Present"
                log_audit_event(
                    "ATTENDANCE_OVERRIDE",
                    reg_no,
                    True,
                    f"User status updated to {new_status} via {sync_slot_type}. Value: 1.0",
                )
            else:
                # No record exists — insert with status 'Present'
                if sync_slot_type == "check_in":
                    cursor.execute(
                        """
                        INSERT INTO daily_attendance_status 
                        (reg_no, name, dept, date, status, in_time, attendance_value, marked_by, marked_at)
                        VALUES (?, ?, ?, ?, 'Present', ?, 1.0, 'Attendance System', CURRENT_TIMESTAMP)
                        ON CONFLICT (reg_no, date) DO UPDATE SET
                            status = 'Present',
                            in_time = COALESCE(daily_attendance_status.in_time, EXCLUDED.in_time),
                            attendance_value = 1.0,
                            marked_by = 'Attendance System',
                            marked_at = CURRENT_TIMESTAMP
                    """,
                        (reg_no, name, dept, current_date, time_now_str),
                    )
                else:
                    cursor.execute(
                        """
                        INSERT INTO daily_attendance_status 
                        (reg_no, name, dept, date, status, out_time, attendance_value, marked_by, marked_at)
                        VALUES (?, ?, ?, ?, 'Present', ?, 1.0, 'Attendance System', CURRENT_TIMESTAMP)
                        ON CONFLICT (reg_no, date) DO UPDATE SET
                            status = 'Present',
                            out_time = COALESCE(daily_attendance_status.out_time, EXCLUDED.out_time),
                            attendance_value = 1.0,
                            marked_by = 'Attendance System',
                            marked_at = CURRENT_TIMESTAMP
                    """,
                        (reg_no, name, dept, current_date, time_now_str),
                    )
                
                # Log attendance with value
                log_audit_event(
                    "ATTENDANCE_MARKED",
                    reg_no,
                    True,
                    f"First attendance of the day via {sync_slot_type}. Value: 1.0"
                )
    except Exception as e:
        print(f"Error syncing daily_attendance_status for {reg_no}: {e}")



    if is_ccl_scan:
        log_audit_event("CCL_EARNED_SCAN", reg_no, True, f"Confidence: {confidence:.3f}")

    # Update user location when attendance is marked successfully
    if form_data is not None:
        try:
            lat_raw = form_data.get("client_lat")
            lng_raw = form_data.get("client_lng")
            accuracy_raw = form_data.get("client_accuracy")

            if lat_raw is not None and lng_raw is not None:
                latitude = float(lat_raw)
                longitude = float(lng_raw)
                accuracy = float(accuracy_raw) if accuracy_raw is not None else None

                # Get user details
                cursor.execute(
                    """
                    SELECT username, name, dept, role FROM users WHERE reg_no = ?
                    UNION ALL
                    SELECT username, name, dept, role FROM other_staff WHERE reg_no = ?
                """,
                    (reg_no, reg_no),
                )

                user_row = cursor.fetchone()
                if user_row:
                    username, user_name, dept, role = user_row

                    # Insert into user_latest_locations (same as /location/update)
                    cursor.execute(
                        """
                        INSERT INTO user_latest_locations
                        (reg_no, username, name, dept, role, latitude, longitude, accuracy_meters, 
                         source, app_state, captured_at, last_seen_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                        ON CONFLICT (reg_no) DO UPDATE SET
                            username = EXCLUDED.username,
                            name = EXCLUDED.name,
                            dept = EXCLUDED.dept,
                            role = EXCLUDED.role,
                            latitude = EXCLUDED.latitude,
                            longitude = EXCLUDED.longitude,
                            accuracy_meters = EXCLUDED.accuracy_meters,
                            source = EXCLUDED.source,
                            app_state = EXCLUDED.app_state,
                            captured_at = EXCLUDED.captured_at,
                            last_seen_at = CURRENT_TIMESTAMP
                        """,
                        (
                            reg_no,
                            username,
                            user_name,
                            dept,
                            role,
                            latitude,
                            longitude,
                            accuracy,
                            "attendance_mark",
                            "foreground",
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        ),
                    )
                    # print(f"Location updated via attendance mark for {reg_no}")
                    pass
        except Exception as e:
            print(f"[ATTENDANCE WARNING] Error updating location during attendance mark for {reg_no}: {e}")

    # Check CCL Crediting logic
    earned_ccl = False
    try:
        # Check if CCL is enabled for the specific slot type
        settings = get_ccl_settings_for_date()
        early_enabled = settings["early_check_in_ccl_enabled"]
        late_enabled = settings["late_check_out_ccl_enabled"]

        if (early_enabled or late_enabled) and is_ccl_scan:
            # Get current time
            now_time = datetime.now()
            now_time_str = now_time.strftime("%H:%M")
            now_date_str = now_time.strftime("%Y-%m-%d")
            
            # Convert to minutes since midnight
            def time_to_min(t_str):
                p = t_str.split(":")
                return int(p[0]) * 60 + int(p[1])
                
            curr_min = now_time.hour * 60 + now_time.minute
            
            is_early_check_in = False
            is_late_check_out = False
            
            early_start_str = settings.get("early_check_in_start", "07:00")
            early_end_str = settings.get("early_check_in_end", "08:00")
            late_start_str = settings.get("late_check_out_start", "17:00")
            late_end_str = settings.get("late_check_out_end", "18:00")
            
            if ins_slot_type == "check_in" and early_enabled:
                # Check early check-in window
                if time_to_min(early_start_str) <= curr_min < time_to_min(early_end_str):
                    is_early_check_in = True
            elif ins_slot_type == "check_out" and late_enabled:
                # Check late check-out window
                if time_to_min(late_start_str) <= curr_min < time_to_min(late_end_str):
                    is_late_check_out = True
                    
            if is_early_check_in or is_late_check_out:
                ccl_slot = "early_check_in" if is_early_check_in else "late_check_out"
                
                # Verify that they have not already earned ANY CCL today (enforce 1 point max limit per day)
                cursor.execute(
                    "SELECT 1 FROM ccl_earned_history WHERE reg_no = ? AND date = ?",
                    (reg_no, now_date_str)
                )
                already_earned = cursor.fetchone()
                
                if not already_earned:
                    # Insert history log
                    cursor.execute("""
                        INSERT INTO ccl_earned_history (reg_no, name, dept, date, time, slot_type, earned_points)
                        VALUES (?, ?, ?, ?, ?, ?, 1.0)
                    """, (reg_no, name, dept, now_date_str, now_time.strftime("%H:%M:%S"), ccl_slot))
                    
                    # Update earned_leave balance
                    cursor.execute("SELECT balance FROM earned_leave WHERE reg_no = ?", (reg_no,))
                    bal_row = cursor.fetchone()
                    if not bal_row:
                        cursor.execute("""
                            INSERT INTO earned_leave (reg_no, user_name, dept, role, balance)
                            VALUES (?, ?, ?, ?, 1.0)
                        """, (reg_no, name, dept, user_type or "staff", 1.0))
                    else:
                        new_balance = float(bal_row[0]) + 1.0
                        cursor.execute("""
                            UPDATE earned_leave 
                            SET balance = ?, updated_at = CURRENT_TIMESTAMP
                            WHERE reg_no = ?
                        """, (new_balance, reg_no))
                        
                    conn.commit()
                    earned_ccl = True
    except Exception as ccl_err:
        print(f"[ATTENDANCE WARNING] Error processing CCL credit for {reg_no}: {ccl_err}")

    # Output consolidated debug log with proper context
    print(f"[ATTENDANCE SUCCESS] RegNo: {reg_no} | Name: {name} | Dept: {dept} | Slot: {ins_slot_type} | Confidence: {confidence:.3f} | Liveness: {'PASS' if is_live else 'WARNING'} ({liveness_reason}) | Quality Score: {analysis['quality_score']:.2f} | CCL Earned: {earned_ccl}")

    if is_ccl_scan:
        msg = "You earned 1.0 CCL point! (Leave balance updated)" if earned_ccl else "You have already earned your CCL point for today."
        session_label = ""
    elif _slot_half == "first_half":
        session_label = "FN (Morning)"
        msg = f"FN Attendance marked successfully — 0.5 credited. You earned 1.0 CCL point!" if earned_ccl else f"FN (Morning) attendance marked successfully — 0.5 credited"
    elif _slot_half == "second_half":
        session_label = "AN (Afternoon)"
        msg = f"AN Attendance marked successfully — 0.5 credited. You earned 1.0 CCL point!" if earned_ccl else f"AN (Afternoon) attendance marked successfully — 0.5 credited"
    else:
        session_label = ""
        msg = "Attendance marked successfully. You earned 1.0 CCL point!" if earned_ccl else "Attendance marked successfully"

    # Compute the attendance_value for the response
    _resp_attendance_value = 0.5 if _slot_half in ("first_half", "second_half") else 1.0

    return {
        "message": msg,
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "time": timestamp,
        "confidence": float(round(confidence, 3)),
        "verification_status": "verified",
        "bbox": [float(x) for x in face.bbox],
        "session": _slot_half or "full_day",
        "session_label": session_label,
        "attendance_value": _resp_attendance_value,
    }


# -------------------------------------------------
# ADMIN & HOD ATTENDANCE MARKING ENDPOINTS
# -------------------------------------------------


@app.post("/admin/face/attendance")
async def admin_mark_attendance(request: Request, image: UploadFile = File(...)):
    """
    Admin can mark their own attendance using face recognition.
    The admin's reg_no is automatically extracted from their authentication token.
    """
    cursor.execute("""
        SELECT slot_number, start_time, duration_minutes, is_enabled, slot_type, slot_half
        FROM attendance_duration_settings
        WHERE is_enabled = 1
        ORDER BY slot_number ASC
    """)
    duration_rows = cursor.fetchall()

    active_slot_type = "check_in"
    active_slot_half = None
    if duration_rows:
        current_time = datetime.now()
        allowed = False

        for row in duration_rows:
            slot_number = row[0]
            start_time = row[1]
            duration_minutes = row[2]
            slot_type = row[4] if len(row) > 4 and row[4] else "check_in"
            slot_half = row[5] if len(row) > 5 and row[5] else "full_day"
            effective_duration, _, _ = _calculate_effective_slot_duration(slot_type, slot_half, duration_minutes)

            start_hour, start_minute = map(int, start_time.split(":"))
            start_datetime = current_time.replace(
                hour=start_hour, minute=start_minute, second=0, microsecond=0
            )
            end_datetime = start_datetime + timedelta(minutes=effective_duration)

            if start_datetime <= current_time < end_datetime:
                allowed = True
                active_slot_type = slot_type
                active_slot_half = slot_half if slot_half in ("first_half", "second_half") else None
                break

        if not allowed:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                allowed = True
                active_slot_type = ccl_slot_type
                
        if not allowed:
            slots_info = ", ".join(
                [f"Slot {row[0]} ({row[4] if len(row) > 4 and row[4] else 'check_in'}): {row[1]} ({row[2]} min)" for row in duration_rows]
            )
            raise HTTPException(
                status_code=403,
                detail=f"Attendance marking is not allowed at this time. Available slots: {slots_info}",
            )
    else:
        settings = get_ccl_settings_for_date()
        early_enabled = settings["early_check_in_ccl_enabled"]
        late_enabled = settings["late_check_out_ccl_enabled"]
        
        if early_enabled or late_enabled:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                active_slot_type = ccl_slot_type
            else:
                raise HTTPException(
                    status_code=403,
                    detail="Attendance marking is not allowed at this time. (Outside CCL window)",
                )

    # Verify admin token
    admin_user = verify_admin_token(request)
    reg_no = admin_user["reg_no"]

    # Rate limiting
    allowed, retry_after = attendance_rate_limiter.is_allowed(reg_no, "admin_mark_attendance")
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail=f"Too many requests. Please wait {retry_after} seconds before trying again.",
        )

    if active_slot_type == "check_out":
        cursor.execute(
            "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
            (reg_no,)
        )
        if cursor.fetchone()[0] == 0:
            raise HTTPException(
                status_code=400,
                detail="You cannot Check-Out without checking in first for today."
            )

    # Enforce geofencing using parsed parameters
    try:
        form_data = await request.form()
    except:
        form_data = None

    platform = request.headers.get("X-Client-Platform", "app").lower()
    lat = request.query_params.get("client_lat")
    lng = request.query_params.get("client_lng")
    if form_data is not None:
        platform = form_data.get("client_platform", platform)
        lat = form_data.get("client_lat", lat)
        lng = form_data.get("client_lng", lng)

    validation_form = {
        "client_platform": platform,
        "client_lat": lat,
        "client_lng": lng,
        "reg_no": reg_no
    }
    # Geofence first — if active and passes, WiFi check is redundant
    if not _enforce_web_geofence(validation_form, get_client_ip(request)):
        check_wifi(request)

    # Get the image
    img_bytes = await image.read()

    if len(img_bytes) == 0:
        raise HTTPException(status_code=400, detail="Empty image")

    # Per-user concurrency guard + thread pool offload
    async with acquire_user_lock(reg_no):
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            _cpu_executor,
            _attendance_sync_work,
            reg_no, "admin", validation_form, active_slot_type, img_bytes, None, active_slot_half
        )


@app.post("/hod/face/attendance")
async def hod_mark_attendance(request: Request, image: UploadFile = File(...)):
    """
    HOD can mark their own attendance using face recognition.
    The HOD's reg_no is automatically extracted from their authentication token.
    """
    cursor.execute("""
        SELECT slot_number, start_time, duration_minutes, is_enabled, slot_type, slot_half
        FROM attendance_duration_settings
        WHERE is_enabled = 1
        ORDER BY slot_number ASC
    """)
    duration_rows = cursor.fetchall()

    active_slot_type = "check_in"
    active_slot_half = None
    if duration_rows:
        current_time = datetime.now()
        allowed = False

        for row in duration_rows:
            slot_number = row[0]
            start_time = row[1]
            duration_minutes = row[2]
            slot_type = row[4] if len(row) > 4 and row[4] else "check_in"
            slot_half = row[5] if len(row) > 5 and row[5] else "full_day"
            effective_duration, _, _ = _calculate_effective_slot_duration(slot_type, slot_half, duration_minutes)

            start_hour, start_minute = map(int, start_time.split(":"))
            start_datetime = current_time.replace(
                hour=start_hour, minute=start_minute, second=0, microsecond=0
            )
            end_datetime = start_datetime + timedelta(minutes=effective_duration)

            if start_datetime <= current_time < end_datetime:
                allowed = True
                active_slot_type = slot_type
                active_slot_half = slot_half if slot_half in ("first_half", "second_half") else None
                break

        if not allowed:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                allowed = True
                active_slot_type = ccl_slot_type
                
        if not allowed:
            slots_info = ", ".join(
                [f"Slot {row[0]} ({row[4] if len(row) > 4 and row[4] else 'check_in'}): {row[1]} ({row[2]} min)" for row in duration_rows]
            )
            raise HTTPException(
                status_code=403,
                detail=f"Attendance marking is not allowed at this time. Available slots: {slots_info}",
            )
    else:
        settings = get_ccl_settings_for_date()
        early_enabled = settings["early_check_in_ccl_enabled"]
        late_enabled = settings["late_check_out_ccl_enabled"]
        
        if early_enabled or late_enabled:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                active_slot_type = ccl_slot_type
            else:
                raise HTTPException(
                    status_code=403,
                    detail="Attendance marking is not allowed at this time. (Outside CCL window)",
                )

    # Verify HOD token
    hod_user = verify_hod_token(request)
    reg_no = hod_user["reg_no"]

    # Rate limiting
    allowed, retry_after = attendance_rate_limiter.is_allowed(reg_no, "hod_mark_attendance")
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail=f"Too many requests. Please wait {retry_after} seconds before trying again.",
        )

    if active_slot_type == "check_out":
        cursor.execute(
            "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
            (reg_no,)
        )
        if cursor.fetchone()[0] == 0:
            raise HTTPException(
                status_code=400,
                detail="You cannot Check-Out without checking in first for today."
            )

    # Enforce geofencing using parsed parameters
    try:
        form_data = await request.form()
    except:
        form_data = None

    platform = request.headers.get("X-Client-Platform", "app").lower()
    lat = request.query_params.get("client_lat")
    lng = request.query_params.get("client_lng")
    if form_data is not None:
        platform = form_data.get("client_platform", platform)
        lat = form_data.get("client_lat", lat)
        lng = form_data.get("client_lng", lng)

    validation_form = {
        "client_platform": platform,
        "client_lat": lat,
        "client_lng": lng,
        "reg_no": reg_no
    }
    # Geofence first — if active and passes, WiFi check is redundant
    if not _enforce_web_geofence(validation_form, get_client_ip(request)):
        check_wifi(request)

    # Get the image
    img_bytes = await image.read()

    if len(img_bytes) == 0:
        raise HTTPException(status_code=400, detail="Empty image")

    # Per-user concurrency guard + thread pool offload
    async with acquire_user_lock(reg_no):
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            _cpu_executor,
            _attendance_sync_work,
            reg_no, "hod", validation_form, active_slot_type, img_bytes, None, active_slot_half
        )


async def _legacy_mark_attendance(img_bytes: bytes):
    """
    Legacy attendance marking - DISABLED for security.
    This function is kept for backwards compatibility but always raises an error.
    """
    today_str = datetime.now().strftime("%Y-%m-%d")
    _, _, is_holiday = _academic_status_for_date(today_str)
    if is_holiday:
        raise HTTPException(
            status_code=403,
            detail="Attendance is blocked today because the date is marked as a holiday in the academic calendar.",
        )

    raise HTTPException(
        status_code=410,
        detail="Legacy attendance marking has been disabled. Please use the secure attendance marking endpoint.",
    )
    try:
        img = preprocess_image_data(img_bytes)
    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Preprocessing error: {e}")
        raise HTTPException(status_code=400, detail="Image processing failed")

    face = extract_face(img)

    if face is None:
        save_debug_image(img, "fail_detection")
        raise HTTPException(
            status_code=400,
            detail="Face is unable to detect. Please adjust your position and ensure good lighting.",
        )

    query_embedding = face.embedding.astype(np.float32)

    # Get all registered staff users for face matching
    cursor.execute(
        "SELECT reg_no, name, dept, embedding FROM users WHERE role IN ('staff', 'hod') AND embedding IS NOT NULL"
    )
    staff_users = cursor.fetchall()

    all_users = list(staff_users)

    if not all_users:
        raise HTTPException(status_code=404, detail="No users registered")

    best_score = 0.0
    best_match = None
    similarity_details = []

    print(f"DEBUG: Comparing face with {len(all_users)} registered users...")

    for reg_no, name, dept, emb_blob in all_users:
        db_embedding = np.frombuffer(emb_blob, dtype=np.float32)

        # Use improved similarity calculation
        combined_sim, cosine_sim, euclidean_dist, manhattan_dist, corr = (
            calculate_similarity(query_embedding, db_embedding)
        )

        similarity_details.append(
            {
                "name": name,
                "combined": combined_sim,
                "cosine": cosine_sim,
                "euclidean": euclidean_dist,
            }
        )

        if combined_sim > best_score:
            best_score = combined_sim
            best_match = (reg_no, name, dept)

    # Sort and show top 3 matches for debugging
    similarity_details.sort(key=lambda x: x["combined"], reverse=True)
    print("DEBUG: Top 3 matches:")
    for i, match in enumerate(similarity_details[:3]):
        print(
            f"  {i + 1}. {match['name']}: {match['combined']:.3f} (cosine: {match['cosine']:.3f})"
        )

    print(
        f"DEBUG: Best match: {best_match[1] if best_match else 'None'} with score {best_score:.3f}"
    )

    # CRITICAL: Much stricter threshold for legacy mode
    # Since legacy mode compares against ALL users, we need higher confidence
    if use_fallback:
        LEGACY_THRESHOLD = 0.95  # Very strict for fallback
        print("⚠️  Using LEGACY FALLBACK threshold: 0.95 (INSECURE mode)")
    else:
        LEGACY_THRESHOLD = 0.75  # Stricter for InsightFace
        print(f"Using LEGACY threshold: {LEGACY_THRESHOLD} (secure mode)")

    if best_match is None or best_score < LEGACY_THRESHOLD:
        save_debug_image(img, "fail_low_score")
        print(f"DEBUG: Score {best_score:.3f} below threshold {LEGACY_THRESHOLD}")
        log_audit_event(
            "LOW_CONFIDENCE",
            best_match[0] if best_match else None,
            False,
            f"Score: {best_score:.3f}",
        )
        raise HTTPException(
            status_code=401,
            detail=f"Face not recognized (confidence={round(best_score, 3)})",
        )

    reg_no, name, dept = best_match
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute(
        """
        INSERT INTO attendance (reg_no, name, dept, "timestamp")
        VALUES (%s, %s, %s, %s)
    """,
        (reg_no, name, dept, timestamp),
    )
    conn.commit()

    # Sync with daily_attendance_status table
    current_date = datetime.now().strftime("%Y-%m-%d")
    try:
        cursor.execute(
            """
            SELECT id, status FROM daily_attendance_status 
            WHERE reg_no = %s AND date = %s
        """,
            (reg_no, current_date),
        )
        existing = cursor.fetchone()
        if existing and existing[1] in [
            "Absent",
            "Leave",
            "OD",
            "casual",
            "earned",
            "od",
        ]:
            cursor.execute(
                """
                UPDATE daily_attendance_status 
                SET status = 'Present', leave_type = NULL, leave_request_id = NULL,
                    marked_by = 'Attendance System', marked_at = CURRENT_TIMESTAMP
                WHERE reg_no = %s AND date = %s
            """,
                (reg_no, current_date),
            )
        else:
            cursor.execute(
                """
                INSERT INTO daily_attendance_status 
                (reg_no, name, dept, date, status, marked_by, marked_at)
                VALUES (%s, %s, %s, %s, 'Present', 'Attendance System', CURRENT_TIMESTAMP)
                ON CONFLICT (reg_no, date) DO UPDATE SET
                    status = 'Present',
                    marked_by = 'Attendance System',
                    marked_at = CURRENT_TIMESTAMP
            """,
                (reg_no, name, dept, current_date),
            )
    except Exception as e:
        print(f"Error syncing daily_attendance_status for {reg_no}: {e}")

    log_audit_event(
        "ATTENDANCE_MARKED_LEGACY", reg_no, True, f"Confidence: {best_score:.3f}"
    )

    return {
        "message": "Attendance marked",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "time": timestamp,
        "confidence": float(round(best_score, 3)),
        "verification_status": "legacy",
    }


# -------------------------------------------------
# AUDIT ENDPOINT
# -------------------------------------------------
@app.get("/audit/logs")
async def get_audit_logs_endpoint(reg_no: str = None, limit: int = 100):
    """Get audit logs"""
    logs = get_audit_logs(reg_no, limit)
    return {"logs": logs, "count": len(logs)}


@app.get("/audit/status/{reg_no}")
async def get_verification_status(reg_no: str):
    """Get verification status for a user"""
    is_locked, remaining = check_lockout(reg_no)
    failed_count = _failed_attempts.get(reg_no, 0)
    return {
        "is_locked": is_locked,
        "remaining_seconds": remaining,
        "failed_attempts": failed_count,
    }


@app.get("/attendance")
async def get_attendance():
    """Get all attendance records"""
    # Use explicit column names to avoid IndexError
    cursor.execute("""
        SELECT id, reg_no, name, dept, class_div, timestamp 
        FROM attendance 
        ORDER BY id DESC 
        LIMIT 100
    """)
    rows = cursor.fetchall()
    return {
        "attendance": [
            {
                "id": row[0],
                "reg_no": row[1],
                "name": row[2],
                "dept": row[3],
                "class_div": row[4] or "",
                "timestamp": _ts(row[5]),
            }
            for row in rows
        ]
    }


# -------------------------------------------------
# ADMIN PANEL ENDPOINTS
# -------------------------------------------------


def verify_admin_token(request: Request) -> dict:
    """Verify admin authentication token"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authorization header"
        )

    token = auth_header.replace("Bearer ", "")

    # Decode token (simple base64 for demo, use proper JWT in production)
    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")

        user = get_user_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check if user is suspended
        if is_user_suspended(username, is_other_staff=False):
            raise HTTPException(status_code=403, detail="Account suspended. Access denied.")

        # Check if user has admin privileges
        if user[6] not in ["admin", "hod"]:
            raise HTTPException(status_code=403, detail="Admin access required")

        return {
            "id": user[0],
            "username": user[1],
            "reg_no": user[3],
            "name": user[4],
            "dept": user[5],
            "role": user[6],
        }
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid token")


@app.post("/admin/login")
async def admin_login(request: Request):
    """Admin login endpoint"""
    try:
        data = await request.json()
        username = data.get("username")
        password = data.get("password")

        if not username or not password:
            raise HTTPException(status_code=400, detail="Missing credentials")

        user = get_user_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check role
        if user[6] not in ["admin", "hod"]:
            raise HTTPException(status_code=403, detail="Admin access required")

        # Create simple token
        import base64

        token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode(
            "utf-8"
        )

        return {
            "message": "Admin login successful",
            "token": token,
            "user": {
                "id": user[0],
                "username": user[1],
                "regNo": user[3],
                "name": user[4],
                "dept": user[5],
                "role": user[6],
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Admin login error: {e}")
        raise HTTPException(status_code=500, detail="Login failed")


@app.get("/admin/dashboard")
async def admin_dashboard(request: Request):
    """Get admin dashboard statistics - includes instant OD sync"""
    admin_user = verify_admin_token(request)

    # Get counts
    cursor.execute("SELECT COUNT(*) FROM users")
    total_users = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'staff'")
    total_staff = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM attendance")
    total_attendance = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(*) FROM attendance WHERE timestamp::date = CURRENT_DATE"
    )
    today_face_scan = cursor.fetchone()[0]

    # Get today's OD count (separate by type)
    cursor.execute("""
        SELECT leave_type, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE date::date = CURRENT_DATE AND status = 'Present' AND leave_type IN ('od', 'earned', 'casual')
        GROUP BY leave_type
    """)
    od_breakdown = cursor.fetchall()
    od_count = sum(row[1] for row in od_breakdown)
    earned_count = sum(row[1] for row in od_breakdown if row[0] == "earned")
    casual_count = sum(row[1] for row in od_breakdown if row[0] == "casual")

    # Total today = face scan + OD
    today_attendance = today_face_scan + od_count

    cursor.execute("SELECT COUNT(*) FROM departments")
    total_departments = cursor.fetchone()[0]

    # Get recent attendance (face scan + OD)
    try:
        cursor.execute("""
            SELECT id, reg_no, name, dept, timestamp, 'face_scan' as source, status
            FROM attendance
            ORDER BY timestamp DESC
            LIMIT 8
        """)
        face_scan_attendance = cursor.fetchall()

        cursor.execute("""
            SELECT id, reg_no, name, dept, date as timestamp, 'od' as source, NULL as status
            FROM daily_attendance_status
            WHERE date::date = CURRENT_DATE AND status = 'Present' AND leave_type IN ('od', 'earned', 'casual')
            ORDER BY date DESC
            LIMIT 5
        """)
        od_records = cursor.fetchall()

        combined = list(face_scan_attendance) + list(od_records)
        combined.sort(key=lambda x: str(x[4]) if x[4] else "", reverse=True)
        recent_attendance = combined[:10]

    except Exception as e:
        print(f"Error getting combined attendance: {e}")
        cursor.execute("""
            SELECT id, reg_no, name, dept, timestamp, 'attendance' as source, status
            FROM attendance 
            ORDER BY id DESC 
            LIMIT 10
        """)
        recent_attendance = cursor.fetchall()

    # --- Pie chart: institution-wide breakdown from daily_attendance_status ---
    cursor.execute("""
        SELECT status, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE date::date = CURRENT_DATE
        GROUP BY status
    """)
    admin_das_breakdown = {row[0]: row[1] for row in cursor.fetchall()}
    admin_pie_full_day  = admin_das_breakdown.get("Present", 0)
    admin_pie_half_day  = admin_das_breakdown.get("Half Day", 0)
    admin_pie_absent    = admin_das_breakdown.get("Absent", 0)
    admin_pie_leave     = admin_das_breakdown.get("Leave", 0)
    admin_pie_holiday   = admin_das_breakdown.get("Holiday", 0)

    response_data = {
        "stats": {
            "total_users": total_users,
            "total_staff": total_staff,
            "total_attendance": total_attendance,
            "today_attendance": today_attendance,
            "today_face_scan": today_face_scan,
            "today_od": od_count,
            "today_earned": earned_count,
            "today_casual": casual_count,
            "total_departments": total_departments,
            # Pie chart breakdown fields
            "today_full_day": admin_pie_full_day,
            "today_half_day": admin_pie_half_day,
            "today_absent_das": admin_pie_absent,
            "today_leave": admin_pie_leave,
            "today_holiday": admin_pie_holiday,
        },
        "recent_attendance": [
            {
                "id": row[0],
                "reg_no": row[1],
                "name": row[2],
                "dept": row[3],
                "timestamp": _ts(row[4]) if row[4] else str(row[4]),
                "source": row[5],
                "status": _format_scan_status(row[6] if len(row) > 6 else None, row[4]) if row[5] == "face_scan" else None,
                "punch_type": (row[6] or "check_in") if row[5] == "face_scan" and len(row) > 6 else None,
            }
            for row in recent_attendance
        ],
        "admin_user": admin_user,
    }

    # Return with no-cache headers to prevent caching
    return JSONResponse(
        content=response_data,
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
            "Vary": "Accept-Encoding",
            "X-Accel-Expires": "0",
        },
    )


@app.get("/admin/recent-attendance")
async def admin_recent_attendance(request: Request):
    """Get recent attendance records for auto-refresh (admin only)"""
    verify_admin_token(request)

    try:
        # Get face scan attendance
        cursor.execute("""
            SELECT id, reg_no, name, dept, timestamp, 'face_scan' as source,
                   NULL as leave_type, NULL as processed_date, status, NULL as absent_reason
            FROM attendance
            ORDER BY timestamp DESC
            LIMIT 10
        """)
        face_scan_attendance = cursor.fetchall()

        # Get OD/Leave/Absent records from daily_attendance_status (today's only)
        cursor.execute("""
            SELECT das.id, das.reg_no, das.name, das.dept, das.date as timestamp, 
                   (CASE WHEN das.status = 'Absent' THEN 'absent_record' ELSE 'od' END) as source, 
                   das.leave_type, lr.processed_date, das.status, das.absent_reason
            FROM daily_attendance_status das
            LEFT JOIN leave_requests lr ON das.leave_request_id = lr.id
            WHERE das.date::date = CURRENT_DATE 
            AND (
                (das.status = 'Present' AND das.leave_type IN ('od', 'earned', 'casual'))
                OR (das.status = 'Absent')
            )
            ORDER BY das.date DESC
            LIMIT 10
        """)
        status_records = cursor.fetchall()

        # Combine and sort
        combined = list(face_scan_attendance) + list(status_records)
        combined.sort(key=lambda x: str(x[4]) if x[4] else "", reverse=True)
        recent_attendance = combined[:10]

    except Exception as e:
        print(f"Error getting combined attendance: {e}")
        recent_attendance = []

    # Get counts after the query block
    cursor.execute(
        "SELECT COUNT(*) FROM attendance WHERE timestamp::date = CURRENT_DATE"
    )
    today_attendance = cursor.fetchone()[0]

    # Get today's OD/Leave count (separate by type)
    cursor.execute("""
        SELECT leave_type, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE date::date = CURRENT_DATE AND status = 'Present' AND leave_type IN ('od', 'earned', 'casual')
        GROUP BY leave_type
    """)
    od_breakdown = cursor.fetchall()
    od_count = sum(row[1] for row in od_breakdown)
    earned_count = sum(row[1] for row in od_breakdown if row[0] == "earned")
    casual_count = sum(row[1] for row in od_breakdown if row[0] == "casual")

    # Build recent_attendance list based on the combined data
    recent_attendance_list = []
    for row in recent_attendance:
        raw_punch = row[8] if row[5] == "face_scan" else None
        recent_attendance_list.append(
            {
                "id": row[0],
                "reg_no": row[1],
                "name": row[2],
                "dept": row[3],
                "timestamp": _ts(row[4]) if row[4] else str(row[4]),
                "source": row[5],
                "leave_type": row[6],
                "approval_date": _ts(row[7]) if row[7] else None,
                "status": _format_scan_status(row[8], row[4]) if row[5] == "face_scan" else row[8],
                "punch_type": raw_punch or ("check_in" if row[5] == "face_scan" else None),
                "absent_reason": row[9],
            }
        )

    return {
        "stats": {
            "today_attendance": today_attendance + od_count,
            "today_face_scan": today_attendance,
            "today_od": od_count,
            "today_earned": earned_count,
            "today_casual": casual_count,
        },
        "recent_attendance": recent_attendance_list,
    }


@app.get("/admin/system/status")
async def admin_system_status(request: Request):
    """Get system status and health information (admin only)"""
    verify_admin_token(request)

    status_info = {"status": "healthy", "checks": [], "errors": []}

    # Check database connectivity
    try:
        cursor.execute("SELECT 1")
        status_info["checks"].append(
            {
                "name": "Database",
                "status": "healthy",
                "message": "Connected successfully",
            }
        )
    except Exception as e:
        status_info["status"] = "error"
        status_info["checks"].append(
            {"name": "Database", "status": "error", "message": str(e)}
        )
        status_info["errors"].append(
            {
                "timestamp": datetime.now().isoformat(),
                "source": "Database",
                "message": f"Database connection error: {str(e)}",
            }
        )

    # Check if tables exist
    try:
        cursor.execute("SELECT tablename FROM pg_tables WHERE schemaname = 'public'")
        tables = cursor.fetchall()
        table_names = [t[0] for t in tables]
        required_tables = ["users", "attendance", "departments"]
        missing_tables = [t for t in required_tables if t not in table_names]

        if missing_tables:
            status_info["status"] = "warning"
            status_info["checks"].append(
                {
                    "name": "Database Tables",
                    "status": "warning",
                    "message": f"Missing tables: {', '.join(missing_tables)}",
                }
            )
        else:
            status_info["checks"].append(
                {
                    "name": "Database Tables",
                    "status": "healthy",
                    "message": "All required tables exist",
                }
            )
    except Exception as e:
        status_info["checks"].append(
            {"name": "Database Tables", "status": "error", "message": str(e)}
        )

    # Check recent errors from attendance failures
    try:
        cursor.execute("""
            SELECT id, reg_no, name, timestamp, error_message 
            FROM attendance 
            WHERE error_message IS NOT NULL AND error_message != ''
            ORDER BY id DESC 
            LIMIT 20
        """)
        error_rows = cursor.fetchall()
        for row in error_rows:
            status_info["errors"].append(
                {
                    "timestamp": str(row[3]) if row[3] else None,
                    "source": "Attendance",
                    "message": row[4] if row[4] else "Unknown error",
                }
            )
    except:
        pass  # Table might not have error_message column

    # Check for users without face embeddings
    try:
        cursor.execute(
            "SELECT COUNT(*) FROM users WHERE role IN ('staff', 'hod') AND (embedding IS NULL OR embedding = '')"
        )
        no_face_count = cursor.fetchone()[0]
        if no_face_count > 0:
            status_info["checks"].append(
                {
                    "name": "Face Registration",
                    "status": "warning",
                    "message": f"{no_face_count} staff/HOD users without face registered",
                }
            )
        else:
            status_info["checks"].append(
                {
                    "name": "Face Registration",
                    "status": "healthy",
                    "message": "All staff/HOD users have face registered",
                }
            )
    except Exception as e:
        status_info["checks"].append(
            {"name": "Face Registration", "status": "unknown", "message": str(e)}
        )

    return status_info


@app.get("/admin/system/logs")
async def admin_system_logs(request: Request, limit: int = 50):
    """Get system error logs (admin only)"""
    verify_admin_token(request)

    logs = []

    # Get attendance errors
    try:
        cursor.execute(
            """
            SELECT id, reg_no, name, dept, timestamp, error_message 
            FROM attendance 
            WHERE error_message IS NOT NULL AND error_message != ''
            ORDER BY id DESC 
            LIMIT ?
        """,
            (limit,),
        )
        error_rows = cursor.fetchall()
        for row in error_rows:
            logs.append(
                {
                    "id": row[0],
                    "type": "attendance_error",
                    "reg_no": row[1],
                    "name": row[2],
                    "department": row[3],
                    "timestamp": str(row[4]) if row[4] else None,
                    "message": row[5] if row[5] else "Unknown error",
                }
            )
    except:
        pass

    # Get failed login attempts (if tracked)
    try:
        cursor.execute(
            """
            SELECT id, username, timestamp, ip_address, status 
            FROM login_logs 
            WHERE status = 'failed'
            ORDER BY id DESC 
            LIMIT ?
        """,
            (limit,),
        )
        login_rows = cursor.fetchall()
        for row in login_rows:
            logs.append(
                {
                    "id": row[0],
                    "type": "login_failed",
                    "username": row[1],
                    "timestamp": str(row[2]) if row[2] else None,
                    "ip_address": row[3],
                    "message": "Failed login attempt",
                }
            )
    except:
        pass  # Table might not exist

    # Sort by timestamp descending
    logs.sort(key=lambda x: x["timestamp"] or "", reverse=True)

    return {"logs": logs[:limit], "count": len(logs)}


@app.get("/admin/users")
async def admin_get_users(request: Request, role: str = None):
    """Get all users (admin only) with face status for staff"""
    verify_admin_token(request)

    try:
        if role:
            cursor.execute(
                "SELECT id, username, reg_no, name, dept, role, created_at, embedding, can_reregister, COALESCE(suspended, FALSE) FROM users WHERE role = ?",
                (role,),
            )
        else:
            cursor.execute(
                "SELECT id, username, reg_no, name, dept, role, created_at, embedding, can_reregister, COALESCE(suspended, FALSE) FROM users"
            )

        rows = cursor.fetchall()
        users_list = []
        for row in rows:
            user = {
                "id": row[0],
                "username": row[1],
                "reg_no": row[2],
                "name": row[3],
                "dept": row[4],
                "role": row[5],
                "created_at": row[6],
                "suspended": bool(row[9]),
            }
            # Add face status for staff and hod
            if row[5] in ["staff", "hod"]:
                user["face_registered"] = (
                    row[7] is not None if row[7] is not None else False
                )
                user["can_reregister"] = row[8] == 1 if row[8] is not None else False
            users_list.append(user)
        return {"users": users_list, "count": len(users_list)}
    except Exception as e:
        print(f"Error fetching users: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch users")


@app.post("/admin/users/create")
async def admin_create_user(request: Request):
    """Create new user (admin only)"""
    verify_admin_token(request)

    try:
        data = await request.json()
        username = data.get("username")
        password = data.get("password")
        reg_no = data.get("reg_no")
        name = data.get("name")
        dept = data.get("dept")
        role = data.get("role")

        if not all([username, password, name, role]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        # Allow admin-defined roles: admin, hod, staff, principal, vice_chancellor, etc.
        allowed_roles = [
            "admin",
            "hod",
            "staff",
            "principal",
            "vice_chancellor",
            "director",
            "dean",
        ]
        if role not in allowed_roles and not role.startswith("custom_"):
            # Allow any role that starts with 'custom_'
            raise HTTPException(
                status_code=400,
                detail=f"Invalid role. Allowed roles: {', '.join(allowed_roles)}",
            )

        # Auto-generate reg_no if not provided or for hod/staff roles
        if (
            reg_no is None
            or reg_no.strip() == ""
            or role
            in ["hod", "staff", "principal", "vice_chancellor", "director", "dean"]
        ):
            if role == "hod":
                prefix = "HOD"
            elif role == "staff":
                prefix = "STAFF"
            elif role == "principal":
                prefix = "PRINCIPAL"
            elif role == "vice_chancellor":
                prefix = "VC"
            elif role == "director":
                prefix = "DIR"
            elif role == "dean":
                prefix = "DEAN"
            elif role.startswith("custom_"):
                prefix = role.replace("custom_", "").upper()[:6]
            else:
                prefix = "USR"

            # Get count for this role
            cursor.execute("SELECT COUNT(*) FROM users WHERE role = ?", (role,))
            count = cursor.fetchone()[0]
            reg_no = f"{prefix}_{str(count + 1).zfill(4)}"

            # Make sure it's unique
            while get_user_by_reg_no(reg_no):
                count += 1
                reg_no = f"{prefix}_{str(count).zfill(4)}"
        else:
            if get_user_by_reg_no(reg_no):
                raise HTTPException(
                    status_code=400, detail="Registration number already exists"
                )

        if get_user_by_username(username):
            raise HTTPException(status_code=400, detail="Username already exists")

        password_hash = hash_password(password)

        cursor.execute(
            """
            INSERT INTO users (username, password_hash, reg_no, name, dept, role, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
            (username, password_hash, reg_no, name, dept, role, "admin"),
        )
        conn.commit()

        log_audit_event(
            "USER_CREATED", reg_no, True, f"User {username} created with role {role}"
        )

        return success_response(
            "User created successfully",
            {
                "username": username,
                "reg_no": reg_no,
                "name": name,
                "dept": dept,
                "role": role,
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Create user error: {e}")
        raise HTTPException(status_code=500, detail="Failed to create user")


@app.post("/admin/users/bulk-create")
async def admin_bulk_create_users(request: Request):
    """
    Create multiple users at once (admin only)
    Accepts an array of user objects in the request body
    """
    verify_admin_token(request)

    try:
        data = await request.json()
        users_data = data.get("users", [])

        if not users_data or not isinstance(users_data, list):
            raise HTTPException(status_code=400, detail="Users array is required")

        if len(users_data) > 100:
            raise HTTPException(
                status_code=400, detail="Maximum 100 users can be created at once"
            )

        created_users = []
        failed_users = []

        for user_data in users_data:
            try:
                username = user_data.get("username")
                password = user_data.get("password", "password123")  # Default password
                name = user_data.get("name")
                dept = user_data.get("dept")
                role = user_data.get("role", "staff")

                if not all([username, name, dept]):
                    failed_users.append(
                        {"username": username, "error": "Missing required fields"}
                    )
                    continue

                if role not in ["admin", "hod", "staff"]:
                    failed_users.append({"username": username, "error": "Invalid role"})
                    continue

                # Check if username exists
                if get_user_by_username(username):
                    failed_users.append(
                        {"username": username, "error": "Username already exists"}
                    )
                    continue

                # Generate reg_no
                if role == "hod":
                    prefix = "HOD"
                elif role == "staff":
                    prefix = "STAFF"
                else:
                    prefix = "USR"

                cursor.execute("SELECT COUNT(*) FROM users WHERE role = ?", (role,))
                count = cursor.fetchone()[0]
                reg_no = f"{prefix}_{str(count + 1).zfill(4)}"

                while get_user_by_reg_no(reg_no):
                    count += 1
                    reg_no = f"{prefix}_{str(count).zfill(4)}"

                password_hash = hash_password(password)

                cursor.execute(
                    """
                    INSERT INTO users (username, password_hash, reg_no, name, dept, role, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                    (username, password_hash, reg_no, name, dept, role, "admin"),
                )

                created_users.append(
                    {
                        "username": username,
                        "reg_no": reg_no,
                        "name": name,
                        "dept": dept,
                        "role": role,
                        "password": password,  # Return plain password for admin to share
                    }
                )

            except Exception as e:
                failed_users.append(
                    {"username": user_data.get("username", "unknown"), "error": str(e)}
                )

        conn.commit()

        log_audit_event(
            "USERS_BULK_CREATED",
            None,
            True,
            f"Created {len(created_users)} users, {len(failed_users)} failed",
        )

        return success_response(
            f"Bulk user creation completed. {len(created_users)} created, {len(failed_users)} failed.",
            {
                "created_count": len(created_users),
                "failed_count": len(failed_users),
                "created_users": created_users,
                "failed_users": failed_users,
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Bulk create users error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create users: {str(e)}")

@app.get("/admin/users/export/excel")
async def admin_export_users_excel(request: Request):
    verify_admin_token(request)
    import openpyxl
    import io
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Users"
    
    headers = ["ID", "Username", "Register No", "Name", "Department", "Role", "Created At", "Face Registered", "Can Reregister", "Suspended"]
    ws.append(headers)
    
    cursor.execute(
        "SELECT id, username, reg_no, name, dept, role, created_at, embedding, can_reregister, COALESCE(suspended, FALSE) FROM users"
    )
    rows = cursor.fetchall()
    for row in rows:
        face_registered = row[7] is not None
        can_reregister = row[8] == 1
        suspended = bool(row[9])
        ws.append([
            row[0], # ID
            row[1], # Username
            row[2], # Register No
            row[3], # Name
            row[4], # Department
            row[5], # Role
            row[6], # Created At
            "Yes" if face_registered else "No",
            "Yes" if can_reregister else "No",
            "Yes" if suspended else "No"
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=all_users.xlsx"}
    )

@app.get("/admin/users/export/json")
async def admin_export_users_json(request: Request):
    verify_admin_token(request)
    import json
    from fastapi import Response
    
    cursor.execute(
        "SELECT id, username, reg_no, name, dept, role, created_at, embedding, can_reregister, COALESCE(suspended, FALSE) FROM users"
    )
    rows = cursor.fetchall()
    users_list = []
    for row in rows:
        face_registered = row[7] is not None
        can_reregister = row[8] == 1
        suspended = bool(row[9])
        user = {
            "id": row[0],
            "username": row[1],
            "reg_no": row[2],
            "name": row[3],
            "dept": row[4],
            "role": row[5],
            "created_at": row[6],
            "face_registered": face_registered,
            "can_reregister": can_reregister,
            "suspended": suspended
        }
        users_list.append(user)
        
    return Response(
        content=json.dumps(users_list, indent=4),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=all_users.json"}
    )


@app.get("/admin/templates/users/excel")
async def get_users_excel_template(request: Request):
    verify_admin_token(request)
    import openpyxl
    import io
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Users"
    ws.append(["Username", "Name", "Department", "Role", "Password"])
    ws["G1"] = "Available Roles:"
    roles = ["hod", "staff"]
    for i, r in enumerate(roles, start=2):
        ws.cell(row=i, column=7, value=r)
    ws.column_dimensions['G'].width = 25

    cursor.execute("SELECT name FROM departments ORDER BY name")
    depts = [row[0] for row in cursor.fetchall() if row[0].strip().lower() != "administration"]
    ws["H1"] = "Available Departments:"
    for i, d in enumerate(depts, start=2):
        ws.cell(row=i, column=8, value=d)
    ws.column_dimensions['H'].width = 35
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_template.xlsx"}
    )

@app.get("/admin/templates/users/json")
async def get_users_json_template(request: Request):
    verify_admin_token(request)
    cursor.execute("SELECT name FROM departments ORDER BY name")
    depts = [row[0] for row in cursor.fetchall() if row[0].strip().lower() != "administration"]
    template_data = [
        {
            "username": "",
            "name": "",
            "dept": "",
            "role": "",
            "password": "",
            "available_roles": ["hod", "staff"],
            "available_departments": depts
        }
    ]
    import json
    from fastapi import Response
    return Response(
        content=json.dumps(template_data, indent=4),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=users_template.json"}
    )

@app.get("/admin/templates/other_staff/excel")
async def get_other_staff_excel_template(request: Request):
    verify_admin_token(request)
    import openpyxl
    import io
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Other Staff"
    ws.append(["Username", "Name", "Department", "Role", "Password", "DOB"])
    
    ws["H1"] = "Available Roles:"
    roles = ["principal", "placement_staff", "lab_technician", "system_admin", "office_staff"]
    for i, r in enumerate(roles, start=2):
        ws.cell(row=i, column=8, value=r)
    ws.column_dimensions['H'].width = 25

    cursor.execute("SELECT DISTINCT dept FROM other_staff WHERE dept IS NOT NULL AND dept != '' ORDER BY dept")
    other_depts = [row[0] for row in cursor.fetchall()]
    cursor.execute("SELECT name FROM departments")
    standard_depts = {row[0].strip().lower() for row in cursor.fetchall()}
    depts = [d for d in other_depts if d.strip().lower() not in standard_depts]
    default_other_depts = ["Administration", "Placement Staff", "Lab Technician", "System Admin", "Office Staff"]
    for d in default_other_depts:
        if d not in depts and d.strip().lower() not in standard_depts:
            depts.append(d)

    ws["I1"] = "Available Departments:"
    for i, d in enumerate(depts, start=2):
        ws.cell(row=i, column=9, value=d)
    ws.column_dimensions['I'].width = 35
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=other_staff_template.xlsx"}
    )

@app.get("/admin/templates/other_staff/json")
async def get_other_staff_json_template(request: Request):
    verify_admin_token(request)
    cursor.execute("SELECT DISTINCT dept FROM other_staff WHERE dept IS NOT NULL AND dept != '' ORDER BY dept")
    other_depts = [row[0] for row in cursor.fetchall()]
    cursor.execute("SELECT name FROM departments")
    standard_depts = {row[0].strip().lower() for row in cursor.fetchall()}
    depts = [d for d in other_depts if d.strip().lower() not in standard_depts]
    default_other_depts = ["Administration", "Placement Staff", "Lab Technician", "System Admin", "Office Staff"]
    for d in default_other_depts:
        if d not in depts and d.strip().lower() not in standard_depts:
            depts.append(d)

    template_data = [
        {
            "username": "",
            "name": "",
            "dept": "",
            "role": "",
            "password": "",
            "dob": "",
            "available_roles": ["principal", "placement_staff", "lab_technician", "system_admin", "office_staff"],
            "available_departments": depts
        }
    ]
    import json
    from fastapi import Response
    return Response(
        content=json.dumps(template_data, indent=4),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=other_staff_template.json"}
    )

@app.get("/admin/templates/departments/excel")
async def get_departments_excel_template(request: Request):
    verify_admin_token(request)
    import openpyxl
    import io
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Departments"
    ws.append(["Username", "Name", "Department", "Role", "Password"])
    ws["G1"] = "Available Roles:"
    roles = ["hod", "staff"]
    for i, r in enumerate(roles, start=2):
        ws.cell(row=i, column=7, value=r)
    ws.column_dimensions['G'].width = 25

    cursor.execute("SELECT name FROM departments ORDER BY name")
    depts = [row[0] for row in cursor.fetchall() if row[0].strip().lower() != "administration"]
    ws["H1"] = "Available Departments:"
    for i, d in enumerate(depts, start=2):
        ws.cell(row=i, column=8, value=d)
    ws.column_dimensions['H'].width = 35
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=departments_template.xlsx"}
    )

@app.get("/admin/templates/departments/json")
async def get_departments_json_template(request: Request):
    verify_admin_token(request)
    cursor.execute("SELECT name FROM departments ORDER BY name")
    depts = [row[0] for row in cursor.fetchall() if row[0].strip().lower() != "administration"]
    template_data = [
        {
            "username": "",
            "name": "",
            "dept": "",
            "role": "",
            "password": "",
            "available_roles": ["hod", "staff"],
            "available_departments": depts
        }
    ]
    import json
    from fastapi import Response
    return Response(
        content=json.dumps(template_data, indent=4),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=departments_template.json"}
    )


@app.post("/admin/users/bulk-upload")
async def admin_bulk_upload_users(request: Request, file: UploadFile = File(...)):
    """
    Bulk upload users via JSON or Excel (.xlsx/.xls) (admin only)
    """
    verify_admin_token(request)

    filename = file.filename or ""
    is_excel = filename.endswith(".xlsx") or filename.endswith(".xls")
    is_json = filename.endswith(".json")

    if not (is_excel or is_json):
        raise HTTPException(status_code=400, detail="Only JSON (.json) and Excel (.xlsx, .xls) files are supported")

    users_data = []
    try:
        if is_json:
            contents = await file.read()
            try:
                users_data = json.loads(contents.decode("utf-8"))
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid JSON format: {str(e)}")
            if not isinstance(users_data, list):
                raise HTTPException(status_code=400, detail="JSON must be an array of user objects")
        else:
            # Excel parsing
            contents = await file.read()
            import io
            from openpyxl import load_workbook
            f = io.BytesIO(contents)
            wb = load_workbook(f, read_only=True, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Excel file is empty")

            headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            
            idx_username = -1
            idx_name = -1
            idx_dept = -1
            idx_role = -1
            idx_password = -1

            for idx, h in enumerate(headers):
                if h in ["username", "user_name", "user"]:
                    idx_username = idx
                elif h in ["name", "full name", "fullname", "full_name"]:
                    idx_name = idx
                elif h in ["dept", "department", "department name", "department_name"]:
                    idx_dept = idx
                elif h in ["role", "user_role"]:
                    idx_role = idx
                elif h in ["password", "pass"]:
                    idx_password = idx

            missing = []
            if idx_username == -1:
                missing.append("username")
            if idx_name == -1:
                missing.append("name")
            if idx_dept == -1:
                missing.append("dept")
            if idx_role == -1:
                missing.append("role")

            if missing:
                raise HTTPException(status_code=400, detail=f"Missing required columns in Excel headers: {', '.join(missing)}")

            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                username = str(row[idx_username]).strip() if idx_username < len(row) and row[idx_username] is not None else None
                name = str(row[idx_name]).strip() if idx_name < len(row) and row[idx_name] is not None else None
                dept = str(row[idx_dept]).strip() if idx_dept < len(row) and row[idx_dept] is not None else None
                role = str(row[idx_role]).strip().lower() if idx_role < len(row) and row[idx_role] is not None else "staff"
                password = str(row[idx_password]).strip() if idx_password != -1 and idx_password < len(row) and row[idx_password] is not None else "password123"

                if not username and not name and not dept:
                    continue

                users_data.append({
                    "username": username,
                    "name": name,
                    "dept": dept,
                    "role": role,
                    "password": password
                })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")

    created_users = []
    failed_users = []
    skipped_users = []
    seen_usernames = set()

    try:
        # Run everything in a transaction
        for idx, user_data in enumerate(users_data, start=2):
            try:
                username = user_data.get("username")
                password = user_data.get("password") or "password123"
                name = user_data.get("name")
                dept = user_data.get("dept")
                role = user_data.get("role") or "staff"

                if not all([username, name, dept]):
                    failed_users.append({
                        "username": username or f"Row {idx}",
                        "error": "Missing required fields (username, name, dept)"
                    })
                    continue

                username = username.strip()
                name = name.strip()
                dept = dept.strip()
                role = role.strip().lower()

                allowed_roles = ["admin", "hod", "staff", "principal", "vice_chancellor", "director", "dean"]
                if role not in allowed_roles and not role.startswith("custom_"):
                    failed_users.append({
                        "username": username,
                        "error": f"Invalid role '{role}'"
                    })
                    continue

                if username in seen_usernames:
                    failed_users.append({
                        "username": username,
                        "error": "Duplicate username within the uploaded file"
                    })
                    continue
                seen_usernames.add(username)

                existing_user = get_user_by_username(username)
                if existing_user:
                    skipped_users.append({
                        "username": username,
                        "reason": "Username already exists in database"
                    })
                    continue

                # Auto-create department if it does not exist
                cursor.execute("SELECT id FROM departments WHERE name = ?", (dept,))
                dept_row = cursor.fetchone()
                if not dept_row:
                    cursor.execute("INSERT INTO departments (name) VALUES (?)", (dept,))

                # Generate reg_no
                if role == "hod":
                    prefix = "HOD"
                elif role == "staff":
                    prefix = "STAFF"
                elif role == "principal":
                    prefix = "PRINCIPAL"
                elif role == "vice_chancellor":
                    prefix = "VC"
                elif role == "director":
                    prefix = "DIR"
                elif role == "dean":
                    prefix = "DEAN"
                elif role.startswith("custom_"):
                    prefix = role.replace("custom_", "").upper()[:6]
                else:
                    prefix = "USR"

                cursor.execute("SELECT COUNT(*) FROM users WHERE role = ?", (role,))
                count = cursor.fetchone()[0]
                reg_no = f"{prefix}_{str(count + 1).zfill(4)}"

                while get_user_by_reg_no(reg_no):
                    count += 1
                    reg_no = f"{prefix}_{str(count).zfill(4)}"

                password_hash = hash_password(password)

                cursor.execute(
                    """
                    INSERT INTO users (username, password_hash, reg_no, name, dept, role, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (username, password_hash, reg_no, name, dept, role, "admin")
                )

                created_users.append({
                    "username": username,
                    "reg_no": reg_no,
                    "name": name,
                    "dept": dept,
                    "role": role,
                    "password": password
                })
            except Exception as row_err:
                failed_users.append({
                    "username": user_data.get("username") or f"Row {idx}",
                    "error": str(row_err)
                })

        conn.commit()

        log_audit_event(
            "USERS_BULK_UPLOADED",
            None,
            True,
            f"Uploaded {len(created_users)} users, {len(failed_users)} failed via file"
        )

        return success_response(
            f"Bulk upload completed. {len(created_users)} created, {len(skipped_users)} skipped, {len(failed_users)} failed.",
            {
                "created_count": len(created_users),
                "skipped_count": len(skipped_users),
                "failed_count": len(failed_users),
                "created_users": created_users,
                "skipped_users": skipped_users,
                "failed_users": failed_users
            }
        )
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database transaction failed: {str(e)}")


@app.post("/admin/other_staff/bulk-upload")
async def admin_bulk_upload_other_staff(request: Request, file: UploadFile = File(...)):
    """
    Bulk upload other staff via JSON or Excel (.xlsx/.xls) (admin only)
    """
    verify_admin_token(request)

    filename = file.filename or ""
    is_excel = filename.endswith(".xlsx") or filename.endswith(".xls")
    is_json = filename.endswith(".json")

    if not (is_excel or is_json):
        raise HTTPException(status_code=400, detail="Only JSON (.json) and Excel (.xlsx, .xls) files are supported")

    staff_data = []
    try:
        if is_json:
            contents = await file.read()
            try:
                staff_data = json.loads(contents.decode("utf-8"))
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid JSON format: {str(e)}")
            if not isinstance(staff_data, list):
                raise HTTPException(status_code=400, detail="JSON must be an array of staff objects")
        else:
            # Excel parsing
            contents = await file.read()
            import io
            from openpyxl import load_workbook
            f = io.BytesIO(contents)
            wb = load_workbook(f, read_only=True, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Excel file is empty")

            headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            
            idx_username = -1
            idx_name = -1
            idx_dept = -1
            idx_role = -1
            idx_password = -1
            idx_dob = -1

            for idx, h in enumerate(headers):
                if h in ["username", "user_name", "user"]:
                    idx_username = idx
                elif h in ["name", "full name", "fullname", "full_name"]:
                    idx_name = idx
                elif h in ["dept", "department", "department name", "department_name"]:
                    idx_dept = idx
                elif h in ["role", "user_role"]:
                    idx_role = idx
                elif h in ["password", "pass"]:
                    idx_password = idx
                elif h in ["dob", "date of birth", "birth_date", "birthdate"]:
                    idx_dob = idx

            missing = []
            if idx_username == -1:
                missing.append("username")
            if idx_name == -1:
                missing.append("name")
            if idx_dept == -1:
                missing.append("dept")
            if idx_role == -1:
                missing.append("role")

            if missing:
                raise HTTPException(status_code=400, detail=f"Missing required columns in Excel headers: {', '.join(missing)}")

            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                username = str(row[idx_username]).strip() if idx_username < len(row) and row[idx_username] is not None else None
                name = str(row[idx_name]).strip() if idx_name < len(row) and row[idx_name] is not None else None
                dept = str(row[idx_dept]).strip() if idx_dept < len(row) and row[idx_dept] is not None else None
                role = str(row[idx_role]).strip().lower() if idx_role < len(row) and row[idx_role] is not None else "office_staff"
                password = str(row[idx_password]).strip() if idx_password != -1 and idx_password < len(row) and row[idx_password] is not None else "password123"
                dob = str(row[idx_dob]).strip() if idx_dob != -1 and idx_dob < len(row) and row[idx_dob] is not None else None

                if not username and not name and not dept:
                    continue

                staff_data.append({
                    "username": username,
                    "name": name,
                    "dept": dept,
                    "role": role,
                    "password": password,
                    "dob": dob
                })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")

    created_staff = []
    failed_staff = []
    skipped_staff = []
    seen_usernames = set()

    try:
        for idx, s_data in enumerate(staff_data, start=2):
            try:
                username = s_data.get("username")
                password = s_data.get("password") or "password123"
                name = s_data.get("name")
                dept = s_data.get("dept")
                role = s_data.get("role") or "office_staff"
                dob = s_data.get("dob")

                if not all([username, name, dept]):
                    failed_staff.append({
                        "username": username or f"Row {idx}",
                        "error": "Missing required fields (username, name, dept)"
                    })
                    continue

                username = username.strip()
                name = name.strip()
                dept = dept.strip()
                role = role.strip().lower()

                allowed_roles = list(OTHER_STAFF_ROLES)
                if role not in allowed_roles:
                    failed_staff.append({
                        "username": username,
                        "error": f"Invalid role '{role}' for other staff"
                    })
                    continue

                if username in seen_usernames:
                    failed_staff.append({
                        "username": username,
                        "error": "Duplicate username within the uploaded file"
                    })
                    continue
                seen_usernames.add(username)

                # Auto-create department if it does not exist
                cursor.execute("SELECT id FROM departments WHERE name = ?", (dept,))
                dept_row = cursor.fetchone()
                if not dept_row:
                    cursor.execute("INSERT INTO departments (name) VALUES (?)", (dept,))

                existing_staff = get_other_staff_by_username(username)
                if existing_staff:
                    skipped_staff.append({
                        "username": username,
                        "reason": "Username already exists in database"
                    })
                    continue

                # Auto-generate reg_no
                if role == "principal":
                    prefix = "PRINCIPAL"
                elif role == "placement_staff":
                    prefix = "PLACE"
                elif role == "lab_technician":
                    prefix = "LAB"
                elif role == "system_admin":
                    prefix = "SYS"
                elif role == "office_staff":
                    prefix = "OFFICE"
                else:
                    prefix = "OS"

                cursor.execute("SELECT COUNT(*) FROM other_staff WHERE role = ?", (role,))
                count = cursor.fetchone()[0]
                reg_no = f"{prefix}_{str(count + 1).zfill(4)}"

                while get_other_staff_by_reg_no(reg_no):
                    count += 1
                    reg_no = f"{prefix}_{str(count).zfill(4)}"

                password_hash = hash_password(password)

                cursor.execute(
                    """
                    INSERT INTO other_staff (username, password_hash, reg_no, name, dob, role, dept, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (username, password_hash, reg_no, name, dob, role, dept, "admin")
                )

                created_staff.append({
                    "username": username,
                    "reg_no": reg_no,
                    "name": name,
                    "dept": dept,
                    "role": role,
                    "password": password
                })
            except Exception as row_err:
                failed_staff.append({
                    "username": s_data.get("username") or f"Row {idx}",
                    "error": str(row_err)
                })

        conn.commit()

        log_audit_event(
            "OTHER_STAFF_BULK_UPLOADED",
            None,
            True,
            f"Uploaded {len(created_staff)} other staff, {len(failed_staff)} failed via file"
        )

        return success_response(
            f"Bulk upload completed. {len(created_staff)} created, {len(skipped_staff)} skipped, {len(failed_staff)} failed.",
            {
                "created_count": len(created_staff),
                "skipped_count": len(skipped_staff),
                "failed_count": len(failed_staff),
                "created_users": created_staff,
                "skipped_users": skipped_staff,
                "failed_users": failed_staff
            }
        )
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Database transaction failed: {str(e)}")


@app.put("/admin/users/{user_id}")
async def admin_update_user(request: Request, user_id: int):
    """
    Update user details (admin only)
    """
    # Use verify_admin_token to get admin info - token verification is sufficient
    verify_admin_token(request)

    try:
        # Parse request - JSON only
        data = await request.json()
        name = data.get("name")
        username = data.get("username")
        dept = data.get("dept")
        role = data.get("role")
        password = data.get("password")
        suspended = data.get("suspended")
    except Exception as e:
        print(f"Parse error: {e}")
        raise HTTPException(status_code=400, detail="Invalid request body")

    if (
        role
        and role
        not in [
            "admin",
            "hod",
            "staff",
            "principal",
            "vice_chancellor",
            "director",
            "dean",
        ]
        and not role.startswith("custom_")
    ):
        raise HTTPException(status_code=400, detail="Invalid role")

    try:
        cursor.execute(
            "SELECT id, username, password_hash, reg_no, name, dept, role, is_active, face_embedding, current_device_id, out_permission_enabled, out_permission_expiry, created_by, created_at, updated_at, embedding, can_reregister, suspended FROM users WHERE id = ?",
            (user_id,),
        )
        existing = cursor.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="User not found")

        if username and username != existing[1]:
            cursor.execute(
                "SELECT id FROM users WHERE username = ? AND id != ?",
                (username, user_id),
            )
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="Username already exists")

        updates = []
        params = []
        if name is not None:
            updates.append("name = ?")
            params.append(name)
        if username is not None:
            updates.append("username = ?")
            params.append(username)
        if dept is not None:
            updates.append("dept = ?")
            params.append(dept)
        if role is not None:
            updates.append("role = ?")
            params.append(role)
        if password:
            updates.append("password_hash = ?")
            params.append(hash_password(password))
        if suspended is not None:
            updates.append("suspended = ?")
            params.append(bool(suspended))

        if not updates:
            return {"message": "No changes to update"}

        params.append(user_id)
        cursor.execute(
            f"UPDATE users SET {', '.join(updates)} WHERE id = ?",
            tuple(params),
        )
        conn.commit()

        log_audit_event(
            "USER_UPDATED", existing[3], True, f"User {existing[1]} updated"
        )

        return {"message": "User updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Update user error: {e}")
        raise HTTPException(status_code=500, detail="Failed to update user")


@app.delete("/admin/users/{user_id}")
async def admin_delete_user(request: Request, user_id: int):
    """
    Delete user (admin only)
    """
    # Use verify_admin_token to get admin info - token verification is sufficient
    admin_user = verify_admin_token(request)

    try:
        cursor.execute(
            "SELECT id, username, password_hash, reg_no, name, dept, role, is_active, face_embedding, current_device_id, out_permission_enabled, out_permission_expiry, created_by, created_at, updated_at, embedding, can_reregister, suspended FROM users WHERE id = ?",
            (user_id,),
        )
        existing = cursor.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="User not found")

        # Prevent deleting admin users (role is at index 6)
        if existing[6] == "admin":
            raise HTTPException(status_code=400, detail="Cannot delete admin user")

        reg_no = existing[3]
        delete_user_data_by_reg_no(reg_no)

        cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()

        log_audit_event(
            "USER_DELETED", existing[3], True, f"User {existing[1]} deleted"
        )

        return {"message": "User deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Delete user error: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete user")
        print(f"Delete user error: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete user")


# -------------------------------------------------
# ADMIN OTHER STAFF MANAGEMENT
# -------------------------------------------------
@app.get("/admin/other_staff")
async def admin_get_other_staff(request: Request, role: str = None):
    """Get all other staff members (Admin only)"""
    verify_admin_token(request)

    try:
        if role:
            cursor.execute(
                "SELECT id, username, reg_no, name, dob, role, dept, embedding, can_reregister, created_at, created_by, COALESCE(suspended, FALSE) FROM other_staff WHERE role = ?",
                (role,),
            )
        else:
            cursor.execute(
                "SELECT id, username, reg_no, name, dob, role, dept, embedding, can_reregister, created_at, created_by, COALESCE(suspended, FALSE) FROM other_staff"
            )

        rows = cursor.fetchall()
        staff_list = []
        for row in rows:
            staff_list.append(
                {
                    "id": row[0],
                    "username": row[1],
                    "reg_no": row[2],  # Use snake_case to match frontend
                    "name": row[3],
                    "dob": row[4],
                    "role": row[5],
                    "dept": row[6],
                    "face_registered": row[7] is not None,
                    "can_reregister": row[8] == 1 if row[8] else False,
                    "created_at": row[9],  # Use snake_case
                    "created_by": row[10],  # Use snake_case
                    "suspended": bool(row[11]),
                }
            )
        return {"other_staff": staff_list}
    except Exception as e:
        print(f"Error fetching other staff: {e}")
        return {"other_staff": []}


@app.post("/admin/other_staff/create")
async def admin_create_other_staff(request: Request):
    """Create a new other staff member (Admin only)"""
    try:
        admin_user = verify_admin_token(request)
        data = await request.json()

        username = data.get("username")
        password = data.get("password")
        reg_no = data.get("reg_no")
        name = data.get("name")
        dob = data.get("dob")
        role = data.get("role")
        dept = data.get("dept")
        default_dept = get_default_department_for_role(role)
        if default_dept:
            dept = default_dept
        elif isinstance(dept, str):
            dept = dept.strip()

        # DEBUG: Log received data
        print(
            f"[DEBUG] Create other_staff request - username: {username}, reg_no: {reg_no}, name: {name}, role: {role}, dept: {dept}"
        )

        # Validate required fields
        if not all([username, password, name, role]):
            print(
                f"[DEBUG] Missing required fields - username: {username}, password: {password}, name: {name}, role: {role}"
            )
            raise HTTPException(status_code=400, detail="Missing required fields")

        # Validate role
        allowed_roles = list(OTHER_STAFF_ROLES)
        print(
            f"[DEBUG] Role validation - received: '{role}', allowed: {allowed_roles}, is_valid: {role in allowed_roles}"
        )
        if role not in allowed_roles:
            raise HTTPException(status_code=400, detail="Invalid role")

        # Check if user already exists
        print(
            f"[DEBUG] Checking if user exists - username: {username}"
        )
        if get_other_staff_by_username(username):
            print(f"[DEBUG] Username already exists: {username}")
            raise HTTPException(status_code=400, detail="Username already exists")

        # Auto-generate reg_no if not provided
        if not reg_no or reg_no.strip() == "":
            if role == "principal":
                prefix = "PRINCIPAL"
            elif role == "placement_staff":
                prefix = "PLACE"
            elif role == "lab_technician":
                prefix = "LAB"
            elif role == "system_admin":
                prefix = "SYS"
            elif role == "office_staff":
                prefix = "OFFICE"
            else:
                prefix = "OS"

            cursor.execute("SELECT COUNT(*) FROM other_staff WHERE role = ?", (role,))
            count = cursor.fetchone()[0]
            reg_no = f"{prefix}_{str(count + 1).zfill(4)}"

            while get_other_staff_by_reg_no(reg_no):
                count += 1
                reg_no = f"{prefix}_{str(count).zfill(4)}"
        else:
            if get_other_staff_by_reg_no(reg_no):
                print(f"[DEBUG] Registration number already exists: {reg_no}")
                raise HTTPException(
                    status_code=400, detail="Registration number already exists"
                )

        # Hash password
        password_hash = hash_password(password)

        # Insert user
        print(
            f"[DEBUG] Inserting user - username: {username}, reg_no: {reg_no}, name: {name}, role: {role}, dob: {dob}, dept: {dept}"
        )
        cursor.execute(
            """
            INSERT INTO other_staff (username, password_hash, reg_no, name, dob, role, dept, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                username,
                password_hash,
                reg_no,
                name,
                dob,
                role,
                dept,
                admin_user["username"],
            ),
        )
        conn.commit()

        log_audit_event(
            "OTHER_STAFF_CREATED",
            reg_no,
            True,
            f"Created by admin {admin_user['username']}",
        )

        return {
            "message": "Other staff created successfully",
            "other_staff": {
                "username": username,
                "reg_no": reg_no,
                "name": name,
                "dob": dob,
                "role": role,
                "dept": dept,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Create other staff error: {e}")
        raise HTTPException(status_code=500, detail="Failed to create other staff")


@app.put("/admin/other_staff/{staff_id}")
async def admin_update_other_staff(request: Request, staff_id: int):
    """Update other staff member (Admin only)"""
    try:
        admin_user = verify_admin_token(request)
        data = await request.json()

        # Get existing staff
        staff = get_other_staff_by_id(staff_id)
        if not staff:
            raise HTTPException(status_code=404, detail="Staff not found")

        # Update fields
        # staff layout: id, username, password_hash, reg_no, name, dob, role, dept, ...
        username = data.get("username", staff[1])
        reg_no = data.get("reg_no", staff[3])
        if isinstance(username, str):
            username = username.strip()
        if isinstance(reg_no, str):
            reg_no = reg_no.strip()
        name = data.get("name", staff[4])
        dob = data.get("dob", staff[5])
        role = data.get("role", staff[6])
        dept = data.get("dept", staff[7])
        default_dept = get_default_department_for_role(role)
        if default_dept:
            dept = default_dept
        elif isinstance(dept, str):
            dept = dept.strip()
        can_reregister = bool(data.get("can_reregister", False))

        # Validate role
        if role not in OTHER_STAFF_ROLES:
            raise HTTPException(status_code=400, detail="Invalid role")

        # Keep existing values if empty strings were sent
        if not username:
            username = staff[1]
        if not reg_no:
            reg_no = staff[3]
        if not name:
            name = staff[4]

        # Check username uniqueness
        if username and (username.strip().lower() != (staff[1] or "").strip().lower()):
            cursor.execute(
                "SELECT id FROM other_staff WHERE LOWER(username) = LOWER(?)",
                (username,),
            )
            existing_user = cursor.fetchone()
            if existing_user and existing_user[0] != staff_id:
                raise HTTPException(status_code=400, detail="Username already exists")

        # Check reg_no uniqueness - strip whitespace and compare
        # staff[3] is the reg_no field based on get_other_staff_by_id query
        reg_no = reg_no.strip() if reg_no else reg_no
        original_reg_no = (staff[3] or "").strip() if staff[3] else ""
        if reg_no and reg_no.lower() != original_reg_no.lower():
            cursor.execute(
                "SELECT id FROM other_staff WHERE LOWER(reg_no) = LOWER(?)",
                (reg_no,),
            )
            existing_reg = cursor.fetchone()
            if existing_reg and existing_reg[0] != staff_id:
                raise HTTPException(
                    status_code=400,
                    detail="Registration number already exists",
                )

        # Update password if provided
        password_hash = staff[2]
        if data.get("password"):
            password_hash = hash_password(data["password"])

        # Get existing suspended status
        cursor.execute("SELECT COALESCE(suspended, FALSE) FROM other_staff WHERE id = ?", (staff_id,))
        existing_suspended = cursor.fetchone()[0]
        suspended = bool(data.get("suspended", existing_suspended))

        cursor.execute(
            """
            UPDATE other_staff 
            SET username = ?, password_hash = ?, reg_no = ?, name = ?, dob = ?, role = ?, dept = ?, can_reregister = ?, suspended = ?
            WHERE id = ?
        """,
            (
                username,
                password_hash,
                reg_no,
                name,
                dob,
                role,
                dept,
                can_reregister,
                suspended,
                staff_id,
            ),
        )
        conn.commit()

        log_audit_event(
            "OTHER_STAFF_UPDATED",
            reg_no,
            True,
            f"Updated by admin {admin_user['username']}",
        )

        return {
            "message": "Other staff updated successfully",
            "other_staff": {
                "id": staff_id,
                "username": username,
                "reg_no": reg_no,
                "name": name,
                "dob": dob,
                "role": role,
                "dept": dept,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        import psycopg2
        if isinstance(e, psycopg2.IntegrityError):
            msg = str(e).lower()
            if "other_staff.username" in msg or "username" in msg:
                raise HTTPException(status_code=400, detail="Username already exists")
            if "other_staff.reg_no" in msg or "reg_no" in msg:
                raise HTTPException(
                    status_code=400,
                    detail="Registration number already exists",
                )
            raise HTTPException(status_code=400, detail="Invalid update data")

        print(f"Update other staff error: {e}")
        raise HTTPException(status_code=500, detail="Failed to update other staff")


@app.delete("/admin/other_staff/{staff_id}")
async def admin_delete_other_staff(request: Request, staff_id: int):
    """Delete other staff member (Admin only)"""
    try:
        admin_user = verify_admin_token(request)

        # Get staff to delete
        staff = get_other_staff_by_id(staff_id)
        if not staff:
            raise HTTPException(status_code=404, detail="Staff not found")

        # Delete staff
        reg_no = staff[3]
        delete_user_data_by_reg_no(reg_no)

        cursor.execute("DELETE FROM other_staff WHERE id = ?", (staff_id,))
        conn.commit()

        log_audit_event(
            "OTHER_STAFF_DELETED",
            staff[3],
            True,
            f"Deleted by admin {admin_user['username']}",
        )

        return {
            "message": "Other staff deleted successfully",
            "deleted_staff": {
                "id": staff[0],
                "username": staff[1],
                "reg_no": staff[3],
                "name": staff[4],
                "role": staff[6],
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Delete other staff error: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete other staff")


@app.get("/admin/other_staff/attendance")
async def admin_get_other_staff_attendance(
    request: Request,
    dept: str = None,
    role: str = None,
    date: str = None,
    reg_no: str = None,
    start_date: str = None,
    end_date: str = None,
):
    """Get other staff attendance with role/department filters (Admin only).
    Includes face scan attendance + OD/Leave from daily_attendance_status.
    """
    verify_admin_token(request)

    try:
        conditions: list[str] = []
        params: list[str] = []

        # Build query for face scan attendance
        if dept and dept.strip() and dept != "All":
            conditions.append("osa.dept = ?")
            params.append(dept.strip())

        if role and role.strip() and role != "All":
            role_val = role.strip()
            if role_val not in OTHER_STAFF_ROLES:
                raise HTTPException(status_code=400, detail="Invalid role filter")
            conditions.append("osa.role = ?")
            params.append(role_val)

        if date and date.strip():
            conditions.append("osa.timestamp::date = ?")
            params.append(date.strip())

        if reg_no and reg_no.strip():
            conditions.append("osa.reg_no = ?")
            params.append(reg_no.strip())

        if start_date and start_date.strip():
            conditions.append("osa.timestamp::date >= ?")
            params.append(start_date.strip())

        if end_date and end_date.strip():
            conditions.append("osa.timestamp::date <= ?")
            params.append(end_date.strip())

        where_clause = (
            f"WHERE {' AND '.join(conditions)}" if conditions else "WHERE 1=1"
        )

        # Get face scan attendance
        cursor.execute(
            f"""
            SELECT osa.id, osa.reg_no, osa.name, osa.dept, osa.role, osa.timestamp, 'face_scan' as source
            FROM other_staff_attendance osa
            {where_clause}
            ORDER BY osa.timestamp DESC
            LIMIT 1000
        """,
            params,
        )
        attendance_rows = cursor.fetchall()

        # Get OD/Leave records from daily_attendance_status for the same filters
        das_conditions = []
        das_params = list(params)  # Copy the params

        if dept and dept.strip() and dept != "All":
            das_conditions.append("das.dept = ?")
        if role and role.strip() and role != "All":
            das_conditions.append(
                "das.name IN (SELECT name FROM other_staff WHERE role = ?)"
            )
            das_params.append(role.strip())
        if reg_no and reg_no.strip():
            das_conditions.append("das.reg_no = ?")
        if date and date.strip():
            das_conditions.append("das.date::date = ?")
        if start_date and start_date.strip():
            das_conditions.append("das.date::date >= ?")
        if end_date and end_date.strip():
            das_conditions.append("das.date::date <= ?")

        das_where = " AND ".join(das_conditions) if das_conditions else "1=1"

        cursor.execute(
            f"""
            SELECT das.id, das.reg_no, das.name, das.dept, 
                   (SELECT role FROM other_staff WHERE reg_no = das.reg_no LIMIT 1) as role,
                   das.date as timestamp, 'od' as source, das.status, das.leave_type
            FROM daily_attendance_status das
            WHERE {das_where} AND das.leave_type IN ('od', 'earned', 'casual')
            ORDER BY das.date DESC
            LIMIT 100
        """,
            das_params,
        )
        od_rows = cursor.fetchall()

        # Combine and build response
        all_attendance = list(attendance_rows) + list(od_rows)

        cursor.execute("""
            SELECT role, dept, COUNT(*) as count
            FROM other_staff
            GROUP BY role, dept
            ORDER BY role
        """)
        staff_rows = cursor.fetchall()

        cursor.execute("""
            SELECT dept, COUNT(*) as count
            FROM other_staff_attendance
            GROUP BY dept
            ORDER BY dept
        """)
        dept_totals = cursor.fetchall()

        cursor.execute("""
            SELECT role, COUNT(*) as count
            FROM other_staff_attendance
            GROUP BY role
            ORDER BY role
        """)
        role_totals = cursor.fetchall()

        result = {
            "attendance": [
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "role": row[4] if len(row) > 4 else row[4],
                    "timestamp": _ts(row[5]) if row[5] else str(row[5]),
                    "source": row[6] if len(row) > 6 else "face_scan",
                    "status": row[7] if len(row) > 7 else None,
                    "leave_type": row[8] if len(row) > 8 else None,
                }
                for row in all_attendance
            ],
            "count": len(all_attendance),
            "staff_counts": [
                {"role": row[0], "dept": row[1], "count": row[2]} for row in staff_rows
            ],
            "attendance_by_dept": [
                {"dept": row[0], "count": row[1]} for row in dept_totals
            ],
            "attendance_by_role": [
                {"role": row[0], "count": row[1]} for row in role_totals
            ],
        }

        # Add holiday-aware stats when querying for a single staff with date range
        if reg_no and start_date and end_date:
            _start, _end = _clamp_to_academic_year(start_date, end_date)
            _end_dt = _cap_end_to_today(datetime.strptime(_end, "%Y-%m-%d"))
            _holidays = _get_holiday_dates_in_range(_start, _end_dt.strftime("%Y-%m-%d"))
            _working_days = 0
            _d = datetime.strptime(_start, "%Y-%m-%d")
            while _d <= _end_dt:
                if _d.weekday() < 5 and _d.strftime("%Y-%m-%d") not in _holidays:
                    _working_days += 1
                _d += timedelta(days=1)
            _present_count = len(_get_full_attendance_dates(reg_no, _start, _end, is_other_staff=True))
            result["working_days"] = _working_days
            result["present_days"] = _present_count
            result["absent_days"] = max(0, _working_days - _present_count)

        return result
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching other staff attendance: {e}")
        raise HTTPException(
            status_code=500, detail="Failed to fetch other staff attendance"
        )


@app.post("/admin/other_staff/face/register")
async def admin_register_other_staff_face(
    request: Request,
    name: str = Form(...),
    reg_no: str = Form(...),
    dept: str = Form(...),
    role: str = Form(...),
    image: UploadFile = File(...),
):
    """Register face for other_staff roles (Admin only)."""
    verify_admin_token(request)

    role = (role or "").strip().lower()
    if role not in OTHER_STAFF_ROLES:
        raise HTTPException(status_code=400, detail="Invalid other staff role")

    cursor.execute("SELECT id, name, role FROM other_staff WHERE reg_no = ?", (reg_no,))
    existing = cursor.fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Other staff not found")

    img_bytes = await image.read()
    img = preprocess_image_data(img_bytes)
    face = extract_face(img)
    if face is None:
        save_debug_image(img, "register_fail")
        raise HTTPException(
            status_code=400,
            detail="Face not detected clearly. Ensure good lighting and face is visible.",
        )

    embedding = face.embedding.astype(np.float32)
    cursor.execute(
        "UPDATE other_staff SET embedding = ? WHERE reg_no = ?",
        (embedding.tobytes(), reg_no),
    )
    conn.commit()
    _cache_update_primary(reg_no, embedding)

    return {
        "message": "Face registered successfully",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "role": role,
        "bbox": [float(x) for x in face.bbox],
    }


@app.post("/admin/other_staff/face/permission/{reg_no}")
async def admin_grant_other_staff_permission(request: Request, reg_no: str):
    """Admin grants face re-registration permission for other_staff user."""
    verify_admin_token(request)

    cursor.execute("SELECT id, name FROM other_staff WHERE reg_no = ?", (reg_no,))
    existing = cursor.fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Other staff not found")

    cursor.execute(
        "UPDATE other_staff SET can_reregister = 1 WHERE reg_no = ?", (reg_no,)
    )
    conn.commit()
    return {
        "message": "Permission granted successfully. User can now re-register their face.",
        "reg_no": reg_no,
        "user_name": existing[1],
    }


@app.delete("/admin/other_staff/face/permission/{reg_no}")
async def admin_revoke_other_staff_permission(request: Request, reg_no: str):
    """Admin revokes face re-registration permission for other_staff user."""
    verify_admin_token(request)

    cursor.execute("SELECT id, name FROM other_staff WHERE reg_no = ?", (reg_no,))
    existing = cursor.fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Other staff not found")

    cursor.execute(
        "UPDATE other_staff SET can_reregister = 0 WHERE reg_no = ?", (reg_no,)
    )
    conn.commit()
    return {
        "message": "Permission revoked successfully.",
        "reg_no": reg_no,
        "user_name": existing[1],
    }


@app.get("/admin/attendance")
async def admin_get_attendance(request: Request, date: str = None, limit: int = 1000):
    """Get attendance records (admin only) - optimized with indexes and LIMIT"""
    verify_admin_token(request)

    try:
        # Validate and limit the query for performance
        limit = min(max(limit, 100), 5000)  # Between 100 and 5000 records

        if date:
            # Uses idx_attendance_timestamp index (partial match)
            cursor.execute(
                """
                SELECT id, reg_no, name, dept, class_div, timestamp, status
                FROM attendance
                WHERE timestamp::date = ?
                ORDER BY timestamp DESC
                LIMIT ?
            """,
                (date, limit),
            )
        else:
            # Uses idx_attendance_timestamp index for ordering
            cursor.execute("""
                SELECT id, reg_no, name, dept, class_div, timestamp, status
                FROM attendance
                ORDER BY timestamp DESC
                LIMIT ?
            """, (limit,))

        rows = cursor.fetchall()
        return {
            "attendance": [
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "class_div": row[4] or "",
                    "timestamp": _ts(row[5]),
                    "status": _format_scan_status(row[6], row[5]),
                }
                for row in rows
            ],
            "count": len(rows),
            "limit": limit,
            "has_more": len(rows) == limit
        }
    except Exception as e:
        print(f"Error fetching attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch attendance")


@app.get("/admin/attendance/staff")
async def admin_get_staff_attendance(
    request: Request,
    date: str = None,
    start_date: str = None,
    end_date: str = None,
    reg_no: str = None,
):
    """Get staff attendance records (admin only)"""
    verify_admin_token(request)

    try:
        query_parts = [
            "SELECT a.id, a.reg_no, a.name, a.dept, a.class_div, a.timestamp, a.status "
        ]
        query_parts.append("FROM attendance a")
        conditions = []
        params = []

        # Filter by single date
        if date:
            conditions.append("a.timestamp::date = ?")
            params.append(date)
        # Filter by date range
        elif start_date and end_date:
            conditions.append("a.timestamp::date >= ?")
            params.append(start_date)
            conditions.append("a.timestamp::date <= ?")
            params.append(end_date)
        # Filter by start_date only (single day)
        elif start_date:
            conditions.append("a.timestamp::date = ?")
            params.append(start_date)
        # Filter by end_date only (single day)
        elif end_date:
            conditions.append("a.timestamp::date = ?")
            params.append(end_date)

        # Filter by reg_no if provided
        if reg_no:
            conditions.append("a.reg_no = ?")
            params.append(reg_no)

        if conditions:
            query_parts.append("WHERE " + " AND ".join(conditions))

        query_parts.append("ORDER BY a.timestamp DESC")
        query = " ".join(query_parts)

        cursor.execute(query, params)
        attendance_rows = cursor.fetchall()

        # Get all users with their roles
        cursor.execute("SELECT reg_no, role FROM users")
        user_rows = cursor.fetchall()
        user_roles = {row[0]: row[1] for row in user_rows}

        # Filter to only staff and hod attendance
        filtered_attendance = []
        for row in attendance_rows:
            reg_no_val = row[1]
            if user_roles.get(reg_no_val) in ["staff", "hod"]:
                filtered_attendance.append(
                    {
                        "id": row[0],
                        "reg_no": row[1],
                        "name": row[2],
                        "dept": row[3],
                        "class_div": row[4] or "",
                        "timestamp": _ts(row[5]),
                        "status": _format_scan_status(row[6], row[5]),
                        "role": user_roles.get(reg_no_val),
                    }
                )

        result = {"attendance": filtered_attendance, "count": len(filtered_attendance)}

        # Add holiday-aware absent_days/working_days when filtering for a single staff
        if reg_no and start_date and end_date:
            _start, _end = _clamp_to_academic_year(start_date, end_date)
            _end_dt = _cap_end_to_today(datetime.strptime(_end, "%Y-%m-%d"))
            _holidays = _get_holiday_dates_in_range(_start, _end_dt.strftime("%Y-%m-%d"))
            _working_days = 0
            _d = datetime.strptime(_start, "%Y-%m-%d")
            while _d <= _end_dt:
                if _d.weekday() < 5 and _d.strftime("%Y-%m-%d") not in _holidays:
                    _working_days += 1
                _d += timedelta(days=1)
            # Count present days requiring check-out
            _present_count = len(_get_full_attendance_dates(reg_no, _start, _end))
            result["working_days"] = _working_days
            result["present_days"] = _present_count
            result["absent_days"] = max(0, _working_days - _present_count)

        return result
    except Exception as e:
        print(f"Error fetching staff attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch staff attendance")


@app.delete("/admin/attendance/{record_id}")
async def admin_delete_attendance(request: Request, record_id: int):
    """Delete attendance record (admin only)"""
    verify_admin_token(request)

    try:
        cursor.execute("SELECT * FROM attendance WHERE id = ?", (record_id,))
        existing = cursor.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Record not found")

        cursor.execute("DELETE FROM attendance WHERE id = ?", (record_id,))
        conn.commit()

        log_audit_event(
            "ATTENDANCE_DELETED", existing[1], True, f"Record {record_id} deleted"
        )

        return {"message": "Attendance record deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Delete attendance error: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete record")


@app.get("/admin/attendance/stats")
async def admin_get_attendance_stats(
    request: Request, start_date: str = None, end_date: str = None
):
    """Get attendance statistics (admin only).
    Combines face-scan attendance with daily_attendance_status (OD/Leave) for accurate stats.
    """
    verify_admin_token(request)

    try:
        # Default to last 30 days if no dates provided
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        # Clamp to academic year
        start_date, end_date = _clamp_to_academic_year(start_date, end_date)

        # Get stats by department - combine face scans + daily_status
        cursor.execute("""
            SELECT dept, COUNT(DISTINCT reg_no) as staff_count
            FROM users
            WHERE role = 'staff'
            GROUP BY dept
        """)
        dept_staff = cursor.fetchall()
        dept_staff_count = {row[0]: row[1] for row in dept_staff}

        # Get face-scan attendance by dept (enforce check-in/check-out logic and daily status exclusions)
        cursor.execute(
            f"""
            SELECT dept, COUNT(*) as scan_count
            FROM (
                SELECT a.dept, a.reg_no, DATE(a.timestamp) as d
                FROM attendance a
                WHERE DATE(a.timestamp) >= %s AND DATE(a.timestamp) <= %s
                  AND NOT EXISTS (
                      SELECT 1 FROM daily_attendance_status das 
                      WHERE das.reg_no = a.reg_no 
                        AND das.date::date = DATE(a.timestamp) 
                        AND das.status = 'Absent'
                  )
                GROUP BY a.dept, a.reg_no, DATE(a.timestamp)
                HAVING (
                    (DATE(a.timestamp) = CURRENT_DATE AND COUNT(*) > 0)
                    OR
                    (DATE(a.timestamp) < CURRENT_DATE AND COUNT(CASE WHEN a.status = 'check_in' THEN 1 END) > 0 AND COUNT(CASE WHEN a.status = 'check_out' THEN 1 END) > 0)
                )
            ) sub
            GROUP BY dept
        """,
            (start_date, end_date),
        )
        scan_by_dept = {row[0]: row[1] for row in cursor.fetchall()}

        # Get daily_attendance_status by dept (OD, Leave, etc.)
        cursor.execute(
            """
            SELECT dept, 
                   COUNT(CASE WHEN status = 'Present' AND leave_type IN ('od', 'earned', 'casual') THEN 1 END) as present_days,
                   COUNT(CASE WHEN status = 'Leave' THEN 1 END) as leave_days,
                   COUNT(CASE WHEN status = 'Present' AND leave_type IS NULL THEN 1 END) as regular_present
            FROM daily_attendance_status
            WHERE date >= %s AND date <= %s
            GROUP BY dept
        """,
            (start_date, end_date),
        )
        status_by_dept = {}
        for row in cursor.fetchall():
            status_by_dept[row[0]] = {
                "present_days": row[1] or 0,
                "leave_days": row[2] or 0,
                "regular_present": row[3] or 0,
            }

        # Build dept stats
        dept_stats = []
        all_depts = (
            set(dept_staff_count.keys())
            | set(scan_by_dept.keys())
            | set(status_by_dept.keys())
        )
        for dept in all_depts:
            staff_count = dept_staff_count.get(dept, 0)
            scan_count = scan_by_dept.get(dept, 0)
            status_info = status_by_dept.get(dept, {})

            # Calculate working days in range (capped to today, exclude holidays)
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = _cap_end_to_today(datetime.strptime(end_date, "%Y-%m-%d"))
            holiday_dates = _get_holiday_dates_in_range(start_date, end_dt.strftime("%Y-%m-%d"))
            working_days = 0
            d = start_dt
            while d <= end_dt:
                if d.weekday() < 5 and d.strftime("%Y-%m-%d") not in holiday_dates:
                    working_days += 1
                d += timedelta(days=1)

            total_working = working_days * staff_count
            present_days = scan_count + status_info.get("present_days", 0)
            leave_days = status_info.get("leave_days", 0)
            absent_days = max(0, total_working - present_days - leave_days)

            dept_stats.append(
                {
                    "dept": dept,
                    "staff_count": staff_count,
                    "present": present_days,
                    "leave": leave_days,
                    "absent": absent_days,
                    "working_days": working_days,
                    "total_working": total_working,
                }
            )

        # Get stats by date (last 7 days or specified range)
        cursor.execute(
            """
            SELECT timestamp::date as date, COUNT(DISTINCT reg_no) as count 
            FROM attendance 
            WHERE timestamp::date >= %s AND timestamp::date <= %s
            GROUP BY timestamp::date
            ORDER BY date DESC
        """,
            (start_date, end_date),
        )
        scan_by_date = {str(row[0]): row[1] for row in cursor.fetchall()}

        # Get daily status by date
        cursor.execute(
            """
            SELECT date, 
                   COUNT(CASE WHEN status = 'Present' AND leave_type IN ('od', 'earned', 'casual') THEN 1 END) as present_with_tag,
                   COUNT(CASE WHEN status = 'Leave' THEN 1 END) as leave
            FROM daily_attendance_status
            WHERE date >= %s AND date <= %s
            GROUP BY date
        """,
            (start_date, end_date),
        )
        status_by_date = {}
        for row in cursor.fetchall():
            status_by_date[str(row[0])] = {
                "present_with_tag": row[1] or 0,
                "leave": row[2] or 0,
            }

        # Combine date stats
        all_dates = sorted(
            set(scan_by_date.keys()) | set(status_by_date.keys()), reverse=True
        )
        date_stats = []
        for date_str in all_dates[:7]:  # Last 7 dates
            scan_count = scan_by_date.get(date_str, 0)
            status_info = status_by_date.get(date_str, {})
            present_with_tag = status_info.get("present_with_tag", 0)
            leave = status_info.get("leave", 0)

            date_stats.append(
                {
                    "date": date_str,
                    "present": scan_count + present_with_tag,
                    "leave": leave,
                }
            )

        return {
            "by_department": dept_stats,
            "by_date": date_stats,
            "start_date": start_date,
            "end_date": end_date,
        }
    except Exception as e:
        print(f"Error fetching stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch statistics")


@app.get("/admin/attendance/export")
async def admin_export_attendance(
    request: Request, format: str = "json", date: str = None, dept: str = None
):
    """
    Export attendance data (admin only)
    Supports JSON, CSV export formats
    """
    verify_admin_token(request)

    try:
        # Build query based on filters
        query = "SELECT reg_no, name, dept, timestamp FROM attendance WHERE 1=1"
        params = []

        if date:
            query += " AND timestamp::date = ?"
            params.append(date)

        if dept:
            query += " AND dept = ?"
            params.append(dept)

        query += " ORDER BY timestamp DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        attendance_data = [
            {
                "reg_no": row[0],
                "name": row[1],
                "department": row[2],
                "timestamp": _ts(row[3]),
            }
            for row in rows
        ]

        if format == "csv":
            # Generate CSV
            csv_content = "Reg No,Name,Department,Timestamp\n"
            for record in attendance_data:
                csv_content += f"{record['reg_no']},{record['name']},{record['department']},{record['timestamp']}\n"

            return success_response(
                "Attendance exported successfully",
                {"format": "csv", "count": len(attendance_data), "data": csv_content},
            )
        else:
            return success_response(
                "Attendance exported successfully",
                {
                    "format": "json",
                    "count": len(attendance_data),
                    "data": attendance_data,
                },
            )

    except Exception as e:
        print(f"Export attendance error: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to export attendance: {str(e)}"
        )


# New endpoint: Get all HODs for admin selection
@app.get("/admin/attendance/hods")
async def admin_get_hods(request: Request):
    """Get all HODs (admin only)"""
    verify_admin_token(request)

    try:
        cursor.execute(
            "SELECT id, username, reg_no, name, dept FROM users WHERE role = 'hod' ORDER BY dept, name"
        )
        rows = cursor.fetchall()
        return {
            "hods": [
                {
                    "id": row[0],
                    "username": row[1],
                    "reg_no": row[2],
                    "name": row[3],
                    "dept": row[4],
                }
                for row in rows
            ]
        }
    except Exception as e:
        print(f"Error fetching HODs: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch HODs")


# New endpoint: Get all Staff for admin selection
@app.get("/admin/attendance/staff-list")
async def admin_get_staff_list(request: Request, dept: str = None):
    """Get all staff members (admin only)"""
    verify_admin_token(request)

    try:
        if dept:
            cursor.execute(
                "SELECT id, username, reg_no, name, dept FROM users WHERE role = 'staff' AND dept = ? ORDER BY name",
                (dept,),
            )
        else:
            cursor.execute(
                "SELECT id, username, reg_no, name, dept FROM users WHERE role = 'staff' ORDER BY dept, name"
            )
        rows = cursor.fetchall()
        return {
            "staff": [
                {
                    "id": row[0],
                    "username": row[1],
                    "reg_no": row[2],
                    "name": row[3],
                    "dept": row[4],
                }
                for row in rows
            ]
        }
    except Exception as e:
        print(f"Error fetching staff: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch staff")


# New endpoint: Get attendance details for a specific person
@app.get("/admin/attendance/person-details")
async def admin_get_person_attendance(
    request: Request, reg_no: str = None, start_date: str = None, end_date: str = None
):
    """Get attendance details for a specific person with date range (admin only).
    Combines face-scan attendance records with daily_attendance_status (approved leaves/OD).
    """
    verify_admin_token(request)

    if not reg_no:
        raise HTTPException(status_code=400, detail="reg_no is required")

    try:
        # 1) Face-scan attendance records
        scan_query = (
            "SELECT reg_no, name, dept, timestamp FROM attendance WHERE reg_no = %s"
        )
        scan_params = [reg_no]

        if start_date:
            scan_query += " AND timestamp::date >= %s"
            scan_params.append(start_date)

        if end_date:
            scan_query += " AND timestamp::date <= %s"
            scan_params.append(end_date)

        scan_query += " ORDER BY timestamp DESC"

        print(f"[DEBUG] Scan Query: {scan_query}")
        print(f"[DEBUG] Scan Params: {scan_params}")

        cursor.execute(scan_query, scan_params)
        scan_rows = cursor.fetchall()
        print(f"[DEBUG] Found {len(scan_rows)} face-scan attendance records")

        # 2) Daily attendance status records (approved leaves / OD)
        status_query = """
            SELECT reg_no, name, dept, date, status, leave_type, leave_request_id
            FROM daily_attendance_status
            WHERE reg_no = %s
        """
        status_params = [reg_no]

        if start_date:
            status_query += " AND date >= %s"
            status_params.append(start_date)

        if end_date:
            status_query += " AND date <= %s"
            status_params.append(end_date)

        status_query += " ORDER BY date DESC"

        cursor.execute(status_query, status_params)
        status_rows = cursor.fetchall()
        print(f"[DEBUG] Found {len(status_rows)} daily_attendance_status records")

        # Collect dates with full attendance (check-in + check-out)
        scan_dates_set = _get_full_attendance_dates(reg_no, start_date or "", end_date or "")

        # Categorize daily_attendance_status into OD (present) and Leave
        od_dates_set = set()
        leave_dates_set = set()
        for row in status_rows:
            if row[3]:
                date_str = str(row[3])
                if row[4] == "Present" and row[5] in ("od", "earned", "casual"):
                    od_dates_set.add(date_str)
                elif row[4] == "Leave":
                    leave_dates_set.add(date_str)

        # Present = face scans + OD dates
        present_dates = scan_dates_set | od_dates_set
        all_dates = sorted(list(present_dates | leave_dates_set), reverse=True)
        present_count = len(present_dates)
        leave_count = len(leave_dates_set)

        # Calculate working days and absent days (holiday-aware)
        if start_date and end_date:
            _start, _end = _clamp_to_academic_year(start_date, end_date)
            _end_dt = _cap_end_to_today(datetime.strptime(_end, "%Y-%m-%d"))
            _holidays = _get_holiday_dates_in_range(_start, _end_dt.strftime("%Y-%m-%d"))
            _working_days = 0
            _d = datetime.strptime(_start, "%Y-%m-%d")
            while _d <= _end_dt:
                if _d.weekday() < 5 and _d.strftime("%Y-%m-%d") not in _holidays:
                    _working_days += 1
                _d += timedelta(days=1)
            _absent_days = max(0, _working_days - present_count - leave_count)
        else:
            _working_days = None
            _absent_days = None

        # Build combined records list
        attendance_records = []

        # Add face-scan records
        for row in scan_rows:
            attendance_records.append(
                {
                    "reg_no": row[0],
                    "name": row[1],
                    "dept": row[2],
                    "timestamp": _ts(row[3]),
                    "date": str(row[3]).split(" ")[0] if row[3] else None,
                    "source": "face_scan",
                    "status": "Present",
                    "leave_type": None,
                }
            )

        # Add daily status records (only for dates not already in face scans)
        for row in status_rows:
            date_str = str(row[3]) if row[3] else None
            if date_str and date_str not in scan_dates_set:
                attendance_records.append(
                    {
                        "reg_no": row[0],
                        "name": row[1],
                        "dept": row[2],
                        "timestamp": date_str,
                        "date": date_str,
                        "source": "leave",
                        "status": row[4],
                        "leave_type": row[5],
                        "leave_request_id": row[6],
                    }
                )

        # Get person info
        cursor.execute("SELECT name, dept FROM users WHERE reg_no = %s", (reg_no,))
        person = cursor.fetchone()

        return {
            "person": {
                "reg_no": reg_no,
                "name": person[0] if person else "Unknown",
                "dept": person[1] if person else "Unknown",
            },
            "total_records": len(attendance_records),
            "total_days_present": present_count,
            "dates_present": all_dates,
            "working_days": _working_days,
            "absent_days": _absent_days,
            "present_days": present_count,
            "leave_days": leave_count,
            "attendance_records": attendance_records,
        }
    except Exception as e:
        print(f"Error fetching person attendance: {e}")
        raise HTTPException(
            status_code=500, detail="Failed to fetch attendance details"
        )


@app.get("/admin/audit/logs")
async def admin_get_audit_logs(request: Request, limit: int = 100):
    """Get audit logs (admin only)"""
    verify_admin_token(request)

    logs = get_audit_logs(None, limit)
    return {"logs": logs, "count": len(logs)}


# -------------------------------------------------
# DEPARTMENT MANAGEMENT ENDPOINTS
# -------------------------------------------------


@app.get("/admin/departments")
async def admin_get_departments(request: Request):
    """Get all departments (admin only)"""
    verify_admin_token(request)

    try:
        cursor.execute("SELECT id, name, created_at FROM departments ORDER BY name")
        rows = cursor.fetchall()
        return {
            "departments": [
                {"id": row[0], "name": row[1], "created_at": row[2]} for row in rows
            ]
        }
    except Exception as e:
        print(f"Error fetching departments: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch departments")


# New endpoint: Get students by department
@app.get("/admin/analytics/students")
async def admin_get_students_by_dept(request: Request, dept: str = None):
    """Get students by department (admin only)"""
    verify_admin_token(request)

    try:
        if dept:
            cursor.execute(
                "SELECT reg_no, name, dept, class_div FROM students WHERE dept = ? ORDER BY reg_no",
                (dept,),
            )
        else:
            cursor.execute(
                "SELECT reg_no, name, dept, class_div FROM students ORDER BY dept, reg_no"
            )
        rows = cursor.fetchall()
        return {
            "students": [
                {"reg_no": row[0], "name": row[1], "dept": row[2], "class_div": row[3]}
                for row in rows
            ]
        }
    except Exception as e:
        print(f"Error fetching students: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch students")


# New endpoint: Get student attendance details
@app.get("/admin/analytics/student-details")
async def admin_get_student_details(
    request: Request, reg_no: str = None, start_date: str = None, end_date: str = None
):
    """Get attendance details for a specific student (admin only)"""
    verify_admin_token(request)

    if not reg_no:
        raise HTTPException(status_code=400, detail="reg_no is required")

    try:
        # Get student info
        cursor.execute(
            "SELECT name, dept, class_div FROM students WHERE reg_no = ?", (reg_no,)
        )
        student = cursor.fetchone()

        if not student:
            raise HTTPException(status_code=404, detail="Student not found")

        # Build attendance query
        query = "SELECT reg_no, name, dept, timestamp FROM attendance WHERE reg_no = ?"
        params = [reg_no]

        if start_date:
            query += " AND timestamp::date >= ?"
            params.append(start_date)

        if end_date:
            query += " AND timestamp::date <= ?"
            params.append(end_date)

        query += " ORDER BY timestamp DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Get unique dates (days present)
        cursor.execute(
            """
            SELECT DISTINCT timestamp::date as attendance_date 
            FROM attendance 
            WHERE reg_no = ? 
            AND (? IS NULL OR timestamp::date >= ?)
            AND (? IS NULL OR timestamp::date <= ?)
            ORDER BY attendance_date DESC
        """,
            (reg_no, start_date, start_date, end_date, end_date),
        )
        dates = cursor.fetchall()

        return {
            "student": {
                "reg_no": reg_no,
                "name": student[0],
                "dept": student[1],
                "class_div": student[2],
            },
            "total_records": len(rows),
            "total_days_present": len(dates),
            "dates_present": [row[0] for row in dates],
            "attendance_records": [
                {
                    "reg_no": row[0],
                    "name": row[1],
                    "dept": row[2],
                    "timestamp": _ts(row[3]),
                }
                for row in rows
            ],
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching student details: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch student details")


@app.post("/admin/departments")
async def admin_create_department(request: Request):
    """Create a new department (admin only)"""
    verify_admin_token(request)

    try:
        data = await request.json()
        dept_name = data.get("name", "").strip()

        if not dept_name:
            raise HTTPException(status_code=400, detail="Department name is required")

        # Check if department already exists
        cursor.execute("SELECT id FROM departments WHERE name = ?", (dept_name,))
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="Department already exists")

        cursor.execute("INSERT INTO departments (name) VALUES (?)", (dept_name,))
        conn.commit()

        print(f"[ADMIN] Created new department: {dept_name}")

        return {
            "message": "Department created successfully",
            "department": {"name": dept_name},
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error creating department: {e}")
        raise HTTPException(status_code=500, detail="Failed to create department")

@app.post("/admin/departments/bulk-upload")
async def admin_bulk_upload_departments(request: Request, file: UploadFile = File(...)):
    """Bulk upload departments via JSON or Excel (admin only)"""
    verify_admin_token(request)
    import io

    filename = file.filename or ""
    content_type = file.content_type or ""
    
    # Read the file contents
    contents = await file.read()
    
    # Check if the file is actually a users file
    is_user_file = False
    
    # Parse JSON to check keys
    if filename.lower().endswith(".json") or "json" in content_type:
        try:
            data = json.loads(contents.decode("utf-8"))
            first_item = None
            if isinstance(data, list) and data:
                first_item = data[0]
            elif isinstance(data, dict):
                users_list = data.get("users", [])
                if users_list and isinstance(users_list, list):
                    first_item = users_list[0]
                    
            if isinstance(first_item, dict) and ("username" in first_item or "role" in first_item or "password" in first_item):
                is_user_file = True
        except Exception:
            pass
            
    # Parse Excel to check headers
    elif filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls") or "spreadsheet" in content_type or "excel" in content_type:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
            sheet = wb.active
            first_row = next(sheet.iter_rows(values_only=True), None)
            if first_row:
                headers = [str(c).strip().lower() for c in first_row if c is not None]
                if "username" in headers or "user_name" in headers or "role" in headers:
                    is_user_file = True
        except Exception:
            pass

    if is_user_file:
        new_file = UploadFile(file=io.BytesIO(contents), filename=filename, headers=file.headers)
        return await admin_bulk_upload_users(request, new_file)

    dept_names = []
    
    # Case 1: JSON file
    if filename.lower().endswith(".json") or "json" in content_type:
        try:
            data = json.loads(contents.decode("utf-8"))
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, str):
                        dept_names.append(item.strip())
                    elif isinstance(item, dict):
                        if "name" in item and ("dept" in item or "department" in item):
                            name = item.get("dept") or item.get("department") or item.get("department_name")
                        else:
                            name = item.get("name") or item.get("department") or item.get("dept") or item.get("department_name")
                            
                        if name and isinstance(name, str):
                            dept_names.append(name.strip())
            else:
                raise HTTPException(status_code=400, detail="JSON must be an array of names or objects")
        except Exception as e:
            if isinstance(e, HTTPException):
                raise
            raise HTTPException(status_code=400, detail=f"Failed to parse JSON file: {e}")
            
    # Case 2: Excel file
    elif filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls") or "spreadsheet" in content_type or "excel" in content_type:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Excel file is empty")
                
            header_row = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            name_idx = -1
            
            # Map column indices
            for idx, h in enumerate(header_row):
                if h in ["name", "department", "dept", "department name", "department_name"]:
                    name_idx = idx
                    break
                    
            if name_idx == -1:
                name_idx = 0
                
            for row in rows[1:]:
                if name_idx < len(row):
                    cell_val = row[name_idx]
                    if cell_val is not None:
                        val_str = str(cell_val).strip()
                        if val_str:
                            dept_names.append(val_str)
                            
        except Exception as e:
            if isinstance(e, HTTPException):
                raise
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")
    else:
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload a .json, .xlsx, or .xls file.")

    # Deduplicate within the uploaded file
    dept_names = list(dict.fromkeys(dept_names))
    
    # Process insertion
    inserted = []
    skipped_duplicates = []
    errors = []
    
    for dept_name in dept_names:
        if not dept_name:
            continue
        try:
            cursor.execute("SELECT id FROM departments WHERE name = ?", (dept_name,))
            if cursor.fetchone():
                skipped_duplicates.append(dept_name)
                continue
                
            cursor.execute("INSERT INTO departments (name) VALUES (?)", (dept_name,))
            inserted.append(dept_name)
        except Exception as e:
            errors.append(f"Failed to insert '{dept_name}': {e}")
            
    conn.commit()
    
    return {
        "success": True,
        "message": f"Successfully imported {len(inserted)} departments.",
        "inserted": inserted,
        "skipped_duplicates": skipped_duplicates,
        "errors": errors
    }


@app.delete("/admin/departments/{dept_name}")
async def admin_delete_department(request: Request, dept_name: str):
    """Delete a department (admin only)"""
    verify_admin_token(request)

    try:
        # Get users belonging to this department
        cursor.execute("SELECT reg_no FROM users WHERE dept = ?", (dept_name,))
        users = cursor.fetchall()
        for u in users:
            delete_user_data_by_reg_no(u[0])

        cursor.execute("DELETE FROM users WHERE dept = ?", (dept_name,))

        # Get other staff belonging to this department
        cursor.execute("SELECT reg_no FROM other_staff WHERE dept = ?", (dept_name,))
        other_staff_members = cursor.fetchall()
        for os_member in other_staff_members:
            delete_user_data_by_reg_no(os_member[0])

        cursor.execute("DELETE FROM other_staff WHERE dept = ?", (dept_name,))

        # Delete department
        cursor.execute("DELETE FROM departments WHERE name = ?", (dept_name,))
        conn.commit()

        print(f"[ADMIN] Deleted department and its users: {dept_name}")

        return {"message": f"Department '{dept_name}' and all associated users/data deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error deleting department: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete department")


@app.put("/admin/departments/{old_dept_name}")
async def admin_update_department(request: Request, old_dept_name: str):
    """Update a department name (admin only)"""
    verify_admin_token(request)

    # URL decode the department name (handle special characters)
    old_dept_name = urllib.parse.unquote(old_dept_name).strip()

    try:
        data = await request.json()
        new_name = data.get("name", "").strip()

        if not new_name:
            raise HTTPException(status_code=400, detail="Department name is required")

        # Check if old department exists
        cursor.execute("SELECT id FROM departments WHERE name = ?", (old_dept_name,))
        old_dept = cursor.fetchone()
        if not old_dept:
            raise HTTPException(status_code=404, detail="Department not found")

        # Check if new name already exists
        cursor.execute("SELECT id FROM departments WHERE name = ?", (new_name,))
        if cursor.fetchone():
            raise HTTPException(
                status_code=400, detail="Department name already exists"
            )

        # Update department name - ensure we only update the specific department
        cursor.execute(
            "UPDATE departments SET name = ? WHERE name = ? AND id = ?",
            (
                new_name,
                old_dept_name,
                old_dept[0],
            ),
        )

        # Also update the department name in users table
        cursor.execute(
            "UPDATE users SET dept = ? WHERE dept = ?",
            (
                new_name,
                old_dept_name,
            ),
        )

        conn.commit()

        print(f"[ADMIN] Updated department: {old_dept_name} -> {new_name}")

        return {
            "message": "Department updated successfully",
            "department": {"name": new_name},
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating department: {e}")
        raise HTTPException(status_code=500, detail="Failed to update department")


# -------------------------------------------------
# HOD PANEL ENDPOINTS
# -------------------------------------------------


def verify_hod_token(request: Request) -> dict:
    """Verify HOD authentication token"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authorization header"
        )

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")

        user = get_user_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check if user is suspended
        if is_user_suspended(username, is_other_staff=False):
            raise HTTPException(status_code=403, detail="Account suspended. Access denied.")

        # Check if user is HOD
        if user[6] != "hod":
            raise HTTPException(status_code=403, detail="HOD access required")

        return {
            "id": user[0],
            "username": user[1],
            "reg_no": user[3],
            "name": user[4],
            "dept": user[5],
            "role": user[6],
        }
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid token")


@app.post("/hod/login")
async def hod_login(request: Request):
    """HOD login endpoint"""
    try:
        data = await request.json()
        username = data.get("username")
        password = data.get("password")

        if not username or not password:
            raise HTTPException(status_code=400, detail="Missing credentials")

        user = get_user_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check role
        if user[6] != "hod":
            raise HTTPException(status_code=403, detail="HOD access required")

        # Create token
        import base64

        token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode(
            "utf-8"
        )

        user_dept = user[5]
        print(
            f"[HOD_LOGIN] HOD '{user[4]}' (username: {username}) logging in with department: {user_dept}"
        )

        return {
            "message": "HOD login successful",
            "token": token,
            "user": {
                "id": user[0],
                "username": user[1],
                "regNo": user[3],
                "name": user[4],
                "dept": user[5],
                "role": user[6],
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"HOD login error: {e}")
        raise HTTPException(status_code=500, detail="Login failed")


@app.get("/hod/dashboard")
async def hod_dashboard(request: Request):
    """Get HOD dashboard statistics (department-specific)"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    # Get counts for this department only
    cursor.execute("SELECT COUNT(*) FROM attendance WHERE dept = ?", (dept,))
    dept_attendance = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(*) FROM attendance WHERE dept = ? AND timestamp::date = CURRENT_DATE",
        (dept,),
    )
    today_attendance = cursor.fetchone()[0]

    cursor.execute(
        "SELECT COUNT(*) FROM users WHERE dept = ? AND role = 'staff'", (dept,)
    )
    dept_staff = cursor.fetchone()[0]

    # Get recent attendance for this department (face scan + OD)
    cursor.execute(
        """
        SELECT id, reg_no, name, dept, class_div, timestamp, 'face_scan' as source, status
        FROM attendance 
        WHERE dept = ? 
        ORDER BY id DESC 
        LIMIT 8
    """,
        (dept,),
    )
    face_scan_attendance = cursor.fetchall()

    # Get OD records for today with approval date
    cursor.execute(
        """
        SELECT das.id, das.reg_no, das.name, das.dept, das.date as timestamp, 'od' as source, 
               das.leave_type, lr.processed_date
        FROM daily_attendance_status das
        LEFT JOIN leave_requests lr ON das.leave_request_id = lr.id
        WHERE das.dept = %s AND das.date::date = CURRENT_DATE AND das.status = 'Present' 
        AND das.leave_type IN ('od', 'earned', 'casual')
        ORDER BY das.date DESC
        LIMIT 5
    """,
        (dept,),
    )
    od_records = cursor.fetchall()

    # Combine face scan + OD
    combined = list(face_scan_attendance) + list(od_records)
    combined.sort(key=lambda x: str(x[5]) if len(x) > 5 and x[5] else "", reverse=True)
    recent_attendance = combined[:10]

    # Get today's OD count for department (separate counts)
    cursor.execute(
        """
        SELECT leave_type, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE dept = %s AND date::date = CURRENT_DATE AND status = 'Present' 
        AND leave_type IN ('od', 'earned', 'casual')
        GROUP BY leave_type
    """,
        (dept,),
    )
    od_breakdown = cursor.fetchall()
    od_count = sum(row[1] for row in od_breakdown)
    earned_count = sum(row[1] for row in od_breakdown if row[0] == "earned")
    casual_count = sum(row[1] for row in od_breakdown if row[0] == "casual")

    # Build recent_attendance list with approval_date
    recent_attendance_list = []
    for row in recent_attendance:
        if len(row) >= 8:
            # OD record with approval date
            recent_attendance_list.append(
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "timestamp": str(row[4]) if row[4] else "",
                    "source": row[5],
                    "leave_type": row[6],
                    "approval_date": _ts(row[7]) if len(row) > 7 and row[7] else None,
                }
            )
        else:
            # Face scan record
            raw_status = row[7] if len(row) > 7 else None
            recent_attendance_list.append(
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "class_div": row[4] or "",
                    "timestamp": _ts(row[5])
                    if len(row) > 5 and row[5]
                    else str(row[5])
                    if len(row) > 5
                    else "",
                    "source": row[6] if len(row) > 6 else "face_scan",
                    "leave_type": None,
                    "status": _format_scan_status(raw_status, row[5] if len(row) > 5 else None),
                    "punch_type": raw_status or "check_in",
                }
            )

    # Calculate present and absent for department
    # Present = face scan + OD, Absent = staff without any attendance/OD today
    total_staff_in_dept = dept_staff
    today_present_count = today_attendance + od_count
    today_absent_count = max(0, total_staff_in_dept - today_present_count)

    # --- Pie chart: breakdown from daily_attendance_status for dept ---
    cursor.execute(
        """
        SELECT status, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE dept = %s AND date::date = CURRENT_DATE
        GROUP BY status
    """,
        (dept,),
    )
    das_breakdown = {row[0]: row[1] for row in cursor.fetchall()}
    pie_full_day  = das_breakdown.get("Present", 0)
    pie_half_day  = das_breakdown.get("Half Day", 0)
    pie_absent    = das_breakdown.get("Absent", 0)
    pie_leave     = das_breakdown.get("Leave", 0)
    pie_holiday   = das_breakdown.get("Holiday", 0)

    return JSONResponse(
        content={
            "stats": {
                "department": dept,
                "total_attendance": dept_attendance,
                "today_attendance": today_attendance + od_count,
                "today_face_scan": today_attendance,
                "today_od": od_count,
                "today_earned": earned_count,
                "today_casual": casual_count,
                "total_staff": dept_staff,
                "today_present": today_present_count,
                "today_absent": today_absent_count,
                # Pie chart breakdown fields
                "today_full_day": pie_full_day,
                "today_half_day": pie_half_day,
                "today_leave": pie_leave,
                "today_holiday": pie_holiday,
            },
            "recent_attendance": recent_attendance_list,
            "hod_user": hod_user,
        },
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.get("/hod/attendance")
async def hod_get_attendance(request: Request, date: str = None):
    """Get attendance records for HOD's department only"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    try:
        # Use explicit column names to avoid IndexError
        if date:
            cursor.execute(
                """
                SELECT id, reg_no, name, dept, class_div, timestamp, status
                FROM attendance 
                WHERE dept = ? AND timestamp::date = ? 
                ORDER BY id DESC
            """,
                (dept, date),
            )
        else:
            cursor.execute(
                """
                SELECT id, reg_no, name, dept, class_div, timestamp, status
                FROM attendance 
                WHERE dept = ? 
                ORDER BY id DESC
            """,
                (dept,),
            )

        rows = cursor.fetchall()
        return {
            "attendance": [
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "class_div": row[4] or "",
                    "timestamp": _ts(row[5]),
                    "status": _format_scan_status(row[6], row[5]),
                }
                for row in rows
            ],
            "count": len(rows),
        }
    except Exception as e:
        print(f"Error fetching attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch attendance")


@app.get("/hod/attendance/staff")
async def hod_get_staff_attendance(request: Request, date: str = None):
    """Get staff attendance records for HOD's department only"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    try:
        # Get all attendance for department, then filter by role from users table in memory
        if date:
            cursor.execute(
                """
                SELECT a.id, a.reg_no, a.name, a.dept, a.timestamp, a.status 
                FROM attendance a
                WHERE a.dept = ? AND a.timestamp::date = ?
                ORDER BY a.id DESC
            """,
                (dept, date),
            )
        else:
            cursor.execute(
                """
                SELECT a.id, a.reg_no, a.name, a.dept, a.timestamp, a.status 
                FROM attendance a
                WHERE a.dept = ?
                ORDER BY a.id DESC
            """,
                (dept,),
            )

        attendance_rows = cursor.fetchall()

        # Get all users in this department with their roles
        cursor.execute("SELECT reg_no, role FROM users WHERE dept = ?", (dept,))
        user_rows = cursor.fetchall()
        user_roles = {row[0]: row[1] for row in user_rows}

        # Filter to only staff and hod attendance
        filtered_attendance = []
        for row in attendance_rows:
            reg_no = row[1]
            if user_roles.get(reg_no) in ["staff", "hod"]:
                filtered_attendance.append(
                    {
                        "id": row[0],
                        "reg_no": row[1],
                        "name": row[2],
                        "dept": row[3],
                        "timestamp": _ts(row[4]),
                        "status": _format_scan_status(row[5], row[4]),
                        "role": user_roles.get(reg_no),
                    }
                )

        return {"attendance": filtered_attendance, "count": len(filtered_attendance)}
    except Exception as e:
        print(f"Error fetching staff attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch staff attendance")


@app.get("/hod/staff")
async def hod_get_staff(request: Request):
    """Get staff members in HOD's department with face status"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]
    hod_name = hod_user["name"]

    print(f"[HOD_STAFF] HOD '{hod_name}' (dept: {dept}) requesting staff list")

    try:
        cursor.execute(
            "SELECT id, username, reg_no, name, dept, role, created_at, embedding, can_reregister FROM users WHERE dept = ? AND role = 'staff'",
            (dept,),
        )
        rows = cursor.fetchall()

        print(f"[HOD_STAFF] Found {len(rows)} staff members for department '{dept}'")
        for row in rows:
            print(f"[HOD_STAFF]   - {row[3]} ({row[2]}) - Dept: {row[4]}")

        return {
            "staff": [
                {
                    "id": row[0],
                    "username": row[1],
                    "reg_no": row[2],
                    "name": row[3],
                    "dept": row[4],
                    "role": row[5],
                    "created_at": row[6],
                    "face_registered": row[7] is not None,
                    "can_reregister": row[8] == 1 if row[8] is not None else False,
                }
                for row in rows
            ],
            "count": len(rows),
            "filter_applied": f"dept={dept}",
        }
    except Exception as e:
        print(f"Error fetching staff: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch staff")


@app.get("/hod/attendance/stats")
async def hod_get_attendance_stats(request: Request):
    """Get attendance statistics for HOD's department"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    try:
        # Get stats by date (last 7 days) for this department
        cursor.execute(
            """
            SELECT timestamp::date as date, COUNT(*) as count 
            FROM attendance 
            WHERE dept = ? AND timestamp >= CURRENT_DATE - INTERVAL '7 days'
            GROUP BY timestamp::date
            ORDER BY date DESC
        """,
            (dept,),
        )
        date_stats = cursor.fetchall()

        return {
            "department": dept,
            "by_date": [{"date": row[0], "count": row[1]} for row in date_stats],
        }
    except Exception as e:
        print(f"Error fetching stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch statistics")


@app.get("/hod/attendance/range-stats")
async def hod_get_attendance_range_stats(
    request: Request, start_date: str = None, end_date: str = None
):
    """Get department-wide attendance statistics for a date range (HOD only).
    Combines face-scan attendance + daily_attendance_status (approved leaves/OD).
    Returns per-staff breakdown + overall summary.
    """
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    if not start_date or not end_date:
        raise HTTPException(
            status_code=400, detail="start_date and end_date are required (YYYY-MM-DD)"
        )

    try:
        # 1) Get all staff in department
        cursor.execute(
            "SELECT reg_no, name, role FROM users WHERE role IN ('staff', 'hod') AND dept = %s ORDER BY CASE WHEN role = 'hod' THEN 0 ELSE 1 END, name",
            (dept,),
        )
        staff_rows = cursor.fetchall()
        staff_list = [{"reg_no": r[0], "name": r[1], "role": r[2]} for r in staff_rows]

        # 2) Face-scan attendance per staff in date range
        cursor.execute(
            """
            SELECT reg_no, COUNT(*) as days_present
            FROM (
                SELECT reg_no, DATE(timestamp) as d
                FROM attendance a
                WHERE dept = %s AND DATE(timestamp) >= %s AND DATE(timestamp) <= %s
                  AND NOT EXISTS (
                      SELECT 1 FROM daily_attendance_status das 
                      WHERE das.reg_no = a.reg_no 
                        AND das.date::date = DATE(a.timestamp) 
                        AND das.status = 'Absent'
                  )
                GROUP BY reg_no, DATE(timestamp)
                HAVING (
                    (DATE(timestamp) = CURRENT_DATE AND COUNT(*) > 0)
                    OR
                    (DATE(timestamp) < CURRENT_DATE AND COUNT(CASE WHEN status = 'check_in' THEN 1 END) > 0 AND COUNT(CASE WHEN status = 'check_out' THEN 1 END) > 0)
                )
            ) sub
            GROUP BY reg_no
            """,
            (dept, start_date, end_date),
        )
        scan_data = {r[0]: r[1] for r in cursor.fetchall()}

        # 3) Daily attendance status (leaves/OD) per staff in date range
        # Group by both status AND leave_type to properly separate OD from regular Present
        cursor.execute(
            """
            SELECT reg_no, status, COALESCE(leave_type, ''), COUNT(*) as cnt
            FROM daily_attendance_status
            WHERE dept = %s AND date >= %s AND date <= %s
            GROUP BY reg_no, status, COALESCE(leave_type, '')
            """,
            (dept, start_date, end_date),
        )
        status_data: dict = {}
        for r in cursor.fetchall():
            reg, status, leave_type, cnt = r
            if reg not in status_data:
                status_data[reg] = {}

            # For OD: status="Present" with leave_type="od"
            # We store it using status_leave_type as key
            if leave_type in ("od", "earned", "casual"):
                key = leave_type
            else:
                key = status

            status_data[reg][key] = cnt

        # 4) Count total working days in range (capped to today, exclude holidays)
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = _cap_end_to_today(datetime.strptime(end_date, "%Y-%m-%d"))
        holiday_dates = _get_holiday_dates_in_range(start_date, end_dt.strftime("%Y-%m-%d"))
        working_days = 0
        d = start_dt
        while d <= end_dt:
            if d.weekday() < 5 and d.strftime("%Y-%m-%d") not in holiday_dates:
                working_days += 1
            d += timedelta(days=1)

        # 5) Build per-staff stats
        staff_stats = []
        total_present = 0
        total_leave = 0
        total_absent = 0
        total_od = 0

        for s in staff_list:
            reg = s["reg_no"]
            scan_days = scan_data.get(reg, 0)
            status_info = status_data.get(reg, {})

            leave_days = status_info.get("Leave", 0)
            od_days = status_info.get("od", 0)
            earned_days = status_info.get("earned", 0)
            casual_days = status_info.get("casual", 0)

            # Total days marked present (face scan + OD/earned/casual)
            present_days = scan_days + od_days + earned_days + casual_days
            # Total absent = working days - present - leave
            absent_days = max(0, working_days - present_days - leave_days)

            # Attendance percentage
            if working_days > 0:
                pct = round((present_days / working_days) * 100, 1)
            else:
                pct = 0

            staff_stats.append(
                {
                    "reg_no": reg,
                    "name": s["name"],
                    "working_days": working_days,
                    "present": present_days,
                    "leave": leave_days,
                    "od": od_days,
                    "earned_leave": earned_days,
                    "casual_leave": casual_days,
                    "absent": absent_days,
                    "attendance_pct": pct,
                }
            )

            total_present += present_days
            total_leave += leave_days
            total_absent += absent_days
            total_od += od_days

        total_working = working_days * len(staff_list)
        overall_pct = (
            round((total_present / total_working * 100), 1) if total_working > 0 else 0
        )

        return {
            "success": True,
            "department": dept,
            "start_date": start_date,
            "end_date": end_date,
            "working_days": working_days,
            "total_staff": len(staff_list),
            "summary": {
                "total_working_days": total_working,
                "total_present": total_present,
                "total_leave": total_leave,
                "total_absent": total_absent,
                "total_od": total_od,
                "overall_attendance_pct": overall_pct,
            },
            "staff": staff_stats,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching range stats: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch range statistics")


# New endpoint: Get staff list for HOD
@app.get("/hod/attendance/staff-list")
async def hod_get_staff_list(request: Request):
    """Get staff members in HOD's department"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    try:
        cursor.execute(
            """
            SELECT id, username, name, reg_no, dept, role
            FROM users
            WHERE role IN ('staff', 'hod') AND dept = ?
            ORDER BY CASE WHEN role = 'hod' THEN 0 ELSE 1 END, name
            """,
            (dept,),
        )
        rows = cursor.fetchall()
        return {
            "staff": [
                {
                    "id": row[0],
                    "username": row[1],
                    "name": row[2],
                    "reg_no": row[3],
                    "dept": row[4],
                    "role": row[5],
                }
                for row in rows
            ]
        }
    except Exception as e:
        print(f"Error fetching staff: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch staff")


# HOD Analytics endpoint
@app.get("/hod/analytics/summary")
async def hod_get_analytics_summary(
    request: Request, start_date: str = None, end_date: str = None
):
    """Get analytics summary for HOD's department.
    Returns department-wide attendance statistics with per-staff breakdown.
    """
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    # Default to academic year if no dates
    now = datetime.now()
    acad_start, acad_end = _get_academic_date_range()
    if not start_date:
        start_date = acad_start or now.strftime("%Y-%m-01")
    if not end_date:
        end_date = acad_end or now.strftime("%Y-%m-%d")
    start_date, end_date = _clamp_to_academic_year(start_date, end_date)

    try:
        # Get all staff in department
        cursor.execute(
            "SELECT reg_no, name, role FROM users WHERE role IN ('staff', 'hod') AND dept = %s ORDER BY CASE WHEN role = 'hod' THEN 0 ELSE 1 END, name",
            (dept,),
        )
        staff_rows = cursor.fetchall()
        staff_list = [{"reg_no": r[0], "name": r[1], "role": r[2]} for r in staff_rows]

        # Get face-scan attendance per staff in date range
        cursor.execute(
            """
            SELECT reg_no, COUNT(*) as days_present
            FROM (
                SELECT reg_no, DATE(timestamp) as d
                FROM attendance a
                WHERE dept = %s AND DATE(timestamp) >= %s AND DATE(timestamp) <= %s
                  AND NOT EXISTS (
                      SELECT 1 FROM daily_attendance_status das 
                      WHERE das.reg_no = a.reg_no 
                        AND das.date::date = DATE(a.timestamp) 
                        AND das.status = 'Absent'
                  )
                GROUP BY reg_no, DATE(timestamp)
                HAVING (
                    (DATE(timestamp) = CURRENT_DATE AND COUNT(*) > 0)
                    OR
                    (DATE(timestamp) < CURRENT_DATE AND COUNT(CASE WHEN status = 'check_in' THEN 1 END) > 0 AND COUNT(CASE WHEN status = 'check_out' THEN 1 END) > 0)
                )
            ) sub
            GROUP BY reg_no
            """,
            (dept, start_date, end_date),
        )
        scan_data = {r[0]: r[1] for r in cursor.fetchall()}

        # Get daily_attendance_status (OD/Leave) per staff
        cursor.execute(
            """
            SELECT reg_no, status, COALESCE(leave_type, ''), COUNT(*) as cnt
            FROM daily_attendance_status
            WHERE dept = %s AND date::date >= %s AND date::date <= %s
            GROUP BY reg_no, status, COALESCE(leave_type, '')
            """,
            (dept, start_date, end_date),
        )
        status_data = {}
        for r in cursor.fetchall():
            reg, status, leave_type, cnt = r
            if reg not in status_data:
                status_data[reg] = {}
            if leave_type in ("od", "earned", "casual"):
                status_data[reg][leave_type] = cnt
            else:
                status_data[reg][status] = cnt

        # Count working days in range (capped to today, exclude holidays)
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = _cap_end_to_today(datetime.strptime(end_date, "%Y-%m-%d"))
        holiday_dates = _get_holiday_dates_in_range(start_date, end_dt.strftime("%Y-%m-%d"))
        working_days = 0
        d = start_dt
        while d <= end_dt:
            if d.weekday() < 5 and d.strftime("%Y-%m-%d") not in holiday_dates:
                working_days += 1
            d += timedelta(days=1)

        # Build per-staff stats
        staff_stats = []
        total_present = 0
        total_leave = 0
        total_absent = 0
        total_od = 0

        for s in staff_list:
            reg = s["reg_no"]
            scan_days = scan_data.get(reg, 0)
            status_info = status_data.get(reg, {})

            leave_days = status_info.get("Leave", 0)
            od_days = status_info.get("od", 0)
            earned_days = status_info.get("earned", 0)
            casual_days = status_info.get("casual", 0)

            present_days = scan_days + od_days + earned_days + casual_days
            absent_days = max(0, working_days - present_days - leave_days)

            pct = (
                round((present_days / working_days) * 100, 1) if working_days > 0 else 0
            )

            staff_stats.append(
                {
                    "reg_no": reg,
                    "name": s["name"],
                    "working_days": working_days,
                    "present": present_days,
                    "leave": leave_days,
                    "od": od_days,
                    "earned_leave": earned_days,
                    "casual_leave": casual_days,
                    "absent": absent_days,
                    "attendance_pct": pct,
                }
            )

            total_present += present_days
            total_leave += leave_days
            total_absent += absent_days
            total_od += od_days

        total_working = working_days * len(staff_list)
        overall_pct = (
            round((total_present / total_working * 100), 1) if total_working > 0 else 0
        )

        return {
            "success": True,
            "department": dept,
            "start_date": start_date,
            "end_date": end_date,
            "working_days": working_days,
            "total_staff": len(staff_list),
            "summary": {
                "total_working_days": total_working,
                "total_present": total_present,
                "total_leave": total_leave,
                "total_absent": total_absent,
                "total_od": total_od,
                "overall_attendance_pct": overall_pct,
            },
            "staff": staff_stats,
        }
    except Exception as e:
        print(f"Error fetching HOD analytics: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch analytics")


# New endpoint: Get attendance details for a specific staff
@app.get("/hod/attendance/staff-details")
async def hod_get_staff_attendance(
    request: Request, reg_no: str = None, start_date: str = None, end_date: str = None
):
    """Get attendance details for a specific staff in HOD's department"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    if not reg_no:
        raise HTTPException(status_code=400, detail="reg_no is required")

    try:
        # First verify the staff belongs to this department
        cursor.execute(
            "SELECT name, dept FROM users WHERE reg_no = ? AND role IN ('staff', 'hod')",
            (reg_no,),
        )
        person = cursor.fetchone()

        if not person:
            raise HTTPException(status_code=404, detail="Staff member not found")

        if person[1] != dept:
            raise HTTPException(
                status_code=403, detail="Staff member not in your department"
            )

        # Build query
        query = "SELECT reg_no, name, dept, timestamp FROM attendance WHERE reg_no = ?"
        params = [reg_no]

        if start_date:
            query += " AND timestamp::date >= ?"
            params.append(start_date)

        if end_date:
            query += " AND timestamp::date <= ?"
            params.append(end_date)

        query += " ORDER BY timestamp DESC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Get unique dates (days present)
        cursor.execute(
            """
            SELECT DISTINCT timestamp::date as attendance_date 
            FROM attendance 
            WHERE reg_no = ? 
            AND (? IS NULL OR timestamp::date >= ?)
            AND (? IS NULL OR timestamp::date <= ?)
            ORDER BY attendance_date DESC
        """,
            (reg_no, start_date, start_date, end_date, end_date),
        )
        dates = cursor.fetchall()

        return {
            "person": {"reg_no": reg_no, "name": person[0], "dept": person[1]},
            "total_records": len(rows),
            "total_days_present": len(dates),
            "dates_present": [row[0] for row in dates],
            "attendance_records": [
                {
                    "reg_no": row[0],
                    "name": row[1],
                    "dept": row[2],
                    "timestamp": _ts(row[3]),
                }
                for row in rows
            ],
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching staff attendance: {e}")
        raise HTTPException(
            status_code=500, detail="Failed to fetch attendance details"
        )


@app.post("/hod/staff/create")
async def hod_create_staff(request: Request):
    """Create new staff member (HOD only - for their department)"""
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    try:
        data = await request.json()
        username = data.get("username")
        password = data.get("password")
        reg_no = data.get("reg_no")
        name = data.get("name")

        if not all([username, password, name]):
            raise HTTPException(status_code=400, detail="Missing required fields")

        # Auto-generate reg_no if not provided
        if reg_no is None or reg_no.strip() == "":
            # Generate staff ID: STAFF_XXXX where XXXX is sequential
            cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'staff'")
            count = cursor.fetchone()[0]
            reg_no = f"STAFF_{str(count + 1).zfill(4)}"

        # Check if user already exists
        if get_user_by_username(username):
            raise HTTPException(status_code=400, detail="Username already exists")

        if get_user_by_reg_no(reg_no):
            # If reg_no exists, generate a new one
            count = 0
            while True:
                count += 1
                reg_no = f"STAFF_{str(count).zfill(4)}"
                if not get_user_by_reg_no(reg_no):
                    break

        password_hash = hash_password(password)

        # Insert staff with HOD's department
        cursor.execute(
            """
            INSERT INTO users (username, password_hash, reg_no, name, dept, role, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
            (
                username,
                password_hash,
                reg_no,
                name,
                dept,
                "staff",
                hod_user["username"],
            ),
        )
        conn.commit()

        log_audit_event(
            "STAFF_CREATED_BY_HOD",
            reg_no,
            True,
            f"Staff {username} created by HOD of {dept}",
        )

        return {
            "message": "Staff created successfully",
            "user": {
                "username": username,
                "reg_no": reg_no,
                "name": name,
                "dept": dept,
                "role": "staff",
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Create staff error: {e}")
        raise HTTPException(status_code=500, detail="Failed to create staff")


@app.delete("/hod/staff/{staff_id}")
async def hod_delete_staff(request: Request, staff_id: int):
    """
    Delete staff member (HOD only - from their department)
    Requires face verification to ensure it's the authorized HOD
    """
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    # Get the HOD's own embedding for verification
    cursor.execute(
        "SELECT embedding FROM users WHERE reg_no = ?", (hod_user["reg_no"],)
    )
    hod_embedding = cursor.fetchone()

    if hod_embedding is None or hod_embedding[0] is None:
        raise HTTPException(
            status_code=403,
            detail="HOD face not registered. Please register your face first.",
        )

    try:
        # Parse multipart form data
        content_type = request.headers.get("content-type", "").lower()

        if "multipart/form-data" in content_type:
            form_data = await request.form()

            # Face verification is required for HOD delete operations
            face_image = form_data.get("face_image")
            if face_image:
                img_bytes = await face_image.read()
                img = preprocess_image_data(img_bytes)

                face = extract_face(img)
                if face is None:
                    save_debug_image(img, f"hod_verify_fail_{hod_user['reg_no']}")
                    raise HTTPException(
                        status_code=400,
                        detail="Could not verify HOD face. Please ensure your face is visible.",
                    )

                query_embedding = face.embedding.astype(np.float32)
                db_embedding = np.frombuffer(hod_embedding[0], dtype=np.float32)

                # Verify HOD's face matches
                combined_sim, cosine_sim, euclidean_dist, manhattan_dist, corr = (
                    calculate_similarity(query_embedding, db_embedding)
                )

                print(
                    f"DEBUG: HOD face verification - Cosine: {cosine_sim:.4f}, Euclidean: {euclidean_dist:.4f}"
                )

                # Use lenient verification for HOD's own face
                if cosine_sim < 0.30 and euclidean_dist > 2.5:
                    raise HTTPException(
                        status_code=401,
                        detail="HOD face verification failed. Please try again.",
                    )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Face verification parse error: {e}")

    try:
        cursor.execute(
            "SELECT id, username, password_hash, reg_no, name, dept, role, is_active, face_embedding, current_device_id, out_permission_enabled, out_permission_expiry, created_by, created_at, updated_at, embedding, can_reregister, suspended FROM users WHERE id = ? AND dept = ? AND role = 'staff'",
            (staff_id, dept),
        )
        existing = cursor.fetchone()
        if not existing:
            raise HTTPException(
                status_code=404, detail="Staff not found in your department"
            )

        reg_no = existing[3]
        delete_user_data_by_reg_no(reg_no)

        cursor.execute("DELETE FROM users WHERE id = ?", (staff_id,))
        conn.commit()

        log_audit_event(
            "STAFF_DELETED_BY_HOD",
            existing[3],
            True,
            f"Staff {existing[1]} deleted by HOD of {dept}",
        )

        return {"message": "Staff deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Delete staff error: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete staff")


@app.put("/hod/staff/{staff_id}")
async def hod_update_staff(request: Request, staff_id: int):
    """
    Update staff member details (HOD only - for their department)
    Requires face verification to ensure it's the authorized HOD
    """
    hod_user = verify_hod_token(request)
    dept = hod_user["dept"]

    # Get the HOD's own embedding for verification
    cursor.execute(
        "SELECT embedding FROM users WHERE reg_no = ?", (hod_user["reg_no"],)
    )
    hod_embedding = cursor.fetchone()

    if hod_embedding is None or hod_embedding[0] is None:
        raise HTTPException(
            status_code=403,
            detail="HOD face not registered. Please register your face first.",
        )

    try:
        # Parse multipart form data
        content_type = request.headers.get("content-type", "").lower()

        name = None
        username = None
        password = None
        verify_face = True  # Default: require face verification

        if "multipart/form-data" in content_type:
            form_data = await request.form()
            name = form_data.get("name")
            username = form_data.get("username")
            password = form_data.get("password")
            verify_face = form_data.get("verify_face", "true").lower() == "true"

            # Get face image if provided and verification is needed
            face_image = form_data.get("face_image")
            if face_image and verify_face:
                img_bytes = await face_image.read()
                img = preprocess_image_data(img_bytes)

                face = extract_face(img)
                if face is None:
                    save_debug_image(img, f"hod_verify_fail_{hod_user['reg_no']}")
                    raise HTTPException(
                        status_code=400,
                        detail="Could not verify HOD face. Please ensure your face is visible.",
                    )

                query_embedding = face.embedding.astype(np.float32)
                db_embedding = np.frombuffer(hod_embedding[0], dtype=np.float32)

                # Verify HOD's face matches
                combined_sim, cosine_sim, euclidean_dist, manhattan_dist, corr = (
                    calculate_similarity(query_embedding, db_embedding)
                )

                print(
                    f"DEBUG: HOD face verification - Cosine: {cosine_sim:.4f}, Euclidean: {euclidean_dist:.4f}"
                )

                # Use lenient verification for HOD's own face
                if cosine_sim < 0.30 and euclidean_dist > 2.5:
                    raise HTTPException(
                        status_code=401,
                        detail="HOD face verification failed. Please try again.",
                    )
        else:
            # JSON request
            data = await request.json()
            name = data.get("name")
            username = data.get("username")
            password = data.get("password")

        # Verify staff exists in HOD's department
        cursor.execute(
            "SELECT id, username, password_hash, reg_no, name, dept, role, is_active, face_embedding, current_device_id, out_permission_enabled, out_permission_expiry, created_by, created_at, updated_at, embedding, can_reregister, suspended FROM users WHERE id = ? AND dept = ? AND role = 'staff'",
            (staff_id, dept),
        )
        existing = cursor.fetchone()
        if not existing:
            raise HTTPException(
                status_code=404, detail="Staff not found in your department"
            )

        # Update fields
        updates = []
        params = []

        if name:
            updates.append("name = ?")
            params.append(name)

        if username:
            # Check if username is already taken by another user
            cursor.execute(
                "SELECT id FROM users WHERE username = ? AND id != ?",
                (username, staff_id),
            )
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="Username already exists")
            updates.append("username = ?")
            params.append(username)

        if password:
            password_hash = hash_password(password)
            updates.append("password_hash = ?")
            params.append(password_hash)

        if not updates:
            raise HTTPException(status_code=400, detail="No fields to update")

        params.append(staff_id)

        cursor.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
        conn.commit()

        log_audit_event(
            "STAFF_UPDATED_BY_HOD",
            existing[3],
            True,
            f"Staff {existing[1]} updated by HOD of {dept}",
        )

        return {"message": "Staff updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Update staff error: {e}")
        raise HTTPException(status_code=500, detail="Failed to update staff")


# -------------------------------------------------
# FACE REGISTRATION ENDPOINTS
# -------------------------------------------------


@app.post("/admin/face/register")
async def admin_register_face(
    request: Request,
    name: str = Form(...),
    reg_no: str = Form(...),
    dept: str = Form(...),
    role: str = Form(...),
    image: UploadFile = File(...),
):
    """Register face for any user (Admin only)"""
    admin_user = verify_admin_token(request)

    img_bytes = await image.read()
    img = preprocess_image_data(img_bytes)

    # Check image quality before face extraction
    quality = check_image_quality(img)
    print(f"📷 Image quality score: {quality['quality_score']}")
    if quality["warnings"]:
        for warning in quality["warnings"]:
            print(warning)

    face = extract_face(img)
    if face is None:
        save_debug_image(img, "register_fail")
        raise HTTPException(
            status_code=400,
            detail="Face not detected clearly. Ensure good lighting and face is visible.",
        )

    embedding = face.embedding.astype(np.float32)

    # Debug: Print embedding statistics
    print(f"DEBUG: Admin registering face for {reg_no}")
    print(f"  Embedding shape: {embedding.shape}")
    print(f"  Embedding norm: {np.linalg.norm(embedding):.4f}")
    print(f"  Embedding mean: {np.mean(embedding):.4f}")
    print(f"  Embedding std: {np.std(embedding):.4f}")

    # Insert or update based on role - staff, hod, or admin allowed
    if role not in ["staff", "hod", "admin"]:
        raise HTTPException(
            status_code=400, detail="Only staff/hod/admin registration is allowed"
        )

    # For staff, update the users table
    try:
        cursor.execute(
            """
            UPDATE users SET embedding = ? WHERE reg_no = ?
        """,
            (embedding.tobytes(), reg_no),
        )
        conn.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to register: {str(e)}")
    _cache_update_primary(reg_no, embedding)

    _save_face_embedding_sample(reg_no, "users", embedding, "registration_admin", 1.0)

    log_audit_event(
        "FACE_REGISTERED_BY_ADMIN", reg_no, True, f"Face registered for {role} by admin"
    )

    return {
        "message": "Face registered successfully",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "role": role,
        "bbox": [float(x) for x in face.bbox],
    }


@app.post("/admin/face/register/self")
async def admin_register_own_face(request: Request, image: UploadFile = File(...)):
    """
    Admin can register their own face for attendance marking.
    This allows admin to mark their own attendance using face recognition.
    Multiple registrations are allowed - it will update the existing face data.
    """
    admin_user = verify_admin_token(request)

    # Get admin's details from token
    reg_no = admin_user["reg_no"]
    name = admin_user["name"]
    dept = admin_user["dept"]

    img_bytes = await image.read()
    img = preprocess_image_data(img_bytes)

    face = extract_face(img)
    if face is None:
        save_debug_image(img, "register_fail")
        raise HTTPException(
            status_code=400,
            detail="Face not detected clearly. Ensure good lighting and face is visible.",
        )

    embedding = face.embedding.astype(np.float32)

    # Debug: Print embedding statistics
    print(f"DEBUG: Admin self-registering face for {reg_no}")
    print(f"  Embedding shape: {embedding.shape}")
    print(f"  Embedding norm: {np.linalg.norm(embedding):.4f}")
    print(f"  Embedding mean: {np.mean(embedding):.4f}")
    print(f"  Embedding std: {np.std(embedding):.4f}")

    # Update admin's face in the users table (no permission needed for admin)
    try:
        cursor.execute(
            """
            UPDATE users SET embedding = ? WHERE reg_no = ?
        """,
            (embedding.tobytes(), reg_no),
        )
        conn.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to register: {str(e)}")
    _cache_update_primary(reg_no, embedding)

    _save_face_embedding_sample(
        reg_no, "users", embedding, "registration_admin_self", 1.0
    )

    log_audit_event(
        "FACE_REGISTERED_ADMIN_SELF", reg_no, True, "Admin registered their own face"
    )

    return {
        "message": "Face registered successfully for admin",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "role": "admin",
        "bbox": [float(x) for x in face.bbox],
    }


@app.post("/admin/face/permission/{reg_no}")
async def admin_grant_permission(request: Request, reg_no: str):
    """Admin grants permission for staff/hod to re-register face"""
    verify_admin_token(request)

    # Verify user exists
    cursor.execute("SELECT id, name, role FROM users WHERE reg_no = ?", (reg_no,))
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(status_code=404, detail="User not found")

    if existing[2] not in ["staff", "hod"]:
        raise HTTPException(
            status_code=400, detail="Only staff/hod can have face permission"
        )

    # Grant permission
    cursor.execute("UPDATE users SET can_reregister = 1 WHERE reg_no = ?", (reg_no,))
    conn.commit()

    log_audit_event(
        "FACE_PERMISSION_GRANTED_BY_ADMIN",
        reg_no,
        True,
        f"Admin granted permission to re-register for {existing[1]}",
    )

    return {
        "message": "Permission granted successfully. User can now re-register their face.",
        "reg_no": reg_no,
        "user_name": existing[1],
    }


@app.delete("/admin/face/permission/{reg_no}")
async def admin_revoke_permission(request: Request, reg_no: str):
    """Admin revokes permission for staff/hod to re-register face"""
    verify_admin_token(request)

    # Verify user exists
    cursor.execute("SELECT id, name, role FROM users WHERE reg_no = ?", (reg_no,))
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(status_code=404, detail="User not found")

    # Revoke permission
    cursor.execute("UPDATE users SET can_reregister = 0 WHERE reg_no = ?", (reg_no,))
    conn.commit()

    log_audit_event(
        "FACE_PERMISSION_REVOKED_BY_ADMIN",
        reg_no,
        True,
        f"Admin revoked permission to re-register for {existing[1]}",
    )

    return {
        "message": "Permission revoked successfully. User can no longer re-register their face.",
        "reg_no": reg_no,
        "user_name": existing[1],
    }


@app.post("/hod/face/register")
async def hod_register_face(
    request: Request,
    name: str = Form(...),
    reg_no: str = Form(...),
    dept: str = Form(...),
    role: str = Form(...),
    image: UploadFile = File(...),
):
    """Register face for staff in HOD's department"""
    hod_user = verify_hod_token(request)

    # Verify the user belongs to HOD's department
    cursor.execute(
        "SELECT dept FROM users WHERE reg_no = ? AND role = ?", (reg_no, role)
    )
    existing = cursor.fetchone()

    if existing and existing[0] != hod_user["dept"]:
        raise HTTPException(
            status_code=403, detail="User does not belong to your department"
        )

    img_bytes = await image.read()
    img = preprocess_image_data(img_bytes)

    face = extract_face(img)
    if face is None:
        save_debug_image(img, "register_fail")
        raise HTTPException(
            status_code=400,
            detail="Face not detected clearly. Ensure good lighting and face is visible.",
        )

    embedding = face.embedding.astype(np.float32)

    # Debug: Print embedding statistics
    print(f"DEBUG: HOD registering face for {reg_no}")
    print(f"  Embedding shape: {embedding.shape}")
    print(f"  Embedding norm: {np.linalg.norm(embedding):.4f}")
    print(f"  Embedding mean: {np.mean(embedding):.4f}")
    print(f"  Embedding std: {np.std(embedding):.4f}")

    # Only staff/hod/admin registration allowed
    if role not in ["staff", "hod", "admin"]:
        raise HTTPException(
            status_code=400, detail="Only staff/hod/admin registration is allowed"
        )

    try:
        cursor.execute(
            """
            UPDATE users SET embedding = ? WHERE reg_no = ?
        """,
            (embedding.tobytes(), reg_no),
        )
        conn.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to register: {str(e)}")
    _cache_update_primary(reg_no, embedding)

    _save_face_embedding_sample(reg_no, "users", embedding, "registration_hod", 1.0)

    log_audit_event(
        "FACE_REGISTERED_BY_HOD",
        reg_no,
        True,
        f"Face registered for {role} by HOD of {hod_user['dept']}",
    )

    return {
        "message": "Face registered successfully",
        "reg_no": reg_no,
        "name": name,
        "dept": hod_user["dept"],
        "role": role,
        "bbox": [float(x) for x in face.bbox],
    }


@app.post("/hod/face/register/self")
async def hod_register_own_face(request: Request, image: UploadFile = File(...)):
    """
    HOD can register their own face for attendance marking.
    This allows HOD to mark their own attendance using face recognition.
    Multiple registrations are allowed - it will update the existing face data.
    """
    hod_user = verify_hod_token(request)

    # Get HOD's details from token
    reg_no = hod_user["reg_no"]
    name = hod_user["name"]
    dept = hod_user["dept"]

    img_bytes = await image.read()
    img = preprocess_image_data(img_bytes)

    face = extract_face(img)
    if face is None:
        save_debug_image(img, "register_fail")
        raise HTTPException(
            status_code=400,
            detail="Face not detected clearly. Ensure good lighting and face is visible.",
        )

    embedding = face.embedding.astype(np.float32)

    # Debug: Print embedding statistics
    print(f"DEBUG: HOD self-registering face for {reg_no}")
    print(f"  Embedding shape: {embedding.shape}")
    print(f"  Embedding norm: {np.linalg.norm(embedding):.4f}")
    print(f"  Embedding mean: {np.mean(embedding):.4f}")
    print(f"  Embedding std: {np.std(embedding):.4f}")

    # Update HOD's face in the users table (no permission needed for HOD)
    try:
        cursor.execute(
            """
            UPDATE users SET embedding = ? WHERE reg_no = ?
        """,
            (embedding.tobytes(), reg_no),
        )
        conn.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to register: {str(e)}")
    _cache_update_primary(reg_no, embedding)

    _save_face_embedding_sample(
        reg_no, "users", embedding, "registration_hod_self", 1.0
    )

    log_audit_event(
        "FACE_REGISTERED_HOD_SELF", reg_no, True, "HOD registered their own face"
    )

    return {
        "message": "Face registered successfully for HOD",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "role": "hod",
        "bbox": [float(x) for x in face.bbox],
    }


def _register_face_sync_work(reg_no: str, name: str, dept: str, img_bytes: bytes) -> dict:
    img = preprocess_image_data(img_bytes)
    face = extract_face(img)
    if face is None:
        save_debug_image(img, "register_fail")
        raise HTTPException(
            status_code=400,
            detail="Face not detected clearly. Ensure good lighting and face is visible.",
        )
    embedding = face.embedding.astype(np.float32)
    print(f"DEBUG: Registering face for {reg_no}")
    print(f"  Embedding shape: {embedding.shape}")
    print(f"  Embedding norm: {np.linalg.norm(embedding):.4f}")
    print(f"  Embedding mean: {np.mean(embedding):.4f}")
    try:
        cursor.execute(
            "UPDATE users SET embedding = ? WHERE reg_no = ?",
            (embedding.tobytes(), reg_no),
        )
        conn.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to register: {str(e)}")
    _cache_update_primary(reg_no, embedding)
    _save_face_embedding_sample(reg_no, "users", embedding, "registration_staff_self", 1.0)
    log_audit_event("FACE_REGISTERED_STAFF", reg_no, True, "Staff registered their face")
    return {
        "message": "Face registered successfully",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "role": "staff",
        "bbox": [float(x) for x in face.bbox],
    }


@app.post("/staff/face/register")
async def staff_register_face(
    request: Request,
    name: str = Form(...),
    reg_no: str = Form(...),
    dept: str = Form(...),
    role: str = Form(...),
    image: UploadFile = File(...),
):
    """
    Staff can register their own face (one time, or with HOD permission)
    SECURITY: This endpoint verifies the staff details match the logged-in user
    """
    # First verify the staff token to get the logged-in user
    logged_in_staff = verify_staff_token(request)

    # CRITICAL: Verify the logged-in staff's reg_no matches the submitted reg_no
    if logged_in_staff["reg_no"] != reg_no:
        log_audit_event(
            "FACE_REG_MISMATCH",
            reg_no,
            False,
            f"Logged in as {logged_in_staff['reg_no']} but tried to register {reg_no}",
        )
        raise HTTPException(
            status_code=403,
            detail="You can only register your own face. Please use your logged-in credentials.",
        )

    # Verify staff exists in database with matching details
    cursor.execute(
        "SELECT dept, role, embedding, name FROM users WHERE reg_no = ? AND role = 'staff'",
        (reg_no,),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(status_code=404, detail="Staff not found")

    db_dept = existing[0]
    db_name = existing[3]

    # CRITICAL: Verify the department matches what's in the database
    if db_dept != dept:
        raise HTTPException(
            status_code=403,
            detail=f"Department mismatch. Your department is {db_dept}.",
        )

    # CRITICAL: Verify the name matches what's in the database
    if db_name != name:
        raise HTTPException(
            status_code=403,
            detail=f"Name mismatch. Please use your registered name: {db_name}",
        )

    current_embedding = existing[2]

    # Check if already registered
    if current_embedding is not None:
        # Already registered - check if they have permission to re-register
        cursor.execute("SELECT can_reregister FROM users WHERE reg_no = ?", (reg_no,))
        perm = cursor.fetchone()

        if perm is None or perm[0] != 1:
            raise HTTPException(
                status_code=403,
                detail="Face already registered. Please contact your HOD for permission to re-register.",
            )

        # Clear the permission after use
        cursor.execute(
            "UPDATE users SET can_reregister = 0 WHERE reg_no = ?", (reg_no,)
        )
        conn.commit()

    # Rate limiting
    allowed, retry_after = face_register_rate_limiter.is_allowed(reg_no, "face_register")
    if not allowed:
        raise HTTPException(
            status_code=429,
            detail=f"Too many requests. Please wait {retry_after} seconds before trying again.",
        )

    # Read image (async I/O – fine on event loop)
    img_bytes = await image.read()
    if len(img_bytes) == 0:
        raise HTTPException(status_code=400, detail="Empty image")

    # Per-user concurrency guard + thread pool offload
    async with acquire_user_lock(reg_no):
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            _cpu_executor,
            _register_face_sync_work,
            reg_no, name, dept, img_bytes
        )


@app.post("/hod/face/permission/{reg_no}")
async def hod_grant_permission(request: Request, reg_no: str):
    """HOD grants permission for staff to re-register face"""
    hod_user = verify_hod_token(request)

    # Verify staff belongs to HOD's department
    cursor.execute(
        "SELECT id, name FROM users WHERE reg_no = ? AND role = 'staff' AND dept = ?",
        (reg_no, hod_user["dept"]),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Staff not found in your department"
        )

    # Grant permission
    cursor.execute("UPDATE users SET can_reregister = 1 WHERE reg_no = ?", (reg_no,))
    conn.commit()

    log_audit_event(
        "FACE_PERMISSION_GRANTED",
        reg_no,
        True,
        f"HOD {hod_user['username']} granted permission to re-register",
    )

    return {
        "message": "Permission granted successfully. Staff can now re-register their face.",
        "reg_no": reg_no,
        "staff_name": existing[1],
    }


@app.delete("/hod/face/permission/{reg_no}")
async def hod_revoke_permission(request: Request, reg_no: str):
    """HOD revokes permission for staff to re-register face"""
    hod_user = verify_hod_token(request)

    # Verify staff belongs to HOD's department
    cursor.execute(
        "SELECT id, name FROM users WHERE reg_no = ? AND role = 'staff' AND dept = ?",
        (reg_no, hod_user["dept"]),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Staff not found in your department"
        )

    # Revoke permission
    cursor.execute("UPDATE users SET can_reregister = 0 WHERE reg_no = ?", (reg_no,))
    conn.commit()

    log_audit_event(
        "FACE_PERMISSION_REVOKED",
        reg_no,
        True,
        f"HOD {hod_user['username']} revoked permission to re-register",
    )

    return {
        "message": "Permission revoked successfully. Staff can no longer re-register their face.",
        "reg_no": reg_no,
        "staff_name": existing[1],
    }


@app.get("/face/status/{reg_no}")
async def get_face_status(request: Request, reg_no: str):
    """Check if a user has face registered and their permission status"""
    # Check users table
    cursor.execute(
        "SELECT name, dept, role, embedding, can_reregister FROM users WHERE reg_no = ?",
        (reg_no,),
    )
    user = cursor.fetchone()

    if user:
        # Check if there's a pending request
        cursor.execute(
            """
            SELECT id, status, hod_approved, admin_approved 
            FROM face_reregister_requests 
            WHERE staff_reg_no = ? AND status = 'pending'
            ORDER BY id DESC LIMIT 1
        """,
            (reg_no,),
        )
        pending_request = cursor.fetchone()

        has_pending_request = pending_request is not None
        request_status = pending_request[1] if pending_request else None

        return {
            "reg_no": reg_no,
            "name": user[0],
            "dept": user[1],
            "role": user[2],
            "face_registered": user[3] is not None,  # Check if embedding exists
            "can_reregister": user[4] == 1 if len(user) > 4 else False,
            "has_pending_request": has_pending_request,
            "request_status": request_status,
        }

    raise HTTPException(status_code=404, detail="User not found")


@app.get("/staff/can_mark_attendance/{reg_no}")
async def check_can_mark_attendance(request: Request, reg_no: str):
    """Check if staff can mark attendance - must have face registered first"""
    # Verify the request has valid authentication
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing authorization")

    # Check if user exists and is a staff member
    cursor.execute(
        "SELECT name, dept, role, embedding FROM users WHERE reg_no = ?", (reg_no,)
    )
    user = cursor.fetchone()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    name, dept, role, embedding = user

    # Check if face is registered
    face_registered = embedding is not None

    return {
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "role": role,
        "can_mark_attendance": face_registered,
        "face_registered": face_registered,
        "message": "Face registered. You can mark attendance."
        if face_registered
        else "Face not registered. Please register your face first.",
    }


@app.get("/admin/face/status")
async def admin_check_face_status(request: Request):
    """Check if admin has registered their face for attendance"""
    admin_user = verify_admin_token(request)

    reg_no = admin_user["reg_no"]

    cursor.execute(
        "SELECT reg_no, name, embedding FROM users WHERE reg_no = ? AND role = 'admin'",
        (reg_no,),
    )
    user = cursor.fetchone()

    if not user:
        raise HTTPException(status_code=404, detail="Admin not found")

    has_face = user[2] is not None

    return {
        "reg_no": reg_no,
        "name": user[1],
        "role": "admin",
        "has_face_registered": has_face,
        "can_mark_attendance": has_face,
        "message": "Face registered. You can mark attendance."
        if has_face
        else "Face not registered. Please register your face first.",
    }


@app.get("/hod/face/status")
async def hod_check_face_status(request: Request):
    """Check if HOD has registered their face for attendance"""
    hod_user = verify_hod_token(request)

    reg_no = hod_user["reg_no"]

    cursor.execute(
        "SELECT reg_no, name, embedding FROM users WHERE reg_no = ? AND role = 'hod'",
        (reg_no,),
    )
    user = cursor.fetchone()

    if not user:
        raise HTTPException(status_code=404, detail="HOD not found")

    has_face = user[2] is not None

    return {
        "reg_no": reg_no,
        "name": user[1],
        "role": "hod",
        "has_face_registered": has_face,
        "can_mark_attendance": has_face,
        "message": "Face registered. You can mark attendance."
        if has_face
        else "Face not registered. Please register your face first.",
    }


# -------------------------------------------------
# -------------------------------------------------
# CONFIG ENDPOINT
# -------------------------------------------------
@app.post("/staff/face/reregister/request")
async def staff_request_reregister(request: Request):
    """Staff submits a request to re-register their face"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing authorization")

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")
    except:
        raise HTTPException(status_code=401, detail="Invalid token format")

    user = get_user_by_username(username)
    if not user or user[6] != "staff":
        raise HTTPException(
            status_code=403, detail="Only staff can request re-registration"
        )

    reg_no = user[3]
    name = user[4]
    dept = user[5]

    # Check if already has a pending request
    cursor.execute(
        """
        SELECT id FROM face_reregister_requests 
        WHERE staff_reg_no = ? AND status = 'pending'
    """,
        (reg_no,),
    )
    existing_request = cursor.fetchone()

    if existing_request:
        raise HTTPException(
            status_code=400, detail="You already have a pending request"
        )

    # Check if face is already registered
    cursor.execute("SELECT embedding FROM users WHERE reg_no = ?", (reg_no,))
    embedding = cursor.fetchone()

    if not embedding or embedding[0] is None:
        raise HTTPException(
            status_code=400,
            detail="Your face is not registered yet. Please register first.",
        )

    # Create new request
    cursor.execute(
        """
        INSERT INTO face_reregister_requests (staff_reg_no, staff_name, dept, status)
        VALUES (?, ?, ?, 'pending')
    """,
        (reg_no, name, dept),
    )
    conn.commit()

    log_audit_event(
        "REREGISTER_REQUEST_SUBMITTED",
        reg_no,
        True,
        f"Staff {name} submitted face re-registration request",
    )

    return {
        "message": "Request submitted successfully. Please wait for HOD and Admin approval.",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
    }


@app.get("/hod/face/reregister/requests")
async def hod_get_reregister_requests(request: Request):
    """HOD gets all pending re-registration requests for their department"""
    hod_user = verify_hod_token(request)

    cursor.execute(
        """
        SELECT id, staff_reg_no, staff_name, dept, request_date, status, hod_approved, admin_approved
        FROM face_reregister_requests
        WHERE dept = ? AND status = 'pending'
        ORDER BY request_date DESC
    """,
        (hod_user["dept"],),
    )

    requests = cursor.fetchall()

    return {
        "requests": [
            {
                "id": r[0],
                "staff_reg_no": r[1],
                "staff_name": r[2],
                "dept": r[3],
                "request_date": r[4],
                "status": r[5],
                "hod_approved": bool(r[6]) if r[6] else False,
                "admin_approved": bool(r[7]) if r[7] else False,
            }
            for r in requests
        ]
    }


@app.post("/hod/face/reregister/approve/{staff_reg_no}")
async def hod_approve_reregister(request: Request, staff_reg_no: str):
    """HOD approves a re-registration request"""
    hod_user = verify_hod_token(request)

    # Verify staff belongs to HOD's department
    cursor.execute(
        "SELECT id, staff_name FROM face_reregister_requests WHERE staff_reg_no = ? AND dept = ? AND status = 'pending'",
        (staff_reg_no, hod_user["dept"]),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Request not found or already processed"
        )

    # Update HOD approval
    cursor.execute(
        "UPDATE face_reregister_requests SET hod_approved = 1 WHERE staff_reg_no = ?",
        (staff_reg_no,),
    )

    # Check if both HOD and Admin approved
    cursor.execute(
        "SELECT admin_approved FROM face_reregister_requests WHERE staff_reg_no = ?",
        (staff_reg_no,),
    )
    admin_approved = cursor.fetchone()

    if admin_approved and admin_approved[0] == 1:
        # Both approved - grant permission
        cursor.execute(
            "UPDATE users SET can_reregister = 1 WHERE reg_no = ?", (staff_reg_no,)
        )
        cursor.execute(
            "UPDATE face_reregister_requests SET status = 'approved', processed_by = ?, processed_date = CURRENT_TIMESTAMP WHERE staff_reg_no = ?",
            (hod_user["username"], staff_reg_no),
        )
        log_audit_event(
            "REREGISTER_APPROVED",
            staff_reg_no,
            True,
            f"HOD {hod_user['username']} approved - full approval granted",
        )
    else:
        log_audit_event(
            "REREGISTER_HOD_APPROVED",
            staff_reg_no,
            True,
            f"HOD {hod_user['username']} approved (waiting for Admin)",
        )

    conn.commit()

    return {
        "message": "HOD approval granted. "
        + (
            "Full approval granted - staff can now re-register."
            if (admin_approved and admin_approved[0] == 1)
            else "Waiting for Admin approval."
        ),
        "staff_reg_no": staff_reg_no,
    }


@app.post("/hod/face/reregister/deny/{staff_reg_no}")
async def hod_deny_reregister(
    request: Request, staff_reg_no: str, reason: str = "Denied by HOD"
):
    """HOD denies a re-registration request"""
    hod_user = verify_hod_token(request)

    cursor.execute(
        "SELECT id FROM face_reregister_requests WHERE staff_reg_no = ? AND dept = ? AND status = 'pending'",
        (staff_reg_no, hod_user["dept"]),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Request not found or already processed"
        )

    cursor.execute(
        """
        UPDATE face_reregister_requests 
        SET status = 'denied', processed_by = ?, processed_date = CURRENT_TIMESTAMP, reason = ?
        WHERE staff_reg_no = ?
    """,
        (hod_user["username"], reason, staff_reg_no),
    )
    conn.commit()

    log_audit_event(
        "REREGISTER_DENIED",
        staff_reg_no,
        True,
        f"HOD {hod_user['username']} denied - reason: {reason}",
    )

    return {"message": "Request denied", "staff_reg_no": staff_reg_no, "reason": reason}


@app.get("/admin/face/reregister/requests")
async def admin_get_reregister_requests(request: Request):
    """Admin gets all pending re-registration requests for staff (not other_staff)"""
    verify_admin_token(request)

    cursor.execute("""
        SELECT id, staff_reg_no, staff_name, dept, request_date, status, hod_approved, admin_approved
        FROM face_reregister_requests
        WHERE status = 'pending'
        ORDER BY request_date DESC
    """)

    all_requests = cursor.fetchall()
    staff_requests = []

    for r in all_requests:
        reg_no = r[1]
        cursor.execute("SELECT id FROM users WHERE reg_no = ?", (reg_no,))
        if cursor.fetchone():
            staff_requests.append(
                {
                    "id": r[0],
                    "staff_reg_no": r[1],
                    "staff_name": r[2],
                    "dept": r[3],
                    "request_date": r[4],
                    "status": r[5],
                    "hod_approved": bool(r[6]) if r[6] else False,
                    "admin_approved": bool(r[7]) if r[7] else False,
                    "user_type": "staff",
                }
            )

    return {"requests": staff_requests}


@app.post("/admin/face/reregister/approve/{staff_reg_no}")
async def admin_approve_reregister(request: Request, staff_reg_no: str):
    """Admin approves a re-registration request (direct approval, no HOD needed)"""
    verify_admin_token(request)

    cursor.execute(
        "SELECT id FROM face_reregister_requests WHERE staff_reg_no = ? AND status = 'pending'",
        (staff_reg_no,),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Request not found or already processed"
        )

    # Grant permission directly
    cursor.execute("SELECT id FROM other_staff WHERE reg_no = ?", (staff_reg_no,))
    if cursor.fetchone():
        cursor.execute(
            "UPDATE other_staff SET can_reregister = 1 WHERE reg_no = ?",
            (staff_reg_no,),
        )
    else:
        cursor.execute(
            "UPDATE users SET can_reregister = 1 WHERE reg_no = ?", (staff_reg_no,)
        )

    cursor.execute(
        "UPDATE face_reregister_requests SET status = 'approved', hod_approved = 1, admin_approved = 1, processed_by = 'admin', processed_date = CURRENT_TIMESTAMP WHERE staff_reg_no = ?",
        (staff_reg_no,),
    )
    conn.commit()

    log_audit_event(
        "REREGISTER_APPROVED",
        staff_reg_no,
        True,
        "Admin approved - full approval granted",
    )

    return {
        "message": "Re-registration approved. Staff can now re-register their face.",
        "staff_reg_no": staff_reg_no,
    }


@app.post("/admin/face/reregister/deny/{staff_reg_no}")
async def admin_deny_reregister(
    request: Request, staff_reg_no: str, reason: str = "Denied by Admin"
):
    """Admin denies a re-registration request"""
    verify_admin_token(request)

    cursor.execute(
        "SELECT id FROM face_reregister_requests WHERE staff_reg_no = ? AND status = 'pending'",
        (staff_reg_no,),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Request not found or already processed"
        )

    cursor.execute(
        """
        UPDATE face_reregister_requests 
        SET status = 'denied', processed_by = 'admin', processed_date = CURRENT_TIMESTAMP, reason = ?
        WHERE staff_reg_no = ?
    """,
        (reason, staff_reg_no),
    )
    conn.commit()

    log_audit_event(
        "REREGISTER_DENIED", staff_reg_no, True, f"Admin denied - reason: {reason}"
    )

    return {"message": "Request denied", "staff_reg_no": staff_reg_no, "reason": reason}


@app.get("/face/verify")
async def verify_face_status(request: Request):
    """Verify face status for current user"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing authorization")

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")

        user = get_user_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        reg_no = user[3]
        role = user[6]

        # Check face status - only staff/hod/admin allowed
        if role not in ["admin", "hod", "staff"]:
            raise HTTPException(status_code=400, detail="Invalid role")

        # Admin, HOD, Staff - embedding stored in users table
        return {
            "registered": user[2]
            is not None,  # password_hash field, but we need embedding
            "name": user[4],
            "dept": user[5],
            "role": role,
        }

    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid token")


# -------------------------------------------------
# CONFIG ENDPOINT
# -------------------------------------------------
@app.get("/config")
async def get_config():
    """Get system configuration"""
    return {
        "confidence_threshold": CONFIDENCE_THRESHOLD,
        "min_cosine_similarity": MIN_COSINE_SIMILARITY,
        "max_euclidean_distance": MAX_EUCLIDEAN_DISTANCE,
        "max_failed_attempts": MAX_FAILED_ATTEMPTS,
        "lockout_duration_minutes": LOCKOUT_DURATION_MINUTES,
    }


# -------------------------------------------------
# USER LEAVE REQUEST ENDPOINTS
# -------------------------------------------------

MAX_REASON_LENGTH = 1000


def verify_user_token(request: Request) -> dict:
    """Verify user authentication token - works for all authenticated users"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authorization header"
        )

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")

        user = get_user_by_username(username)
        if not user:
            other_staff = get_other_staff_by_username(username)
            if other_staff:
                user = (
                    other_staff[0],
                    other_staff[1],
                    other_staff[2],
                    other_staff[3],
                    other_staff[4],
                    other_staff[7],
                    other_staff[6],
                )

        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        return {
            "id": user[0],
            "username": user[1],
            "reg_no": user[3],
            "name": user[4],
            "dept": user[5],
            "role": user[6],
        }
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


@app.post("/user/change_password")
async def user_change_password(request: Request):
    """Change password for any authenticated user"""
    try:
        user = verify_user_token(request)
        data = await request.json()
        current_password = (data.get("current_password") or "").strip()
        new_password = (data.get("new_password") or "").strip()
        confirm_password = (data.get("confirm_password") or "").strip()

        if not current_password or not new_password or not confirm_password:
            raise HTTPException(
                status_code=400, detail="All password fields are required"
            )
        if len(new_password) < 6:
            raise HTTPException(
                status_code=400, detail="New password must be at least 6 characters long"
            )
        if new_password != confirm_password:
            raise HTTPException(status_code=400, detail="New passwords do not match")
        if current_password == new_password:
            raise HTTPException(
                status_code=400, detail="New password must be different"
            )

        user_row = get_user_by_username(user["username"])
        table = "users"
        if not user_row:
            user_row = get_other_staff_by_username(user["username"])
            table = "other_staff"
        if not user_row:
            raise HTTPException(status_code=404, detail="User not found")

        if not verify_password(current_password, user_row[2]):
            raise HTTPException(status_code=400, detail="Current password is incorrect")

        cursor.execute(
            f"UPDATE {table} SET password_hash = %s WHERE id = %s",
            (hash_password(new_password), user_row[0]),
        )

        log_audit_event(
            "PASSWORD_CHANGED",
            user.get("reg_no"),
            True,
            f"Password changed for {user.get('username')}",
        )

        return {"message": "Password changed successfully. Please login again."}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Change password error: {e}")
        raise HTTPException(status_code=500, detail="Failed to change password")


@app.post("/leave/submit")
async def submit_leave_request(request: Request):
    """Submit a leave request with reason - requires authentication.
    For Casual Leave: validates CL balance and restricts date range.
    For Earned Leave / On Duty: no CL deduction, marks as Present/OD.
    """
    try:
        data = await request.json()

        user = verify_any_user_token(request)

        leave_type = data.get("leave_type")
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        reason = data.get("reason", "").strip()
        # Half-day leave fields (optional — only when half_day_enabled in leave_settings)
        is_half_day = bool(data.get("is_half_day", False))
        which_half = (data.get("which_half") or "").strip().lower()  # 'first' | 'second'
        if is_half_day and which_half not in ("first", "second"):
            raise HTTPException(
                status_code=400,
                detail="which_half must be 'first' or 'second' when is_half_day is true",
            )
        # Half-day leave only valid for a single day
        if is_half_day and start_date and start_date != end_date:
            raise HTTPException(
                status_code=400,
                detail="Half-day leave can only be applied for a single day (start_date must equal end_date)",
            )

        if not all([leave_type, start_date, end_date]):
            raise HTTPException(
                status_code=400,
                detail="Missing required fields: leave_type, start_date, end_date",
            )

        if not reason:
            raise HTTPException(status_code=400, detail="Reason is required")

        if len(reason) > MAX_REASON_LENGTH:
            raise HTTPException(
                status_code=400,
                detail=f"Reason exceeds maximum length of {MAX_REASON_LENGTH} characters",
            )

        if len(reason) < 10:
            raise HTTPException(
                status_code=400, detail="Reason must be at least 10 characters"
            )

        # Map frontend leave type names to internal codes
        leave_type_map = {
            # Internal codes (already lowercase)
            "casual": "casual",
            "earned": "earned",
            "od": "od",
            "sick": "sick",
            "maternity": "maternity",
            "paternity": "paternity",
            "unpaid": "unpaid",
            "other": "other",
            # Human-readable names from frontend
            "casual leave": "casual",
            "earned leave": "earned",
            "on duty": "od",
            "sick leave": "sick",
        }
        leave_type_lower = leave_type.lower().strip() if leave_type else ""
        internal_type = leave_type_map.get(leave_type_lower, leave_type_lower)

        allowed_internal_types = [
            "sick",
            "casual",
            "earned",
            "od",
            "maternity",
            "paternity",
            "unpaid",
            "other",
        ]
        if internal_type not in allowed_internal_types:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid leave type. Allowed: Casual Leave, Earned Leave, On Duty, Sick Leave, Maternity, Paternity, Unpaid, Other",
            )

        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Invalid date format. Use YYYY-MM-DD"
            )

        if end < start:
            raise HTTPException(
                status_code=400, detail="End date must be on or after start date"
            )

        # Prevent requesting leave 1 day before/after or spanning across an Academic Calendar holiday (sandwich leave policy)
        # 1. Check if any day within the selected date range is a holiday
        curr = start
        while curr <= end:
            curr_str = curr.strftime("%Y-%m-%d")
            status_curr, reason_curr, is_holiday_curr = _academic_status_for_date(curr_str)
            if is_holiday_curr and status_curr == "holiday":
                reason_str = f" ({reason_curr})" if reason_curr else ""
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot request leave: The selected date range includes an academic holiday on {curr_str}{reason_str}."
                )
            curr += timedelta(days=1)

        # 2. Check if the day before start_date is a holiday
        day_before = (start - timedelta(days=1)).strftime("%Y-%m-%d")
        status_before, reason_before, is_holiday_before = _academic_status_for_date(day_before)
        if is_holiday_before and status_before == "holiday":
            reason_str = f" ({reason_before})" if reason_before else ""
            raise HTTPException(
                status_code=400,
                detail=f"Cannot request leave adjacent to an academic holiday. The day before your leave start date ({day_before}) is a holiday in the academic calendar{reason_str}."
            )

        # 3. Check if the day after end_date is a holiday
        day_after = (end + timedelta(days=1)).strftime("%Y-%m-%d")
        status_after, reason_after, is_holiday_after = _academic_status_for_date(day_after)
        if is_holiday_after and status_after == "holiday":
            reason_str = f" ({reason_after})" if reason_after else ""
            raise HTTPException(
                status_code=400,
                detail=f"Cannot request leave adjacent to an academic holiday. The day after your leave end date ({day_after}) is a holiday in the academic calendar{reason_str}."
            )

        # For Casual Leave: validate CL balance before allowing submission
        if internal_type == "casual":
            current_month = datetime.now().strftime("%Y-%m")
            cursor.execute(
                """
                SELECT current_month_cl_available, accumulated_cl
                FROM casual_leave
                WHERE reg_no = %s AND current_month = %s
            """,
                (user["reg_no"], current_month),
            )
            cl_record = cursor.fetchone()

            if not cl_record:
                # Initialize CL for user
                initialize_cl_for_user(
                    user["reg_no"], user["name"], user["dept"], user["role"]
                )
                cursor.execute(
                    """
                    SELECT current_month_cl_available, accumulated_cl
                    FROM casual_leave
                    WHERE reg_no = %s AND current_month = %s
                """,
                    (user["reg_no"], current_month),
                )
                cl_record = cursor.fetchone()

            if not cl_record:
                raise HTTPException(
                    status_code=400, detail="Unable to fetch CL balance"
                )

            cl_available, accumulated = cl_record
            total_cl = cl_available + accumulated

            # Count working days in the date range (exclude weekends); half-day = 0.5
            day_count = 0.0
            current = start
            while current <= end:
                if current.weekday() < 5:  # Mon-Fri
                    day_count += 0.5 if is_half_day else 1.0
                current += timedelta(days=1)

            if day_count <= 0:
                raise HTTPException(
                    status_code=400, detail="No working days in the selected date range"
                )

            if day_count > total_cl:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient CL balance. You have {total_cl} CL available but requested {day_count} working day(s). "
                    f"(Current month: {cl_available}, Accumulated: {accumulated})",
                )

        # For Earned Leave: validate CCL balance before allowing submission
        elif internal_type == "earned":
            cursor.execute("SELECT balance FROM earned_leave WHERE reg_no = %s", (user["reg_no"],))
            record = cursor.fetchone()
            if not record:
                # Initialize balance
                cursor.execute("""
                    INSERT INTO earned_leave (reg_no, user_name, dept, role, balance)
                    VALUES (%s, %s, %s, %s, 0.0)
                """, (user["reg_no"], user["name"], user["dept"], user["role"]))
                conn.commit()
                total_el = 0.0
            else:
                total_el = float(record[0])

            # Count working days in the date range (exclude weekends); half-day = 0.5
            day_count = 0.0
            current = start
            while current <= end:
                if current.weekday() < 5:  # Mon-Fri
                    day_count += 0.5 if is_half_day else 1.0
                current += timedelta(days=1)

            if day_count <= 0:
                raise HTTPException(
                    status_code=400, detail="No working days in the selected date range"
                )

            if day_count > total_el:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient Earned Leave balance. You have {total_el} EL/CCL available but requested {day_count} working day(s).",
                )

        cursor.execute(
            """
            INSERT INTO leave_requests 
            (user_reg_no, user_name, dept, leave_type, start_date, end_date, reason, is_half_day, which_half)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """,
            (
                user["reg_no"],
                user["name"],
                user["dept"],
                leave_type,
                start_date,
                end_date,
                reason,
                is_half_day,
                which_half if is_half_day else None,
            ),
        )
        result = cursor.fetchone()
        request_id = result[0] if result else None

        cursor.execute(
            """
            INSERT INTO leave_request_audit_log 
            (leave_request_id, action, performed_by, performed_by_name, new_status)
            VALUES (%s, %s, %s, %s, %s)
        """,
            (request_id, "SUBMITTED", user["reg_no"], user["name"], "pending"),
        )

        cursor.execute(
            """
            INSERT INTO admin_notifications 
            (notification_type, title, message, related_id, created_for)
            VALUES (%s, %s, %s, %s, %s)
        """,
            (
                "leave_request",
                f"New Leave Request from {user['name']}",
                f"{user['name']} submitted a {leave_type} leave request from {start_date} to {end_date}",
                request_id,
                "admin",
            ),
        )

        log_audit_event(
            "LEAVE_SUBMITTED",
            user["reg_no"],
            True,
            f"Leave request {request_id} submitted",
        )

        return {
            "success": True,
            "message": "Leave request submitted successfully",
            "request_id": request_id,
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error submitting leave request: {e}")
        raise HTTPException(status_code=500, detail="Unable to send leave request")


@app.get("/leave/my-requests")
async def get_my_leave_requests(request: Request):
    """Get all leave requests for the authenticated user"""
    try:
        user = verify_user_token(request)

        cursor.execute(
            """
            SELECT id, leave_type, start_date, end_date, reason, submission_date, 
                   status, processed_by, processed_date, admin_comment
            FROM leave_requests 
            WHERE user_reg_no = %s
            ORDER BY submission_date DESC
        """,
            (user["reg_no"],),
        )

        rows = cursor.fetchall()
        requests_list = []
        for row in rows:
            requests_list.append(
                {
                    "id": row[0],
                    "leave_type": row[1],
                    "start_date": row[2],
                    "end_date": row[3],
                    "reason": row[4],
                    "submission_date": _ts(row[5]),
                    "status": row[6],
                    "processed_by": row[7],
                    "processed_date": _ts(row[8]),
                    "admin_comment": row[9],
                }
            )

        return {"success": True, "requests": requests_list}

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching leave requests: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch leave requests")


@app.get("/leave/request/{request_id}")
async def get_leave_request_details(request: Request, request_id: int):
    """Get details of a specific leave request"""
    try:
        user = verify_user_token(request)

        cursor.execute(
            """
            SELECT id, user_reg_no, user_name, dept, leave_type, start_date, end_date, 
                   reason, submission_date, status, processed_by, processed_date, admin_comment
            FROM leave_requests 
            WHERE id = %s
        """,
            (request_id,),
        )

        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Leave request not found")

        if row[1] != user["reg_no"] and user["role"] not in ["admin", "hod"]:
            raise HTTPException(status_code=403, detail="Access denied")

        return {
            "success": True,
            "request": {
                "id": row[0],
                "user_reg_no": row[1],
                "user_name": row[2],
                "dept": row[3],
                "leave_type": row[4],
                "start_date": row[5],
                "end_date": row[6],
                "reason": row[7],
                "submission_date": _ts(row[8]),
                "status": row[9],
                "processed_by": row[10],
                "processed_date": _ts(row[11]),
                "admin_comment": row[12],
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching leave request: {e}")
        raise HTTPException(
            status_code=500, detail="Failed to fetch leave request details"
        )


@app.put("/leave/request/{request_id}")
async def update_leave_request(request: Request, request_id: int):
    """Update a pending leave request (only if still pending)"""
    try:
        user = verify_user_token(request)
        data = await request.json()

        cursor.execute(
            "SELECT user_reg_no, status FROM leave_requests WHERE id = %s",
            (request_id,),
        )
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Leave request not found")

        if row[0] != user["reg_no"]:
            raise HTTPException(status_code=403, detail="Access denied")

        if row[1] != "pending":
            raise HTTPException(
                status_code=400, detail="Cannot update a processed leave request"
            )

        leave_type = data.get("leave_type")
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        reason = data.get("reason", "").strip()

        if reason and len(reason) > MAX_REASON_LENGTH:
            raise HTTPException(
                status_code=400,
                detail=f"Reason exceeds maximum length of {MAX_REASON_LENGTH} characters",
            )

        if reason and len(reason) < 10:
            raise HTTPException(
                status_code=400, detail="Reason must be at least 10 characters"
            )

        updates = []
        params = []

        if leave_type:
            updates.append("leave_type = %s")
            params.append(leave_type)
        if start_date:
            updates.append("start_date = %s")
            params.append(start_date)
        if end_date:
            updates.append("end_date = %s")
            params.append(end_date)
        if reason:
            updates.append("reason = %s")
            params.append(reason)

        if not updates:
            raise HTTPException(status_code=400, detail="No fields to update")

        params.append(request_id)

        cursor.execute(
            f"UPDATE leave_requests SET {', '.join(updates)} WHERE id = %s", params
        )

        cursor.execute(
            """
            INSERT INTO leave_request_audit_log 
            (leave_request_id, action, performed_by, performed_by_name, comments)
            VALUES (%s, %s, %s, %s, %s)
        """,
            (
                request_id,
                "UPDATED",
                user["reg_no"],
                user["name"],
                "Request updated by user",
            ),
        )

        return {"success": True, "message": "Leave request updated successfully"}

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating leave request: {e}")
        raise HTTPException(status_code=500, detail="Failed to update leave request")


@app.delete("/leave/request/{request_id}")
async def cancel_leave_request(request: Request, request_id: int):
    """Cancel a pending leave request"""
    try:
        user = verify_user_token(request)

        cursor.execute(
            "SELECT user_reg_no, status FROM leave_requests WHERE id = %s",
            (request_id,),
        )
        row = cursor.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Leave request not found")

        if row[0] != user["reg_no"]:
            raise HTTPException(status_code=403, detail="Access denied")

        if row[1] != "pending":
            raise HTTPException(
                status_code=400, detail="Cannot cancel a processed leave request"
            )

        cursor.execute(
            "UPDATE leave_requests SET status = 'rejected', processed_by = %s, processed_date = CURRENT_TIMESTAMP WHERE id = %s",
            (user["name"], request_id),
        )

        cursor.execute(
            """
            INSERT INTO leave_request_audit_log 
            (leave_request_id, action, performed_by, performed_by_name, previous_status, new_status, comments)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
            (
                request_id,
                "CANCELLED",
                user["reg_no"],
                user["name"],
                "pending",
                "rejected",
                "Cancelled by user",
            ),
        )

        return {"success": True, "message": "Leave request cancelled successfully"}

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error cancelling leave request: {e}")
        raise HTTPException(status_code=500, detail="Failed to cancel leave request")


# -------------------------------------------------
# ADMIN LEAVE REQUEST MANAGEMENT ENDPOINTS
# -------------------------------------------------


@app.get("/admin/leave/requests")
async def admin_get_leave_requests(
    request: Request,
    status: str = None,
    dept: str = None,
    search: str = None,
    start_date: str = None,
    end_date: str = None,
    sort_by: str = "submission_date",
    sort_order: str = "desc",
    page: int = 1,
    limit: int = 50,
):
    """Get all leave requests with filtering, sorting, and pagination - Admin only"""
    verify_admin_token(request)

    # Build query
    query = """
        SELECT lr.id, lr.user_reg_no, lr.user_name, lr.dept, lr.leave_type, 
               lr.start_date, lr.end_date, lr.reason, lr.submission_date, 
               lr.status, lr.processed_by, lr.processed_date, lr.admin_comment,
               lr.is_read_by_admin
        FROM leave_requests lr
        WHERE 1=1
    """
    params = []

    if status:
        query += " AND lr.status = ?"
        params.append(status)

    if dept:
        query += " AND lr.dept = ?"
        params.append(dept)

    if search:
        query += (
            " AND (lr.user_name LIKE ? OR lr.user_reg_no LIKE ? OR lr.reason LIKE ?)"
        )
        search_term = f"%{search}%"
        params.extend([search_term, search_term, search_term])

    if start_date:
        query += " AND lr.submission_date >= ?"
        params.append(start_date)

    if end_date:
        query += " AND lr.submission_date <= ?"
        params.append(end_date)

    # Sorting
    allowed_sort_fields = [
        "submission_date",
        "start_date",
        "end_date",
        "status",
        "user_name",
        "dept",
    ]
    if sort_by not in allowed_sort_fields:
        sort_by = "submission_date"

    sort_order = "DESC" if sort_order.lower() == "desc" else "ASC"
    query += f" ORDER BY lr.{sort_by} {sort_order}"

    # Pagination
    offset = (page - 1) * limit
    query += " LIMIT ? OFFSET ?"
    params.extend([limit, offset])

    cursor.execute(query, params)
    rows = cursor.fetchall()

    requests_list = []
    for row in rows:
        requests_list.append(
            {
                "id": row[0],
                "user_reg_no": row[1],
                "user_name": row[2],
                "dept": row[3],
                "leave_type": row[4],
                "start_date": row[5],
                "end_date": row[6],
                "reason": row[7],
                "submission_date": row[8],
                "status": row[9],
                "processed_by": row[10],
                "processed_date": row[11],
                "admin_comment": row[12],
                "is_read": bool(row[13]) if row[13] is not None else False,
            }
        )

    # Get total count for pagination
    count_query = "SELECT COUNT(*) FROM leave_requests lr WHERE 1=1"
    count_params = []
    if status:
        count_query += " AND lr.status = ?"
        count_params.append(status)
    if dept:
        count_query += " AND lr.dept = ?"
        count_params.append(dept)
    if search:
        count_query += (
            " AND (lr.user_name LIKE ? OR lr.user_reg_no LIKE ? OR lr.reason LIKE ?)"
        )
        search_term = f"%{search}%"
        count_params.extend([search_term, search_term, search_term])
    if start_date:
        count_query += " AND lr.submission_date >= ?"
        count_params.append(start_date)
    if end_date:
        count_query += " AND lr.submission_date <= ?"
        count_params.append(end_date)

    cursor.execute(count_query, count_params)
    total_count = cursor.fetchone()[0]

    return {
        "success": True,
        "requests": requests_list,
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total_count,
            "pages": (total_count + limit - 1) // limit,
        },
    }


@app.get("/admin/leave/requests/pending")
async def admin_get_pending_leave_requests(request: Request):
    """Get all pending leave requests - Admin only"""
    verify_admin_token(request)

    cursor.execute("""
        SELECT lr.id, lr.user_reg_no, lr.user_name, lr.dept, lr.leave_type, 
               lr.start_date, lr.end_date, lr.reason, lr.submission_date, 
               lr.status, lr.is_read_by_admin
        FROM leave_requests lr
        WHERE lr.status = 'pending'
        ORDER BY lr.submission_date ASC
    """)

    rows = cursor.fetchall()
    requests_list = []

    for row in rows:
        requests_list.append(
            {
                "id": row[0],
                "user_reg_no": row[1],
                "user_name": row[2],
                "dept": row[3],
                "leave_type": row[4],
                "start_date": row[5],
                "end_date": row[6],
                "reason": row[7],
                "submission_date": row[8],
                "status": row[9],
                "is_read": bool(row[10]) if row[10] is not None else False,
            }
        )

    return {"success": True, "requests": requests_list, "count": len(requests_list)}


@app.post("/admin/leave/request/{request_id}/approve")
async def admin_approve_leave_request(request: Request, request_id: int):
    """Approve a leave request - Admin only.
    - Casual Leave / Earned Leave / On Duty → marks as Present with tag
    - Other leave types (sick, maternity, etc.) → marks as Leave
    - For Casual Leave: deducts CL balance (accumulated first, then current month)
    - Instantly syncs to daily_attendance_status for all dates in range
    """
    admin_user = verify_admin_token(request)

    try:
        data = await request.json()
        comment = data.get("comment", "")
    except:
        comment = ""

    # Check if request exists and get full details
    cursor.execute(
        "SELECT status, leave_type, start_date, end_date, user_reg_no, user_name, dept, is_half_day, which_half FROM leave_requests WHERE id = %s",
        (request_id,),
    )
    row = cursor.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Leave request not found")

    current_status, leave_type, start_date, end_date, user_reg_no, user_name, dept, is_half_day, which_half = row
    is_half_day = bool(is_half_day) if is_half_day is not None else False
    which_half = (which_half or "").strip().lower()  # 'first' | 'second' | ''

    if current_status != "pending":
        raise HTTPException(
            status_code=400, detail="Only pending requests can be approved"
        )

    # Normalize leave_type to lowercase for consistent handling
    leave_type_lower = leave_type.lower().strip() if leave_type else ""

    # Map frontend leave type names to internal codes
    leave_type_map = {
        "casual": "casual",
        "casual leave": "casual",
        "earned": "earned",
        "earned leave": "earned",
        "od": "od",
        "on duty": "od",
        "sick": "sick",
        "sick leave": "sick",
        "maternity": "maternity",
        "paternity": "paternity",
        "unpaid": "unpaid",
        "other": "other",
    }
    internal_type = leave_type_map.get(leave_type_lower, leave_type_lower)

    # Determine attendance status based on leave type
    # Casual Leave, Earned Leave, On Duty → Present with tag
    # Other leave types → Leave
    if internal_type in ("casual", "earned", "od"):
        attendance_status = "Present"
        status_tag = internal_type  # Tag: "casual", "earned", or "od"
    else:
        attendance_status = "Leave"
        status_tag = internal_type

    # Create/update daily attendance status entries for each day in the date range
    # This instantly syncs the approved leave to the analysis tab
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")

    # Determine if half-day mode is active
    _hd_settings = _get_half_day_settings()
    _half_day_mode = _hd_settings["half_day_enabled"]

    synced_dates = []
    current = start
    while current <= end:
        date_str = current.strftime("%Y-%m-%d")
        try:
            if is_half_day and which_half and _half_day_mode:
                # ── HALF-DAY LEAVE APPROVAL ────────────────────────────────────
                # Fetch current half statuses
                cursor.execute(
                    "SELECT first_half_status, second_half_status FROM daily_attendance_status WHERE reg_no = %s AND date = %s",
                    (user_reg_no, date_str),
                )
                existing = cursor.fetchone()
                cur_fh = existing[0] if existing else None
                cur_sh = existing[1] if existing else None

                # Apply leave to the correct half
                new_fh = "Leave" if which_half == "first" else cur_fh
                new_sh = "Leave" if which_half == "second" else cur_sh
                new_overall = _compute_daily_status(new_fh, new_sh)
                # Calculate attendance value (leave doesn't count as present)
                attendance_value = _compute_attendance_value_from_halves(new_fh, new_sh)

                cursor.execute(
                    """
                    INSERT INTO daily_attendance_status
                    (reg_no, name, dept, date, status,
                     first_half_status, second_half_status,
                     attendance_value, leave_request_id, leave_type, marked_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (reg_no, date) DO UPDATE SET
                        status = EXCLUDED.status,
                        first_half_status = EXCLUDED.first_half_status,
                        second_half_status = EXCLUDED.second_half_status,
                        attendance_value = EXCLUDED.attendance_value,
                        leave_request_id = EXCLUDED.leave_request_id,
                        leave_type = EXCLUDED.leave_type,
                        marked_by = EXCLUDED.marked_by
                """,
                    (
                        user_reg_no, user_name, dept, date_str,
                        new_overall, new_fh, new_sh, attendance_value,
                        request_id, status_tag, admin_user["name"],
                    ),
                )
            else:
                # ── FULL-DAY LEAVE APPROVAL ────────────────────────────────────
                # Calculate attendance value for full-day leave
                attendance_value = 0.0 if attendance_status == "Leave" else 1.0
                
                cursor.execute(
                    """
                    INSERT INTO daily_attendance_status 
                    (reg_no, name, dept, date, status, attendance_value, leave_request_id, leave_type, marked_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (reg_no, date) DO UPDATE SET
                        status = EXCLUDED.status,
                        attendance_value = EXCLUDED.attendance_value,
                        leave_request_id = EXCLUDED.leave_request_id,
                        leave_type = EXCLUDED.leave_type,
                        marked_by = EXCLUDED.marked_by
                """,
                    (
                        user_reg_no,
                        user_name,
                        dept,
                        date_str,
                        attendance_status,
                        attendance_value,
                        request_id,
                        status_tag,
                        admin_user["name"],
                    ),
                )
            synced_dates.append(date_str)

        except Exception as e:
            print(
                f"ERROR inserting daily_attendance_status for {user_reg_no} on {date_str}: {e}"
            )
            # Try a simpler INSERT without ON CONFLICT as fallback
            try:
                cursor.execute(
                    "DELETE FROM daily_attendance_status WHERE reg_no = %s AND date = %s",
                    (user_reg_no, date_str),
                )
                cursor.execute(
                    """
                    INSERT INTO daily_attendance_status 
                    (reg_no, name, dept, date, status, leave_request_id, leave_type, marked_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                    (
                        user_reg_no,
                        user_name,
                        dept,
                        date_str,
                        attendance_status,
                        request_id,
                        status_tag,
                        admin_user["name"],
                    ),
                )
                synced_dates.append(date_str)
                print(f"FALLBACK INSERT succeeded for {user_reg_no} on {date_str}")
            except Exception as e2:
                print(f"FALLBACK also failed for {user_reg_no} on {date_str}: {e2}")
        current += timedelta(days=1)

    # Verify the records were actually created
    if synced_dates:
        date_conditions = " OR ".join([f"date = %s" for _ in synced_dates])
        cursor.execute(
            f"""
            SELECT date, status, leave_type FROM daily_attendance_status
            WHERE reg_no = %s AND ({date_conditions})
            ORDER BY date
        """,
            [user_reg_no] + synced_dates,
        )
        verified = cursor.fetchall()
        print(
            f"Leave approved: {user_reg_no} ({user_name}) - {leave_type} from {start_date} to {end_date} "
            f"→ Status: {attendance_status}, Tag: {status_tag}, Dates synced: {len(synced_dates)}, Verified: {len(verified)}"
        )
        for v in verified:
            print(f"  Verified: {v[0]} | {v[1]} | {v[2]}")
    else:
        print(
            f"WARNING: Leave approved but NO dates were synced for {user_reg_no} ({user_name})"
        )

    # For Casual Leave: deduct CL balance
    if internal_type == "casual":
        current_month = datetime.now().strftime("%Y-%m")
        cursor.execute(
            """
            SELECT current_month_cl_available, accumulated_cl, cl_used_current_month
            FROM casual_leave
            WHERE reg_no = %s AND current_month = %s
        """,
            (user_reg_no, current_month),
        )
        cl_record = cursor.fetchone()

        if cl_record:
            cl_available, accumulated, used = cl_record
            # Count working days (exclude weekends); half-day = 0.5
            day_count = 0.0
            d = start
            while d <= end:
                if d.weekday() < 5:
                    day_count += 0.5 if is_half_day else 1.0
                d += timedelta(days=1)

            if day_count > 0:
                # Deduct from accumulated first, then current month
                remaining_to_deduct = day_count
                new_accumulated = float(accumulated)
                new_cl_available = float(cl_available)

                if new_accumulated >= remaining_to_deduct:
                    new_accumulated -= remaining_to_deduct
                    remaining_to_deduct = 0
                else:
                    remaining_to_deduct -= new_accumulated
                    new_accumulated = 0
                    if new_cl_available >= remaining_to_deduct:
                        new_cl_available -= remaining_to_deduct
                        remaining_to_deduct = 0

                new_used = float(used) + day_count

                cursor.execute(
                    """
                    UPDATE casual_leave
                    SET current_month_cl_available = %s, accumulated_cl = %s,
                        cl_used_current_month = %s, last_updated = CURRENT_TIMESTAMP
                    WHERE reg_no = %s AND current_month = %s
                """,
                    (
                        new_cl_available,
                        new_accumulated,
                        new_used,
                        user_reg_no,
                        current_month,
                    ),
                )

                print(
                    f"CL deducted: {user_reg_no} used {day_count} days. "
                    f"New balance: current_month={new_cl_available}, accumulated={new_accumulated}, used={new_used}"
                )

    # For Earned Leave: deduct EL balance
    elif internal_type == "earned":
        cursor.execute("SELECT balance FROM earned_leave WHERE reg_no = %s", (user_reg_no,))
        el_record = cursor.fetchone()
        if el_record:
            current_el = float(el_record[0])
            # Count working days (exclude weekends); half-day = 0.5
            day_count = 0.0
            d = start
            while d <= end:
                if d.weekday() < 5:
                    day_count += 0.5 if is_half_day else 1.0
                d += timedelta(days=1)

            if day_count > 0:
                new_el = max(0.0, current_el - day_count)
                cursor.execute("""
                    UPDATE earned_leave
                    SET balance = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE reg_no = %s
                """, (new_el, user_reg_no))
                print(f"EL deducted: {user_reg_no} used {day_count} days. New EL balance: {new_el}")


    # Update request status

    cursor.execute(
        """
        UPDATE leave_requests 
        SET status = 'approved', processed_by = %s, processed_date = CURRENT_TIMESTAMP, admin_comment = %s, is_read_by_admin = 1
        WHERE id = %s
    """,
        (admin_user["name"], comment, request_id),
    )

    # Create audit log entry
    cursor.execute(
        """
        INSERT INTO leave_request_audit_log 
        (leave_request_id, action, performed_by, performed_by_name, previous_status, new_status, comments)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """,
        (
            request_id,
            "APPROVED",
            admin_user["reg_no"],
            admin_user["name"],
            "pending",
            "approved",
            comment,
        ),
    )

    log_audit_event(
        "LEAVE_APPROVED",
        admin_user["reg_no"],
        True,
        f"Leave request {request_id} approved - Type: {leave_type}, Status: {attendance_status}, Tag: {status_tag}",
    )

    # Commit the transaction to ensure daily_attendance_status is saved
    conn.commit()

    return {
        "success": True,
        "message": "Leave request approved successfully",
        "attendance_status": attendance_status,
        "status_tag": status_tag,
        "leave_type_display": leave_type,
        "synced_dates": synced_dates,
    }


@app.post("/admin/leave/request/{request_id}/reject")
async def admin_reject_leave_request(request: Request, request_id: int):
    """Reject a leave request - Admin only"""
    admin_user = verify_admin_token(request)

    try:
        data = await request.json()
        comment = data.get("comment", "")
    except:
        comment = ""

    if not comment:
        raise HTTPException(
            status_code=400, detail="Rejection reason/comment is required"
        )

    # Check if request exists
    cursor.execute("SELECT status FROM leave_requests WHERE id = %s", (request_id,))
    row = cursor.fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Leave request not found")

    if row[0] != "pending":
        raise HTTPException(
            status_code=400, detail="Only pending requests can be rejected"
        )

    # Update request status
    cursor.execute(
        """
        UPDATE leave_requests 
        SET status = 'rejected', processed_by = %s, processed_date = CURRENT_TIMESTAMP, admin_comment = %s
        WHERE id = %s
    """,
        (admin_user["name"], comment, request_id),
    )

    # Create audit log entry
    cursor.execute(
        """
        INSERT INTO leave_request_audit_log 
        (leave_request_id, action, performed_by, performed_by_name, previous_status, new_status, comments)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """,
        (
            request_id,
            "REJECTED",
            admin_user["reg_no"],
            admin_user["name"],
            "pending",
            "rejected",
            comment,
        ),
    )

    log_audit_event(
        "LEAVE_REJECTED",
        admin_user["reg_no"],
        True,
        f"Leave request {request_id} rejected",
    )

    return {
        "success": True,
        "message": "Leave request rejected successfully",
    }


@app.get("/admin/leave/export")
async def admin_export_leave_requests(
    request: Request,
    status: str = None,
    dept: str = None,
    start_date: str = None,
    end_date: str = None,
):
    """Export leave requests as CSV - Admin only"""
    verify_admin_token(request)

    # Build query
    query = """
        SELECT lr.user_reg_no, lr.user_name, lr.dept, lr.leave_type, 
               lr.start_date, lr.end_date, lr.reason, lr.submission_date, 
               lr.status, lr.processed_by, lr.processed_date, lr.admin_comment
        FROM leave_requests lr
        WHERE 1=1
    """
    params = []

    if status:
        query += " AND lr.status = ?"
        params.append(status)

    if dept:
        query += " AND lr.dept = ?"
        params.append(dept)

    if start_date:
        query += " AND lr.submission_date >= ?"
        params.append(start_date)

    if end_date:
        query += " AND lr.submission_date <= ?"
        params.append(end_date)

    query += " ORDER BY lr.submission_date DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()

    # Create CSV content
    csv_content = "Reg No,Name,Department,Leave Type,Start Date,End Date,Reason,Submission Date,Status,Processed By,Processed Date,Admin Comment\n"

    for row in rows:
        # Escape quotes in reason and comments
        reason = str(row[6]).replace('"', '""')
        admin_comment = str(row[11] or "").replace('"', '""')

        csv_content += f'"{row[0]}","{row[1]}","{row[2]}","{row[3]}","{row[4]}","{row[5]}","{reason}","{row[7]}","{row[8]}","{row[9] or ""}","{row[10] or ""}","{admin_comment}"\n'

    return {
        "success": True,
        "data": csv_content,
        "count": len(rows),
        "filename": f"leave_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
    }


@app.put("/admin/leave/request/{request_id}/mark-read")
async def admin_mark_leave_request_read(request: Request, request_id: int):
    """Mark a leave request as read - Admin only"""
    verify_admin_token(request)

    cursor.execute(
        "UPDATE leave_requests SET is_read_by_admin = 1 WHERE id = %s", (request_id,)
    )
    conn.commit()

    return {"success": True, "message": "Request marked as read"}


@app.post("/admin/attendance/sync-daily-status")
async def admin_sync_daily_attendance_status(request: Request, date: str = None):
    """Sync attendance records into daily_attendance_status table (admin only).
    This populates daily_attendance_status from the attendance table for any
    records that were marked before the sync was automatic.
    """
    verify_admin_token(request)

    try:
        _hd_settings = _get_half_day_settings()
        # 1) Get all scans from attendance table
        if date:
            cursor.execute(
                """
                SELECT reg_no, name, dept, timestamp, status
                FROM attendance
                WHERE timestamp::date = ?
                """,
                (date,),
            )
        else:
            cursor.execute(
                """
                SELECT reg_no, name, dept, timestamp, status
                FROM attendance
                """
            )
        att_scans = cursor.fetchall()

        # 2) Get all scans from other_staff_attendance table
        if date:
            cursor.execute(
                """
                SELECT reg_no, name, dept, timestamp, status
                FROM other_staff_attendance
                WHERE timestamp::date = ?
                """,
                (date,),
            )
        else:
            cursor.execute(
                """
                SELECT reg_no, name, dept, timestamp, status
                FROM other_staff_attendance
                """
            )
        osa_scans = cursor.fetchall()

        all_scans = att_scans + osa_scans
        
        # Process unique (reg_no, date) pairs
        processed_pairs = set()
        synced = 0
        
        for reg_no, name, dept, timestamp, status in all_scans:
            if not timestamp:
                continue
            if hasattr(timestamp, "date"):
                scan_date = timestamp.date()
                scan_time_str = timestamp.time().strftime("%H:%M:%S")
            else:
                ts_str = str(timestamp)
                date_str = ts_str.split("T")[0] if "T" in ts_str else ts_str.split(" ")[0]
                scan_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                scan_time_str = ts_str.split("T")[1] if "T" in ts_str else ts_str.split(" ")[1] if " " in ts_str else "09:00:00"

            scan_date_str = str(scan_date)
            
            # Determine effective half based on the configured FN/AN boundary
            try:
                fh_end_str = _hd_settings.get("first_half_end", "13:00")
                fh_end_time = datetime.strptime(fh_end_str, "%H:%M").time()
                
                scan_time = datetime.strptime(scan_time_str, "%H:%M:%S").time()
                effective_half = "first_half" if scan_time <= fh_end_time else "second_half"
            except Exception:
                effective_half = "first_half"

            # Insert/update in separate morning/evening table
            if effective_half == "first_half":
                cursor.execute(
                    """
                    INSERT INTO morning_attendance (reg_no, name, dept, date, in_time, status, attendance_value, marked_by, marked_at)
                    VALUES (?, ?, ?, ?, ?, 'Present', 0.5, 'Sync', CURRENT_TIMESTAMP)
                    ON CONFLICT (reg_no, date) DO UPDATE SET
                        in_time = COALESCE(morning_attendance.in_time, EXCLUDED.in_time),
                        status = 'Present',
                        attendance_value = 0.5,
                        marked_by = 'Sync',
                        marked_at = CURRENT_TIMESTAMP
                    """,
                    (reg_no, name, dept, scan_date_str, scan_time_str),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO evening_attendance (reg_no, name, dept, date, in_time, status, attendance_value, marked_by, marked_at)
                    VALUES (?, ?, ?, ?, ?, 'Present', 0.5, 'Sync', CURRENT_TIMESTAMP)
                    ON CONFLICT (reg_no, date) DO UPDATE SET
                        in_time = COALESCE(evening_attendance.in_time, EXCLUDED.in_time),
                        status = 'Present',
                        attendance_value = 0.5,
                        marked_by = 'Sync',
                        marked_at = CURRENT_TIMESTAMP
                    """,
                    (reg_no, name, dept, scan_date_str, scan_time_str),
                )
            
            processed_pairs.add((reg_no, scan_date_str))

        # 3) Synchronize aggregated records into daily_attendance_status
        for r_no, d_str in processed_pairs:
            # Query current state from both session tables
            cursor.execute("SELECT status, attendance_value, in_time FROM morning_attendance WHERE reg_no = ? AND date = ?", (r_no, d_str))
            _m_row = cursor.fetchone()
            cursor.execute("SELECT status, attendance_value, in_time FROM evening_attendance WHERE reg_no = ? AND date = ?", (r_no, d_str))
            _e_row = cursor.fetchone()

            new_fh = _m_row[0] if _m_row else None
            new_sh = _e_row[0] if _e_row else None
            m_val = float(_m_row[1]) if _m_row and _m_row[1] is not None else 0.0
            e_val = float(_e_row[1]) if _e_row and _e_row[1] is not None else 0.0
            attendance_value = m_val + e_val

            new_status = _compute_daily_status(new_fh, new_sh, target_date=d_str)
            fh_time = _m_row[2] if _m_row else None
            sh_time = _e_row[2] if _e_row else None

            # Get user info
            cursor.execute("SELECT name, dept FROM users WHERE reg_no = ? UNION ALL SELECT name, dept FROM other_staff WHERE reg_no = ?", (r_no, r_no))
            user_row = cursor.fetchone()
            u_name = user_row[0] if user_row else "Unknown"
            u_dept = user_row[1] if user_row else "Unknown"

            cursor.execute(
                """
                INSERT INTO daily_attendance_status
                (reg_no, name, dept, date, status,
                 first_half_status, second_half_status,
                 first_half_in_time, second_half_in_time,
                 attendance_value, marked_by, marked_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Sync', CURRENT_TIMESTAMP)
                ON CONFLICT (reg_no, date) DO UPDATE SET
                    first_half_status = EXCLUDED.first_half_status,
                    second_half_status = EXCLUDED.second_half_status,
                    first_half_in_time = COALESCE(EXCLUDED.first_half_in_time, daily_attendance_status.first_half_in_time),
                    second_half_in_time = COALESCE(EXCLUDED.second_half_in_time, daily_attendance_status.second_half_in_time),
                    status = EXCLUDED.status,
                    attendance_value = EXCLUDED.attendance_value,
                    marked_by = 'Sync',
                    marked_at = CURRENT_TIMESTAMP
                """,
                (r_no, u_name, u_dept, d_str, new_status, new_fh, new_sh, fh_time, sh_time, attendance_value)
            )
            synced += 1

        conn.commit()

        return {
            "success": True,
            "message": f"Successfully processed scans and synced {synced} daily status records.",
            "synced_count": synced,
        }
    except Exception as e:
        print(f"Error syncing daily attendance status: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to sync daily status: {str(e)}")


@app.get("/admin/attendance/daily-status")
async def admin_get_daily_attendance_status(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    dept: str = None,
    status: str = None,
):
    """Get daily attendance status including leave/OD/absent status - Admin only.
    Generates Absent records for users who have no attendance or leave record.
    """
    verify_admin_token(request)

    try:
        # Determine date range
        if start_date and end_date:
            start = start_date
            end = end_date
        elif start_date:
            start = end = start_date
        elif end_date:
            start = end = end_date
        else:
            from datetime import date as date_type

            today = date_type.today().isoformat()
            start = end = today
        # Clamp to academic year
        start, end = _clamp_to_academic_year(start, end)
        # Expand to every date in range
        from datetime import date as _dt, timedelta as _td
        _s = _dt.fromisoformat(start)
        _e = _dt.fromisoformat(end)
        dates = [(_s + _td(days=i)).isoformat() for i in range((_e - _s).days + 1)]

        # Get all users (staff + hod + admin)
        user_query = "SELECT reg_no, name, dept, role FROM users WHERE role IN ('staff', 'hod', 'admin')"
        user_params = []
        if dept:
            user_query += " AND dept = %s"
            user_params.append(dept)
        cursor.execute(user_query, user_params)
        users = cursor.fetchall()

        # Get all other_staff
        os_query = "SELECT reg_no, name, dept, role FROM other_staff"
        os_params = []
        if dept:
            os_query += " AND dept = %s"
            os_params.append(dept)
        cursor.execute(os_query, os_params)
        other_staff = cursor.fetchall()

        all_users = []
        for u in users:
            all_users.append(
                {
                    "reg_no": u[0],
                    "name": u[1],
                    "dept": u[2],
                    "role": u[3],
                    "source": "users",
                }
            )
        for u in other_staff:
            all_users.append(
                {
                    "reg_no": u[0],
                    "name": u[1],
                    "dept": u[2],
                    "role": u[3],
                    "source": "other_staff",
                }
            )

        # Get existing daily_attendance_status records for the date range
        status_query = """
            SELECT reg_no, name, dept, date, status, leave_request_id, leave_type, marked_by, marked_at, absent_reason,
                   first_half_status, second_half_status, first_half_in_time, first_half_out_time, second_half_in_time, second_half_out_time, attendance_value
            FROM daily_attendance_status
            WHERE date >= ? AND date <= ?
        """
        status_params = [start, end]
        if dept:
            status_query += " AND dept = ?"
            status_params.append(dept)
        if status:
            status_query += " AND status = ?"
            status_params.append(status)
        status_query += " ORDER BY date DESC, name ASC"
        cursor.execute(status_query, status_params)
        status_rows = cursor.fetchall()

        # Build a lookup: (reg_no, date) -> status record
        status_map = {}
        for row in status_rows:
            key = (row[0], str(row[3]))
            status_map[key] = {
                "reg_no": row[0],
                "name": row[1],
                "dept": row[2],
                "date": _date_str(row[3]),
                "status": row[4],
                "leave_request_id": row[5],
                "leave_type": row[6],
                "marked_by": row[7],
                "marked_at": _ts(row[8]),
                "absent_reason": row[9] if len(row) > 9 else None,
                "first_half_status": row[10] if len(row) > 10 else None,
                "second_half_status": row[11] if len(row) > 11 else None,
                "first_half_in_time": str(row[12]) if len(row) > 12 and row[12] else None,
                "first_half_out_time": str(row[13]) if len(row) > 13 and row[13] else None,
                "second_half_in_time": str(row[14]) if len(row) > 14 and row[14] else None,
                "second_half_out_time": str(row[15]) if len(row) > 15 and row[15] else None,
                "attendance_value": float(row[16]) if len(row) > 16 and row[16] is not None else None,
            }


        # Generate results: only return records that exist (no synthetic Absent)
        results = list(status_map.values())

        # Sort by date desc, name asc (stable sort)
        results.sort(key=lambda x: (x["date"], x["name"]))
        results.sort(key=lambda x: x["date"], reverse=True)

        return {"success": True, "data": results, "count": len(results)}
    except Exception as e:
        print(f"Error fetching daily attendance status: {e}")
        raise HTTPException(
            status_code=500, detail="Failed to fetch daily attendance status"
        )


@app.get("/admin/attendance/daily-status/export")
async def admin_export_daily_attendance_status(
    request: Request, start_date: str = None, end_date: str = None, dept: str = None
):
    """Export daily attendance status including leave/OD as CSV - Admin only"""
    verify_admin_token(request)

    try:
        query = """
            SELECT reg_no, name, dept, date, status, leave_request_id, leave_type, marked_by, marked_at
            FROM daily_attendance_status
            WHERE 1=1
        """
        params = []

        if start_date:
            query += " AND date >= ?"
            params.append(start_date)

        if end_date:
            query += " AND date <= ?"
            params.append(end_date)

        if dept:
            query += " AND dept = ?"
            params.append(dept)

        query += " ORDER BY date DESC, name ASC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Create CSV content
        csv_content = "Reg No,Name,Department,Date,Status,Leave Request ID,Leave Type,Marked By,Marked At\n"

        for row in rows:
            csv_content += f'"{row[0]}","{row[1]}","{row[2]}","{row[3]}","{row[4]}","{row[5] or ""}","{row[6] or ""}","{row[7] or ""}","{row[8] or ""}"\n'

        return {
            "success": True,
            "data": csv_content,
            "count": len(rows),
            "filename": f"daily_attendance_status_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        }
    except Exception as e:
        print(f"Error exporting daily attendance status: {e}")
        raise HTTPException(
            status_code=500, detail="Failed to export daily attendance status"
        )


# -------------------------------------------------
# ADMIN NOTIFICATIONS ENDPOINTS
# -------------------------------------------------


@app.get("/admin/notifications")
async def admin_get_notifications(request: Request, unread_only: bool = False):
    """Get admin notifications - Admin only"""
    verify_admin_token(request)

    query = "SELECT id, notification_type, title, message, related_id, is_read, created_at FROM admin_notifications WHERE created_for = 'admin'"

    if unread_only:
        query += " AND is_read = 0"

    query += " ORDER BY created_at DESC LIMIT 100"

    cursor.execute(query)
    rows = cursor.fetchall()

    notifications = []
    for row in rows:
        notifications.append(
            {
                "id": row[0],
                "type": row[1],
                "title": row[2],
                "message": row[3],
                "related_id": row[4],
                "is_read": bool(row[5]),
                "created_at": row[6],
            }
        )

    # Get unread count
    cursor.execute(
        "SELECT COUNT(*) FROM admin_notifications WHERE created_for = 'admin' AND is_read = 0"
    )
    unread_count = cursor.fetchone()[0]

    return {
        "success": True,
        "notifications": notifications,
        "unread_count": unread_count,
    }


@app.put("/admin/notifications/{notification_id}/read")
async def admin_mark_notification_read(request: Request, notification_id: int):
    """Mark a notification as read - Admin only"""
    verify_admin_token(request)

    cursor.execute(
        "UPDATE admin_notifications SET is_read = 1 WHERE id = ? AND created_for = 'admin'",
        (notification_id,),
    )
    conn.commit()

    return {"success": True, "message": "Notification marked as read"}


@app.put("/admin/notifications/read-all")
async def admin_mark_all_notifications_read(request: Request):
    """Mark all notifications as read - Admin only"""
    verify_admin_token(request)

    cursor.execute(
        "UPDATE admin_notifications SET is_read = 1 WHERE created_for = 'admin'"
    )
    conn.commit()

    return {"success": True, "message": "All notifications marked as read"}


# -------------------------------------------------
# INIT AUDIT
# -------------------------------------------------
log_audit_event("SYSTEM_START", None, True, "Face Attendance System started")

# -------------------------------------------------
# STAFF PANEL ENDPOINTS
# -------------------------------------------------


@app.post("/staff/login")
async def staff_login(request: Request):
    """Staff login endpoint"""
    try:
        data = await request.json()
        username = data.get("username")
        password = data.get("password")

        if not username or not password:
            raise HTTPException(status_code=400, detail="Missing credentials")

        user = get_user_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check if user is suspended
        if is_user_suspended(username, is_other_staff=False):
            raise HTTPException(
                status_code=403,
                detail="Account suspended. Please contact the administrator.",
            )

        # Check role
        if user[6] != "staff":
            raise HTTPException(status_code=403, detail="Staff access required")

        device_id = data.get("device_id")
        cursor.execute("UPDATE users SET current_device_id = ? WHERE username = ?", (device_id, user[1]))
        conn.commit()

        # Create token
        import base64

        token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode(
            "utf-8"
        )

        return {
            "message": "Staff login successful",
            "token": token,
            "user": {
                "id": user[0],
                "username": user[1],
                "regNo": user[3],
                "name": user[4],
                "dept": user[5],
                "role": user[6],
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Staff login error: {e}")
        raise HTTPException(status_code=500, detail="Login failed")


def verify_staff_token(request: Request) -> dict:
    """Verify staff authentication token"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authorization header"
        )

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")

        user = get_user_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check if user is suspended
        if is_user_suspended(username, is_other_staff=False):
            raise HTTPException(status_code=403, detail="Account suspended. Access denied.")

        # Check if user is staff
        if user[6] != "staff":
            raise HTTPException(status_code=403, detail="Staff access required")

        return {
            "id": user[0],
            "username": user[1],
            "reg_no": user[3],
            "name": user[4],
            "dept": user[5],
            "role": user[6],
        }
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid token")


def verify_user_token(request: Request) -> dict:
    """Verify general user token for any role (including custom roles like principal, vice_chancellor, etc.)"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authorization header"
        )

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")

        # First, check the users table (for regular staff, HOD, admin)
        user = get_user_by_username(username)
        if user:
            if not verify_password(password, user[2]):
                raise HTTPException(status_code=401, detail="Invalid credentials")

            # Check if user is suspended
            if is_user_suspended(username, is_other_staff=False):
                raise HTTPException(status_code=403, detail="Account suspended. Access denied.")

            # Return user info from users table
            return {
                "id": user[0],
                "username": user[1],
                "reg_no": user[3],
                "name": user[4],
                "dept": user[5],
                "role": user[6],
            }

        # If not found in users table, check other_staff table (for principal, placement_staff, lab_technician, etc.)
        other_staff = get_other_staff_by_username(username)
        if other_staff:
            if not verify_password(password, other_staff[2]):
                raise HTTPException(status_code=401, detail="Invalid credentials")

            # Check if user is suspended
            if is_user_suspended(username, is_other_staff=True):
                raise HTTPException(status_code=403, detail="Account suspended. Access denied.")

            # Return user info from other_staff table
            return {
                "id": other_staff[0],
                "username": other_staff[1],
                "reg_no": other_staff[3],
                "name": other_staff[4],
                "dept": other_staff[7],  # dept is at index 7 in other_staff
                "role": other_staff[6],  # role is at index 6 in other_staff
            }

        # User not found in either table
        raise HTTPException(status_code=401, detail="Invalid credentials")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid token")


# -------------------------------------------------
# OTHER STAFF AUTHENTICATION
# -------------------------------------------------
@app.post("/other_staff/login")
async def other_staff_login(request: Request):
    """Login endpoint for other staff (principal, placement_staff, lab_technician, system_admin, office_staff)"""
    try:
        data = await request.json()
        username = data.get("username")
        password = data.get("password")

        if not username or not password:
            raise HTTPException(status_code=400, detail="Missing credentials")

        # Get user by username
        user = get_other_staff_by_username(username)

        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Verify password
        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check if user is suspended
        if is_user_suspended(username, is_other_staff=True):
            raise HTTPException(
                status_code=403,
                detail="Account suspended. Please contact the administrator.",
            )

        device_id = data.get("device_id")
        cursor.execute("UPDATE other_staff SET current_device_id = ? WHERE username = ?", (device_id, user[1]))
        conn.commit()

        # Create token
        import base64

        token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode(
            "utf-8"
        )

        # Check if face is registered (embedding is at index 8)
        face_registered = user[8] is not None if len(user) > 8 else False

        return {
            "message": "Login successful",
            "token": token,
            "user": {
                "id": user[0],
                "username": user[1],
                "regNo": user[3],
                "name": user[4],
                "dob": user[5],
                "role": user[6],
                "dept": user[7] if len(user) > 7 else None,
                "face_registered": face_registered,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Other staff login error: {e}")
        raise HTTPException(status_code=500, detail="Login failed")


def verify_other_staff_token(request: Request) -> dict:
    """Verify other_staff authentication token"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authorization header"
        )

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":")

        user = get_other_staff_by_username(username)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        if not verify_password(password, user[2]):
            raise HTTPException(status_code=401, detail="Invalid credentials")

        # Check if user is suspended
        if is_user_suspended(username, is_other_staff=True):
            raise HTTPException(status_code=403, detail="Account suspended. Access denied.")

        # Check if face registered (embedding is at index 8)
        face_registered = user[8] is not None if len(user) > 8 else False

        return {
            "id": user[0],
            "username": user[1],
            "reg_no": user[3],
            "name": user[4],
            "dob": user[5],
            "role": user[6],
            "dept": user[7] if len(user) > 7 else None,
            "face_registered": face_registered,
        }
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid token")


@app.get("/staff/dashboard")
async def staff_dashboard(request: Request):
    """Get staff dashboard statistics - includes face scan + OD"""
    staff_user = verify_staff_token(request)
    reg_no = staff_user["reg_no"]
    dept = staff_user["dept"]

    # Get face scan attendance count
    cursor.execute("SELECT COUNT(*) FROM attendance WHERE reg_no = ?", (reg_no,))
    total_face_scan = cursor.fetchone()[0]

    # Get today's face scan
    cursor.execute(
        "SELECT COUNT(*) FROM attendance WHERE reg_no = %s AND timestamp::date = CURRENT_DATE",
        (reg_no,),
    )
    today_face_scan = cursor.fetchone()[0]

    # Get today's OD from daily_attendance_status (separate counts)
    cursor.execute(
        """
        SELECT leave_type, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE reg_no = %s AND date::date = CURRENT_DATE AND status = 'Present' 
        AND leave_type IN ('od', 'earned', 'casual')
        GROUP BY leave_type
    """,
        (reg_no,),
    )
    od_breakdown = cursor.fetchall()
    od_count = sum(row[1] for row in od_breakdown)
    earned_count = sum(row[1] for row in od_breakdown if row[0] == "earned")
    casual_count = sum(row[1] for row in od_breakdown if row[0] == "casual")
    od_only_count = sum(row[1] for row in od_breakdown if row[0] == "od")

    # Total attendance = face scan + OD (od + earned + casual)
    total_attendance = total_face_scan + od_count
    today_attendance = today_face_scan + od_count

    # Get recent attendance (face scan + OD records with approval date)
    cursor.execute(
        """
        SELECT id, reg_no, name, dept, class_div, timestamp, 'face_scan' as source, status
        FROM attendance 
        WHERE reg_no = %s 
        ORDER BY id DESC 
        LIMIT 8
    """,
        (reg_no,),
    )
    face_scan_records = cursor.fetchall()

    # Get OD records with approval date
    cursor.execute(
        """
        SELECT das.id, das.reg_no, das.name, das.dept, das.date as timestamp, 'od' as source, 
               das.leave_type, lr.processed_date
        FROM daily_attendance_status das
        LEFT JOIN leave_requests lr ON das.leave_request_id = lr.id
        WHERE das.reg_no = %s AND das.date::date = CURRENT_DATE AND das.status = 'Present' 
        AND das.leave_type IN ('od', 'earned', 'casual')
        ORDER BY das.date DESC
        LIMIT 5
    """,
        (reg_no,),
    )
    od_records = cursor.fetchall()

    # Combine and sort
    combined = list(face_scan_records) + list(od_records)
    combined.sort(key=lambda x: str(x[5]) if len(x) > 5 and x[5] else "", reverse=True)
    recent_attendance = combined[:10]

    # Build recent_attendance list with approval_date
    recent_attendance_list = []
    for row in recent_attendance:
        if len(row) >= 9: # Has status and leave_type
            # OD record with approval date
            recent_attendance_list.append(
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "class_div": "",
                    "timestamp": str(row[4]) if row[4] else "",
                    "source": row[5],
                    "leave_type": row[6],
                    "approval_date": _ts(row[7]) if len(row) > 7 and row[7] else None,
                }
            )
        else:
            # Face scan record
            raw_status = row[7] if len(row) > 7 else None
            recent_attendance_list.append(
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "class_div": row[4] or "",
                    "timestamp": _ts(row[5])
                    if len(row) > 5 and row[5]
                    else str(row[5])
                    if len(row) > 5
                    else "",
                    "source": row[6] if len(row) > 6 else "face_scan",
                    "status": _format_scan_status(raw_status, row[5]) if raw_status else "Present",
                    "punch_type": raw_status or "check_in",
                    "leave_type": None,
                }
            )

    # --- Pie chart: personal today status + half-split + historical breakdown ---
    cursor.execute(
        """
        SELECT status, first_half_status, second_half_status
        FROM daily_attendance_status
        WHERE reg_no = %s AND date::date = CURRENT_DATE
        LIMIT 1
    """,
        (reg_no,),
    )
    today_das_row = cursor.fetchone()
    today_status_str    = today_das_row[0] if today_das_row else None
    today_first_half    = today_das_row[1] if today_das_row else None
    today_second_half   = today_das_row[2] if today_das_row else None

    # Historical full-day vs half-day counts from daily_attendance_status
    cursor.execute(
        """
        SELECT status, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE reg_no = %s
        GROUP BY status
    """,
        (reg_no,),
    )
    hist_breakdown = {row[0]: row[1] for row in cursor.fetchall()}
    hist_full_day_count = hist_breakdown.get("Present", 0)
    hist_half_day_count = hist_breakdown.get("Half Day", 0)
    hist_absent_count   = hist_breakdown.get("Absent", 0)
    hist_leave_count    = hist_breakdown.get("Leave", 0)

    return JSONResponse(
        content={
            "stats": {
                "department": dept,
                "total_attendance": total_attendance,
                "today_attendance": today_attendance,
                "today_face_scan": today_face_scan,
                "today_od": od_count,
                "today_earned": earned_count,
                "today_casual": casual_count,
                # Present = face scan + OD (treated as present)
                "today_present": today_face_scan + od_count,
                "today_holiday_reason": _academic_status_for_date(datetime.now().strftime("%Y-%m-%d"))[1],
                # Absent = 0 if holiday or attendance present, else 1
                "today_absent": 0 if (_academic_status_for_date(datetime.now().strftime("%Y-%m-%d"))[2] or (today_face_scan + od_count) > 0) else 1,
                # Pie chart fields
                "today_status": today_status_str,
                "today_first_half": today_first_half,
                "today_second_half": today_second_half,
                "hist_full_day_count": hist_full_day_count,
                "hist_half_day_count": hist_half_day_count,
                "hist_absent_count": hist_absent_count,
                "hist_leave_count": hist_leave_count,
            },
            "recent_attendance": recent_attendance_list,
            "staff_user": staff_user,
        },
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


def verify_any_user_token(request: Request) -> dict:
    """Verify authorization header and return user info and role for any valid user (admin, HOD, staff, other_staff)"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authorization header"
        )
    token = auth_header.replace("Bearer ", "")
    try:
        import base64
        decoded = base64.b64decode(token).decode("utf-8")
        parts = decoded.split(":")
        if len(parts) >= 2:
            username = parts[0]
            password = parts[1]
            # Check in regular users (admin, HOD, staff)
            user = get_user_by_username(username)
            if user and verify_password(password, user[2]):
                if is_user_suspended(username, is_other_staff=False):
                    raise HTTPException(status_code=403, detail="Account suspended. Access denied.")
                return {
                    "id": user[0],
                    "username": user[1],
                    "reg_no": user[3],
                    "name": user[4],
                    "dept": user[5],
                    "role": user[6],
                    "source": "users"
                }
            # Check in other staff
            cursor.execute("SELECT id, username, password_hash, reg_no, name, role, dept, suspended FROM other_staff WHERE username = ?", (username,))
            os_row = cursor.fetchone()
            if os_row and verify_password(password, os_row[2]):
                if os_row[7]:
                    raise HTTPException(status_code=403, detail="Account suspended. Access denied.")
                return {
                    "id": os_row[0],
                    "username": os_row[1],
                    "reg_no": os_row[3],
                    "name": os_row[4],
                    "role": os_row[5],
                    "dept": os_row[6],
                    "source": "other_staff"
                }
    except Exception as e:
        print(f"Token verification error: {e}")
    raise HTTPException(status_code=401, detail="Invalid credentials")


@app.get("/api/attendance/personal")
async def get_personal_attendance_log(
    request: Request, start_date: str = None, end_date: str = None
):
    """Get personal attendance log with date range filter for any logged in user"""
    user_info = verify_any_user_token(request)
    reg_no = user_info["reg_no"]
    is_other_staff = (user_info["source"] == "other_staff")

    try:
        now = datetime.now()
        acad_start, acad_end = _get_academic_date_range()
        if not start_date:
            start_date = acad_start or now.strftime("%Y-%m-01")
        if not end_date:
            end_date = acad_end or now.strftime("%Y-%m-%d")
        start_date, end_date = _clamp_to_academic_year(start_date, end_date)

        table = "other_staff_attendance" if is_other_staff else "attendance"
        
        # 1) Get face scan logs
        cursor.execute(
            f"""
            SELECT id, reg_no, name, dept, timestamp, status
            FROM {table}
            WHERE reg_no = ? AND timestamp::date >= ? AND timestamp::date <= ?
            ORDER BY timestamp DESC
            """,
            (reg_no, start_date, end_date),
        )
        rows = cursor.fetchall()

        # 2) Get daily_attendance_status (OD, Leave, Absent)
        cursor.execute(
            """
            SELECT date, status, leave_type, absent_reason,
                   first_half_status, second_half_status, first_half_in_time, first_half_out_time, second_half_in_time, second_half_out_time, attendance_value
            FROM daily_attendance_status
            WHERE reg_no = ? AND date::date >= ? AND date::date <= ?
            """,
            (reg_no, start_date, end_date),
        )
        status_rows = cursor.fetchall()
        status_map = {}
        for row in status_rows:
            date_str = str(row[0])
            status_map[date_str] = {
                "status": row[1],
                "leave_type": row[2],
                "absent_reason": row[3],
                "first_half_status": row[4],
                "second_half_status": row[5],
                "first_half_in_time": str(row[6]) if row[6] else None,
                "first_half_out_time": str(row[7]) if row[7] else None,
                "second_half_in_time": str(row[8]) if row[8] else None,
                "second_half_out_time": str(row[9]) if row[9] else None,
                "attendance_value": float(row[10]) if row[10] is not None else None,
            }

        scan_dates = set()
        attendance_list = []
        for row in rows:
            ts = row[4]
            date_str = None
            if ts:
                if hasattr(ts, "strftime"):
                    date_str = ts.strftime("%Y-%m-%d")
                else:
                    ts_str = str(ts)
                    date_str = ts_str.split("T")[0] if "T" in ts_str else ts_str.split(" ")[0]
                if date_str:
                    scan_dates.add(date_str)
            
            raw_punch_type = row[5] if row[5] else "check_in"  # check_in or check_out
            day_status = "Present"  # default daily status
            absent_reason = None
            day_leave_type = None
            day_first_half_status = None
            day_second_half_status = None
            day_first_half_in_time = None
            day_first_half_out_time = None
            day_second_half_in_time = None
            day_second_half_out_time = None
            day_attendance_value = None

            if date_str in status_map:
                daily_info = status_map[date_str]
                day_status = daily_info["status"] or "Present"
                day_leave_type = daily_info["leave_type"]
                day_first_half_status = daily_info["first_half_status"]
                day_second_half_status = daily_info["second_half_status"]
                day_first_half_in_time = daily_info["first_half_in_time"]
                day_first_half_out_time = daily_info["first_half_out_time"]
                day_second_half_in_time = daily_info["second_half_in_time"]
                day_second_half_out_time = daily_info["second_half_out_time"]
                day_attendance_value = daily_info["attendance_value"]
                if daily_info["status"] == "Absent":
                    absent_reason = daily_info["absent_reason"]
            
            attendance_list.append({
                "id": row[0],
                "reg_no": row[1],
                "name": row[2],
                "dept": row[3],
                "timestamp": _ts(ts),
                "status": day_status,
                "punch_type": raw_punch_type,
                "source": "face_scan",
                "absent_reason": absent_reason,
                "leave_type": day_leave_type,
                "first_half_status": day_first_half_status,
                "second_half_status": day_second_half_status,
                "first_half_in_time": day_first_half_in_time,
                "first_half_out_time": day_first_half_out_time,
                "second_half_in_time": day_second_half_in_time,
                "second_half_out_time": day_second_half_out_time,
                "attendance_value": day_attendance_value,
            })

        # Add status-only records for dates without face scans
        for date_str, daily_info in status_map.items():
            if date_str not in scan_dates:
                status = daily_info["status"]
                leave_type = daily_info["leave_type"]
                absent_reason = daily_info["absent_reason"]
                if status in ("Leave", "Absent") or (status == "Present" and leave_type in ("od", "earned", "casual")):
                    attendance_list.append({
                        "id": None,
                        "reg_no": reg_no,
                        "name": user_info["name"],
                        "dept": user_info["dept"],
                        "timestamp": date_str + "T00:00:00",
                        "status": status,
                        "source": "leave" if status == "Leave" else ("absent" if status == "Absent" else "od"),
                        "leave_type": leave_type,
                        "absent_reason": absent_reason,
                        "first_half_status": daily_info["first_half_status"],
                        "second_half_status": daily_info["second_half_status"],
                        "first_half_in_time": daily_info["first_half_in_time"],
                        "first_half_out_time": daily_info["first_half_out_time"],
                        "second_half_in_time": daily_info["second_half_in_time"],
                        "second_half_out_time": daily_info["second_half_out_time"],
                        "attendance_value": daily_info["attendance_value"],
                    })


        attendance_list.sort(key=lambda x: x["timestamp"], reverse=True)

        return {
            "success": True,
            "attendance": attendance_list,
            "count": len(attendance_list)
        }
    except Exception as e:
        print(f"Error fetching personal attendance log: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch personal attendance log")


@app.get("/staff/attendance")
async def staff_get_attendance(
    request: Request, date: str = None, start_date: str = None, end_date: str = None
):
    """Get attendance records for the specific staff member.
    Combines face-scan attendance with daily_attendance_status (OD/Leave).
    """
    staff_user = verify_staff_token(request)
    reg_no = staff_user["reg_no"]

    try:
        # Default to academic year if no dates provided
        now = datetime.now()
        acad_start, acad_end = _get_academic_date_range()
        if not start_date:
            start_date = acad_start or now.strftime("%Y-%m-01")
        if not end_date:
            end_date = acad_end or now.strftime("%Y-%m-%d")
        # Clamp to academic year even if explicit dates provided
        start_date, end_date = _clamp_to_academic_year(start_date, end_date)
        # If a specific date isn't provided, use the range

        # 1) Get face-scan attendance records
        if date:
            cursor.execute(
                """
                SELECT id, reg_no, name, dept, class_div, timestamp, status
                FROM attendance 
                WHERE reg_no = ? AND timestamp::date = ? 
                ORDER BY id DESC
            """,
                (reg_no, date),
            )
        else:
            cursor.execute(
                """
                SELECT id, reg_no, name, dept, class_div, timestamp, status
                FROM attendance 
                WHERE reg_no = ? AND timestamp::date >= ? AND timestamp::date <= ?
                ORDER BY id DESC
            """,
                (reg_no, start_date, end_date),
            )

        rows = cursor.fetchall()

        # Collect scan dates
        scan_dates = set()
        for row in rows:
            ts = row[5]
            if ts:
                if hasattr(ts, "strftime"):
                    scan_dates.add(ts.strftime("%Y-%m-%d"))
                else:
                    ts_str = str(ts)
                    if "T" in ts_str:
                        scan_dates.add(ts_str.split("T")[0])
                    elif " " in ts_str:
                        scan_dates.add(ts_str.split(" ")[0])

        # 2) Get daily_attendance_status (OD, Leave, etc.)
        if date:
            cursor.execute(
                """
                SELECT date, status, leave_type, absent_reason
                FROM daily_attendance_status
                WHERE reg_no = %s AND date::date = %s
            """,
                (reg_no, date),
            )
        else:
            cursor.execute(
                """
                SELECT date, status, leave_type, absent_reason
                FROM daily_attendance_status
                WHERE reg_no = %s AND date::date >= %s AND date::date <= %s
            """,
                (reg_no, start_date, end_date),
            )

        status_rows = cursor.fetchall()

        # Categorize status dates
        od_dates = set()
        leave_dates = set()
        earned_dates = set()
        casual_dates = set()
        absent_dates = {}

        for row in status_rows:
            date_str = str(row[0]) if row[0] else None
            status = row[1]
            leave_type = row[2]
            absent_reason = row[3] if len(row) > 3 else None

            if (
                date_str
                and status == "Present"
                and leave_type in ("od", "earned", "casual")
            ):
                if leave_type == "od":
                    od_dates.add(date_str)
                elif leave_type == "earned":
                    earned_dates.add(date_str)
                elif leave_type == "casual":
                    casual_dates.add(date_str)
            elif date_str and status == "Leave":
                leave_dates.add(date_str)
            elif date_str and status == "Absent":
                absent_dates[date_str] = absent_reason

        # Build attendance list including status records
        attendance_list = []
        for row in rows:
            ts = row[5]
            date_str = None
            if ts:
                if hasattr(ts, "strftime"):
                    date_str = ts.strftime("%Y-%m-%d")
                else:
                    ts_str = str(ts)
                    date_str = ts_str.split("T")[0] if "T" in ts_str else ts_str.split(" ")[0]
            
            day_status = row[6] if len(row) > 6 and row[6] else "check_in"
            absent_reason = None
            if date_str in absent_dates:
                day_status = "Absent"
                absent_reason = absent_dates[date_str]
            elif date_str in leave_dates:
                day_status = "Leave"

            attendance_list.append(
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "class_div": row[4] or "",
                    "timestamp": _ts(row[5]),
                    "status": day_status,
                    "source": "face_scan",
                    "absent_reason": absent_reason,
                }
            )

        # Add status records for dates not in face scans
        all_status_dates = od_dates | earned_dates | casual_dates | leave_dates | set(absent_dates.keys())
        for date_str in all_status_dates:
            if date_str not in scan_dates:
                # Find the status for this date
                for row in status_rows:
                    if str(row[0]) == date_str:
                        status = row[1]
                        leave_type = row[2]
                        absent_reason = row[3] if len(row) > 3 else None
                        break
                else:
                    status = "Unknown"
                    leave_type = None
                    absent_reason = None

                attendance_list.append(
                    {
                        "id": None,
                        "reg_no": reg_no,
                        "name": staff_user["name"],
                        "dept": staff_user["dept"],
                        "class_div": "",
                        "timestamp": date_str,
                        "source": "leave" if status == "Leave" else ("absent" if status == "Absent" else "od"),
                        "status": status,
                        "leave_type": leave_type,
                        "absent_reason": absent_reason,
                    }
                )

        # Calculate present days (Half Day counts as 0.5, Present/OD counts as 1.0)
        hd = _get_half_day_settings()
        present_days = 0.0
        if hd["half_day_enabled"]:
            cursor.execute(
                """
                SELECT SUM(COALESCE(attendance_value, 0.0))
                FROM daily_attendance_status
                WHERE reg_no = %s AND date::date >= %s AND date::date <= %s
                """,
                (reg_no, start_date, end_date),
            )
            val_sum = cursor.fetchone()[0]
            present_days = float(val_sum) if val_sum is not None else 0.0
        else:
            full_att_dates = _get_full_attendance_dates(reg_no, start_date, end_date)
            present_days = float(len(full_att_dates | od_dates | earned_dates | casual_dates))

        if date:
            # Single day - only count if not in future
            today = datetime.now().strftime("%Y-%m-%d")
            days_in_range = 1 if date <= today else 0
        else:
            # Calculate working days in range (capped to today, exclude holidays)
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = _cap_end_to_today(datetime.strptime(end_date, "%Y-%m-%d"))
            holiday_dates = _get_holiday_dates_in_range(start_date, end_dt.strftime("%Y-%m-%d"))
            working_days = 0
            d = start_dt
            while d <= end_dt:
                if d.weekday() < 5 and d.strftime("%Y-%m-%d") not in holiday_dates:
                    working_days += 1
                d += timedelta(days=1)
            days_in_range = working_days

        absent_days = max(0, days_in_range - present_days - len(leave_dates))

        return {
            "attendance": attendance_list,
            "count": len(attendance_list),
            "present_days": present_days,
            "absent_days": absent_days,
            "leave_days": len(leave_dates),
            "od_days": len(od_dates),
            "earned_days": len(earned_dates),
            "casual_days": len(casual_dates),
        }
    except Exception as e:
        print(f"Error fetching attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch attendance")


@app.get("/staff/attendance/{reg_no}")
async def staff_get_attendance_by_regno(request: Request, reg_no: str):
    """Get attendance records for a specific staff member by reg_no.
    Combines face-scan attendance with daily_attendance_status (OD/Leave).
    """
    staff_user = verify_staff_token(request)

    try:
        # Get date range from academic year settings, fallback to current month
        now = datetime.now()
        acad_start, acad_end = _get_academic_date_range()
        start_date = acad_start or now.strftime("%Y-%m-01")
        end_date = acad_end or now.strftime("%Y-%m-%d")
        start_date, end_date = _clamp_to_academic_year(start_date, end_date)

        # 1) Get face-scan attendance records
        cursor.execute(
            """
            SELECT id, reg_no, name, dept, class_div, timestamp 
            FROM attendance 
            WHERE reg_no = %s AND timestamp::date >= %s AND timestamp::date <= %s
            ORDER BY id DESC
            LIMIT 100
        """,
            (reg_no, start_date, end_date),
        )

        rows = cursor.fetchall()
        attendance_list = [
            {
                "id": row[0],
                "reg_no": row[1],
                "name": row[2],
                "dept": row[3],
                "class_div": row[4] or "",
                "timestamp": _ts(row[5]),
            }
            for row in rows
        ]

        # Collect scan dates
        scan_dates = set()
        for row in rows:
            ts = row[5]
            if ts:
                if hasattr(ts, "strftime"):
                    scan_dates.add(ts.strftime("%Y-%m-%d"))
                else:
                    ts_str = str(ts)
                    if "T" in ts_str:
                        scan_dates.add(ts_str.split("T")[0])
                    elif " " in ts_str:
                        scan_dates.add(ts_str.split(" ")[0])

        # 2) Get daily_attendance_status records (OD, Leave, etc.)
        cursor.execute(
            """
            SELECT date, status, leave_type
            FROM daily_attendance_status
            WHERE reg_no = %s AND date::date >= %s AND date::date <= %s
            ORDER BY date DESC
        """,
            (reg_no, start_date, end_date),
        )
        status_rows = cursor.fetchall()

        # Collect status dates by type
        od_dates = set()
        leave_dates = set()
        earned_dates = set()
        casual_dates = set()

        for row in status_rows:
            date_str = str(row[0]) if row[0] else None
            status = row[1]
            leave_type = row[2]

            if (
                date_str
                and status == "Present"
                and leave_type in ("od", "earned", "casual")
            ):
                if leave_type == "od":
                    od_dates.add(date_str)
                elif leave_type == "earned":
                    earned_dates.add(date_str)
                elif leave_type == "casual":
                    casual_dates.add(date_str)
            elif date_str and status == "Leave":
                leave_dates.add(date_str)

        # Calculate present days (Half Day counts as 0.5, Present/OD counts as 1.0)
        hd = _get_half_day_settings()
        present_days = 0.0
        if hd["half_day_enabled"]:
            # Query status values directly from daily_attendance_status
            cursor.execute(
                """
                SELECT date, status FROM daily_attendance_status
                WHERE reg_no = %s AND date >= %s AND date <= %s
            """,
                (reg_no, start_date, end_date),
            )
            for r in cursor.fetchall():
                status_val = r[1]
                if status_val == "Present":
                    present_days += 1.0
                elif status_val == "Half Day":
                    present_days += 0.5
        else:
            full_att_dates = _get_full_attendance_dates(reg_no, start_date, end_date)
            present_days = float(len(full_att_dates | od_dates | earned_dates | casual_dates))

        # Calculate absent days using calendar days in range (capped to today, exclude holidays)
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = _cap_end_to_today(datetime.strptime(end_date, "%Y-%m-%d"))
        holiday_dates = _get_holiday_dates_in_range(start_date, end_dt.strftime("%Y-%m-%d"))
        total_days = max(0, (end_dt - start_dt).days + 1 - len(holiday_dates))
        leave_days = len(leave_dates)
        absent_days = max(0, total_days - present_days - leave_days)

        return {
            "attendance": attendance_list,
            "count": len(rows),
            "present_days": present_days,
            "absent_days": absent_days,
            "leave_days": leave_days,
            "od_days": len(od_dates),
            "earned_days": len(earned_dates),
            "casual_days": len(casual_dates),
            "scan_dates": sorted(list(scan_dates)),
            "od_dates": sorted(list(od_dates)),
        }
    except Exception as e:
        print(f"Error fetching attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch attendance")


# -------------------------------------------------
# OTHER STAFF ENDPOINTS (principal, placement_staff, lab_technician, system_admin, office_staff)
# -------------------------------------------------
@app.get("/other_staff/dashboard")
async def other_staff_dashboard(request: Request):
    """Get other staff dashboard statistics - includes face scan + OD/Leave"""
    staff_user = verify_other_staff_token(request)
    reg_no = staff_user["reg_no"]

    # Get counts for this staff member
    cursor.execute(
        "SELECT COUNT(*) FROM other_staff WHERE role = ?", (staff_user["role"],)
    )
    role_count = cursor.fetchone()[0]

    # Get face scan attendance count
    cursor.execute(
        "SELECT COUNT(*) FROM other_staff_attendance WHERE reg_no = ?",
        (reg_no,),
    )
    total_face_scan = cursor.fetchone()[0]

    # Get today's face scan
    cursor.execute(
        "SELECT COUNT(*) FROM other_staff_attendance WHERE reg_no = %s AND timestamp::date = CURRENT_DATE",
        (reg_no,),
    )
    today_face_scan = cursor.fetchone()[0]

    # Get OD/Leave status from daily_attendance_status (separate counts)
    cursor.execute(
        """
        SELECT leave_type, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE reg_no = %s AND date::date = CURRENT_DATE AND status = 'Present' 
        AND leave_type IN ('od', 'earned', 'casual')
        GROUP BY leave_type
    """,
        (reg_no,),
    )
    od_breakdown = cursor.fetchall()
    od_count = sum(row[1] for row in od_breakdown)
    earned_count = sum(row[1] for row in od_breakdown if row[0] == "earned")
    casual_count = sum(row[1] for row in od_breakdown if row[0] == "casual")

    # Total attendance = face scan + OD
    total_attendance = total_face_scan + od_count
    today_attendance = today_face_scan + od_count

    # Get recent attendance (face scan + OD records)
    cursor.execute(
        """
        SELECT id, reg_no, name, dept, role, timestamp, 'face_scan' as source, status
        FROM other_staff_attendance 
        WHERE reg_no = %s 
        ORDER BY id DESC 
        LIMIT 8
    """,
        (reg_no,),
    )
    face_scan_records = cursor.fetchall()

    # Get OD records with approval date
    cursor.execute(
        """
        SELECT das.id, das.reg_no, das.name, das.dept, das.date as timestamp, 'od' as source, 
               das.leave_type, lr.processed_date
        FROM daily_attendance_status das
        LEFT JOIN leave_requests lr ON das.leave_request_id = lr.id
        WHERE das.reg_no = %s AND das.date::date = CURRENT_DATE AND das.status = 'Present' 
        AND das.leave_type IN ('od', 'earned', 'casual')
        ORDER BY das.date DESC
        LIMIT 5
    """,
        (reg_no,),
    )
    od_records = cursor.fetchall()

    # Combine
    combined = list(face_scan_records) + list(od_records)
    combined.sort(key=lambda x: str(x[5]) if len(x) > 5 and x[5] else "", reverse=True)
    recent_attendance = combined[:10]

    # Build recent_attendance list with approval_date
    recent_attendance_list = []
    for row in recent_attendance:
        if len(row) >= 9: # Has status and leave_type
            # OD record with approval date
            recent_attendance_list.append(
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "role": row[4],
                    "timestamp": str(row[4]) if row[4] else "",
                    "source": row[5],
                    "leave_type": row[6],
                    "approval_date": _ts(row[7]) if len(row) > 7 and row[7] else None,
                }
            )
        else:
            # Face scan record
            raw_status = row[7] if len(row) > 7 else None
            recent_attendance_list.append(
                {
                    "id": row[0],
                    "reg_no": row[1],
                    "name": row[2],
                    "dept": row[3],
                    "role": row[4],
                    "timestamp": _ts(row[5])
                    if len(row) > 5 and row[5]
                    else str(row[5])
                    if len(row) > 5
                    else "",
                    "source": row[6] if len(row) > 6 else "face_scan",
                    "status": _format_scan_status(raw_status, row[5]) if raw_status else "Present",
                    "punch_type": raw_status or "check_in",
                    "leave_type": None,
                }
            )

    # --- Pie chart: personal today status + half-split + historical breakdown ---
    cursor.execute(
        """
        SELECT status, first_half_status, second_half_status
        FROM daily_attendance_status
        WHERE reg_no = %s AND date::date = CURRENT_DATE
        LIMIT 1
    """,
        (reg_no,),
    )
    os_today_das_row    = cursor.fetchone()
    os_today_status     = os_today_das_row[0] if os_today_das_row else None
    os_today_first_half = os_today_das_row[1] if os_today_das_row else None
    os_today_second_half= os_today_das_row[2] if os_today_das_row else None

    cursor.execute(
        """
        SELECT status, COUNT(*) as cnt
        FROM daily_attendance_status
        WHERE reg_no = %s
        GROUP BY status
    """,
        (reg_no,),
    )
    os_hist_breakdown     = {row[0]: row[1] for row in cursor.fetchall()}
    os_hist_full_day      = os_hist_breakdown.get("Present", 0)
    os_hist_half_day      = os_hist_breakdown.get("Half Day", 0)
    os_hist_absent        = os_hist_breakdown.get("Absent", 0)
    os_hist_leave         = os_hist_breakdown.get("Leave", 0)

    return JSONResponse(
        content={
            "stats": {
                "name": staff_user["name"],
                "role": staff_user["role"],
                "dept": staff_user["dept"],
                "total_staff_in_role": role_count,
                "total_attendance": total_attendance,
                "today_attendance": today_attendance,
                "today_face_scan": today_face_scan,
                "today_od": od_count,
                "today_earned": earned_count,
                "today_casual": casual_count,
                # Present = face scan + OD (treated as present)
                "today_present": today_face_scan + od_count,
                "today_holiday_reason": _academic_status_for_date(datetime.now().strftime("%Y-%m-%d"))[1],
                # Absent = 0 if holiday or attendance present, else 1
                "today_absent": 0 if (_academic_status_for_date(datetime.now().strftime("%Y-%m-%d"))[2] or (today_face_scan + od_count) > 0) else 1,
                # Pie chart fields
                "today_status": os_today_status,
                "today_first_half": os_today_first_half,
                "today_second_half": os_today_second_half,
                "hist_full_day_count": os_hist_full_day,
                "hist_half_day_count": os_hist_half_day,
                "hist_absent_count": os_hist_absent,
                "hist_leave_count": os_hist_leave,
            },
            "recent_attendance": recent_attendance_list,
        },
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.post("/other_staff/attendance")
async def other_staff_mark_attendance(request: Request):
    """Mark attendance for other staff"""
    staff_user = verify_other_staff_token(request)

    # Check for VPN before processing attendance
    client_ip = get_client_ip(request)
    if _app_settings.get("enforce_vpn_blocking", True) and is_vpn_ip(client_ip):
        raise HTTPException(
            status_code=403,
            detail="VPN detected. Please turn off your VPN connection to continue.",
        )

    # Safe parsing for request data
    form_data = None
    try:
        form_data = await request.form()
    except:
        pass

    body = None
    try:
        body = await request.json()
    except:
        pass

    platform = request.headers.get("X-Client-Platform", "app").lower()
    lat = request.query_params.get("client_lat")
    lng = request.query_params.get("client_lng")

    if form_data is not None:
        platform = form_data.get("client_platform") or platform
        lat = form_data.get("client_lat") or lat
        lng = form_data.get("client_lng") or lng
    elif body is not None:
        platform = body.get("client_platform") or platform
        lat = body.get("client_lat") or lat
        lng = body.get("client_lng") or lng

    validation_form = {
        "client_platform": platform,
        "client_lat": lat,
        "client_lng": lng,
        "reg_no": staff_user["reg_no"]
    }

    # Geofence first — if active and passes, WiFi check is redundant
    if not _enforce_web_geofence(validation_form, client_ip):
        check_wifi(request)

    cursor.execute("""
        SELECT slot_number, start_time, duration_minutes, is_enabled, slot_type, slot_half
        FROM attendance_duration_settings
        WHERE is_enabled = 1
        ORDER BY slot_number ASC
    """)
    duration_rows = cursor.fetchall()

    active_slot_type = "check_in"
    if duration_rows:
        current_time = datetime.now()
        allowed = False

        for row in duration_rows:
            slot_number = row[0]
            start_time = row[1]
            duration_minutes = row[2]
            slot_type = row[4] if len(row) > 4 and row[4] else "check_in"
            slot_half = row[5] if len(row) > 5 and row[5] else "full_day"
            effective_duration, _, _ = _calculate_effective_slot_duration(slot_type, slot_half, duration_minutes)

            start_hour, start_minute = map(int, start_time.split(":"))
            start_datetime = current_time.replace(
                hour=start_hour, minute=start_minute, second=0, microsecond=0
            )
            end_datetime = start_datetime + timedelta(minutes=effective_duration)

            if start_datetime <= current_time < end_datetime:
                allowed = True
                active_slot_type = slot_type
                break

        if not allowed:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                allowed = True
                active_slot_type = ccl_slot_type
                
        if not allowed:
            slots_info = ", ".join(
                [f"Slot {row[0]}: {row[1]} ({row[2]} min)" for row in duration_rows]
            )
            raise HTTPException(
                status_code=403,
                detail=f"Attendance marking is not allowed at this time. Available slots: {slots_info}",
            )
    else:
        settings = get_ccl_settings_for_date()
        early_enabled = settings["early_check_in_ccl_enabled"]
        late_enabled = settings["late_check_out_ccl_enabled"]
        
        if early_enabled or late_enabled:
            ccl_slot_type = get_active_ccl_slot_type()
            if ccl_slot_type:
                active_slot_type = ccl_slot_type
            else:
                raise HTTPException(
                    status_code=403,
                    detail="Attendance marking is not allowed at this time. (Outside CCL window)",
                )

    if active_slot_type == "check_out":
        cursor.execute(
            "SELECT COUNT(*) FROM other_staff_attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
            (staff_user["reg_no"],)
        )
        if cursor.fetchone()[0] == 0:
            cursor.execute(
                "SELECT COUNT(*) FROM attendance WHERE reg_no = ? AND DATE(timestamp) = CURRENT_DATE AND status = 'check_in'",
                (staff_user["reg_no"],)
            )
            if cursor.fetchone()[0] == 0:
                raise HTTPException(
                    status_code=400,
                    detail="You cannot Check-Out without checking in first for today."
                )

    # Parse request body for extra fields (lat/lng/location)
    if form_data is not None:
        data = dict(form_data)
    else:
        try:
            data = await request.json()
        except Exception:
            data = {}

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        cursor.execute(
            """
            INSERT INTO other_staff_attendance (reg_no, name, dept, role, "timestamp", status)
            VALUES (%s, %s, %s, %s, %s, %s)
        """,
            (
                staff_user["reg_no"],
                staff_user["name"],
                staff_user["dept"],
                staff_user["role"],
                timestamp,
                active_slot_type,
            ),
        )
        conn.commit()

        # Sync with daily_attendance_status table
        current_date = datetime.now().strftime("%Y-%m-%d")
        try:
            cursor.execute(
                """
                SELECT id, status FROM daily_attendance_status 
                WHERE reg_no = %s AND date = %s
            """,
                (staff_user["reg_no"], current_date),
            )
            existing = cursor.fetchone()
            if existing and existing[1] in [
                "Absent",
                "Leave",
                "OD",
                "casual",
                "earned",
                "od",
            ]:
                cursor.execute(
                    """
                    UPDATE daily_attendance_status 
                    SET status = 'Present', leave_type = NULL, leave_request_id = NULL,
                        marked_by = 'Attendance System', marked_at = CURRENT_TIMESTAMP
                    WHERE reg_no = %s AND date = %s
                """,
                    (staff_user["reg_no"], current_date),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO daily_attendance_status 
                    (reg_no, name, dept, date, status, marked_by, marked_at)
                    VALUES (%s, %s, %s, %s, 'Present', 'Attendance System', CURRENT_TIMESTAMP)
                    ON CONFLICT (reg_no, date) DO UPDATE SET
                        status = 'Present',
                        marked_by = 'Attendance System',
                        marked_at = CURRENT_TIMESTAMP
                """,
                    (
                        staff_user["reg_no"],
                        staff_user["name"],
                        staff_user["dept"],
                        current_date,
                    ),
                )
        except Exception as e:
            print(
                f"Error syncing daily_attendance_status for {staff_user['reg_no']}: {e}"
            )

        log_audit_event(
            "OTHER_STAFF_ATTENDANCE",
            staff_user["reg_no"],
            True,
            f"Attendance marked for {staff_user['role']}",
        )

        # Update user location when attendance is marked successfully
        try:
            lat_raw = data.get("latitude") or data.get("client_lat")
            lng_raw = data.get("longitude") or data.get("client_lng")
            accuracy_raw = data.get("accuracy") or data.get("client_accuracy")

            if lat_raw is not None and lng_raw is not None:
                latitude = float(lat_raw)
                longitude = float(lng_raw)
                accuracy = float(accuracy_raw) if accuracy_raw is not None else None

                # Insert into user_latest_locations
                cursor.execute(
                    """
                    INSERT INTO user_latest_locations
                    (reg_no, username, name, dept, role, latitude, longitude, accuracy_meters, 
                     source, app_state, captured_at, last_seen_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT (reg_no) DO UPDATE SET
                        username = EXCLUDED.username,
                        name = EXCLUDED.name,
                        dept = EXCLUDED.dept,
                        role = EXCLUDED.role,
                        latitude = EXCLUDED.latitude,
                        longitude = EXCLUDED.longitude,
                        accuracy_meters = EXCLUDED.accuracy_meters,
                        source = EXCLUDED.source,
                        app_state = EXCLUDED.app_state,
                        captured_at = EXCLUDED.captured_at,
                        last_seen_at = CURRENT_TIMESTAMP
                    """,
                    (
                        staff_user["reg_no"],
                        staff_user["username"],
                        staff_user["name"],
                        staff_user["dept"],
                        staff_user["role"],
                        latitude,
                        longitude,
                        accuracy,
                        "attendance_mark",
                        "foreground",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )
                print(
                    f"Location updated via other staff attendance for {staff_user['reg_no']}"
                )
        except Exception as e:
            print(f"Error updating location during other staff attendance mark: {e}")

        return {
            "message": "Attendance marked successfully",
            "reg_no": staff_user["reg_no"],
            "name": staff_user["name"],
            "role": staff_user["role"],
            "timestamp": timestamp,
        }
    except Exception as e:
        print(f"Error marking attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to mark attendance")


@app.get("/other_staff/attendance")
async def other_staff_get_attendance(
    request: Request, date: str = None, start_date: str = None, end_date: str = None
):
    """Get attendance records for other staff - includes face scan + OD from daily_attendance_status"""
    staff_user = verify_other_staff_token(request)
    reg_no = staff_user["reg_no"]

    try:
        # Default to academic year if no dates provided
        now = datetime.now()
        acad_start, acad_end = _get_academic_date_range()
        if not start_date:
            start_date = acad_start or now.strftime("%Y-%m-01")
        if not end_date:
            end_date = acad_end or now.strftime("%Y-%m-%d")
        # Clamp to academic year
        start_date, end_date = _clamp_to_academic_year(start_date, end_date)
        # If a specific date isn't provided, use the range (default current month)

        # Get face-scan attendance records
        if date:
            cursor.execute(
                """
                SELECT id, reg_no, name, dept, role, timestamp, status
                FROM other_staff_attendance 
                WHERE reg_no = ? AND timestamp::date = ? 
                ORDER BY id DESC
            """,
                (reg_no, date),
            )
        else:
            cursor.execute(
                """
                SELECT id, reg_no, name, dept, role, timestamp, status
                FROM other_staff_attendance 
                WHERE reg_no = ? AND timestamp::date >= ? AND timestamp::date <= ?
                ORDER BY id DESC
            """,
                (reg_no, start_date, end_date),
            )

        rows = cursor.fetchall()

        # Collect scan dates
        scan_dates = set()
        for row in rows:
            ts = row[5]
            if ts:
                if hasattr(ts, "strftime"):
                    scan_dates.add(ts.strftime("%Y-%m-%d"))
                else:
                    ts_str = str(ts)
                    if "T" in ts_str:
                        scan_dates.add(ts_str.split("T")[0])
                    elif " " in ts_str:
                        scan_dates.add(ts_str.split(" ")[0])

        # Get daily_attendance_status records (OD, Leave, etc.)
        if date:
            cursor.execute(
                """
                SELECT date, status, leave_type, absent_reason
                FROM daily_attendance_status
                WHERE reg_no = %s AND date::date = %s
            """,
                (reg_no, date),
            )
        else:
            cursor.execute(
                """
                SELECT date, status, leave_type, absent_reason
                FROM daily_attendance_status
                WHERE reg_no = %s AND date::date >= %s AND date::date <= %s
            """,
                (reg_no, start_date, end_date),
            )

        status_rows = cursor.fetchall()

        # Categorize status dates
        od_dates = set()
        leave_dates = set()
        earned_dates = set()
        casual_dates = set()
        absent_dates = {}

        for row in status_rows:
            date_str = str(row[0]) if row[0] else None
            status = row[1]
            leave_type = row[2]
            absent_reason = row[3] if len(row) > 3 else None

            if (
                date_str
                and status == "Present"
                and leave_type in ("od", "earned", "casual")
            ):
                if leave_type == "od":
                    od_dates.add(date_str)
                elif leave_type == "earned":
                    earned_dates.add(date_str)
                elif leave_type == "casual":
                    casual_dates.add(date_str)
            elif date_str and status == "Leave":
                leave_dates.add(date_str)
            elif date_str and status == "Absent":
                absent_dates[date_str] = absent_reason

        # Build attendance list including status records
        attendance_list = [
            {
                "id": row[0],
                "reg_no": row[1],
                "name": row[2],
                "dept": row[3],
                "role": row[4],
                "timestamp": _ts(row[5]),
                "status": row[6] if len(row) > 6 and row[6] else "check_in",
                "source": "face_scan",
                "absent_reason": None,
            }
            for row in rows
        ]

        # Add status records for dates not in face scans
        all_status_dates = od_dates | earned_dates | casual_dates | leave_dates | set(absent_dates.keys())
        for date_str in all_status_dates:
            if date_str not in scan_dates:
                # Find the status for this date
                for srow in status_rows:
                    if str(srow[0]) == date_str:
                        status = srow[1]
                        leave_type = srow[2]
                        absent_reason = srow[3] if len(srow) > 3 else None
                        break
                else:
                    status = "Unknown"
                    leave_type = None
                    absent_reason = None

                attendance_list.append(
                    {
                        "id": None,
                        "reg_no": reg_no,
                        "name": staff_user["name"],
                        "dept": staff_user["dept"],
                        "role": staff_user["role"],
                        "timestamp": date_str,
                        "source": "leave" if status == "Leave" else ("absent" if status == "Absent" else "od"),
                        "status": status,
                        "leave_type": leave_type,
                        "absent_reason": absent_reason,
                    }
                )

        # Calculate stats (Half Day counts as 0.5, Present/OD counts as 1.0)
        hd = _get_half_day_settings()
        present_days = 0.0
        if hd["half_day_enabled"]:
            # Query status values directly from daily_attendance_status
            cursor.execute(
                """
                SELECT date, status FROM daily_attendance_status
                WHERE reg_no = %s AND date >= %s AND date <= %s
            """,
                (reg_no, start_date, end_date),
            )
            for r in cursor.fetchall():
                status_val = r[1]
                if status_val == "Present":
                    present_days += 1.0
                elif status_val == "Half Day":
                    present_days += 0.5
        else:
            full_att_dates = _get_full_attendance_dates(reg_no, start_date, end_date, is_other_staff=True)
            present_days = float(len(full_att_dates | od_dates | earned_dates | casual_dates))

        if date:
            today = datetime.now().strftime("%Y-%m-%d")
            days_in_range = 1 if date <= today else 0
        else:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = _cap_end_to_today(datetime.strptime(end_date, "%Y-%m-%d"))
            holiday_dates = _get_holiday_dates_in_range(start_date, end_dt.strftime("%Y-%m-%d"))
            # Other staff panels expect calendar days (not working days)
            days_in_range = max(0, (end_dt - start_dt).days + 1 - len(holiday_dates))

        absent_days = max(0, days_in_range - present_days - len(leave_dates))

        return {
            "attendance": attendance_list,
            "count": len(attendance_list),
            "present_days": present_days,
            "absent_days": absent_days,
            "leave_days": len(leave_dates),
            "od_days": len(od_dates),
            "earned_days": len(earned_dates),
            "casual_days": len(casual_dates),
        }
    except Exception as e:
        print(f"Error fetching attendance: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch attendance")


@app.post("/other_staff/face/register")
async def other_staff_register_face(
    request: Request,
    name: str = Form(...),
    reg_no: str = Form(...),
    dept: str = Form(...),
    role: str = Form(...),
    image: UploadFile = File(...),
):
    """
    Other staff can register their own face (principal, placement_staff, lab_technician, system_admin, office_staff)
    SECURITY: This endpoint verifies the staff details match the logged-in user
    """
    # First try to verify as other_staff, then fall back to general user
    logged_in_staff = None
    is_other_staff = True

    try:
        logged_in_staff = verify_other_staff_token(request)
    except:
        # Try general user token if other_staff verification fails
        try:
            logged_in_staff = verify_user_token(request)
            is_other_staff = False
        except Exception as e:
            raise HTTPException(status_code=401, detail="Invalid token")

    # CRITICAL: Verify the logged-in staff's reg_no matches the submitted reg_no
    if logged_in_staff["reg_no"] != reg_no:
        log_audit_event(
            "FACE_REG_MISMATCH",
            reg_no,
            False,
            f"Logged in as {logged_in_staff['reg_no']} but tried to register {reg_no}",
        )
        raise HTTPException(
            status_code=403,
            detail="You can only register your own face. Please use your logged-in credentials.",
        )

    # Verify staff exists in database with matching details
    if is_other_staff:
        cursor.execute(
            "SELECT dept, role, embedding, name FROM other_staff WHERE reg_no = ?",
            (reg_no,),
        )
    else:
        cursor.execute(
            "SELECT dept, role, embedding, name FROM users WHERE reg_no = ?", (reg_no,)
        )

    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(status_code=404, detail="Staff not found")

    db_dept = (existing[0] or "").strip()
    db_role = (existing[1] or "").strip().lower()
    db_name = existing[3]
    submitted_dept = (dept or "").strip()

    # Ensure principal always has a department; backfill legacy blank values.
    if not db_dept:
        default_dept = get_default_department_for_role(db_role)
        if default_dept:
            if is_other_staff:
                cursor.execute(
                    "UPDATE other_staff SET dept = ? WHERE reg_no = ?",
                    (default_dept, reg_no),
                )
            else:
                cursor.execute(
                    "UPDATE users SET dept = ? WHERE reg_no = ?", (default_dept, reg_no)
                )
            conn.commit()
            db_dept = default_dept

    if not submitted_dept:
        submitted_dept = db_dept

    # CRITICAL: Verify the department matches what's in the database
    if db_dept != submitted_dept:
        raise HTTPException(
            status_code=403,
            detail=f"Department mismatch. Your department is {db_dept}.",
        )

    # CRITICAL: Verify the name matches what's in the database
    if db_name != name:
        raise HTTPException(
            status_code=403,
            detail=f"Name mismatch. Please use your registered name: {db_name}",
        )

    current_embedding = existing[2]

    # Check if already registered
    if current_embedding is not None:
        # Already registered - check if they have permission to re-register
        if is_other_staff:
            cursor.execute(
                "SELECT can_reregister FROM other_staff WHERE reg_no = ?", (reg_no,)
            )
        else:
            cursor.execute(
                "SELECT can_reregister FROM users WHERE reg_no = ?", (reg_no,)
            )
        perm = cursor.fetchone()

        if perm is None or perm[0] != 1:
            raise HTTPException(
                status_code=403,
                detail="Face already registered. Please contact admin for permission to re-register.",
            )

        # Clear the permission after use
        if is_other_staff:
            cursor.execute(
                "UPDATE other_staff SET can_reregister = 0 WHERE reg_no = ?", (reg_no,)
            )
        else:
            cursor.execute(
                "UPDATE users SET can_reregister = 0 WHERE reg_no = ?", (reg_no,)
            )
        conn.commit()

    img_bytes = await image.read()
    img = preprocess_image_data(img_bytes)

    face = extract_face(img)
    if face is None:
        save_debug_image(img, "register_fail")
        raise HTTPException(
            status_code=400,
            detail="Face not detected clearly. Ensure good lighting and face is visible.",
        )

    embedding = face.embedding.astype(np.float32)

    # Debug: Print embedding statistics
    print(f"DEBUG: Registering face for other_staff {reg_no}")
    print(f"  Embedding shape: {embedding.shape}")
    print(f"  Embedding norm: {np.linalg.norm(embedding):.4f}")
    print(f"  Embedding mean: {np.mean(embedding):.4f}")

    try:
        if is_other_staff:
            cursor.execute(
                """
                UPDATE other_staff SET embedding = ? WHERE reg_no = ?
            """,
                (embedding.tobytes(), reg_no),
            )
        else:
            cursor.execute(
                """
                UPDATE users SET embedding = ? WHERE reg_no = ?
            """,
                (embedding.tobytes(), reg_no),
            )
        conn.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to register: {str(e)}")
    _cache_update_primary(reg_no, embedding)

    _save_face_embedding_sample(
        reg_no,
        "other_staff" if is_other_staff else "users",
        embedding,
        "registration_other_staff_self",
        1.0,
    )

    log_audit_event(
        "FACE_REGISTERED_OTHER_STAFF", reg_no, True, f"{role} registered their face"
    )

    return {
        "message": "Face registered successfully",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
        "role": role,
        "bbox": [float(x) for x in face.bbox],
    }


@app.get("/other_staff/face/status")
async def other_staff_check_face_status(request: Request):
    """Check if other staff has face registered and their permission status"""
    # Try other_staff first, then fall back to general user
    try:
        staff_user = verify_other_staff_token(request)
        is_other_staff = True
    except:
        try:
            staff_user = verify_user_token(request)
            is_other_staff = False
        except Exception as e:
            raise HTTPException(status_code=401, detail="Invalid token")

    reg_no = staff_user["reg_no"]

    # Check the appropriate table
    if is_other_staff:
        cursor.execute(
            "SELECT name, dept, role, embedding, can_reregister FROM other_staff WHERE reg_no = ?",
            (reg_no,),
        )
    else:
        cursor.execute(
            "SELECT name, dept, role, embedding, can_reregister FROM users WHERE reg_no = ?",
            (reg_no,),
        )

    user = cursor.fetchone()

    if user:
        cursor.execute(
            """
            SELECT id, status FROM face_reregister_requests 
            WHERE staff_reg_no = ? AND status = 'pending'
            ORDER BY id DESC LIMIT 1
        """,
            (reg_no,),
        )
        pending_request = cursor.fetchone()

        return {
            "reg_no": reg_no,
            "name": user[0],
            "dept": user[1],
            "role": user[2],
            "face_registered": user[3] is not None,
            "can_reregister": user[4] == 1 if len(user) > 4 else False,
            "has_pending_request": pending_request is not None,
        }

    raise HTTPException(status_code=404, detail="User not found")


# -------------------------------------------------
# OTHER STAFF FACE RE-REGISTRATION REQUESTS
# -------------------------------------------------
@app.post("/other_staff/face/reregister/request")
async def other_staff_request_reregister(request: Request):
    """Other staff submits a request to re-register their face (goes directly to admin)"""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing authorization")

    token = auth_header.replace("Bearer ", "")

    try:
        import base64

        decoded = base64.b64decode(token).decode("utf-8")
        username, password = decoded.split(":", 1)
    except:
        raise HTTPException(status_code=401, detail="Invalid token format")

    user = get_user_by_username(username)
    is_other_staff = False
    if user:
        role = user[6]
        if role not in [
            "principal",
            "placement_staff",
            "lab_technician",
            "system_admin",
            "office_staff",
        ]:
            raise HTTPException(
                status_code=403, detail="Only other staff can request re-registration"
            )
        reg_no = user[3]
        name = user[4]
        dept = user[5]
    else:
        cursor.execute("SELECT * FROM other_staff WHERE username = ?", (username,))
        os_user = cursor.fetchone()
        if not os_user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        reg_no = os_user[3]
        name = os_user[4]
        dept = os_user[7] if len(os_user) > 7 else ""
        is_other_staff = True

    cursor.execute(
        """
        SELECT id FROM face_reregister_requests 
        WHERE staff_reg_no = ? AND status = 'pending'
    """,
        (reg_no,),
    )
    existing_request = cursor.fetchone()

    if existing_request:
        raise HTTPException(
            status_code=400, detail="You already have a pending request"
        )

    if is_other_staff:
        cursor.execute("SELECT embedding FROM other_staff WHERE reg_no = ?", (reg_no,))
    else:
        cursor.execute("SELECT embedding FROM users WHERE reg_no = ?", (reg_no,))
    embedding = cursor.fetchone()

    if not embedding or embedding[0] is None:
        raise HTTPException(
            status_code=400,
            detail="Your face is not registered yet. Please register first.",
        )

    cursor.execute(
        """
        INSERT INTO face_reregister_requests (staff_reg_no, staff_name, dept, status, hod_approved)
        VALUES (?, ?, ?, 'pending', 1)
    """,
        (reg_no, name, dept),
    )
    conn.commit()

    log_audit_event(
        "REREGISTER_REQUEST_SUBMITTED",
        reg_no,
        True,
        f"Other staff {name} submitted face re-registration request",
    )

    return {
        "message": "Request submitted successfully. Please wait for Admin approval.",
        "reg_no": reg_no,
        "name": name,
        "dept": dept,
    }


@app.get("/admin/other_staff/face/reregister/requests")
async def admin_get_other_staff_reregister_requests(request: Request):
    """Admin gets all pending re-registration requests from other staff"""
    verify_admin_token(request)

    cursor.execute("""
        SELECT id, staff_reg_no, staff_name, dept, request_date, status, hod_approved, admin_approved
        FROM face_reregister_requests
        WHERE status = 'pending'
        ORDER BY request_date DESC
    """)

    all_requests = cursor.fetchall()
    other_staff_requests = []

    for r in all_requests:
        reg_no = r[1]
        cursor.execute("SELECT id FROM other_staff WHERE reg_no = ?", (reg_no,))
        if cursor.fetchone():
            other_staff_requests.append(
                {
                    "id": r[0],
                    "staff_reg_no": r[1],
                    "staff_name": r[2],
                    "dept": r[3],
                    "request_date": r[4],
                    "status": r[5],
                    "hod_approved": bool(r[6]) if r[6] else False,
                    "admin_approved": bool(r[7]) if r[7] else False,
                    "user_type": "other_staff",
                }
            )

    return {"requests": other_staff_requests}


@app.post("/admin/other_staff/face/reregister/approve/{staff_reg_no}")
async def admin_approve_other_staff_reregister(request: Request, staff_reg_no: str):
    """Admin approves an other staff re-registration request"""
    verify_admin_token(request)

    cursor.execute(
        "SELECT id FROM face_reregister_requests WHERE staff_reg_no = ? AND status = 'pending'",
        (staff_reg_no,),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Request not found or already processed"
        )

    cursor.execute(
        "UPDATE face_reregister_requests SET admin_approved = 1, status = 'approved', processed_by = 'admin', processed_date = CURRENT_TIMESTAMP WHERE staff_reg_no = ?",
        (staff_reg_no,),
    )

    cursor.execute("SELECT id FROM other_staff WHERE reg_no = ?", (staff_reg_no,))
    if cursor.fetchone():
        cursor.execute(
            "UPDATE other_staff SET can_reregister = 1 WHERE reg_no = ?",
            (staff_reg_no,),
        )
    else:
        cursor.execute(
            "UPDATE users SET can_reregister = 1 WHERE reg_no = ?", (staff_reg_no,)
        )

    conn.commit()

    log_audit_event(
        "REREGISTER_APPROVED",
        staff_reg_no,
        True,
        "Admin approved other staff re-registration",
    )

    return {
        "message": "Re-registration approved. Staff can now re-register their face.",
        "staff_reg_no": staff_reg_no,
    }


@app.post("/admin/other_staff/face/reregister/deny/{staff_reg_no}")
async def admin_deny_other_staff_reregister(
    request: Request, staff_reg_no: str, reason: str = "Denied by Admin"
):
    """Admin denies an other staff re-registration request"""
    verify_admin_token(request)

    cursor.execute(
        "SELECT id FROM face_reregister_requests WHERE staff_reg_no = ? AND status = 'pending'",
        (staff_reg_no,),
    )
    existing = cursor.fetchone()

    if not existing:
        raise HTTPException(
            status_code=404, detail="Request not found or already processed"
        )

    cursor.execute(
        """
        UPDATE face_reregister_requests 
        SET status = 'denied', processed_by = 'admin', processed_date = CURRENT_TIMESTAMP, reason = ?
        WHERE staff_reg_no = ?
    """,
        (reason, staff_reg_no),
    )
    conn.commit()

    log_audit_event(
        "REREGISTER_DENIED",
        staff_reg_no,
        True,
        f"Admin denied other staff - reason: {reason}",
    )

    return {"message": "Request denied", "staff_reg_no": staff_reg_no, "reason": reason}


# -------------------------------------------------
# SERVER STARTUP
# -------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    import socket

    # Get the local IP address of the computer
    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "YOUR_IP_HERE"

    local_ip = get_local_ip()

    # Run the server on all network interfaces (0.0.0.0) instead of localhost
    # This allows the app to connect from any device on ANY network
    print("=" * 60)
    print("SERVER STARTING...")
    print("=" * 60)
    print(f"Server running on: http://0.0.0.0:8001")
    print(f"Local access:       http://localhost:8001")
    print(f"Network access:    http://$local_ip:8001")
    print("=" * 60)
    print("TO CONNECT FROM APP:")
    print("1. Find your computer's IP using 'ipconfig' in CMD")
    print("2. Update customServerURL in college_ip_config.dart")
    print("3. Example: http://$local_ip:8001")
    print("=" * 60)
    # NOTE: On Windows, --workers flag is unstable (OSError socket issue).
    # Use the thread pool executors (20 CPU + 30 DB threads) for concurrency instead.
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8001,
        loop="asyncio",
        limit_concurrency=100,
        timeout_keep_alive=30,
    )
